import sys
import openpyxl

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'C:/Users/DELL/Desktop/Education-DB-Analytic/docs/PTIT_Chiso.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['KS25_Python_Web']

for r in range(5, 17):
    c1 = sheet.cell(row=r, column=1).value
    c2 = sheet.cell(row=r, column=2).value
    c3 = sheet.cell(row=r, column=3).value
    c4 = sheet.cell(row=r, column=4).value # Day CC dau tien
    print(f"Row {r}: Col1={c1}, Col2={c2}, Col3={c3}, Col4={c4}")
