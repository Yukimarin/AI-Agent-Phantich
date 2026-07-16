import mysql.connector
import openpyxl
import os
import json
import numpy as np
from datetime import datetime, date
from collections import defaultdict

# Setup normalization
def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip().upper()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    name_str = name_str.replace('KS24', 'K24').replace('KS25', 'K25')
    name_str = name_str.lower()
    for word in ['hk2', 'hk1', 'hl', 'cu', 'retake', 'old']:
        name_str = name_str.replace(word, '')
    import re
    name_str = re.sub(r'[^a-z0-9]', '', name_str)
    return name_str

def get_excel_chot_data(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    class_data = defaultdict(dict)
    
    for sheetname in wb.sheetnames:
        if sheetname == 'Sheet1' or 'SKL' in sheetname:
            continue
        sheet = wb[sheetname]
        max_r = sheet.max_row
        max_c = sheet.max_column
        if max_r < 5:
            continue
            
        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
        
        dates_list = []
        for c_idx in range(3, max_c):
            val3 = row3[c_idx]
            val4 = row4[c_idx]
            if val3:
                # Parse date
                d_str = str(val3).strip()
                parts = d_str.split('/')
                try:
                    if len(parts) == 2:
                        d = date(2026, int(parts[1]), int(parts[0]))
                    elif len(parts) == 3:
                        y = int(parts[2])
                        if y < 100: y += 2000
                        d = date(y, int(parts[1]), int(parts[0]))
                    else:
                        d = None
                except:
                    d = None
                if d:
                    dates_list.append((c_idx, d, val4))
            elif dates_list:
                dates_list.append((c_idx, dates_list[-1][1], val4))
                
        if not dates_list:
            continue
            
        unique_dates = sorted(list(set(d for idx, d, sub in dates_list)))
        last_date = unique_dates[-1] if unique_dates else None
        last_cols = [idx for idx, d, sub in dates_list if d == last_date] if last_date else []
        
        rp_col_idx = None
        for c_idx in range(max_c - 1, 2, -1):
            vals = []
            for r in range(5, max_r + 1):
                val = sheet.cell(row=r, column=c_idx + 1).value
                if val is not None:
                    try:
                        vals.append(float(val))
                    except:
                        pass
            if len(vals) >= 2:
                avg_val = np.mean(vals)
                if 30.0 <= avg_val <= 115.0:
                    h3 = row3[c_idx]
                    if not h3:
                        rp_col_idx = c_idx
                        break
        
        for r in range(5, max_r + 1):
            cname = sheet.cell(row=r, column=2).value
            teacher = sheet.cell(row=r, column=3).value
            if cname:
                norm_name = normalize_class_name(cname)
                cc_val, bt_val, el_val = 0.0, 0.0, 0.0
                for c_idx in last_cols:
                    sub = row4[c_idx]
                    val = sheet.cell(row=r, column=c_idx + 1).value
                    if val is not None:
                        try:
                            val_f = float(val)
                        except:
                            val_f = 0.0
                        if sub == 'Chuyên cần': cc_val = val_f
                        elif sub == 'Bài tập': bt_val = val_f
                        elif sub == 'Elearning': el_val = val_f
                
                rp_val = None
                if rp_col_idx is not None:
                    cell_val = sheet.cell(row=r, column=rp_col_idx + 1).value
                    if cell_val is not None:
                        try:
                            rp_val = float(cell_val)
                        except:
                            pass
                            
                class_data[norm_name][sheetname] = {
                    'cc': cc_val,
                    'bt': bt_val,
                    'el': el_val,
                    'rp': rp_val,
                    'teacher': str(teacher).strip() if teacher else "Ẩn danh",
                    'date': last_date
                }
    return class_data

def run_query(cursor, query, params=None):
    cursor.execute(query, params or ())
    columns = [desc[0] for desc in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def main():
    conn = mysql.connector.connect(
        host="localhost",
        port=3306,
        user="root",
        password="",
        database="qldt_el"
    )
    cursor = conn.cursor()
    
    excel_path = r"C:\Users\DELL\Desktop\Backup\PTIT\PTIT_Chiso.xlsx"
    excel_data = get_excel_chot_data(excel_path)
    
    # Define Target Batches and Classes
    ks24_classes = [
        {'id': 48, 'name': 'HN-KS24-CNTT1', 'sheet': 'KS24_AI', 'prev_course_id': 194}, # JWS
        {'id': 49, 'name': 'HN-KS24-CNTT2', 'sheet': 'KS24_AI', 'prev_course_id': 194},
        {'id': 156, 'name': 'HN-KS24-CNTT3', 'sheet': 'KS24_AI', 'prev_course_id': 194},
        {'id': 51, 'name': 'HN-KS24-CNTT4', 'sheet': 'KS24_AI', 'prev_course_id': 194},
        {'id': 63, 'name': 'HCM-KS24-CNTT1', 'sheet': 'KS24_AI', 'prev_course_id': 194},
        {'id': 64, 'name': 'HCM-KS24-CNTT2', 'sheet': 'KS24_AI', 'prev_course_id': 162}, # Java Adv
    ]
    
    ks25_classes = [
        {'id': 77, 'name': 'HN-KS25-CNTT1', 'sheet': 'KS25_Python_Web', 'course_id': 193, 'prev_course_id': 124}, # JS
        {'id': 76, 'name': 'HN-KS25-CNTT2', 'sheet': 'KS25_Python_Web', 'course_id': 193, 'prev_course_id': 124},
        {'id': 75, 'name': 'HN-KS25-CNTT3', 'sheet': 'KS25_Python_Web', 'course_id': 193, 'prev_course_id': 124},
        {'id': 74, 'name': 'HN-KS25-CNTT4', 'sheet': 'KS25_Python_Web', 'course_id': 193, 'prev_course_id': 124},
        {'id': 73, 'name': 'HN-KS25-CNTT5', 'sheet': 'KS25_Python_Web', 'course_id': 193, 'prev_course_id': 124},
        {'id': 72, 'name': 'HN-KS25-CNTT6', 'sheet': 'KS25_Python_Web', 'course_id': 193, 'prev_course_id': 124},
    ]
    
    qtkd_classes = [
        {'id': 84, 'name': 'HN-K25-QTKD1', 'sheet': 'KS25_QTKD_PRJ302', 'course_id': 178, 'prev_course_id': 188}, # DTB201
        {'id': 83, 'name': 'HN-K25-QTKD2', 'sheet': 'KS25_QTKD_PRJ302', 'course_id': 178, 'prev_course_id': 188},
        {'id': 82, 'name': 'HN-K25-QTKD3', 'sheet': 'KS25_QTKD_PRJ302', 'course_id': 178, 'prev_course_id': 188},
    ]
    
    all_target_classes = ks24_classes + ks25_classes + qtkd_classes
    
    student_pass_history = {}
    pass_results = run_query(cursor, "SELECT class_id, course_id, student_id, pass, homework, elearning, attendance, hackathon_1, hackathon_2, rpoints, project FROM qldt_el.final_results WHERE pass IS NOT NULL;")
    for r in pass_results:
        sid = int(r['student_id'])
        co_id = int(r['course_id'])
        # Convert Decimals
        student_pass_history[(sid, co_id)] = {
            'homework': float(r['homework']) if r['homework'] is not None else None,
            'elearning': float(r['elearning']) if r['elearning'] is not None else None,
            'attendance': float(r['attendance']) if r['attendance'] is not None else None,
            'hackathon_1': float(r['hackathon_1']) if r['hackathon_1'] is not None else None,
            'hackathon_2': float(r['hackathon_2']) if r['hackathon_2'] is not None else None,
            'rpoints': float(r['rpoints']) if r['rpoints'] is not None else None,
            'project': float(r['project']) if r['project'] is not None else None,
            'pass': int(r['pass'])
        }
        
    risk_results = {}
    
    for c_info in all_target_classes:
        cid = c_info['id']
        cname = c_info['name']
        sheetname = c_info['sheet']
        prev_co_id = c_info['prev_course_id']
        
        norm_cname = normalize_class_name(cname)
        excel_disc = excel_data.get(norm_cname, {}).get(sheetname)
        if cid == 156:
            # Gộp lớp 156 từ 156, 50, 52
            names_to_try = ['hnk24cntt3', 'hnk24cntt5', 'hnk24cntt3cu']
            cc_list, bt_list, el_list, rp_list = [], [], [], []
            teachers = []
            for n in names_to_try:
                disc = excel_data.get(n, {}).get(sheetname)
                if disc:
                    cc_list.append(disc['cc'])
                    bt_list.append(disc['bt'])
                    el_list.append(disc['el'])
                    if disc['rp'] is not None: rp_list.append(disc['rp'])
                    teachers.append(disc['teacher'])
            if cc_list:
                excel_disc = {
                    'cc': np.mean(cc_list),
                    'bt': np.mean(bt_list),
                    'el': np.mean(el_list),
                    'rp': np.mean(rp_list) if rp_list else None,
                    'teacher': " / ".join(list(set(teachers)))
                }
        
        # Determine current course
        is_ks24 = c_info in ks24_classes
        is_ks25 = c_info in ks25_classes
        is_qtkd = c_info in qtkd_classes
        
        # Query students list
        if is_ks24:
            students_db = run_query(cursor, """
                SELECT f.student_id, s.full_name as student_name, s.student_code,
                       f.attendance, f.homework, f.elearning, f.hackathon_1, f.hackathon_2, f.rpoints, f.project
                FROM qldt_el.final_results f
                JOIN qldt_el.students s ON f.student_id = s.id
                WHERE f.class_id = %s AND f.course_id = %s AND s.status = 'ĐANG HỌC';
            """, (cid, prev_co_id))
            if cid == 156 and not students_db:
                students_db = run_query(cursor, """
                    SELECT f.student_id, s.full_name as student_name, s.student_code,
                           f.attendance, f.homework, f.elearning, f.hackathon_1, f.hackathon_2, f.rpoints, f.project
                    FROM qldt_el.final_results f
                    JOIN qldt_el.students s ON f.student_id = s.id
                    WHERE f.class_id IN (156, 50, 52) AND f.course_id = %s AND s.status = 'ĐANG HỌC';
                """, (prev_co_id,))
        else:
            curr_co_id = c_info['course_id']
            students_db = run_query(cursor, """
                SELECT f.student_id, s.full_name as student_name, s.student_code,
                       f.attendance, f.homework, f.elearning, f.hackathon_1, f.hackathon_2, f.rpoints, f.project
                FROM qldt_el.final_results f
                JOIN qldt_el.students s ON f.student_id = s.id
                WHERE f.class_id = %s AND f.course_id = %s AND s.status = 'ĐANG HỌC';
            """, (cid, curr_co_id))
            
        if not students_db:
            students_db = run_query(cursor, """
                SELECT sc.student_id, s.full_name as student_name, s.student_code,
                       0.0 as attendance, 100.0 as homework, 0.0 as elearning, NULL as hackathon_1, NULL as hackathon_2, 100.0 as rpoints, NULL as project
                FROM qldt_el.student_class sc
                JOIN qldt_el.students s ON sc.student_id = s.id
                WHERE sc.class_id = %s AND s.status = 'ĐANG HỌC';
            """, (cid,))
            
        if not students_db:
            continue
            
        # Calibrate students (and force float types to prevent Decimal errors)
        att_list = [float(s['attendance']) if s['attendance'] is not None else 0.0 for s in students_db]
        hw_list = [float(s['homework']) if s['homework'] is not None else 100.0 for s in students_db]
        el_list = [float(s['elearning']) if s['elearning'] is not None else 0.0 for s in students_db]
        rp_list = [float(s['rpoints']) if s['rpoints'] is not None else 100.0 for s in students_db]
        
        db_att_avg = np.mean(att_list) if att_list else 0.0
        db_hw_avg = np.mean(hw_list) if hw_list else 100.0
        db_el_avg = np.mean(el_list) if el_list else 0.0
        db_rp_avg = np.mean(rp_list) if rp_list else 100.0
        
        excel_cc = float(excel_disc['cc']) if excel_disc and excel_disc['cc'] is not None else db_att_avg
        excel_bt_err = float(excel_disc['bt']) if excel_disc and excel_disc['bt'] is not None else (100.0 - db_hw_avg)
        excel_hw = 100.0 - excel_bt_err
        excel_el = float(excel_disc['el']) if excel_disc and excel_disc['el'] is not None else db_el_avg
        excel_rp = float(excel_disc['rp']) if excel_disc and excel_disc['rp'] is not None else max(0.0, 100.0 - excel_cc - excel_bt_err - excel_el)
        
        calibrated = []
        for i, s in enumerate(students_db):
            # Shift calibration
            rp_cal = rp_list[i] + (excel_rp - db_rp_avg) if db_rp_avg > 0 else excel_rp
            rp_cal = min(120.0, max(0.0, rp_cal))
            
            att_cal = att_list[i] + (excel_cc - db_att_avg) if db_att_avg > 0 else excel_cc
            att_cal = min(100.0, max(0.0, att_cal))
            
            hw_cal = hw_list[i] + (excel_hw - db_hw_avg) if db_hw_avg > 0 else excel_hw
            hw_cal = min(100.0, max(0.0, hw_cal))
            
            el_cal = 0.0 if excel_disc and excel_disc.get('el') == 0.0 else el_list[i]
            
            calibrated.append({
                'student_id': s['student_id'],
                'student_name': s['student_name'],
                'student_code': s['student_code'],
                'attendance': att_cal,
                'homework': hw_cal,
                'elearning': el_cal,
                'rpoints': rp_cal,
                'hackathon_1': float(s['hackathon_1']) if s['hackathon_1'] is not None else None,
                'hackathon_2': float(s['hackathon_2']) if s['hackathon_2'] is not None else None,
                'project': float(s['project']) if s['project'] is not None else None
            })
            
        # Predict failure risk and classify reasons
        class_risk_list = []
        for s in calibrated:
            sid = s['student_id']
            att_val = s['attendance']
            hw_val = s['homework']
            el_val = s['elearning']
            rp_val = s['rpoints']
            
            # Prereq GPA and Hackathon history
            prev_record = student_pass_history.get((sid, prev_co_id))
            if prev_record:
                prev_grades = [v for v in [prev_record['homework'], prev_record['rpoints']] if v is not None]
                prev_gpa = np.mean(prev_grades) if prev_grades else 75.0
            else:
                prev_gpa = 75.0
                
            # Hackathon history
            h1 = s['hackathon_1'] if s['hackathon_1'] is not None else (prev_record['hackathon_1'] if prev_record else None)
            h2 = s['hackathon_2'] if s['hackathon_2'] is not None else (prev_record['hackathon_2'] if prev_record else None)
            h_vals = [v for v in [h1, h2] if v is not None]
            s_hack = np.mean(h_vals) if h_vals else 65.0
            p_hack = min(100.0, s_hack * 1.25)
            
            # Prediction eligible score
            if not is_qtkd:
                p_eligible = 0.42 * prev_gpa + 0.58 * p_hack
            else:
                p_eligible = prev_gpa
            p_eligible = min(100.0, max(0.0, p_eligible))
            
            # Criteria checks
            reasons = []
            is_failed = False
            
            # 1. Rpoints
            if rp_val < 80.0:
                is_failed = True
                reasons.append(f"Rpoint th\u1ea5p ({rp_val:.1f}/80)")
            # 2. Attendance (vắng > 20%)
            if att_val > 20.0:
                is_failed = True
                reasons.append(f"V\u1eafng h\u1ecdc nhi\u1ec1u ({att_val:.1f}%)")
            # 3. Homework (hoàn thành < 80%)
            if hw_val < 80.0:
                is_failed = True
                reasons.append(f"N\u1ee3 b\u00e0i t\u1eadp ({hw_val:.1f}%)")
            # 4. Elearning
            if el_val > 3.0:
                is_failed = True
                reasons.append(f"Vi ph\u1ea1m Elearning ({el_val:.0f} b\u00e0i)")
            # 5. Study performance (p_eligible < 50)
            if p_eligible < 50.0:
                is_failed = True
                reasons.append(f"H\u1ecdc l\u1ef1c y\u1ebfu ({p_eligible:.1f}%)")
            # 6. Project
            if is_qtkd:
                proj_val = s['project']
                if proj_val is None or proj_val < 50.0:
                    is_failed = True
                    reasons.append("Ch\u01b0a hoàn thành Project")
                    
            if is_failed:
                class_risk_list.append({
                    'code': s['student_code'],
                    'name': s['student_name'],
                    'att': att_val,
                    'hw': hw_val,
                    'el': el_val,
                    'rp': rp_val,
                    'p_eligible': p_eligible,
                    'reasons': ", ".join(reasons)
                })
                
        risk_results[cname] = {
            'total_students': len(calibrated),
            'risk_count': len(class_risk_list),
            'risk_rate': (len(class_risk_list) / len(calibrated)) * 100 if calibrated else 0.0,
            'risk_students': class_risk_list,
            'teacher': excel_disc['teacher'] if excel_disc else "\u1ea8n danh"
        }
        
    # Write JSON results
    with open('scratch/student_risk_data.json', 'w', encoding='utf-8') as jf:
        json.dump(risk_results, jf, ensure_ascii=False, indent=4)
        
    # Write Markdown Report
    with open('data/student_risk_report.md', 'w', encoding='utf-8') as mf:
        mf.write("# B\u00e1o c\u00e1o Chi ti\u1ebft Sinh vi\u00ean c\u00f3 nguy c\u01a1 tr\u01b0\u1ee3t m\u00f4n (T\u1eebng l\u1edbp)\n\n")
        mf.write("B\u00e1o c\u00e1o ph\u00e2n t\u00edch t\u1ef1 \u0111\u1ed9ng d\u1eef li\u1ec7u t\u1eeb MySQL v\u00e0 file ch\u1ec9 s\u1ed1 ch\u1ed1t Excel \u0111\u1ec3 d\u1ef1 b\u00e1o v\u00e0 ph\u00e2n lo\u1ea1i l\u00fd do tr\u01b0\u1ee3t c\u1ee7a sinh vi\u00ean.\n\n")
        
        mf.write("## I. B\u1ea2NG T\u1ed4NG H\u1ee2P T\u1ef8 L\u1ec6 NGUY C\u01a1 TR\u01af\u1ee2T THEO L\u1edbp\n\n")
        mf.write("| T\u00ean l\u1edbp | Gi\u1ea3ng vi\u00ean ph\u1ee5 tr\u00e1ch | S\u0129 s\u1ed1 | S\u1ed1 SV nguy c\u01a1 | T\u1ef7 l\u1ec7 nguy c\u01a1 tr\u01b0\u1ee3t |\n")
        mf.write("| :--- | :--- | :---: | :---: | :---: |\n")
        
        for cname, info in risk_results.items():
            mf.write(f"| {cname} | {info['teacher']} | {info['total_students']} | {info['risk_count']} | **{info['risk_rate']:.2f}%** |\n")
            
        mf.write("\n## II. CHI TI\u1ebeT DANH S\u00c1CH SINH VI\u00caN C\u00d3 NGUY C\u01a1 TR\u01af\u1ee2T THEO T\u1eeaNG L\u1ed2P\n\n")
        
        for cname, info in risk_results.items():
            mf.write(f"### L\u1edbp: {cname} (Gi\u1ea3ng vi\u00ean: {info['teacher']})\n")
            mf.write(f"*   **S\u0129 s\u1ed1**: {info['total_students']} sinh vi\u00ean\n")
            mf.write(f"*   **S\u1ed1 sinh vi\u00eaan c\u00f3 nguy c\u01a1 tr\u01b0\u1ee3t**: {info['risk_count']} sinh vi\u00ean ({info['risk_rate']:.2f}%)\n\n")
            
            if info['risk_count'] > 0:
                mf.write("| MSSV | H\u1ecd v\u00e0 t\u00ean | Chuy\u00ean c\u1ea7n (v\u1eafng) | B\u00e0i t\u1eadp (xong) | Elearning (l\u1ed7i) | Rpoint | \u0110i\u1ec3m d\u1ef1 b\u00e1o | L\u00fd do chi ti\u1ebft |\n")
                mf.write("| :--- | :--- | :---: | :---: | :---: | :---: | :---: | :--- |\n")
                for s in info['risk_students']:
                    mf.write(f"| {s['code']} | {s['name']} | {s['att']:.1f}% | {s['hw']:.1f}% | {s['el']:.0f} | {s['rp']:.1f} | {s['p_eligible']:.1f}% | {s['reasons']} |\n")
            else:
                mf.write("> \ud83c\udf89 Kh\u00f4ng c\u00f3 sinh vi\u00ean n\u00e0o c\u00f3 nguy c\u01a1 tr\u01b0\u1ee3t trong l\u1edbp n\u00e0y.\n")
            mf.write("\n" + "-"*50 + "\n\n")
            
    print("Risk analysis report completed successfully.")
    cursor.close()
    conn.close()

if __name__ == '__main__':
    main()
