import sys
import os
import openpyxl

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

files = {
    "old": r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG.xlsx",
    "v1": r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG (1).xlsx",
    "v2": r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG (2).xlsx"
}

for name, path in files.items():
    print(f"=== File: {name} ===")
    if not os.path.exists(path):
        print("Does not exist!")
        continue
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb['Trọng số đánh giá']
    print(f"Row 3: {list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]}")
    print(f"Row 4: {list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]}")
    print("All rows in weights:")
    for r_idx, row in enumerate(sheet.iter_rows(min_row=5, values_only=True)):
        val_strs = [str(x) if x is not None else "" for x in row]
        if any(val_strs):
            print(f"  Row {r_idx+5}: {val_strs}")
    wb.close()
    print()
