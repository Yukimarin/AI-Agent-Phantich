import openpyxl
import sys

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'docs/PTIT_Chiso.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

print("=== KIỂM TRA CHẤT LƯỢNG DỮ LIỆU TRÊN FILE EXCEL PTIT_Chiso.xlsx ===")

errors_found = 0
for sheetname in wb.sheetnames:
    if 'KS' not in sheetname and 'SKL' not in sheetname:
        continue
    sheet = wb[sheetname]
    
    # Đọc các dòng từ 5 trở đi
    student_records = []
    for r in range(5, sheet.max_row + 1):
        cval = sheet.cell(row=r, column=2).value
        pval = sheet.cell(row=r, column=3).value
        
        # Nếu dòng trống
        if not cval and not pval:
            continue
            
        # Kiểm tra dữ liệu số trong các cột điểm vi phạm (từ cột 4 trở đi)
        row_vals = []
        for col_idx in range(4, sheet.max_column + 1):
            val = sheet.cell(row=r, column=col_idx).value
            if val is not None:
                # Nếu là chuỗi, kiểm tra xem có phải số không
                if isinstance(val, str):
                    try:
                        val = float(val)
                    except ValueError:
                        print(f"[{sheetname}] Dòng {r}, Cột {col_idx}: Giá trị không phải số: '{val}'")
                        errors_found += 1
                        continue
                if isinstance(val, (int, float)):
                    if val < 0 or val > 100:
                        print(f"[{sheetname}] Dòng {r}, Cột {col_idx}: Giá trị điểm vi phạm ngoài khoảng [0,100]: {val}")
                        errors_found += 1

print(f"\nTổng số lỗi phát hiện trong dữ liệu Excel: {errors_found}")
