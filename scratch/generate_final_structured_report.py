import sys
import openpyxl
from datetime import datetime, date
from collections import defaultdict
import numpy as np

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'docs/PTIT_Chiso.xlsx'
output_path = 'data/kpi_report.md'

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

wb = openpyxl.load_workbook(excel_path, data_only=True)

# ----------------------------------------------------
# 1. PARSE DATA
# ----------------------------------------------------
class_course_data = defaultdict(dict)
teacher_stats = {}

completed_sheets = [
    'KS24-JavaAdvance', 'KS24_JavaWeb', 'KS24_JWS',
    'KS25_Javascript', 'KS25_Database', 'KS25_Python',
    'KS25_QTKD_M103', 'KS25_QTKD_M104', 'KS25_QTKD_DTB201'
]

for sheetname in completed_sheets:
    if sheetname not in wb.sheetnames:
        continue
    sheet = wb[sheetname]
    max_r = sheet.max_row
    max_c = sheet.max_column
    if max_r < 5:
        continue
        
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    dates_list = []
    current_date = None
    for c_idx in range(3, max_c):
        val3 = row3[c_idx]
        val4 = row4[c_idx]
        if val3:
            current_date = parse_date(val3)
        if current_date:
            dates_list.append((c_idx, current_date, val4))
            
    current_class = None
    gv_cc_vals = []
    gv_bt_vals = []
    gv_el_vals = []
    
    for r in range(5, max_r + 1):
        cname = sheet.cell(row=r, column=2).value
        teacher = sheet.cell(row=r, column=3).value
        
        if cname:
            current_class = normalize_class_name(cname)
            gv_cc_vals = []
            gv_bt_vals = []
            gv_el_vals = []
            for c_idx, d, sub in dates_list:
                val = sheet.cell(row=r, column=c_idx + 1).value
                if isinstance(val, (int, float)):
                    if sub == 'Chuyên cần':
                        gv_cc_vals.append(val)
                    elif sub == 'Bài tập':
                        gv_bt_vals.append(val)
                    elif sub == 'Elearning':
                        gv_el_vals.append(val)
                        
        if not current_class or not teacher or str(teacher).strip() in ['', 'None', 'Chưa phân công', 'Giảng viên/Trợ giảng']:
            continue
            
        teacher_name = str(teacher).strip()
        role = 'GV' if cname is not None else 'TG'
        dept = 'QTKD' if 'QTKD' in sheetname or 'QTKD' in current_class else 'CNTT'
        
        cc_vals = gv_cc_vals.copy()
        bt_vals = gv_bt_vals.copy()
        el_vals = gv_el_vals.copy()
        
        cc_avg = np.mean(cc_vals) if cc_vals else 0
        bt_avg = np.mean(bt_vals) if bt_vals else 0
        el_avg = np.mean(el_vals) if el_vals else 0
        
        if cc_vals or bt_vals or el_vals:
            class_course_data[(current_class, role)][sheetname] = {
                'cc_avg': cc_avg,
                'bt_avg': bt_avg,
                'el_avg': el_avg,
                'teacher': teacher_name,
                'min_date': min(d for c, d, s in dates_list) if dates_list else date.min
            }
            
        if teacher_name not in teacher_stats:
            teacher_stats[teacher_name] = {
                'role': role,
                'department': dept,
                'cc': [],
                'bt': [],
                'el': [],
                'classes': set()
            }
            
        teacher_stats[teacher_name]['classes'].add(current_class)
        if cc_vals: teacher_stats[teacher_name]['cc'].extend(cc_vals)
        if bt_vals: teacher_stats[teacher_name]['bt'].extend(bt_vals)
        if el_vals: teacher_stats[teacher_name]['el'].extend(el_vals)

# Reconstruct class transitions for completed courses only
transitions = []
for (cname, role), courses in sorted(class_course_data.items()):
    if len(courses) < 2:
        continue
    sorted_courses = sorted(courses.items(), key=lambda x: x[1]['min_date'])
    for i in range(1, len(sorted_courses)):
        prev_sheet, prev_data = sorted_courses[i-1]
        curr_sheet, curr_data = sorted_courses[i]
        
        delta_cc = prev_data['cc_avg'] - curr_data['cc_avg']
        delta_bt = prev_data['bt_avg'] - curr_data['bt_avg']
        delta_el = prev_data['el_avg'] - curr_data['el_avg']
        
        transitions.append({
            'class': cname,
            'role': role,
            'prev_sheet': prev_sheet,
            'prev_teacher': prev_data['teacher'],
            'prev_cc': prev_data['cc_avg'],
            'prev_bt': prev_data['bt_avg'],
            'prev_el': prev_data['el_avg'],
            'curr_sheet': curr_sheet,
            'curr_teacher': curr_data['teacher'],
            'curr_cc': curr_data['cc_avg'],
            'curr_bt': curr_data['bt_avg'],
            'curr_el': curr_data['el_avg'],
            'delta_cc': delta_cc,
            'delta_bt': delta_bt,
            'delta_el': delta_el
        })

# Calculate CMI
evaluated_staff = []
watchlist_staff = []

for name, stats in teacher_stats.items():
    cc_mean = np.mean(stats['cc']) if stats['cc'] else 0.0
    bt_mean = np.mean(stats['bt']) if stats['bt'] else 0.0
    el_mean = np.mean(stats['el']) if stats['el'] else 0.0
    
    t_imps = [item for item in transitions if item['curr_teacher'] == name and item['role'] == stats['role']]
    
    if t_imps:
        cmi_values = []
        for class_item in t_imps:
            cmi_cc = 0.5 * class_item['prev_cc'] - class_item['curr_cc'] + 15.0
            cmi_bt = 0.5 * class_item['prev_bt'] - class_item['curr_bt'] + 15.0
            cmi_el = 0.5 * class_item['prev_el'] - class_item['curr_el'] + 15.0
            cmi_values.append((cmi_cc + cmi_bt + cmi_el) / 3.0)
            
        mgmt_score = np.mean(cmi_values)
        
        if mgmt_score > 12.0:
            classification = "Rescuers (Giải cứu xuất sắc)"
        elif mgmt_score < 0.0 or cc_mean > 25.0 or bt_mean > 20.0 or el_mean > 20.0:
            classification = "Cần Hỗ Trợ (Needs Support)"
        else:
            classification = "Duy trì (Maintainers)"
            
        avg_delta_cc = np.mean([item['delta_cc'] for item in t_imps])
        avg_delta_bt = np.mean([item['delta_bt'] for item in t_imps])
        avg_delta_el = np.mean([item['delta_el'] for item in t_imps])
        
        evaluated_staff.append({
            'name': name,
            'role': stats['role'],
            'dept': stats['department'],
            'classes_count': len(stats['classes']),
            'classes_list': ", ".join(list(stats['classes'])),
            'cc': cc_mean,
            'bt': bt_mean,
            'el': el_mean,
            'delta_cc': avg_delta_cc,
            'delta_bt': avg_delta_bt,
            'delta_el': avg_delta_el,
            'cmi': mgmt_score,
            'class': classification
        })
    else:
        watchlist_staff.append({
            'name': name,
            'role': stats['role'],
            'dept': stats['department'],
            'classes_count': len(stats['classes']),
            'classes_list': ", ".join(list(stats['classes'])),
            'cc': cc_mean,
            'bt': bt_mean,
            'el': el_mean
        })

# ----------------------------------------------------
# 3. GET LATEST COMPLETED COURSE DATA FOR EACH CLASS
# ----------------------------------------------------
class_latest_completed = {}

for (cname, role), courses in class_course_data.items():
    if role != 'GV':
        continue # Only take GV row metrics
    # Sort sheets by min_date
    sorted_courses = sorted(courses.items(), key=lambda x: x[1]['min_date'])
    latest_sheet, latest_data = sorted_courses[-1]
    
    # Get TG name if available for this sheet
    tg_data = class_course_data.get((cname, 'TG'), {}).get(latest_sheet)
    tg_name = tg_data['teacher'] if tg_data else "Không có"
    
    class_latest_completed[cname] = {
        'sheet': latest_sheet,
        'gv': latest_data['teacher'],
        'tg': tg_name,
        'cc': latest_data['cc_avg'],
        'bt': latest_data['bt_avg'],
        'el': latest_data['el_avg']
    }

# ----------------------------------------------------
# 4. WRITE THE REPORT IN EXACT REQUESTED STRUCTURE
# ----------------------------------------------------
with open(output_path, 'w', encoding='utf-8') as f:
    f.write("# BÁO CÁO ĐÁNH GIÁ CHỈ SỐ ĐÀO TẠO & NĂNG LỰC QUẢN TRỊ LỚP (MÔN HỌC ĐÃ HOÀN THÀNH)\n\n")
    f.write("> [!IMPORTANT]\n")
    f.write("> Báo cáo này loại bỏ hoàn toàn các môn học mới bắt đầu (`KS24 Java AI`, `KS25 Python Web`, và `QTKD DTB202`) để tránh sai lệch thông tin do chưa đủ dữ liệu.\n")
    f.write("> Điểm Quản trị lớp áp dụng **Chỉ số CMI (Composite Management Index)** tích lũy từ toàn bộ các chuyển tiếp lịch sử của môn học đã kết thúc.\n\n")
    
    # ====================================================
    # 1. ĐÁNH GIÁ KS24
    # ====================================================
    f.write("## 1. ĐÁNH GIÁ KHÓA KS24\n\n")
    
    # Class Table
    f.write("### 📊 Các chỉ số vi phạm và đánh giá tình trạng từng lớp:\n\n")
    f.write("| Tên Lớp | Môn học gần nhất | Giảng viên | Trợ giảng | Vi phạm CC | Vi phạm BT | Vi phạm EL | Tình trạng Kỷ luật |\n")
    f.write("| :--- | :--- | :--- | :--- | :---: | :---: | :---: | :--- |\n")
    
    ks24_classes = {k: v for k, v in class_latest_completed.items() if 'K24' in k}
    for cname, d in sorted(ks24_classes.items()):
        status = ""
        if d['cc'] > 25.0 or d['bt'] > 20.0 or d['el'] > 20.0:
            status = "🚨 Báo động vi phạm cao"
        elif d['cc'] > 15.0 or d['bt'] > 10.0 or d['el'] > 10.0:
            status = "⚠️ Cần kiểm soát chặt chẽ"
        else:
            status = "✅ An toàn / Kỷ luật tốt"
        f.write(f"| **{cname}** | {d['sheet']} | {d['gv']} | {d['tg']} | {d['cc']:.2f}% | {d['bt']:.2f}% | {d['el']:.2f}% | {status} |\n")
    f.write("\n")
    
    # Conclusion
    f.write("### 📝 Kết luận: Tình hình chung các lớp KS24\n")
    f.write("- **Cơ sở Hà Nội**: 4 trên 5 lớp kiểm soát tốt chuyên cần và bài tập (CC vắng dưới 12%, BT nợ dưới 9%). Tuy nhiên, lớp **HN-K24-CNTT5** (thầy Hồ Xuân Hùng dạy JavaWeb) là ngoại lệ nghiêm trọng khi để tỷ lệ vắng học lên đến **48.33%** và nợ bài tập **27.00%**. Đồng thời, chỉ số vi phạm Elearning chậm trễ toàn cơ sở Hà Nội ở mức cao (trung bình **26.29%**).\n")
    f.write("- **Cơ sở TP. HCM**: Lớp **HCM-K24-CNTT1** duy trì kỷ luật tuyệt vời với tỷ lệ vi phạm chuyên cần chỉ **2.70%**, bài tập **0.14%** và Elearning **2.84%**.\n\n")
    
    # Solutions
    f.write("### 🛠️ Giải pháp khắc phục:\n")
    f.write("1. **Chấn chỉnh khẩn cấp lớp HN-K24-CNTT5**: Tổ chức họp 3 bên giữa PM, Ban Đào tạo và Giảng viên/Trợ giảng lớp này để chấn chỉnh kỷ luật, gọi điện trực tiếp cho từng học viên vi phạm để đôn đốc.\n")
    f.write("2. **Áp dụng chính sách phạt điểm chuyên cần theo Elearning**: Đóng cổng Elearning hàng tuần và trừ thẳng điểm chuyên cần nếu học viên nộp trễ hạn.\n\n")
    
    f.write("---\n\n")
    
    # ====================================================
    # 2. ĐÁNH GIÁ KS25
    # ====================================================
    f.write("## 2. ĐÁNH GIÁ KHÓA KS25\n\n")
    
    # 2.1 Khối CNTT
    f.write("### 💻 Khối Công nghệ Thông tin (CNTT)\n\n")
    
    # Class Table
    f.write("#### 📊 Các chỉ số vi phạm và đánh giá tình trạng từng lớp:\n\n")
    f.write("| Tên Lớp | Môn học gần nhất | Giảng viên | Trợ giảng | Vi phạm CC | Vi phạm BT | Vi phạm EL | Tình trạng Kỷ luật |\n")
    f.write("| :--- | :--- | :--- | :--- | :---: | :---: | :---: | :--- |\n")
    
    ks25_cntt_classes = {k: v for k, v in class_latest_completed.items() if 'K25' in k and 'QTKD' not in k}
    for cname, d in sorted(ks25_cntt_classes.items()):
        status = ""
        if d['cc'] > 25.0 or d['bt'] > 20.0 or d['el'] > 20.0:
            status = "🚨 Báo động vi phạm cao"
        elif d['cc'] > 15.0 or d['bt'] > 10.0 or d['el'] > 10.0:
            status = "⚠️ Cần kiểm soát chặt chẽ"
        else:
            status = "✅ An toàn / Kỷ luật tốt"
        f.write(f"| **{cname}** | {d['sheet']} | {d['gv']} | {d['tg']} | {d['cc']:.2f}% | {d['bt']:.2f}% | {d['el']:.2f}% | {status} |\n")
    f.write("\n")
    
    # Conclusion
    f.write("#### 📝 Kết luận: Tình hình chung các lớp KS25 Khối CNTT\n")
    f.write("- **Cơ sở Hà Nội**: Kỷ luật lớp học tương đối ổn định (CC vắng 15.92%, BT nợ 11.90%). Riêng lớp **HN-K25-CNTT7** (môn Javascript) có chỉ số vi phạm rất cao: CC vắng **55.94%**, nợ bài tập **40.32%**, chậm Elearning **30.94%**.\n")
    f.write("- **Cơ sở TP. HCM**: Có dấu hiệu đi xuống nghiêm trọng về kỷ luật chuyên cần ở môn Python gần nhất: Chuyên cần vắng trung bình **20.88%** (lớp **HCM-K25-CNTT8** vắng đến **52.15%**, nợ bài tập **19.98%**). Chỉ số Elearning chậm trễ của HCM ở mức **23.77%**.\n\n")
    
    # Solutions
    f.write("#### 🛠️ Giải pháp khắc phục:\n")
    f.write("1. **Kiểm soát chặt chẽ các lớp Javascript & Python**: Bố trí thêm Trợ giảng phụ đạo bài tập cho học viên yếu để tránh chán nản dẫn đến bỏ học.\n")
    f.write("2. **Bắt buộc gọi điện thoại (Call-to-Action) hàng ngày**: Trợ giảng phải gọi điện nhắc nhở ngay trong ngày đối với các học viên vắng học không lý do hoặc nợ bài tập quá 24h.\n\n")
    
    # 2.2 Khối QTKD
    f.write("### 📈 Khối Quản trị Kinh doanh (QTKD)\n\n")
    
    # Class Table
    f.write("#### 📊 Các chỉ số vi phạm và đánh giá tình trạng từng lớp:\n\n")
    f.write("| Tên Lớp | Môn học gần nhất | Giảng viên | Trợ giảng | Vi phạm CC | Vi phạm BT | Vi phạm EL | Tình trạng Kỷ luật |\n")
    f.write("| :--- | :--- | :--- | :--- | :---: | :---: | :---: | :--- |\n")
    
    ks25_qtkd_classes = {k: v for k, v in class_latest_completed.items() if 'QTKD' in k}
    for cname, d in sorted(ks25_qtkd_classes.items()):
        status = ""
        if d['cc'] > 25.0 or d['bt'] > 20.0 or d['el'] > 20.0:
            status = "🚨 Báo động vi phạm cao"
        elif d['cc'] > 15.0 or d['bt'] > 10.0 or d['el'] > 10.0:
            status = "⚠️ Cần kiểm soát chặt chẽ"
        else:
            status = "✅ An toàn / Kỷ luật tốt"
        f.write(f"| **{cname}** | {d['sheet']} | {d['gv']} | {d['tg']} | {d['cc']:.2f}% | {d['bt']:.2f}% | {d['el']:.2f}% | {status} |\n")
    f.write("\n")
    
    # Conclusion
    f.write("#### 📝 Kết luận: Tình hình chung các lớp KS25 Khối QTKD\n")
    f.write("- **Chuyên cần**: Ở mức trung bình khá (vắng 17.84%), lớp **HN-K25-QTKD1** vắng cao nhất (**26.08%**).\n")
    f.write("- **Bài tập**: Rất tốt (nợ bài trung bình chỉ 8.49%).\n")
    f.write("- **Elearning**: Điểm yếu đặc thù của khối QTKD với tỷ lệ vi phạm chậm trễ trung bình lên tới **58.67%** (lớp **HN-K25-QTKD3** chậm trễ đến **70.73%**). Sinh viên QTKD chưa có ý thức tự học online.\n\n")
    
    # Solutions
    f.write("#### 🛠️ Giải pháp khắc phục:\n")
    f.write("1. **Tổ chức workshop 'Phương pháp tự học hiệu quả'**: Hướng dẫn học viên khối QTKD cách thao tác và hoàn thành bài tập Elearning.\n")
    f.write("2. **Đồng hành trực tiếp trên nhóm chat lớp**: Hàng ngày TG chụp ảnh tiến độ Elearning gửi vào nhóm chat của lớp để tạo áp lực học tập đồng hành giữa các học viên.\n\n")
    
    f.write("---\n\n")
    
    # ====================================================
    # 3. ĐÁNH GIÁ NĂNG LỰC QUẢN TRỊ LỚP
    # ====================================================
    f.write("## 3. ĐÁNH GIÁ NĂNG LỰC QUẢN TRỊ LỚP CỦA GV VÀ TG (CMI RANKING)\n\n")
    f.write("Bảng xếp hạng năng lực quản trị lớp của giảng viên (GV) và trợ giảng (TG) dựa trên chỉ số quản trị hợp phần CMI trung bình tích lũy từ các môn đã hoàn thành:\n\n")
    
    # GV CNTT Table
    f.write("### 💻 Khối Công nghệ Thông tin (CNTT) - Giảng viên (GV)\n\n")
    f.write("| Hạng | Giảng viên | Số lớp dạy | CC Vi phạm | BT Vi phạm | EL Vi phạm | Delta CC | Delta BT | Delta EL | Chỉ số CMI | Phân loại |\n")
    f.write("| :---: | :--- | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :--- |\n")
    gv_cntt = [t for t in evaluated_staff if t['dept'] == 'CNTT' and t['role'] == 'GV']
    for idx, t in enumerate(sorted(gv_cntt, key=lambda x: x['cmi'], reverse=True), 1):
        f.write(f"| {idx} | **{t['name']}** | {t['classes_count']} | {t['cc']:.2f}% | {t['bt']:.2f}% | {t['el']:.2f}% | {t['delta_cc']:+.2f}% | {t['delta_bt']:+.2f}% | {t['delta_el']:+.2f}% | **{t['cmi']:+.2f}%** | {t['class']} |\n")
    f.write("\n")
    
    # TG CNTT Table
    f.write("### 💻 Khối Công nghệ Thông tin (CNTT) - Trợ giảng (TG)\n\n")
    f.write("| Hạng | Trợ giảng | Số lớp phụ trách | CC Vi phạm | BT Vi phạm | EL Vi phạm | Delta CC | Delta BT | Delta EL | Chỉ số CMI | Phân loại |\n")
    f.write("| :---: | :--- | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :--- |\n")
    tg_cntt = [t for t in evaluated_staff if t['dept'] == 'CNTT' and t['role'] == 'TG']
    for idx, t in enumerate(sorted(tg_cntt, key=lambda x: x['cmi'], reverse=True), 1):
        f.write(f"| {idx} | **{t['name']}** | {t['classes_count']} | {t['cc']:.2f}% | {t['bt']:.2f}% | {t['el']:.2f}% | {t['delta_cc']:+.2f}% | {t['delta_bt']:+.2f}% | {t['delta_el']:+.2f}% | **{t['cmi']:+.2f}%** | {t['class']} |\n")
    f.write("\n")
    
    # QTKD Staff Table
    f.write("### 📈 Khối Quản trị Kinh doanh (QTKD) - Giảng viên (GV) & Trợ giảng (TG)\n\n")
    f.write("| Hạng | GV/TG | Vai trò | Số lớp dạy | CC Vi phạm | BT Vi phạm | EL Vi phạm | Delta CC | Delta BT | Delta EL | Chỉ số CMI | Phân loại |\n")
    f.write("| :---: | :--- | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :--- |\n")
    staff_qtkd = [t for t in evaluated_staff if t['dept'] == 'QTKD']
    for idx, t in enumerate(sorted(staff_qtkd, key=lambda x: x['cmi'], reverse=True), 1):
        f.write(f"| {idx} | **{t['name']}** | {t['role']} | {t['classes_count']} | {t['cc']:.2f}% | {t['bt']:.2f}% | {t['el']:.2f}% | {t['delta_cc']:+.2f}% | {t['delta_bt']:+.2f}% | {t['delta_el']:+.2f}% | **{t['cmi']:+.2f}%** | {t['class']} |\n")
    f.write("\n")
    
    # Watchlist Table
    f.write("### 📋 Danh sách Giảng viên/Trợ giảng mới (Đang theo dõi)\n")
    f.write("Các nhân sự chỉ dạy 1 lớp hoặc chưa có dữ liệu đối chiếu chéo:\n\n")
    f.write("| Họ và Tên | Vai trò | Khối | Lớp phụ trách | CC Vi phạm | BT Vi phạm | EL Vi phạm |\n")
    f.write("| :--- | :---: | :---: | :--- | :---: | :---: | :---: |\n")
    for t in sorted(watchlist_staff, key=lambda x: x['cc']):
        f.write(f"| **{t['name']}** | {t['role']} | {t['dept']} | {t['classes_list']} | {t['cc']:.2f}% | {t['bt']:.2f}% | {t['el']:.2f}% |\n")

print("Final structured completed KPI report generated successfully.")
