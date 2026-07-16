import openpyxl
import re
import sys

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'docs/PTIT_Chiso.xlsx'
wb = openpyxl.load_workbook(excel_path)

print("=== STARTING EXCEL DATA CLEANING ===")

def clean_value(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if val_str == '' or val_str.lower() == 'nan':
        return None
        
    # Nếu là dấu gạch ngang
    if val_str == '-':
        return 0.0
        
    # Thay thế dấu phẩy thập phân thành dấu chấm (ví dụ 54,55 -> 54.55)
    # Nhưng phải cẩn thận không làm hỏng các chuỗi khác
    if ',' in val_str and not ('(' in val_str):
        # Kiểm tra xem có phải định dạng số dùng dấu phẩy không
        temp = val_str.replace(',', '.')
        try:
            return float(temp)
        except ValueError:
            pass

    # Nếu có chứa ngoặc đơn (ví dụ: 21.88(16), 0(11), 37.93\n(21))
    if '(' in val_str:
        # Lấy phần trước dấu ngoặc đơn
        prefix = val_str.split('(')[0].strip()
        # Thay thế dấu phẩy thập phân nếu có trong phần prefix
        prefix = prefix.replace(',', '.')
        try:
            return float(prefix)
        except ValueError:
            pass
            
    # Thử convert trực tiếp sang float
    try:
        return float(val_str)
    except ValueError:
        pass
        
    # Nếu là các chữ nhập nhầm (ví dụ: An, Tuấn)
    # Ta trả về None để bỏ qua trong tính toán trung bình
    if val_str.lower() in ['an', 'tuấn', 'tuan', 'n/a', 'unknown']:
        return None
        
    return val

cleaned_count = 0
for sheetname in wb.sheetnames:
    if 'KS' not in sheetname and 'SKL' not in sheetname:
        continue
    sheet = wb[sheetname]
    print(f"Cleaning sheet: {sheetname}...")
    
    for r in range(5, sheet.max_row + 1):
        for col_idx in range(4, sheet.max_column + 1):
            cell = sheet.cell(row=r, column=col_idx)
            val = cell.value
            if val is not None:
                new_val = clean_value(val)
                if new_val != val:
                    cell.value = new_val
                    cleaned_count += 1

wb.save(excel_path)
print(f"Excel data cleaning completed. Total cells cleaned/modified: {cleaned_count}")
