import sys
import os
import openpyxl

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

files = {
    "old": r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG.xlsx",
    "v1": r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG (1).xlsx",
    "v2": r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG (2).xlsx"
}

def load_sheet_data(path, sheet_name):
    if not os.path.exists(path):
        return None
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None
    sheet = wb[sheet_name]
    data = []
    for r in sheet.iter_rows(values_only=True):
        data.append(list(r))
    wb.close()
    return data

sheets_to_compare = ['Phân loại cấp bậc', 'Cơ chế đánh giá', 'Trọng số đánh giá', 'Rubric tính điểm']

for sheet_name in sheets_to_compare:
    print(f"\n==================== COMPARE SHEET: {sheet_name} ====================")
    data_old = load_sheet_data(files["old"], sheet_name)
    data_v1 = load_sheet_data(files["v1"], sheet_name)
    data_v2 = load_sheet_data(files["v2"], sheet_name)
    
    if data_old is None or data_v2 is None:
        print(f"Sheet {sheet_name} not found in one of the files.")
        continue
        
    # Compare dimensions
    print(f"Dimensions: Old ({len(data_old)}x{len(data_old[0]) if data_old else 0}), V1 ({len(data_v1)}x{len(data_v1[0]) if data_v1 else 0}), V2 ({len(data_v2)}x{len(data_v2[0]) if data_v2 else 0})")
    
    # Check for cell differences between Old and V2
    max_r = max(len(data_old), len(data_v2))
    diff_count = 0
    for r_idx in range(max_r):
        row_old = data_old[r_idx] if r_idx < len(data_old) else []
        row_v2 = data_v2[r_idx] if r_idx < len(data_v2) else []
        row_v1 = data_v1[r_idx] if r_idx < len(data_v1) else []
        
        max_c = max(len(row_old), len(row_v2), len(row_v1))
        for c_idx in range(max_c):
            val_old = row_old[c_idx] if c_idx < len(row_old) else None
            val_v1 = row_v1[c_idx] if c_idx < len(row_v1) else None
            val_v2 = row_v2[c_idx] if c_idx < len(row_v2) else None
            
            # Standardize comparisons
            if val_old != val_v2:
                diff_count += 1
                if diff_count <= 25: # Show first 25 differences
                    col_letter = openpyxl.utils.get_column_letter(c_idx + 1)
                    print(f"  Row {r_idx+1} Col {col_letter}:")
                    print(f"    Old: {repr(val_old)}")
                    print(f"    V1 : {repr(val_v1)}")
                    print(f"    V2 : {repr(val_v2)}")
                    
    print(f"Total cell differences between Old and V2 in '{sheet_name}': {diff_count}")
