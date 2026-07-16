import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

def main():
    wb = openpyxl.load_workbook("docs/PTIT_Chiso.xlsx", data_only=True)
    sheet = wb['KS25_Database']
    
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    print("=== KS25_Database Headers ===")
    for c_idx in range(len(row3)):
        print(f"Col {c_idx+1} (Letter {openpyxl.utils.get_column_letter(c_idx+1)}): Row3={row3[c_idx]} | Row4={row4[c_idx]}")
        
    print("\n=== Values for HCM-K25-CNTT5 ===")
    for r in range(5, sheet.max_row + 1):
        cname = sheet.cell(row=r, column=2).value
        if cname and 'HCM-K25-CNTT5' in str(cname):
            print(f"Row {r} Class: {cname}")
            for c_idx in range(2, sheet.max_column):
                val = sheet.cell(row=r, column=c_idx + 1).value
                print(f"  Col {c_idx+1} ({openpyxl.utils.get_column_letter(c_idx+1)}): {val}")
            break

if __name__ == "__main__":
    main()
