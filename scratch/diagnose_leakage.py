import mysql.connector
import sys
import openpyxl
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

def mean(lst):
    return sum(lst) / len(lst) if lst else 0.0

def main():
    conn = mysql.connector.connect(
        host="localhost",
        port=3307,
        user="root",
        password="",
        database="qldt_el"
    )
    cursor = conn.cursor(dictionary=True)
    
    excel_path = "docs/PTIT_Chiso.xlsx"
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Trace HN-KS24-CNTT4 and IT202 (Cơ sở dữ liệu - ID 123)
    cid = 51 # HN-KS24-CNTT4
    co_id = 123 # IT202 - K24 - Cơ sở dữ liệu
    
    # 1. Xem dữ liệu thô trong DB
    cursor.execute("""
        SELECT student_id, homework, elearning, attendance, hackathon_1, hackathon_2, rpoints, project, pass
        FROM final_results
        WHERE class_id = %s AND course_id = %s;
    """, (cid, co_id))
    students_db = cursor.fetchall()
    print(f"Total students in DB for IT202: {len(students_db)}")
    
    # 2. Xem dữ liệu trong Excel
    sheet = wb['KS24_JavaWeb'] # Wait, Cơ sở dữ liệu của KS24 map sang sheet nào?
    # Trong code, 'cơ sở dữ liệu' map sang 'KS25_Database'
    sheetname = 'KS25_Database'
    sheet = wb[sheetname]
    print(f"Excel sheet {sheetname} mapping:")
    
    # Tìm lớp HN-KS24-CNTT4 trong sheet KS25_Database
    excel_row = None
    for r in range(5, sheet.max_row + 1):
        cname = sheet.cell(row=r, column=2).value
        if cname and 'HN-KS24-CNTT4' in str(cname):
            excel_row = r
            print(f"  Found class row in Excel: {r} | Value: {cname}")
            break
            
    if excel_row:
        # Đọc dữ liệu Excel chốt
        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
        # In ra các cột cuối cùng
        print("  Excel Row 3 headers (last 10):", row3[-10:])
        print("  Excel Row 4 headers (last 10):", row4[-10:])
        print("  Excel values (last 10):", [sheet.cell(row=excel_row, column=c+1).value for c in range(sheet.max_column - 10, sheet.max_column)])
        
    cursor.close()
    conn.close()

if __name__ == "__main__":
    main()
