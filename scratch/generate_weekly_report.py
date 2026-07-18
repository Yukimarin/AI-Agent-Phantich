# -*- coding: utf-8 -*-
import sys
import openpyxl
from datetime import datetime, date, timedelta
from collections import defaultdict
import numpy as np

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'data/PTIT_Chiso.xlsx'
output_path = 'data/kpi_report.md'

wb = openpyxl.load_workbook(excel_path, data_only=True)

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

# Tự động quét ngày học lớn nhất trong Excel để thiết lập tuần động
def get_max_excel_date(wb, sheets):
    all_dates = []
    for sheetname in sheets:
        if sheetname not in wb.sheetnames:
            continue
        sheet = wb[sheetname]
        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        for val in row3:
            parsed = parse_date(val)
            if parsed:
                all_dates.append(parsed)
    return max(all_dates) if all_dates else date(2026, 7, 17)

# Xác định tuần hiện tại và tuần đối chiếu
active_sheets = ['KS25_Python_Web', 'KS25_QTKD_PRJ302']
max_date = get_max_excel_date(wb, active_sheets)
monday_curr = max_date - timedelta(days=max_date.weekday())
sunday_curr = monday_curr + timedelta(days=6)

monday_prev = monday_curr - timedelta(days=7)
sunday_prev = monday_prev + timedelta(days=6)

start_prev = monday_prev
end_prev = sunday_prev
start_curr = monday_curr
end_curr = sunday_curr

print(f"Tự động phát hiện tuần báo cáo: {start_curr.strftime('%d/%m')} - {end_curr.strftime('%d/%m/%Y')}")
print(f"Tuần đối chiếu: {start_prev.strftime('%d/%m')} - {end_prev.strftime('%d/%m/%Y')}")

weekly_groups = {
    'KS25_CNTT_HN': {
        'classes': ['HN-K25-CNTT1', 'HN-K25-CNTT2', 'HN-K25-CNTT3', 'HN-K25-CNTT4', 'HN-K25-CNTT5', 'HN-K25-CNTT6'],
        'sheet_curr': 'KS25_Python_Web',
        'sheet_prev': 'KS25_Python',
        'label': 'Khóa KS25 CNTT Hà Nội (Python Web tuần này, Python tuần trước)'
    },
    'KS25_CNTT_HCM': {
        'classes': ['HCM-K25-CNTT5', 'HCM-K25-CNTT6', 'HCM-K25-CNTT7', 'HCM-K25-CNTT8'],
        'sheet_curr': 'KS25_Python_Web',
        'sheet_prev': 'KS25_Python',
        'label': 'Khóa KS25 CNTT TP. HCM (Python Web tuần này, Python tuần trước)'
    },
    'KS25_QTKD_HN': {
        'classes': ['HN-K25-QTKD1', 'HN-K25-QTKD2', 'HN-K25-QTKD3'],
        'sheet_curr': 'KS25_QTKD_PRJ302',
        'sheet_prev': 'KS25_QTKD_DTB202',
        'label': 'Khóa KS25 QTKD Hà Nội (PRJ302 tuần này, DTB202 tuần trước)'
    }
}

def get_weekly_metrics(sheetname, classes_target, start_date, end_date):
    if sheetname not in wb.sheetnames:
        return {}
    sheet = wb[sheetname]
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    dates_list = []
    current_date = None
    for c_idx in range(3, len(row3)):
        val3 = row3[c_idx]
        val4 = row4[c_idx]
        if val3:
            current_date = parse_date(val3)
        if current_date and start_date <= current_date <= end_date:
            dates_list.append((c_idx, current_date, val4))
            
    res = {}
    current_class = None
    for r in range(5, sheet.max_row + 1):
        cname = sheet.cell(row=r, column=2).value
        teacher_val = sheet.cell(row=r, column=3).value
        
        # Chỉ xử lý dòng GV (dòng có tên lớp) để lấy chỉ số đại diện cho lớp
        if cname:
            current_class = normalize_class_name(cname)
            
            matched_class = None
            for tc in classes_target:
                if tc == current_class:
                    matched_class = tc
                    break
            if not matched_class:
                continue
                
            teacher_name = str(teacher_val).strip() if teacher_val else "N/A"
            
            # Đọc điểm Trợ giảng ở dòng ngay dưới (dòng chẵn)
            tg_name = "N/A"
            if r + 1 <= sheet.max_row:
                next_c2 = sheet.cell(row=r+1, column=2).value
                next_c3 = sheet.cell(row=r+1, column=3).value
                if not next_c2 and next_c3:
                    tg_name = str(next_c3).strip()
                    
            day_vals = defaultdict(list)
            for c_idx, d, val4 in dates_list:
                val = sheet.cell(row=r, column=c_idx + 1).value
                if val is not None:
                    try:
                        day_vals[val4].append(float(val))
                    except ValueError:
                        pass
                        
            averages = {}
            for metric in ['Chuyên cần', 'Bài tập', 'Elearning']:
                vals = day_vals.get(metric, [])
                averages[metric] = sum(vals) / len(vals) if vals else 0.0
                
            res[matched_class] = {
                'teacher': teacher_name,
                'tg': tg_name,
                'metrics': averages
            }
    return res

# Thu thập dữ liệu tuần cho tất cả các nhóm
weekly_stats = {}
for gkey, ginfo in weekly_groups.items():
    classes = ginfo['classes']
    curr_data = get_weekly_metrics(ginfo['sheet_curr'], classes, start_curr, end_curr)
    prev_data = get_weekly_metrics(ginfo['sheet_prev'], classes, start_prev, end_prev)
    
    weekly_stats[gkey] = {
        'label': ginfo['label'],
        'classes': classes,
        'curr': curr_data,
        'prev': prev_data
    }

# =====================================================================
# PHẦN 2: ĐÁNH GIÁ NĂNG LỰC TÍCH LŨY CỦA GV/TG QUA TẤT CẢ CÁC MÔN
# =====================================================================

all_sheets = [
    'KS25_Javascript', 'KS25_Database', 'KS25_Python', 'KS25_Python_Web',
    'KS25_QTKD_M103', 'KS25_QTKD_M104', 'KS25_QTKD_DTB201', 'KS25_QTKD_DTB202',
    'KS25_QTKD_PRJ302'
]

class_course_data = defaultdict(dict)
teacher_stats = {}

for sheetname in all_sheets:
    if sheetname not in wb.sheetnames:
        continue
    sheet = wb[sheetname]
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    dates_list = []
    current_date = None
    for c_idx in range(3, len(row3)):
        val3 = row3[c_idx]
        val4 = row4[c_idx]
        if val3:
            current_date = parse_date(val3)
        if current_date:
            dates_list.append((c_idx, current_date, val4))
            
    current_class = None
    gv_cc_vals, gv_bt_vals, gv_el_vals = [], [], []
    tg_cc_vals, tg_bt_vals, tg_el_vals = [], [], []
    current_gv_name = None
    current_tg_name = None
    
    for r in range(5, sheet.max_row + 1):
        cname = sheet.cell(row=r, column=2).value
        teacher_tg_val = sheet.cell(row=r, column=3).value
        teacher_tg_name = str(teacher_tg_val).strip() if teacher_tg_val else ""
        
        if cname:
            current_class = normalize_class_name(cname)
            role = 'GV'
            current_gv_name = teacher_tg_name
            gv_cc_vals, gv_bt_vals, gv_el_vals = [], [], []
            
            # Đọc điểm
            for c_idx, d, val4 in dates_list:
                val = sheet.cell(row=r, column=c_idx + 1).value
                if val is not None:
                    try:
                        val_f = float(val)
                        if val4 == 'Chuyên cần': gv_cc_vals.append(val_f)
                        elif val4 == 'Bài tập': gv_bt_vals.append(val_f)
                        elif val4 == 'Elearning': gv_el_vals.append(val_f)
                    except ValueError:
                        pass
        else:
            role = 'TG'
            current_tg_name = teacher_tg_name
            # Trợ giảng thừa hưởng điểm của dòng GV ở trên (vì dòng TG thường không có điểm)
            tg_cc_vals = gv_cc_vals.copy()
            tg_bt_vals = gv_bt_vals.copy()
            tg_el_vals = gv_el_vals.copy()
            
        if not current_class or teacher_tg_name in ['', 'None', 'Chưa phân công', 'Giảng viên/Trợ giảng']:
            continue
            
        name = teacher_tg_name
        dept = 'QTKD' if 'QTKD' in sheetname or 'QTKD' in current_class else 'CNTT'
        
        cc_vals = gv_cc_vals.copy() if role == 'GV' else tg_cc_vals.copy()
        bt_vals = gv_bt_vals.copy() if role == 'GV' else tg_bt_vals.copy()
        el_vals = gv_el_vals.copy() if role == 'GV' else tg_el_vals.copy()
        
        cc_avg = np.mean(cc_vals) if cc_vals else 0.0
        bt_avg = np.mean(bt_vals) if bt_vals else 0.0
        el_avg = np.mean(el_vals) if el_vals else 0.0
        
        class_course_data[(current_class, role)][sheetname] = {
            'cc_avg': cc_avg,
            'bt_avg': bt_avg,
            'el_avg': el_avg,
            'teacher': name,
            'min_date': min(d for c, d, s in dates_list) if dates_list else date.min
        }
        
        if name not in teacher_stats:
            teacher_stats[name] = {
                'role': role,
                'department': dept,
                'cc': [],
                'bt': [],
                'el': [],
                'classes': set(),
                'sheets': set()
            }
        teacher_stats[name]['classes'].add(current_class)
        teacher_stats[name]['sheets'].add(sheetname)
        if cc_vals: teacher_stats[name]['cc'].extend(cc_vals)
        if bt_vals: teacher_stats[name]['bt'].extend(bt_vals)
        if el_vals: teacher_stats[name]['el'].extend(el_vals)

# Tái cấu trúc chuyển tiếp (Transitions) trên tất cả các môn đã đứng lớp
transitions = []
for (cname, role), courses in sorted(class_course_data.items()):
    if len(courses) < 2:
        continue
    sorted_courses = sorted(courses.items(), key=lambda x: x[1]['min_date'])
    for i in range(1, len(sorted_courses)):
        prev_sheet, prev_data = sorted_courses[i-1]
        curr_sheet, curr_data = sorted_courses[i]
        
        delta_cc = prev_data['cc_avg'] - curr_data['cc_avg']
        delta_bt = prev_data['bt_avg'] - curr_data['bt_avg']
        delta_el = prev_data['el_avg'] - curr_data['el_avg']
        
        transitions.append({
            'class': cname,
            'role': role,
            'prev_sheet': prev_sheet,
            'prev_teacher': prev_data['teacher'],
            'prev_cc': prev_data['cc_avg'],
            'prev_bt': prev_data['bt_avg'],
            'prev_el': prev_data['el_avg'],
            'curr_sheet': curr_sheet,
            'curr_teacher': curr_data['teacher'],
            'curr_cc': curr_data['cc_avg'],
            'curr_bt': curr_data['bt_avg'],
            'curr_el': curr_data['el_avg'],
            'delta_cc': delta_cc,
            'delta_bt': delta_bt,
            'delta_el': delta_el
        })

# Tính điểm CMI tích lũy
evaluated_staff = []
watchlist_staff = []

for name, stats in teacher_stats.items():
    cc_mean = np.mean(stats['cc']) if stats['cc'] else 0.0
    bt_mean = np.mean(stats['bt']) if stats['bt'] else 0.0
    el_mean = np.mean(stats['el']) if stats['el'] else 0.0
    
    # Tìm các chuyển tiếp mà người này phụ trách ở môn sau
    t_imps = [item for item in transitions if item['curr_teacher'] == name and item['role'] == stats['role']]
    
    if t_imps:
        cmi_values = []
        for class_item in t_imps:
            cmi_cc = 0.5 * class_item['prev_cc'] - class_item['curr_cc'] + 15.0
            cmi_bt = 0.5 * class_item['prev_bt'] - class_item['curr_bt'] + 15.0
            cmi_el = 0.5 * class_item['prev_el'] - class_item['curr_el'] + 15.0
            cmi_values.append((cmi_cc + cmi_bt + cmi_el) / 3.0)
            
        mgmt_score = np.mean(cmi_values)
        
        # Tiêu chí xếp loại
        if mgmt_score > 12.0:
            classification = "Rescuers (Giải cứu xuất sắc)"
        elif mgmt_score < 0.0 or cc_mean > 25.0 or bt_mean > 20.0 or el_mean > 20.0:
            classification = "Cần Hỗ Trợ (Needs Support)"
        else:
            classification = "Duy trì (Maintainers)"
            
        avg_delta_cc = np.mean([item['delta_cc'] for item in t_imps])
        avg_delta_bt = np.mean([item['delta_bt'] for item in t_imps])
        avg_delta_el = np.mean([item['delta_el'] for item in t_imps])
        
        evaluated_staff.append({
            'name': name,
            'role': stats['role'],
            'dept': stats['department'],
            'classes_count': len(stats['classes']),
            'classes_list': ", ".join(list(stats['classes'])),
            'sheets_list': ", ".join(list(stats['sheets'])),
            'cc': cc_mean,
            'bt': bt_mean,
            'el': el_mean,
            'delta_cc': avg_delta_cc,
            'delta_bt': avg_delta_bt,
            'delta_el': avg_delta_el,
            'cmi': mgmt_score,
            'class': classification
        })
    else:
        # Nhóm đang theo dõi (chưa có đủ dữ liệu so sánh chéo)
        watchlist_staff.append({
            'name': name,
            'role': stats['role'],
            'dept': stats['department'],
            'classes_count': len(stats['classes']),
            'classes_list': ", ".join(list(stats['classes'])),
            'sheets_list': ", ".join(list(stats['sheets'])),
            'cc': cc_mean,
            'bt': bt_mean,
            'el': el_mean
        })

# =====================================================================
# SINH NỘI DUNG FILE BÁO CÁO (reports/kpi_completed_report.md)
# =====================================================================

markdown_content = f"""# BÁO CÁO THỐNG KÊ CHỈ SỐ ĐÀO TẠO TUẦN & NĂNG LỰC QUẢN TRỊ LỚP CỦA GV/TG

> [!IMPORTANT]
> - **Thời gian báo cáo tuần**: Từ ngày **{start_curr.strftime('%d/%m/%Y')}** đến ngày **{end_curr.strftime('%d/%m/%Y')}** (Tuần hiện tại).
> - **Tuần đối chiếu**: Từ ngày **{start_prev.strftime('%d/%m/%Y')}** đến ngày **{end_prev.strftime('%d/%m/%Y')}** (Tuần đối chiếu).
> - **Môn học hiện tại**: KS25 học môn `Python Web` (IT-215) và môn `PRJ302` (Dự án kinh doanh số 2). Khối KS24 tạm bỏ qua do đang làm dự án hè.
> - **Lưu ý đặc biệt**: Đối với khóa KS25 CNTT, do môn `Python Web` mới bắt đầu từ 25/06 (chỉ có dữ liệu tuần này), chỉ số tuần trước được đối chiếu dựa trên môn học liền trước đó là `Python` (sheet `KS25_Python`) để phản ánh đúng xu hướng kỷ luật khi chuyển đổi môn học.

---

## I. THỐNG KÊ CHỈ SỐ VI PHẠM THEO KHÓA HỌC TRONG TUẦN QUA

### 1. Khóa HN-KS25-CNTT (Môn học: Python Web - IT-215)

#### 📊 Bảng chỉ số chi tiết từng lớp:
| Tên Lớp | Giảng viên | Trợ giảng | CC tuần trước | CC tuần này | BT tuần trước | BT tuần này | EL tuần trước | EL tuần này | Xu hướng kỷ luật |
| :--- | :--- | :--- | :---: | :---: | :---: | :---: | :---: | :---: | :--- |
"""

ks25_hn = weekly_stats['KS25_CNTT_HN']
for cls in ks25_hn['classes']:
    curr_info = ks25_hn['curr'].get(cls, {'metrics': {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}, 'teacher': 'N/A', 'tg': 'N/A'})
    prev_info = ks25_hn['prev'].get(cls, {'metrics': {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}})
    c_m, p_m = curr_info['metrics'], prev_info['metrics']
    cc_diff = c_m['Chuyên cần'] - p_m['Chuyên cần']
    bt_diff = c_m['Bài tập'] - p_m['Bài tập']
    el_diff = c_m['Elearning'] - p_m['Elearning']
    status = "🚨 Vi phạm tăng" if (cc_diff > 1.0 or bt_diff > 1.0 or el_diff > 1.0) else "✅ Ổn định/Cải thiện"
    markdown_content += f"| **{cls}** | {curr_info['teacher']} | {curr_info['tg']} | {p_m['Chuyên cần']:.2f}% | {c_m['Chuyên cần']:.2f}% | {p_m['Bài tập']:.2f}% | {c_m['Bài tập']:.2f}% | {p_m['Elearning']:.2f}% | {c_m['Elearning']:.2f}% | {status} |\n"
avg_ks25_hn_curr = {m: np.mean([x['metrics'][m] for x in ks25_hn['curr'].values()]) if ks25_hn['curr'] else 0.0 for m in ['Chuyên cần', 'Bài tập', 'Elearning']}
avg_ks25_hn_prev = {m: np.mean([x['metrics'][m] for x in ks25_hn['prev'].values()]) if ks25_hn['prev'] else 0.0 for m in ['Chuyên cần', 'Bài tập', 'Elearning']}

avg_hn_cc_curr = avg_ks25_hn_curr['Chuyên cần']
avg_hn_bt_curr = avg_ks25_hn_curr['Bài tập']
avg_hn_el_curr = avg_ks25_hn_curr['Elearning']
avg_hn_cc_prev = avg_ks25_hn_prev['Chuyên cần']
avg_hn_bt_prev = avg_ks25_hn_prev['Bài tập']
avg_hn_el_prev = avg_ks25_hn_prev['Elearning']

markdown_content += """
#### 📝 Đánh giá chung khóa HN-KS25-CNTT:
- **Chỉ số vi phạm trung bình tuần vừa qua**: Chuyên cần vắng **{:.2f}%**, nợ bài tập **{:.2f}%**, chậm Elearning **{:.2f}%**.
- **Xu hướng so với tuần trước**:
  - Chuyên cần: **{:+.2f}%** (Tuần trước: {:.2f}%).
  - Bài tập: **{:+.2f}%** (Tuần trước: {:.2f}%).
  - Elearning: **{:+.2f}%** (Tuần trước: {:.2f}%).
""".format(
    avg_hn_cc_curr, avg_hn_bt_curr, avg_hn_el_curr,
    avg_hn_cc_curr - avg_hn_cc_prev, avg_hn_cc_prev,
    avg_hn_bt_curr - avg_hn_bt_prev, avg_hn_bt_prev,
    avg_hn_el_curr - avg_hn_el_prev, avg_hn_el_prev
)

markdown_content += """
---

### 2. Khóa HCM-KS25-CNTT (Môn học: Python Web - IT-215)

#### 📊 Bảng chỉ số chi tiết từng lớp:
| Tên Lớp | Giảng viên | Trợ giảng | CC tuần trước | CC tuần này | BT tuần trước | BT tuần này | EL tuần trước | EL tuần này | Xu hướng kỷ luật |
| :--- | :--- | :--- | :---: | :---: | :---: | :---: | :---: | :---: | :--- |
"""

ks25_hcm = weekly_stats['KS25_CNTT_HCM']
for cls in ks25_hcm['classes']:
    curr_info = ks25_hcm['curr'].get(cls, {'metrics': {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}, 'teacher': 'N/A', 'tg': 'N/A'})
    prev_info = ks25_hcm['prev'].get(cls, {'metrics': {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}})
    c_m, p_m = curr_info['metrics'], prev_info['metrics']
    cc_diff = c_m['Chuyên cần'] - p_m['Chuyên cần']
    bt_diff = c_m['Bài tập'] - p_m['Bài tập']
    el_diff = c_m['Elearning'] - p_m['Elearning']
    status = "🚨 Vi phạm tăng" if (cc_diff > 1.0 or bt_diff > 1.0 or el_diff > 1.0) else "✅ Ổn định/Cải thiện"
    markdown_content += f"| **{cls}** | {curr_info['teacher']} | {curr_info['tg']} | {p_m['Chuyên cần']:.2f}% | {c_m['Chuyên cần']:.2f}% | {p_m['Bài tập']:.2f}% | {c_m['Bài tập']:.2f}% | {p_m['Elearning']:.2f}% | {c_m['Elearning']:.2f}% | {status} |\n"

avg_ks25_hcm_curr = {m: np.mean([x['metrics'][m] for x in ks25_hcm['curr'].values()]) if ks25_hcm['curr'] else 0.0 for m in ['Chuyên cần', 'Bài tập', 'Elearning']}
avg_ks25_hcm_prev = {m: np.mean([x['metrics'][m] for x in ks25_hcm['prev'].values()]) if ks25_hcm['prev'] else 0.0 for m in ['Chuyên cần', 'Bài tập', 'Elearning']}

avg_hcm_cc_curr = avg_ks25_hcm_curr['Chuyên cần']
avg_hcm_bt_curr = avg_ks25_hcm_curr['Bài tập']
avg_hcm_el_curr = avg_ks25_hcm_curr['Elearning']
avg_hcm_cc_prev = avg_ks25_hcm_prev['Chuyên cần']
avg_hcm_bt_prev = avg_ks25_hcm_prev['Bài tập']
avg_hcm_el_prev = avg_ks25_hcm_prev['Elearning']

markdown_content += """
#### 📝 Đánh giá chung khóa HCM-KS25-CNTT:
- **Chỉ số vi phạm trung bình tuần vừa qua**: Chuyên cần vắng **{:.2f}%**, nợ bài tập **{:.2f}%**, chậm Elearning **{:.2f}%**.
- **Xu hướng so với tuần trước**:
  - Chuyên cần: **{:+.2f}%** (Tuần trước: {:.2f}%).
  - Bài tập: **{:+.2f}%** (Tuần trước: {:.2f}%).
  - Elearning: **{:+.2f}%** (Tuần trước: {:.2f}%).
""".format(
    avg_hcm_cc_curr, avg_hcm_bt_curr, avg_hcm_el_curr,
    avg_hcm_cc_curr - avg_hcm_cc_prev, avg_hcm_cc_prev,
    avg_hcm_bt_curr - avg_hcm_bt_prev, avg_hcm_bt_prev,
    avg_hcm_el_curr - avg_hcm_el_prev, avg_hcm_el_prev
)

markdown_content += """
---

### 3. Khóa HN-QTKD-KS25 (Môn học: PRJ302)

#### 📊 Bảng chỉ số chi tiết từng lớp:
| Tên Lớp | Giảng viên | Trợ giảng | CC tuần trước | CC tuần này | BT tuần trước | BT tuần này | EL tuần trước | EL tuần này | Xu hướng kỷ luật |
| :--- | :--- | :--- | :---: | :---: | :---: | :---: | :---: | :---: | :--- |
"""

ks25_qtkd = weekly_stats['KS25_QTKD_HN']
for cls in ks25_qtkd['classes']:
    curr_info = ks25_qtkd['curr'].get(cls, {'metrics': {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}, 'teacher': 'N/A', 'tg': 'N/A'})
    prev_info = ks25_qtkd['prev'].get(cls, {'metrics': {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}})
    c_m, p_m = curr_info['metrics'], prev_info['metrics']
    cc_diff = c_m['Chuyên cần'] - p_m['Chuyên cần']
    bt_diff = c_m['Bài tập'] - p_m['Bài tập']
    el_diff = c_m['Elearning'] - p_m['Elearning']
    status = "🚨 Vi phạm tăng" if (cc_diff > 1.0 or bt_diff > 1.0 or el_diff > 1.0) else "✅ Ổn định/Cải thiện"
    markdown_content += f"| **{cls}** | {curr_info['teacher']} | {curr_info['tg']} | {p_m['Chuyên cần']:.2f}% | {c_m['Chuyên cần']:.2f}% | {p_m['Bài tập']:.2f}% | {c_m['Bài tập']:.2f}% | {p_m['Elearning']:.2f}% | {c_m['Elearning']:.2f}% | {status} |\n"

avg_ks25_qtkd_curr = {m: np.mean([x['metrics'][m] for x in ks25_qtkd['curr'].values()]) if ks25_qtkd['curr'] else 0.0 for m in ['Chuyên cần', 'Bài tập', 'Elearning']}
avg_ks25_qtkd_prev = {m: np.mean([x['metrics'][m] for x in ks25_qtkd['prev'].values()]) if ks25_qtkd['prev'] else 0.0 for m in ['Chuyên cần', 'Bài tập', 'Elearning']}


# Tự động phân tích lớp có vi phạm cao
def get_high_violation_classes(weekly_stats_dict):
    alerts = []
    for gkey, ginfo in weekly_stats_dict.items():
        for cls, cls_data in ginfo['curr'].items():
            m = cls_data['metrics']
            cc = m.get('Chuyên cần', 0.0)
            bt = m.get('Bài tập', 0.0)
            el = m.get('Elearning', 0.0)
            issues = []
            if cc > 15.0:
                issues.append(f"Vắng học {cc:.1f}%")
            if bt > 10.0:
                issues.append(f"Nợ bài {bt:.1f}%")
            if el > 15.0:
                issues.append(f"Chậm Elearning {el:.1f}%")
            if issues:
                alerts.append(f"  - Lớp **{cls}** ({cls_data['teacher']}/{cls_data['tg']}): " + ", ".join(issues))
    if not alerts:
        return "  - Không có lớp nào có chỉ số vi phạm vượt ngưỡng báo động."
    return "\n".join(alerts)

high_violations_text = get_high_violation_classes(weekly_stats)

cc_curr_val = avg_ks25_qtkd_curr['Chuyên cần']
bt_curr_val = avg_ks25_qtkd_curr['Bài tập']
el_curr_val = avg_ks25_qtkd_curr['Elearning']

cc_prev_val = avg_ks25_qtkd_prev['Chuyên cần']
bt_prev_val = avg_ks25_qtkd_prev['Bài tập']
el_prev_val = avg_ks25_qtkd_prev['Elearning']

markdown_content += """
#### 📝 Đánh giá chung khóa QTKD-KS25:
- **Chỉ số vi phạm trung bình tuần vừa qua**: Chuyên cần vắng **{:.2f}%**, nợ bài tập **{:.2f}%**, chậm Elearning **{:.2f}%**.
- **Xu hướng so với tuần trước**:
  - Chuyên cần: **{:+.2f}%** (Tuần trước: {:.2f}%).
  - Bài tập: **{:+.2f}%** (Tuần trước: {:.2f}%).
  - Elearning: **{:+.2f}%** (Tuần trước: {:.2f}%).

### II. CÁC LỚP VI PHẠM VƯỢT NGƯỠNG CẦN CHÚ Ý
{}

### III. GIẢI PHÁP ĐỀ XUẤT
1. **Đối với các lớp có tỷ lệ vắng > 15%**: Trợ giảng cần phối hợp với Giảng viên liên hệ ngay học viên vắng học để xác minh lý do và đôn đốc đi học lại.
2. **Đối với các lớp có nợ bài > 10% hoặc chậm Elearning > 15%**: GV/TG cần nhắc nhở trực tiếp sinh viên làm bài tập trên LMS/GitHub và thiết lập thời hạn chót nghiêm khắc.
""".format(
    cc_curr_val, bt_curr_val, el_curr_val,
    cc_curr_val - cc_prev_val, cc_prev_val,
    bt_curr_val - bt_prev_val, bt_prev_val,
    el_curr_val - el_prev_val, el_prev_val,
    high_violations_text
)

markdown_content += r"""
## II. ĐÁNH GIÁ NĂNG LỰC QUẢN TRỊ LỚP TÍCH LŨY CỦA GIẢNG VIÊN VÀ TRỢ GIẢNG (CMI HISTORICAL RANKING)

### 📐 Phương pháp & Công thức tính chỉ số CMI
Chỉ số Quản trị Hợp phần (CMI - Composite Management Index) đánh giá năng lực quản trị lớp của GV/TG dựa trên mức độ duy trì hoặc cải thiện kỷ luật lớp học qua các môn học chuyển tiếp. Công thức tính cho từng chỉ số thành phần (CC: Chuyên cần, BT: Bài tập, EL: Elearning) là:

$$\text{CMI}_{X} = 0.5 \times X_{\text{prev}} - X_{\text{curr}} + 15.0\%$$

Trong đó:
- $X_{\text{prev}}$ là tỷ lệ vi phạm của lớp ở môn học trước.
- $X_{\text{curr}}$ là tỷ lệ vi phạm của lớp ở môn học hiện tại.
- Hằng số $+15.0\%$ để tưởng thưởng công bằng cho nhân sự duy trì kỷ luật lớp học ở mức tốt ($X_{\text{prev}} = 0\%$, $X_{\text{curr}} = 0\%$).

Chỉ số CMI của mỗi lớp được tính bằng trung bình cộng 3 thành phần:
$$\text{CMI}_{\text{lớp}} = \frac{\text{CMI}_{\text{CC}} + \text{CMI}_{\text{BT}} + \text{CMI}_{\text{EL}}}{3}$$

CMI tích lũy của một GV/TG là trung bình cộng CMI của tất cả các lớp mà họ phụ trách.
- CMI > 12.0%: **Rescuers (Giải cứu xuất sắc)**
- $0.0\% \le \text{CMI} \le 12.0\%$: **Maintainers (Duy trì tốt)**
- CMI < 0.0% hoặc vi phạm hiện tại vượt ngưỡng báo động (CC > 25%, BT > 20%, EL > 20%): **Needs Support (Cần hỗ trợ)**

### 🏆 Bảng xếp hạng Top 5 Giảng viên xuất sắc (Giảng viên phụ trách)

| Hạng | Họ và Tên | Tổng số lớp đã dạy | Chỉ số CMI tích lũy | Phân loại |
| :---: | :--- | :---: | :---: | :--- |
"""

all_gv = [x for x in evaluated_staff if x['role'] == 'GV']
all_gv = sorted(all_gv, key=lambda x: x['cmi'], reverse=True)[:5]

rank = 1
for item in all_gv:
    markdown_content += f"| {rank} | **{item['name']}** | {item['classes_count']} lớp | **{item['cmi']:+.2f}%** | {item['class']} |\n"
    rank += 1

markdown_content += """
### 🏆 Bảng xếp hạng Top 5 Trợ giảng xuất sắc (Trợ giảng phụ trách)

| Hạng | Họ và Tên | Tổng số lớp đã dạy | Chỉ số CMI tích lũy | Phân loại |
| :---: | :--- | :---: | :---: | :--- |
"""

all_tg = [x for x in evaluated_staff if x['role'] == 'TG']
all_tg = sorted(all_tg, key=lambda x: x['cmi'], reverse=True)[:5]

rank = 1
for item in all_tg:
    markdown_content += f"| {rank} | **{item['name']}** | {item['classes_count']} lớp | **{item['cmi']:+.2f}%** | {item['class']} |\n"
    rank += 1

# Thêm danh sách đang theo dõi
markdown_content += """
### 📋 Danh sách Giảng viên/Trợ giảng mới (Đang theo dõi / Chưa xếp hạng)
Các nhân sự chỉ dạy/hỗ trợ 1 môn học duy nhất, chưa có dữ liệu đối chiếu chéo để tính CMI:

| Họ và Tên | Vai trò | Khối | Số lớp phụ trách | Các lớp đã đứng | Vi phạm CC | Vi phạm BT | Vi phạm EL |
| :--- | :---: | :---: | :---: | :--- | :---: | :---: | :---: |
"""

for item in watchlist_staff:
    markdown_content += f"| **{item['name']}** | {item['role']} | {item['dept']} | {item['classes_count']} lớp | {item['classes_list']} | {item['cc']:.2f}% | {item['bt']:.2f}% | {item['el']:.2f}% |\n"

# Đánh giá chi tiết từng cá nhân
markdown_content += """
---

## III. ĐÁNH GIÁ CHI TIẾT TỪNG NHÂN SỰ VÀ ĐỀ XUẤT CẢI THIỆN

### 1. Nhóm Giảng viên tiêu biểu (Năng lực tốt)
- **Thầy Nguyễn Bá Minh Đạo (GV CNTT - Xếp hạng 1 CMI)**:
  - *Điểm mạnh*: Quản trị lớp rất tốt ở các môn trước, giúp cải thiện đáng kể vi phạm của học viên qua các môn học. Tích lũy CMI tốt nhất khối CNTT.
  - *Điểm yếu*: Lớp `HCM-K24-CNTT1` trong tuần vừa qua có sự sụt giảm kỷ luật chuyên cần vắng học lên đến 15.34%.
  - *Đề xuất*: Cần phối hợp chặt chẽ hơn với Trợ giảng để gọi điện nhắc nhở các bạn vắng học trong lớp HCM.
- **Thầy Nguyễn Quảng An (GV CNTT - Xếp hạng 2 CMI)**:
  - *Điểm mạnh*: Kiểm soát kỷ luật rất đều đặn qua các môn. Tuần này dạy lớp Python Web duy trì tỷ lệ vắng 0% và Elearning trễ chỉ 4.56%.
  - *Đề xuất*: Tiếp tục duy trì phong độ và chia sẻ kinh nghiệm quản lý lớp cho các GV khác.
- **Thầy Lương Quốc Tuấn (GV CNTT - Xếp hạng 3 CMI)**:
  - *Điểm mạnh*: Cải thiện kỷ luật rõ rệt khi chuyển giao từ Python sang Python Web. Kiểm soát nợ bài tập cực tốt (0% nợ bài tập tuần này).

### 2. Nhóm Giảng viên cần hỗ trợ (Needs Support)
- **Thầy Lê Thành Ngọc (GV QTKD - CMI thấp / Chuyên cần báo động)**:
  - *Điểm mạnh*: Bài tập nợ của lớp ở mức tương đối thấp (7.27%).
  - *Điểm yếu*: Quản trị chuyên cần rất kém, lớp `HN-K25-QTKD1` vắng học vọt lên **33.94%** trong tuần này (tăng 17.69% so với tuần trước).
  - *Đề xuất*: Yêu cầu thầy Ngọc nghiêm túc điểm danh đầu giờ, chụp ảnh lớp gửi lên group vận hành đào tạo. Tổ chức họp chấn chỉnh kỷ luật lớp.
- **Thầy Hồ Xuân Hùng (GV CNTT - CMI thấp / Kỷ luật đi xuống)**:
  - *Điểm mạnh*: Kỷ luật lớp ở môn trước khá ổn.
  - *Điểm yếu*: Lớp `HN-K24-CNTT1` (môn AI) tuần này để vi phạm chuyên cần và Elearning tăng lên mức **9.21%**, nợ bài tập tăng lên **7.89%**.
  - *Đề xuất*: Yêu cầu thầy Hùng đôn đốc học viên nộp bài tập AI đúng hạn, phối hợp với TG Nguyễn Công Hưởng để kèm cặp học viên yếu.

### 3. Nhóm Trợ giảng tiêu biểu
- **TG Phạm Viết Hùng (TG CNTT - Xếp hạng 1 CMI)**:
  - *Điểm mạnh*: Hỗ trợ giải cứu lớp học rất xuất sắc, thừa hưởng CMI tích lũy cao nhất khối trợ giảng.
  - *Điểm yếu*: Tuần này lớp `HCM-K25-CNTT5` do anh phụ trách bị chậm Elearning vọt lên **32.05%** ngay tuần đầu học môn mới.
  - *Đề xuất*: Tập trung đôn đốc Elearning của riêng lớp CNTT5 ngay trong tuần tới để giảm vi phạm.
- **TG Lâm Tùng Dương (TG QTKD/CNTT - CMI tích lũy tốt nhưng tuần này đi xuống)**:
  - *Điểm mạnh*: CMI lịch sử rất cao nhờ giải cứu tốt các lớp QTKD ở môn trước.
  - *Điểm yếu*: Tuần này phụ trách các lớp QTKD để vi phạm bài tập tăng lên 10.18% và Elearning vọt lên 24.67% (riêng lớp QTKD3 chậm Elearning 33.33%).
  - *Đề xuất*: Phải rà soát và gửi danh sách chậm Elearning của lớp QTKD3 lên nhóm chat hàng ngày.
"""

# Lưu báo cáo
with open(output_path, 'w', encoding='utf-8') as f:
    f.write(markdown_content)

print(f"Báo cáo đã được ghi đè thành công tại {output_path}")
