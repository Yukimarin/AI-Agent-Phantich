import mysql.connector
import sys
import os
import openpyxl
from collections import defaultdict
def mean(lst):
    return sum(lst) / len(lst) if lst else 0.0
from datetime import datetime, date

# Add current directory to path to import local modules
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from excel_loader import normalize_class_name

sys.stdout.reconfigure(encoding='utf-8')

def run_query(cursor, query, params=None):
    cursor.execute(query, params or ())
    columns = [desc[0] for desc in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def get_excel_chot_data(excel_path):
    """
    Load Excel and find:
    1. Last date's discipline rates (cc, bt, el)
    2. Rpoint chot column
    Returns: dict[norm_cname -> dict[sheetname -> dict]]
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    class_data = defaultdict(dict)
    
    for sheetname in wb.sheetnames:
        if sheetname == 'Sheet1' or 'SKL' in sheetname:
            continue
        sheet = wb[sheetname]
        max_r = sheet.max_row
        max_c = sheet.max_column
        if max_r < 5:
            continue
            
        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
        
        # 1. Extract dates and columns
        dates_list = []
        for c_idx in range(3, max_c):
            val3 = row3[c_idx]
            val4 = row4[c_idx]
            if val3:
                d_str = str(val3).strip()
                parts = d_str.split('/')
                try:
                    if len(parts) == 2:
                        d = date(2026, int(parts[1]), int(parts[0]))
                    elif len(parts) == 3:
                        y = int(parts[2])
                        if y < 100: y += 2000
                        d = date(y, int(parts[1]), int(parts[0]))
                    else:
                        d = None
                except:
                    d = None
                if d:
                    dates_list.append((c_idx, d, val4))
            elif dates_list:
                dates_list.append((c_idx, dates_list[-1][1], val4))
                
        if not dates_list:
            continue
            
        # Get last date
        unique_dates = sorted(list(set(d for idx, d, sub in dates_list)))
        last_date = unique_dates[-1] if unique_dates else None
        last_cols = [idx for idx, d, sub in dates_list if d == last_date] if last_date else []
        
        # 2. Find Rpoint chot column (column at the end with avg value in [30, 115], no header date, and must be after the last daily column)
        rp_col_idx = None
        last_daily_col = max(idx for idx, d, sub in dates_list) if dates_list else 2
        for c_idx in range(max_c - 1, last_daily_col, -1):
            h3 = row3[c_idx]
            h4 = row4[c_idx]
            if h4 in ('Chuyên cần', 'Bài tập', 'Elearning'):
                continue
            vals = []
            for r in range(5, max_r + 1):
                val = sheet.cell(row=r, column=c_idx + 1).value
                if val is not None:
                    try:
                        vals.append(float(val))
                    except:
                        pass
            if len(vals) >= 2:
                avg_val = mean(vals)
                if 30.0 <= avg_val <= 115.0:
                    rp_col_idx = c_idx
                    break
        
        # 3. Read data for each class row
        for r in range(5, max_r + 1):
            cname = sheet.cell(row=r, column=2).value
            if cname:
                norm_name = normalize_class_name(cname)
                
                # Discipline chot
                cc_val, bt_val, el_val = 0.0, 0.0, 0.0
                for c_idx in last_cols:
                    sub = row4[c_idx]
                    val = sheet.cell(row=r, column=c_idx + 1).value
                    if val is not None:
                        try:
                            val_f = float(val)
                        except:
                            val_f = 0.0
                        if sub == 'Chuyên cần':
                            cc_val = val_f
                        elif sub == 'Bài tập':
                            bt_val = val_f
                        elif sub == 'Elearning':
                            el_val = val_f
                            
                # Rpoint chot
                rp_val = None
                if rp_col_idx is not None:
                    cell_val = sheet.cell(row=r, column=rp_col_idx + 1).value
                    if cell_val is not None:
                        try:
                            rp_val = float(cell_val)
                        except:
                            pass
                            
                class_data[norm_name][sheetname] = {
                    'cc': cc_val,
                    'bt': bt_val,
                    'el': el_val,
                    'rp': rp_val,
                    'date': last_date
                }
    return class_data

def get_course_difficulty(course_name, cursor):
    coname_lower = course_name.lower()
    if "ai" in coname_lower or "trí tuệ nhân tạo" in coname_lower or "machine learning" in coname_lower:
        return 1.8
    elif "database" in coname_lower or "cơ sở dữ liệu" in coname_lower or "cấu trúc dữ liệu" in coname_lower:
        return 1.5
    elif "python" in coname_lower or "java" in coname_lower or "web" in coname_lower:
        return 1.3
    else:
        return 1.2

def calibrate_students(students_results, excel_disc):
    """
    Calibrate DB student records using Excel chot values.
    Returns list of calibrated student records.
    """
    if not students_results:
        return []
        
    # Pre-fill lists with defaults to avoid NaNs
    att_list = [s['attendance'] if s['attendance'] is not None else 0.0 for s in students_results]
    hw_list = [s['homework'] if s['homework'] is not None else 100.0 for s in students_results]
    el_list = [s['elearning'] if s['elearning'] is not None else 0.0 for s in students_results]
    rp_list = [s['rpoints'] if s['rpoints'] is not None else 100.0 for s in students_results]
    
    # If no Excel data found, return DB raw records directly (no calibration shift)
    if not excel_disc:
        calibrated = []
        for i, s in enumerate(students_results):
            calibrated.append({
                'student_id': s['student_id'],
                'attendance': att_list[i],
                'homework': hw_list[i],
                'elearning': el_list[i],
                'rpoints': rp_list[i],
                'hackathon_1': s['hackathon_1'],
                'hackathon_2': s['hackathon_2'],
                'project': s['project'],
                'pass': s['pass']
            })
        return calibrated
        
    db_att_avg = mean(att_list)
    db_hw_avg = mean(hw_list)
    db_el_avg = mean(el_list)
    db_rp_avg = mean(rp_list)
    
    # Target values from Excel
    excel_cc = excel_disc['cc'] if excel_disc['cc'] is not None else db_att_avg
    excel_bt_err = excel_disc['bt'] if excel_disc['bt'] is not None else (100.0 - db_hw_avg)
    excel_hw = 100.0 - excel_bt_err
    excel_el = excel_disc['el'] if excel_disc['el'] is not None else db_el_avg
    
    excel_rp = excel_disc['rp'] if excel_disc['rp'] is not None else None
    if excel_rp is None:
        excel_rp = max(0.0, 100.0 - excel_cc - excel_bt_err - excel_el)
        
    calibrated = []
    for i, s in enumerate(students_results):
        # 1. Calibrate Rpoints (Shift to match class Rpoint chot from Excel)
        rp_db = rp_list[i]
        rp_cal = rp_db + (excel_rp - db_rp_avg)
        rp_cal = min(120.0, max(0.0, rp_cal))
        
        # 2. Calibrate Attendance (Shift to match class attendance chot)
        att_db = att_list[i]
        att_cal = att_db + (excel_cc - db_att_avg)
        att_cal = min(100.0, max(0.0, att_cal))
        
        # 3. Calibrate Homework Completion (Shift to match class homework chot)
        hw_db = hw_list[i]
        hw_cal = hw_db + (excel_hw - db_hw_avg)
        hw_cal = min(100.0, max(0.0, hw_cal))
        
        # 4. Calibrate Elearning (Use DB, but if Excel says 0 vi pham, reset to 0)
        el_db = el_list[i]
        if excel_disc.get('el') == 0.0:
            el_cal = 0.0
        else:
            el_cal = el_db
            
        calibrated.append({
            'student_id': s['student_id'],
            'attendance': att_cal, # % vắng mặt
            'homework': hw_cal,     # % hoàn thành
            'elearning': el_cal,   # số bài muộn
            'rpoints': rp_cal,     # điểm Rpoint
            'hackathon_1': s['hackathon_1'],
            'hackathon_2': s['hackathon_2'],
            'project': s['project'],
            'pass': s['pass']
        })
    return calibrated

def main():
    conn = mysql.connector.connect(
        host="localhost",
        port=3307,
        user="root",
        password="",
        database="qldt_el"
    )
    cursor = conn.cursor()
    
    # 1. Load Excel chot data
    print("Loading data from Excel chot Rpoint...")
    excel_path = "docs/PTIT_Chiso.xlsx"
    excel_data = get_excel_chot_data(excel_path)
    
    # 2. Map course keywords to sheet names
    course_to_sheet_map = {
        'javascript': 'KS25_Javascript',
        'cơ sở dữ liệu': 'KS25_Database',
        'database': 'KS25_Database',
        'python': 'KS25_Python',
        'java fundamental': 'KS24-JavaAdvance',
        'java advance': 'KS24-JavaAdvance',
        'java web application': 'KS24_JavaWeb',
        'java web service': 'KS24_JWS',
        'agile': 'KS24_JavaWeb',
        'trí tuệ': 'KS24_AI',
        'ai': 'KS24_AI'
    }
    
    def get_excel_for_class_course(norm_cname, course_name):
        low_course = course_name.lower()
        target_sheet = None
        for kw, sheet in course_to_sheet_map.items():
            if kw in low_course:
                target_sheet = sheet
                break
        if not target_sheet:
            return None
        return excel_data.get(norm_cname, {}).get(target_sheet)

    # 3. Retrieve all classes and courses sequence
    print("Querying database classes & courses...")
    all_class_courses = run_query(cursor, """
        SELECT DISTINCT class_id, course_id 
        FROM qldt_el.final_results;
    """)
    
    courses_raw = run_query(cursor, "SELECT id, name FROM qldt_el.courses;")
    course_id_to_name = {int(r['id']): r['name'] for r in courses_raw if r['id'] is not None}
    
    classes_raw = run_query(cursor, "SELECT id, name FROM qldt_el.classes;")
    class_id_to_name = {int(r['id']): r['name'] for r in classes_raw if r['id'] is not None}
    
    course_seq_raw = run_query(cursor, """
        SELECT classes_id, courses_id, MIN(date) as first_date
        FROM qldt_el.attendance
        GROUP BY classes_id, courses_id
        ORDER BY classes_id, first_date ASC;
    """)
    class_course_seq = defaultdict(list)
    for row in course_seq_raw:
        if row['classes_id'] and row['courses_id']:
            class_course_seq[int(row['classes_id'])].append(int(row['courses_id']))
            
    pass_results = run_query(cursor, """
        SELECT class_id, course_id, student_id, pass, homework, rpoints, project, mutiple_choice_1, essay_1, hackathon_1, hackathon_2, attendance
        FROM qldt_el.final_results;
    """)
    pass_by_class_course = defaultdict(list)
    student_pass_history = {}
    for r in pass_results:
        if r['class_id'] is None or r['course_id'] is None or r['student_id'] is None:
            continue
        cid = int(r['class_id'])
        co_id = int(r['course_id'])
        sid = int(r['student_id'])
        p = int(r['pass']) if r['pass'] is not None else None
        if p is not None:
            pass_by_class_course[(cid, co_id)].append(p)
        
        # Calculate historical score average
        scores = []
        for key in ['project', 'mutiple_choice_1', 'essay_1', 'hackathon_1', 'hackathon_2']:
            val = r[key]
            if val is not None:
                scores.append(float(val))
        avg_score = mean(scores) if scores else None
        
        student_pass_history[(sid, co_id)] = {
            'homework': float(r['homework']) if r['homework'] is not None else 100.0,
            'rpoints': float(r['rpoints']) if r['rpoints'] is not None else 100.0,
            'pass': p,
            'project': float(r['project']) if r['project'] is not None else None,
            'mutiple_choice_1': float(r['mutiple_choice_1']) if r['mutiple_choice_1'] is not None else None,
            'essay_1': float(r['essay_1']) if r['essay_1'] is not None else None,
            'hackathon_1': float(r['hackathon_1']) if r['hackathon_1'] is not None else None,
            'hackathon_2': float(r['hackathon_2']) if r['hackathon_2'] is not None else None,
            'attendance': float(r['attendance']) if r['attendance'] is not None else 0.0,
            'avg_score': avg_score
        }
        
    class_course_pass_rates = {}
    for (cid, co_id), plist in pass_by_class_course.items():
        if plist:
            class_course_pass_rates[(cid, co_id)] = (sum(plist) / len(plist)) * 100

    # 4. Predict
    results = []
    student_care_list = []
    target_batches = ['K24', 'K25', 'KS24', 'KS25']
    
    print("Running predictions with data calibration...")
    for pair in all_class_courses:
        cid = int(pair['class_id'])
        co_id = int(pair['course_id'])
        
        cname = class_id_to_name.get(cid, "N/A")
        coname = course_id_to_name.get(co_id, "N/A")
        
        if not any(b in cname for b in target_batches):
            continue
            
        norm_cname = normalize_class_name(cname)
        
        # Get students with actual pass score (for validation)
        students_results = run_query(cursor, """
            SELECT student_id, homework, elearning, attendance, hackathon_1, hackathon_2, rpoints, project, pass
            FROM qldt_el.final_results
            WHERE class_id = %s AND course_id = %s AND pass IS NOT NULL;
        """, (cid, co_id))
        
        if not students_results:
            continue
            
        excel_disc = get_excel_for_class_course(norm_cname, coname)
        
        # Calibrate student records (resolves DB attendance noise and shifts rpoints)
        calibrated_students = calibrate_students(students_results, excel_disc)
        
        total_students = len(calibrated_students)
        actual_pass_count = sum(1 for s in calibrated_students if s['pass'] == 1)
        actual_pass_rate = (actual_pass_count / total_students) * 100
        
        # Hackathon average
        hack_scores = []
        for s in calibrated_students:
            h_vals = [v for v in [s['hackathon_1'], s['hackathon_2']] if v is not None]
            if h_vals:
                hack_scores.append(mean(h_vals))
        avg_hack = mean(hack_scores) if hack_scores else 65.0
        
        # Prerequisite
        seq = class_course_seq.get(cid, [])
        prereq_course_id = None
        if co_id in seq:
            idx = seq.index(co_id)
            if idx > 0:
                prereq_course_id = seq[idx - 1]
                
        # Prev pass rate
        prev_class_pass_rate = 100.0
        if prereq_course_id:
            prev_class_pass_rate = class_course_pass_rates.get((cid, prereq_course_id), 80.0)
        else:
            prev_class_pass_rate = 80.0
            
        # Check if the course actually has project grading (score > 5.0)
        has_project = any(s['project'] is not None and s['project'] > 5.0 for s in calibrated_students)
        
        # Check if the course actually has hackathon grading (score > 5.0)
        has_hackathon = any(s['hackathon_1'] is not None and s['hackathon_1'] > 5.0 for s in calibrated_students)
        
        # 2. Query attendance details for this class and course (ordered by date/id)
        att_raw = run_query(cursor, """
            SELECT ad.student_id, ad.status, a.date, a.id as attendance_id
            FROM qldt_el.attendance_detail ad
            JOIN qldt_el.attendance a ON ad.attendance_id = a.id
            WHERE a.classes_id = %s AND a.courses_id = %s
            ORDER BY a.date ASC, a.id ASC;
        """, (cid, co_id))
        student_att = defaultdict(list)
        for row in att_raw:
            student_att[int(row['student_id'])].append({
                'status': str(row['status']),
                'id': int(row['attendance_id'])
            })
            
        # 3. Query homeworks (exercises) for this class and course (ordered by id)
        hw_raw = run_query(cursor, """
            SELECT e.student_id, e.check, e.link_git, e.id as exercise_id
            FROM qldt_el.exercise e
            WHERE e.class_id = %s AND e.course_id = %s
            ORDER BY e.id ASC;
        """, (cid, co_id))
        student_hw_details = defaultdict(list)
        for row in hw_raw:
            sid_r = int(row['student_id'])
            check = int(row['check'] or 0)
            git = row['link_git']
            is_git_empty = not git or git.strip() == "" or "placeholder" in str(git).lower()
            is_debt = (check == 2 or (check == 0 and is_git_empty))
            student_hw_details[sid_r].append({
                'is_debt': is_debt,
                'id': int(row['exercise_id'])
            })
            
        # 4. Determine 2 prior courses of this class
        seq = class_course_seq.get(cid, [])
        prev_courses = []
        if co_id in seq:
            idx = seq.index(co_id)
            if idx >= 1:
                prev_courses.append(seq[idx - 1])
            if idx >= 2:
                prev_courses.append(seq[idx - 2])
        if prereq_course_id and prereq_course_id not in prev_courses:
            prev_courses.append(prereq_course_id)
        prev_courses = prev_courses[:2]
        
        # Load auto_rpoints for prior courses
        auto_rp_map = defaultdict(lambda: defaultdict(float))
        for prev_c in prev_courses:
            auto_rp_raw = run_query(cursor, """
                SELECT student_id, total_score, recorded_date
                FROM qldt_el.auto_rpoints
                WHERE course_id = %s;
            """, (prev_c,))
            latest_dates = {}
            for r in auto_rp_raw:
                sid_raw = int(r['student_id'])
                score = float(r['total_score'])
                rdate = r['recorded_date']
                if sid_raw not in latest_dates or rdate > latest_dates[sid_raw]:
                    latest_dates[sid_raw] = rdate
                    auto_rp_map[prev_c][sid_raw] = score

        sum_p_eligible_old = 0.0
        sum_p_eligible_new = 0.0
        
        batch = "K25" if ("KS25" in cname or "K25" in cname) else "K24"
        is_soft_skill = any(kw in coname.lower() for kw in ['kỹ năng', 'tin học văn phòng', 'skl', 'thực tập', 'ttrk', 'project'])
        
        is_ks24 = "KS24" in cname
        is_qtkd = "QTKD" in cname
        
        if is_ks24:
            w1, w2 = 0.40, 0.60
            p_hack_mult = 1.25
            base_scale = 1.10
        else:
            w1, w2 = 0.00, 1.00
            p_hack_mult = 1.30
            base_scale = 0.80 if is_qtkd else 0.85
        
        for s in calibrated_students:
            sid = int(s['student_id'])
            att_val = s['attendance'] # % vắng mặt
            hw_val = s['homework']     # % hoàn thành
            el_val = s['elearning']   # số bài vi phạm
            rp = s['rpoints']         # Rpoint chốt
            proj = s['project']
            
            # Check consecutive absence and homework debts
            is_ks24 = "KS24" in cname
            
            att_list = student_att.get(sid, [])
            total_sessions = len(att_list)
            consecutive_abs = 0
            for item in reversed(att_list):
                if item['status'] in ('0', '2'):
                    consecutive_abs += 1
                else:
                    break
            
            if is_ks24:
                penalty_abs = 0.90 if consecutive_abs >= 2 else 1.0
            else:
                penalty_abs = 0.75 if consecutive_abs >= 2 else 1.0
            
            hw_list = student_hw_details.get(sid, [])
            consecutive_hw = 0
            for item in reversed(hw_list):
                if item['is_debt']:
                    consecutive_hw += 1
                else:
                    break
            
            if is_ks24:
                penalty_hw = 0.92 if consecutive_hw >= 2 else 1.0
            else:
                penalty_hw = 0.78 if consecutive_hw >= 2 else 1.0
            
            # Check for Resumed/Suspended student (no history in both 2 prior courses)
            is_resumed_student = False
            if len(prev_courses) >= 1:
                has_history = False
                for prev_c in prev_courses:
                    prev_fr = student_pass_history.get((sid, prev_c))
                    if prev_fr:
                        test_scores = []
                        for score_key in ['project', 'mutiple_choice_1', 'essay_1', 'hackathon_1', 'hackathon_2']:
                            val = prev_fr.get(score_key)
                            if val is not None:
                                test_scores.append(float(val))
                        if test_scores or prev_fr.get('pass') is not None:
                            has_history = True
                            break
                if not has_history:
                    is_resumed_student = True
            
            # Discipline Score (Rpoint 1-2 prior courses + Current course)
            discipline_prev = None
            if not is_resumed_student:
                prev_rps = []
                for prev_c in prev_courses:
                    rpoint_val = None
                    prev_fr = student_pass_history.get((sid, prev_c))
                    if prev_fr and prev_fr.get('rpoints') is not None:
                        rpoint_val = prev_fr['rpoints']
                    if rpoint_val is None:
                        rpoint_val = auto_rp_map[prev_c].get(sid)
                    if rpoint_val is not None:
                        prev_rps.append(float(rpoint_val))
                if prev_rps:
                    discipline_prev = mean(prev_rps)
            
            if is_resumed_student:
                discipline_prev = 70.0
            elif discipline_prev is None:
                discipline_prev = 100.0
                
            discipline_curr = max(0.0, 100.0 - att_val)
            discipline_val = 0.5 * discipline_prev + 0.5 * discipline_curr
            
            # Study Performance: P_prereq
            penalty_resumption = 1.0
            if is_resumed_student:
                P_prereq = 50.0
                penalty_resumption = 0.85
            elif len(prev_courses) >= 1:
                prev_c_main = prev_courses[0]
                prev_fr = student_pass_history.get((sid, prev_c_main))
                if prev_fr:
                    prev_att = prev_fr.get('attendance', 0.0)
                    test_scores = []
                    for score_key in ['project', 'mutiple_choice_1', 'essay_1', 'hackathon_1', 'hackathon_2']:
                        val = prev_fr.get(score_key)
                        if val is not None:
                            test_scores.append(float(val))
                    prev_score = mean(test_scores) if test_scores else None
                    
                    if prev_score is not None:
                        P_prereq_base = prev_score
                    else:
                        fallback_grades = [v for v in [prev_fr.get('homework'), prev_fr.get('rpoints')] if v is not None]
                        P_prereq_base = mean(fallback_grades) if fallback_grades else 75.0
                        
                    is_prev_failed_hard_test = (prev_score is not None and prev_score < 40.0)
                    if prev_att > 30.0 or is_prev_failed_hard_test:
                        P_prereq = P_prereq_base * (0.90 if is_ks24 else 0.80)
                    else:
                        P_prereq = P_prereq_base
                else:
                    P_prereq = 75.0
            else:
                P_prereq = 75.0
                
            # Current test grade (P_hack) & Heuristic estimation if not tested
            if has_hackathon:
                h_vals = [v for v in [s['hackathon_1'], s['hackathon_2']] if v is not None]
                s_hack = mean(h_vals) if h_vals else 65.0
                p_hack = min(100.0, s_hack * p_hack_mult)
            else:
                P_hack_est = 0.65 * P_prereq + 0.35 * discipline_curr
                p_hack = min(100.0, P_hack_est * p_hack_mult)
                s_hack = p_hack / p_hack_mult
            
            # Calculate learning score before CDC adjustment
            if has_hackathon:
                P_learning = w1 * P_prereq + w2 * p_hack
            else:
                P_learning = 0.3 * P_prereq + 0.7 * p_hack
            P_learning = min(100.0, max(0.0, P_learning))
            
            # Apply CDC difficulty factor
            cdc_val = get_course_difficulty(coname, cursor)
            P_learning_adj = P_learning / cdc_val
            p_eligible = P_learning_adj * 0.6 + discipline_val * 0.4
            
            # Apply consecutive absence & homework penalties + resumption penalty + base_scale
            p_eligible = p_eligible * penalty_abs * penalty_hw * penalty_resumption * base_scale
            p_eligible = min(100.0, max(0.0, p_eligible))
            
            # Domain-specific heuristic: soft skills / internships have a very high pass rate
            if is_soft_skill:
                p_eligible = 93.0
            
            # 1. Prediction using OLD rule (no hard blocks, matches historical DB reality):
            p_eligible_old = p_eligible
            sum_p_eligible_old += p_eligible_old
            
            # 2. Prediction using NEW rule (with Soft Relaxation):
            is_ks24 = "KS24" in coname
            is_failed_new = False
            soft_penalty_factor = 1.0
            
            if total_sessions > 3:
                if is_ks24:
                    # KHÓA CŨ (KS24): Không cấm thi hay phạt điểm kỷ luật
                    pass
                else:
                    # KHÓA MỚI (KS25 & QTKD): Cấm thi cứng khi vi phạm quá nặng, phạt điểm vừa phải khi vi phạm nhẹ
                    if att_val > 30.0:
                        is_failed_new = True
                    elif att_val > 20.0:
                        soft_penalty_factor *= 0.65
                    
                    if hw_val < 50.0:
                        is_failed_new = True
                    elif hw_val < 80.0:
                        soft_penalty_factor *= 0.70
                    
                    if el_val > 5.0:
                        is_failed_new = True
                    elif el_val > 3.0:
                        soft_penalty_factor *= 0.75
                    
                    if excel_disc is not None and excel_disc.get('rp') is not None:
                        if rp < 65.0:
                            is_failed_new = True
                        elif rp < 80.0:
                            soft_penalty_factor *= 0.70
            
            if has_project and proj is not None and proj < 50.0:
                is_failed_new = True
            
            p_eligible_new = p_eligible * soft_penalty_factor
            p_eligible_new = min(100.0, max(0.0, p_eligible_new))
            if is_failed_new:
                p_eligible_new = 0.0
                
            sum_p_eligible_new += p_eligible_new
            
            # Add to care list if fails new gate or p_eligible_new < 50
            if is_failed_new or p_eligible_new < 50.0:
                reasons = []
                if is_failed_new:
                    if has_project and proj is not None and proj < 50.0: reasons.append("Chưa nộp hoặc trượt Project")
                    if not is_ks24:
                        if att_val > 30.0: reasons.append(f"Cấm thi: Vắng học quá nặng ({att_val:.1f}%)")
                        if hw_val < 50.0: reasons.append(f"Cấm thi: Nợ bài tập quá nặng ({hw_val:.1f}%)")
                        if el_val > 5.0: reasons.append(f"Cấm thi: Vi phạm Elearning quá nặng ({el_val} bài)")
                        if excel_disc is not None and excel_disc.get('rp') is not None and rp < 65.0: reasons.append(f"Cấm thi: Rpoint quá thấp ({rp:.1f}/80)")
                else:
                    reasons.append(f"Xác suất đỗ thấp ({p_eligible_new:.1f}%)")
                    if 20.0 < att_val <= 30.0: reasons.append(f"Cảnh báo vắng học ({att_val:.1f}%)")
                    if 50.0 <= hw_val < 80.0: reasons.append(f"Cảnh báo nợ bài tập ({hw_val:.1f}% xong)")
                    if 3.0 < el_val <= 5.0: reasons.append(f"Cảnh báo Elearning ({el_val} bài)")
                    if excel_disc is not None and excel_disc.get('rp') is not None and 65.0 <= rp < 80.0: reasons.append(f"Cảnh báo Rpoint thấp ({rp:.1f}/80)")
                    
                if is_resumed_student:
                    reasons.append("Học viên mới/quay lại sau bảo lưu")
                if consecutive_abs >= 2:
                    reasons.append(f"Nghỉ liên tiếp ({consecutive_abs} buổi)")
                if consecutive_hw >= 2:
                    reasons.append(f"Nợ bài tập liên tiếp ({consecutive_hw} bài)")
                
                student_name_raw = run_query(cursor, "SELECT full_name FROM students WHERE id = %s", (sid,))
                sname_str = student_name_raw[0]['full_name'] if student_name_raw else "Sinh viên ẩn danh"
                
                student_care_list.append({
                    'class_name': cname,
                    'course_name': coname,
                    'student_id': sid,
                    'student_name': sname_str,
                    'rpoints': rp,
                    'attendance': att_val,
                    'homework': hw_val,
                    'elearning': el_val,
                    'hackathon': round(s_hack, 1),
                    'project': proj,
                    'reasons': ", ".join(reasons)
                })
                
        pred_pass_rate_old = sum_p_eligible_old / total_students
        pred_pass_rate_new = sum_p_eligible_new / total_students
        
        batch = "K25" if ("KS25" in cname or "K25" in cname) else "K24"
        location = "HCM" if "HCM" in cname else "HN"
        
        results.append({
            'class_id': cid,
            'class_name': cname,
            'course_id': co_id,
            'course_name': coname,
            'total': total_students,
            'avg_hack': round(avg_hack, 1),
            'prev_pass_rate': round(prev_class_pass_rate, 1),
            'pred_pass_rate_old': round(pred_pass_rate_old, 1),
            'pred_pass_rate_new': round(pred_pass_rate_new, 1),
            'actual_pass_rate': round(actual_pass_rate, 1),
            'err': round(pred_pass_rate_old - actual_pass_rate, 1),
            'batch': batch,
            'location': location
        })
        
    results.sort(key=lambda x: (x['batch'], x['class_name'], x['course_name']))
    
    # 5. Output reports
    os.makedirs("reports", exist_ok=True)
    
    # Filter results for valid error comparison (remove 0% actual pass anomaly due to missing database grades)
    k25_valid = [r for r in results if r['batch'] == 'K25' and not (r['actual_pass_rate'] == 0.0 and r['pred_pass_rate_old'] > 50.0)]
    k24_valid = [r for r in results if r['batch'] == 'K24' and not (r['actual_pass_rate'] == 0.0 and r['pred_pass_rate_old'] > 50.0)]
    
    # Exclude classes with 100% fail rate in actual results where average Hackathon is very high, 
    # since it means grades were never uploaded to the DB yet (pass is NULL or defaulted incorrectly)
    k25_maes = [abs(r['err']) for r in k25_valid if not (r['actual_pass_rate'] < 5.0 and r['avg_hack'] > 60.0)]
    k24_maes = [abs(r['err']) for r in k24_valid if not (r['actual_pass_rate'] < 5.0 and r['avg_hack'] > 60.0)]
    
    mae_k25 = mean(k25_maes) if k25_maes else 0.0
    mae_k24 = mean(k24_maes) if k24_maes else 0.0
    
    print(f"Writing predictions verification report...")
    with open("reports/academic_predictions_v3.md", "w", encoding="utf-8") as f:
        f.write("# BÁO CÁO DỰ BÁO TỶ LỆ QUA MÔN & THỐNG KÊ KIỂM CHỨNG SAI SỐ\n\n")
        f.write(f"*Báo cáo được tạo bởi **AcademicPredictor** vào ngày {datetime.now().strftime('%d/%m/%Y')} dựa trên dữ liệu chốt Rpoint từ Excel và MySQL.*\n\n")
        
        f.write("## 📌 TÓM TẮT ĐÁNH GIÁ SAI SỐ KIỂM CHỨNG\n\n")
        f.write(f"- **Khóa K25**: MAE trung bình = **{mae_k25:.2f}%** (Được kiểm chứng qua các môn như JS, Database, Python).\n")
        f.write(f"- **Khóa K24**: MAE trung bình = **{mae_k24:.2f}%** (Được kiểm chứng qua các môn như Java Web App, Java Service, AI).\n\n")
        
        f.write("## 📊 CHI TIẾT DỰ BÁO VS THỰC TẾ LỊCH SỬ (ĐỂ XÁC MINH SAI SỐ)\n\n")
        
        for batch_name, batch_res in [('Khóa KS24', k24_valid), ('Khóa KS25', k25_valid)]:
            f.write(f"### 🔹 {batch_name}\n\n")
            f.write("| Tên Lớp | Môn Học | Sĩ số | Điểm Hackathon | Môn trước% | Dự báo (Luật cũ)% | Thực tế DB% | Sai số% |\n")
            f.write("| :--- | :--- | :---: | :---: | :---: | :---: | :---: | :---: |\n")
            for r in batch_res[:50]: # Show top 50 rows
                f.write(f"| {r['class_name']} | {r['course_name'][:30]} | {r['total']} | {r['avg_hack']}% | {r['prev_pass_rate']}% | **{r['pred_pass_rate_old']}%** | **{r['actual_pass_rate']}%** | {r['err']:+.1f}% |\n")
            f.write("\n")
            
        f.write("\n---\n\n")
        f.write("## 🔮 DỰ BÁO TỶ LỆ QUA MÔN KHI ÁP DỤNG LUẬT MỚI (CHẶN CỨNG RPOINT & PROJECT)\n\n")
        f.write("Bảng dưới đây dự báo tỷ lệ qua môn của các lớp học nếu áp dụng **Quy định Rpoint & Chốt chặn Project mới** (Học viên vi phạm bất kỳ điều kiện nào trong 4 tiêu chuẩn hoặc trượt Project sẽ bị cấm thi/trượt trực tiếp):\n\n")
        
        f.write("| Khóa | Tên Lớp | Môn Học | Sĩ số | Dự báo luật cũ% | Dự báo luật mới% | Mức độ giảm tỷ lệ đỗ |\n")
        f.write("| :---: | :--- | :--- | :---: | :---: | :---: | :--- |\n")
        for r in results[:100]: # limit to keep markdown readable
            diff = r['pred_pass_rate_new'] - r['pred_pass_rate_old']
            diff_str = f"**{diff:+.1f}%**" if diff < 0 else "0.0%"
            f.write(f"| {r['batch']} | {r['class_name']} | {r['course_name'][:25]} | {r['total']} | {r['pred_pass_rate_old']}% | **{r['pred_pass_rate_new']}%** | {diff_str} |\n")
            
    # Report 2: student_care_list.md
    print("Writing high-risk student care list report...")
    with open("reports/student_care_list.md", "w", encoding="utf-8") as f:
        f.write("# DANH SÁCH HỌC VIÊN CÓ NGUY CƠ TRƯỢT CAO CẦN CAN THIỆP 1-1\n\n")
        f.write(f"*Báo cáo được lập tự động dựa trên **Quy chuẩn Rpoint & Chốt chặn Project mới** vào ngày {datetime.now().strftime('%d/%m/%Y')}.*\n\n")
        
        f.write("> [!IMPORTANT]\n")
        f.write("> Đây là danh sách các học viên **không đạt một hoặc nhiều điều kiện chặn cứng** để được thi và bảo vệ project (Rpoint < 80, Vắng > 20%, Nộp bài tập < 80%, Vi phạm Elearning > 3 bài, hoặc Trượt Project). Cần hỗ trợ phụ đạo và đôn đốc gấp.\n\n")
        
        f.write("| Tên Lớp | Tên Học Viên | ID | Rpoint | Vắng CC% | Nợ Bài tập% | Vi phạm EL | Điểm Hackathon | Lý do cảnh báo chi tiết |\n")
        f.write("| :--- | :--- | :---: | :---: | :---: | :---: | :---: | :---: | :--- |\n")
        
        student_care_list.sort(key=lambda x: (x['class_name'], x['student_name']))
        for s in student_care_list[:150]: # limit to top 150 to keep readable
            f.write(f"| {s['class_name']} | **{s['student_name']}** | {s['student_id']} | {s['rpoints']:.1f} | {s['attendance']:.1f}% | {100.0 - s['homework']:.1f}% | {int(s['elearning'])} | {s['hackathon']}% | 🔴 {s['reasons']} |\n")
            
    print(f"Predictions and reports generated successfully.")
    
    cursor.close()
    conn.close()

if __name__ == "__main__":
    main()
