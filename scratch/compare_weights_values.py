import sys
import os
import openpyxl

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

old_path = r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG.xlsx"
v2_path = r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG (2).xlsx"

wb_old = openpyxl.load_workbook(old_path, data_only=True)
wb_v2 = openpyxl.load_workbook(v2_path, data_only=True)

ws_old = wb_old['Trọng số đánh giá']
ws_v2 = wb_v2['Trọng số đánh giá']

# Read all criteria and weights from OLD
# Cột: STT (A), Tiêu chí (B), Rank 1-2 (C), Rank 3-4 (D), Rank 5-6 (E), Rank 7-8 (F)
weights_old = {}
group_name = ""
for row in ws_old.iter_rows(min_row=4, max_row=29, values_only=True):
    stt, name, r12, r34, r56, r78, _ = row[:7]
    if stt in ('I', 'II', 'III', 'IV', 'V', 'VI'):
        group_name = name
        continue
    if not name or name == 'TỔNG TRỌNG SỐ':
        continue
    # Normalize name for matching
    clean_name = name.strip().replace('\n', ' ')
    weights_old[clean_name] = {
        'group': group_name,
        'r12': r12 or 0.0,
        'r34': r34 or 0.0,
        'r56': r56 or 0.0,
        'r78': r78 or 0.0
    }

# Read all criteria and weights from V2
# Cột: STT (A), Tiêu chí (B), Rank 1-2 (C), Rank 3-4 (D), Rank 5-6 QL (E), Rank 5-6 TH (F), Rank 7-8 (G)
weights_v2 = {}
group_name = ""
for row in ws_v2.iter_rows(min_row=4, max_row=30, values_only=True):
    stt, name, r12, r34, r56_ql, r56_th, r78, _ = row[:8]
    if stt in ('I', 'II', 'III', 'IV', 'V', 'VI'):
        group_name = name
        continue
    if not name or name == 'TỔNG TRỌNG SỐ':
        continue
    clean_name = name.strip().replace('\n', ' ')
    weights_v2[clean_name] = {
        'group': group_name,
        'r12': r12 or 0.0,
        'r34': r34 or 0.0,
        'r56_ql': r56_ql or 0.0,
        'r56_th': r56_th or 0.0,
        'r78': r78 or 0.0
    }

print("=== SO SÁNH TRỌNG SỐ ĐÁNH GIÁ (OLD vs V2) ===")
# Find matching criteria or diffs
all_criteria = sorted(list(set(list(weights_old.keys()) + list(weights_v2.keys()))))
for c in all_criteria:
    old_data = weights_old.get(c)
    v2_data = weights_v2.get(c)
    
    if old_data and v2_data:
        # Check if weights differ
        diff = (
            old_data['r12'] != v2_data['r12'] or
            old_data['r34'] != v2_data['r34'] or
            old_data['r56'] != v2_data['r56_ql'] or  # Assuming OLD r56 maps to V2 r56_ql
            old_data['r78'] != v2_data['r78']
        )
        if diff:
            print(f"Criteria: {c[:80]}")
            print(f"  Group OLD: {old_data['group']} -> V2: {v2_data['group']}")
            print(f"  Rank 1-2: OLD {old_data['r12']} -> V2 {v2_data['r12']}")
            print(f"  Rank 3-4: OLD {old_data['r34']} -> V2 {v2_data['r34']}")
            print(f"  Rank 5-6 QL: OLD {old_data['r56']} -> V2 {v2_data['r56_ql']}")
            print(f"  Rank 5-6 TH (New): V2 {v2_data['r56_th']}")
            print(f"  Rank 7-8: OLD {old_data['r78']} -> V2 {v2_data['r78']}")
            print()
    elif old_data:
        print(f"[-] Removed / Modified in V2: {c[:80]}")
        print(f"  OLD: {old_data}")
        print()
    elif v2_data:
        print(f"[+] Added / Modified in V2: {c[:80]}")
        print(f"  V2: {v2_data}")
        print()

wb_old.close()
wb_v2.close()
