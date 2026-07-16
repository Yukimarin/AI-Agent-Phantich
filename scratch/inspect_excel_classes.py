import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

def main():
    wb = openpyxl.load_workbook("docs/PTIT_Chiso.xlsx", data_only=True)
    print("=== EXCEL SHEETS ===")
    print(wb.sheetnames)
    
    # In ra một số dòng của các sheet KS25
    for sheetname in wb.sheetnames:
        if 'KS25' not in sheetname:
            continue
        sheet = wb[sheetname]
        print(f"\n--- Sheet: {sheetname} ---")
        
        # Đọc 10 tên lớp đầu tiên ở cột B (cột 2)
        classes_found = []
        for r in range(5, min(sheet.max_row + 1, 30)):
            val = sheet.cell(row=r, column=2).value
            if val and val.strip():
                classes_found.append(val.strip())
        print("Classes found:", list(set(classes_found)))

if __name__ == "__main__":
    main()
