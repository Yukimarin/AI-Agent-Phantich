import openpyxl
import sys
from excel_loader import normalize_class_name

sys.stdout.reconfigure(encoding='utf-8')

def main():
    wb = openpyxl.load_workbook("docs/PTIT_Chiso.xlsx", data_only=True)
    
    for sheetname in wb.sheetnames:
        if 'KS24' not in sheetname and 'KS25' not in sheetname:
            continue
        sheet = wb[sheetname]
        max_r = sheet.max_row
        max_c = sheet.max_column
        
        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
        
        for r in range(5, max_r + 1):
            cname = sheet.cell(row=r, column=2).value
            if cname and 'HCM' in str(cname):
                # Print the last few cells
                vals = []
                for c_idx in range(max(0, max_c - 5), max_c):
                    val = sheet.cell(row=r, column=c_idx + 1).value
                    h3 = row3[c_idx]
                    h4 = row4[c_idx]
                    vals.append(f"col_{c_idx}(h3={h3}, h4={h4}): {val}")
                print(f"Sheet: {sheetname} | Class: {cname}")
                print(f"  Last cells: {vals}")

if __name__ == "__main__":
    main()
