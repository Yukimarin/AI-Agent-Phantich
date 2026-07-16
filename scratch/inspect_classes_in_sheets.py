import openpyxl

excel_path = 'docs/PTIT_Chiso.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

sheets_to_check = ['KS24_AI', 'KS25_Python_Web', 'KS25_QTKD_DTB202', 'KS25_QTKD_PRJ302']

for sname in sheets_to_check:
    if sname not in wb.sheetnames:
        print(f"Sheet {sname} not found")
        continue
    sheet = wb[sname]
    print(f"\n--- Sheet: {sname} ---")
    classes = set()
    for r in range(5, sheet.max_row + 1):
        cval = sheet.cell(row=r, column=2).value
        if cval:
            classes.add(str(cval).strip())
    print("Classes found:", sorted(list(classes)))
