import mysql.connector
import sys
import openpyxl
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

def mean(lst):
    return sum(lst) / len(lst) if lst else 0.0

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = (name_str
                .replace("KS25", "K25")
                .replace("KS24", "K24")
                .replace("KS23", "K23"))
    return name_str

course_to_sheet_map = {
    'javascript': 'KS25_Javascript',
    'cơ sở dữ liệu': 'KS25_Database',
    'database': 'KS25_Database',
    'python': 'KS25_Python',
    'java fundamental': 'KS24-JavaAdvance',
    'java advance': 'KS24-JavaAdvance',
    'java web application': 'KS24_JavaWeb',
    'java web service': 'KS24_JWS',
    'agile': 'KS24_JavaWeb',
    'trí tuệ': 'KS24_AI',
    'ai': 'KS24_AI'
}

def get_excel_chot_data(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    class_data = defaultdict(dict)
    for sheetname in wb.sheetnames:
        if sheetname == 'Sheet1' or 'SKL' in sheetname:
            continue
        sheet = wb[sheetname]
        max_r = sheet.max_row
        max_c = sheet.max_column
        if max_r < 5:
            continue
        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
        dates_list = []
        for c_idx in range(3, max_c):
            val3 = row3[c_idx]
            val4 = row4[c_idx]
            if val3:
                dates_list.append((c_idx, val3, val4))
            elif dates_list:
                dates_list.append((c_idx, dates_list[-1][1], val4))
        
        # Find Rpoint col
        rp_col_idx = None
        for c_idx in range(max_c - 1, 2, -1):
            vals = []
            for r in range(5, max_r + 1):
                val = sheet.cell(row=r, column=c_idx + 1).value
                if val is not None:
                    try: vals.append(float(val))
                    except: pass
            if len(vals) >= 2:
                avg_val = mean(vals)
                if 30.0 <= avg_val <= 115.0:
                    h3 = row3[c_idx]
                    if not h3:
                        rp_col_idx = c_idx
                        break
        
        # Read class rows
        for r in range(5, max_r + 1):
            cname = sheet.cell(row=r, column=2).value
            if cname:
                norm_name = normalize_class_name(cname)
                # Find CC, BT, EL last values
                cc_val, bt_val, el_val = None, None, None
                for c_idx, d, sub in reversed(dates_list):
                    val = sheet.cell(row=r, column=c_idx + 1).value
                    if val is not None:
                        try:
                            val_f = float(val)
                            if sub == 'Chuyên cần' and cc_val is None: cc_val = val_f
                            elif sub == 'Bài tập' and bt_val is None: bt_val = val_f
                            elif sub == 'Elearning' and el_val is None: el_val = val_f
                        except: pass
                
                rp_val = None
                if rp_col_idx is not None:
                    val = sheet.cell(row=r, column=rp_col_idx + 1).value
                    if val is not None:
                        try: rp_val = float(val)
                        except: pass
                        
                class_data[norm_name][sheetname] = {
                    'cc': cc_val,
                    'bt': bt_val,
                    'el': el_val,
                    'rp': rp_val
                }
    return class_data

def main():
    conn = mysql.connector.connect(
        host="localhost",
        port=3307,
        user="root",
        password="",
        database="qldt_el"
    )
    cursor = conn.cursor(dictionary=True)
    
    # Load metadata
    cursor.execute("SELECT id, name FROM classes;")
    classes = cursor.fetchall()
    class_id_to_name = {c['id']: c['name'] for c in classes}
    
    cursor.execute("SELECT id, name FROM courses;")
    courses = cursor.fetchall()
    course_id_to_name = {c['id']: c['name'] for c in courses}
    
    cursor.execute("SELECT DISTINCT class_id, course_id FROM final_results;")
    all_class_courses = cursor.fetchall()
    
    # Load excel data
    excel_data = get_excel_chot_data("docs/PTIT_Chiso.xlsx")
    
    # Class course sequence
    class_course_seq = defaultdict(list)
    cursor.execute("""
        SELECT classes_id, courses_id, MIN(date) as first_date
        FROM attendance
        GROUP BY classes_id, courses_id
        ORDER BY classes_id, first_date;
    """)
    for row in cursor.fetchall():
        if row['classes_id'] and row['courses_id']:
            class_course_seq[int(row['classes_id'])].append(int(row['courses_id']))
        
    print("=== HN-KS24-CNTT4 DETAILED RESULTS ===")
    for pair in all_class_courses:
        cid = int(pair['class_id'])
        co_id = int(pair['course_id'])
        
        cname = class_id_to_name.get(cid, "N/A")
        coname = course_id_to_name.get(co_id, "N/A")
        
        if 'HN-KS24-CNTT4' not in cname:
            continue
            
        norm_cname = normalize_class_name(cname)
        
        cursor.execute("""
            SELECT student_id, homework, elearning, attendance, hackathon_1, hackathon_2, rpoints, project, pass
            FROM final_results
            WHERE class_id = %s AND course_id = %s AND pass IS NOT NULL;
        """, (cid, co_id))
        students_results = cursor.fetchall()
        
        if not students_results:
            continue
            
        # Get Excel data
        low_course = coname.lower()
        target_sheet = None
        for kw, sheet in course_to_sheet_map.items():
            if kw in low_course:
                target_sheet = sheet
                break
        excel_disc = excel_data.get(norm_cname, {}).get(target_sheet) if target_sheet else None
        
        # Prerequisite
        seq = class_course_seq.get(cid, [])
        prereq_course_id = None
        if co_id in seq:
            idx = seq.index(co_id)
            if idx > 0:
                prereq_course_id = seq[idx - 1]
                
        # Query prerequisite pass rate
        prev_class_pass_rate = 80.0
        if prereq_course_id:
            cursor.execute("""
                SELECT COUNT(*) as total, SUM(CASE WHEN pass = 1 THEN 1 ELSE 0 END) as passed
                FROM final_results
                WHERE class_id = %s AND course_id = %s AND pass IS NOT NULL;
            """, (cid, prereq_course_id))
            prev_res = cursor.fetchone()
            if prev_res and prev_res['total'] > 0:
                prev_class_pass_rate = (prev_res['passed'] / prev_res['total']) * 100
                
        # Calculate rates
        total_students = len(students_results)
        actual_pass_count = sum(1 for s in students_results if s['pass'] == 1)
        actual_pass_rate = (actual_pass_count / total_students) * 100
        
        # Calculate avg hackathon
        hack_scores = []
        for s in students_results:
            h_vals = [v for v in [s['hackathon_1'], s['hackathon_2']] if v is not None]
            if h_vals:
                hack_scores.append(mean(h_vals))
        avg_hack = mean(hack_scores) if hack_scores else 65.0
        
        print(f"Course: {coname[:25]} (ID {co_id}) | Sĩ số: {total_students} | Excel mapping found? {excel_disc is not None}")
        print(f"  Avg Hack: {avg_hack:.1f}% | Prereq Pass: {prev_class_pass_rate:.1f}% | Actual Pass: {actual_pass_rate:.1f}%")

    cursor.close()
    conn.close()

if __name__ == "__main__":
    main()
