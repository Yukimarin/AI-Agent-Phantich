import sys
import os
import openpyxl
import json
from collections import defaultdict

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

# Constants & Paths
EXCEL_PATH = r"data/PTIT_Chiso.xlsx"
DAILY_LOG_JSON_PATH = r"data/daily_log_analysis.json"
PRED_JSON_PATH = r"scratch/predictions_cv_data.json"
VIOLATIONS_JSON_PATH = r"data/vi_pham_gvtg.json"

# Whitelist Staff of CNTT & QTKD
WHITELIST_STAFF = {
    # A. Khối CNTT-KS24-HN
    "hồ xuân hùng": {"name": "Hồ Xuân Hùng", "group": "Khối CNTT-KS24-HN", "role": "Quản lý", "rank": "5"},
    "bùi thanh hải": {"name": "Bùi Thanh Hải", "group": "Khối CNTT-KS24-HN", "role": "Giảng viên", "rank": "3"},
    "mai xuân chinh": {"name": "Mai Xuân Chinh", "group": "Khối CNTT-KS24-HN", "role": "Giảng viên", "rank": "3"},
    "đinh thành nam": {"name": "Đinh Thành Nam", "group": "Khối CNTT-KS24-HN", "role": "Giảng viên", "rank": "3"},
    "nguyễn công hưởng": {"name": "Nguyễn Công Hưởng", "group": "Khối CNTT-KS24-HN", "role": "Giảng viên", "rank": "3"},
    "phạm tuấn bình": {"name": "Phạm Tuấn Bình", "group": "Khối CNTT-KS24-HN", "role": "Giảng viên", "rank": "3"},

    # B. Khối CNTT-KS25-HN
    "trịnh quốc hai": {"name": "Trịnh Quốc Hai", "group": "Khối CNTT-KS25-HN", "role": "Quản lý", "rank": "5"},
    "lương quốc tuấn": {"name": "Lương Quốc Tuấn", "group": "Khối CNTT-KS25-HN", "role": "Giảng viên", "rank": "3"},
    "nguyễn quảng an": {"name": "Nguyễn Quảng An", "group": "Khối CNTT-KS25-HN", "role": "Giảng viên", "rank": "3"},
    "lại trung lâm": {"name": "Lại Trung Lâm", "group": "Khối CNTT-KS25-HN", "role": "Giảng viên", "rank": "3"},
    "phạm ngọc kiên": {"name": "Phạm Ngọc Kiên", "group": "Khối CNTT-KS25-HN", "role": "Giảng viên", "rank": "3"},
    "ngọ văn quý": {"name": "Ngọ Văn Quý", "group": "Khối CNTT-KS25-HN", "role": "Giảng viên", "rank": "3"},
    "lâm tùng dương": {"name": "Lâm Tùng Dương", "group": "Khối CNTT-KS25-HN", "role": "Giảng viên", "rank": "3"},
    "trần minh cường": {"name": "Trần Minh Cường", "group": "Khối CNTT-KS25-HN", "role": "Quản lý", "rank": "5"},

    # C. CNTT-HCM
    "nguyễn bá minh đạo": {"name": "Nguyễn Bá Minh Đạo", "group": "Khối CNTT-HCM", "role": "Quản lý", "rank": "5"},
    "trần quốc tuấn": {"name": "Trần Quốc Tuấn", "group": "Khối CNTT-HCM", "role": "Giảng viên", "rank": "3"},
    "lê hà thanh sang": {"name": "Lê Hà Thanh Sang", "group": "Khối CNTT-HCM", "role": "Giảng viên", "rank": "3"},
    "phạm viết hùng": {"name": "Phạm Viết Hùng", "group": "Khối CNTT-HCM", "role": "Trợ giảng", "rank": "2"}, # Rank 2 as requested / TA
    "lưu xuân hoàng nguyên": {"name": "Lưu Xuân Hoàng Nguyên", "group": "Khối CNTT-HCM", "role": "Giảng viên", "rank": "3"},
    "nguyễn đức minh": {"name": "Nguyễn Đức Minh", "group": "Khối CNTT-HCM", "role": "Giảng viên", "rank": "3"},
    "phan ngọc tài": {"name": "Phan Ngọc Tài", "group": "Khối CNTT-HCM", "role": "Giảng viên", "rank": "3"},
    "đặng minh luân": {"name": "Đặng Minh Luân", "group": "Khối CNTT-HCM", "role": "Giảng viên", "rank": "3"},
    "nguyễn ngọc sơn": {"name": "Nguyễn Ngọc Sơn", "group": "Khối CNTT-HCM", "role": "Giảng viên", "rank": "3"},

    # D. QTKD - HN
    "hoàng thị kim oanh": {"name": "Hoàng Thị Kim Oanh", "group": "Khối QTKD-HN", "role": "Quản lý", "rank": "5"},
    "lê thành ngọc": {"name": "Lê Thành Ngọc", "group": "Khối QTKD-HN", "role": "Giảng viên", "rank": "3"}, # Set rank 3 as requested
    "hoàng thị hậu": {"name": "Hoàng Thị Hậu", "group": "Khối QTKD-HN", "role": "Giảng viên", "rank": "3"},
    "đặng quỳnh trang": {"name": "Đặng Quỳnh Trang", "group": "Khối QTKD-HN", "role": "Giảng viên", "rank": "3"},
    "nguyễn ngọc vân khanh": {"name": "Nguyễn Ngọc Vân Khanh", "group": "Khối QTKD-HN", "role": "Giảng viên", "rank": "3"},
    "nguyễn thị hồng minh": {"name": "Nguyễn Thị Hồng Minh", "group": "Khối QTKD-HN", "role": "Giảng viên", "rank": "3"},
    "triệu thị thanh tâm": {"name": "Triệu Thị Thanh Tâm", "group": "Khối QTKD-HN", "role": "Trợ giảng", "rank": "1"},
    "nguyễn thị như quỳnh": {"name": "Nguyễn Thị Như Quỳnh", "group": "Khối QTKD-HN", "role": "Trợ giảng", "rank": "1"},

    # E. QTKD - HCM
    "lê nhựt mi": {"name": "Lê Nhựt Mi", "group": "Khối QTKD-HCM", "role": "Giảng viên", "rank": "3"},
    "lê thị bảo yến": {"name": "Lê Thị Bảo Yến", "group": "Khối QTKD-HCM", "role": "Giảng viên", "rank": "3"}
}

# Target Staff names for matching
TARGETS = [
    "Lương Quốc Tuấn",
    "Lâm Tùng Dương",
    "Trần Quốc Tuấn",
    "Phạm Viết Hùng",
    "Nguyễn Ngọc Vân Khanh",
    "Lê Thành Ngọc"
]

# New weights from V2 Excel standard
WEIGHTS_V2 = {
    'TA': { # Rank 1-2
        'tuanchu': 0.10, 'kyluat': 0.10, 'hoclieu': 0.10, 'phoihop': 0.05,
        'siso': 0.15, 'el': 0.10, 'bt': 0.05,
        'giaiquyet': 0.05, 'pass_rate': 0.05, 'kha_gioi': 0.05,
        'cc_cn': 0.02, 'bang_cap': 0.03,
        'truyendat_dugio': 0.0, 'truyendat_uprank': 0.0,
        'csat': 0.05, 'hoclieu_duyet': 0.10,
        'toiuu_ct': 0.0, 'hieusuat_duoi': 0.0, 'thuonghieu': 0.0, 'sangkien': 0.0
    },
    'GV': { # Rank 3-4
        'tuanchu': 0.10, 'kyluat': 0.05, 'hoclieu': 0.05, 'phoihop': 0.05,
        'siso': 0.10, 'el': 0.05, 'bt': 0.05,
        'giaiquyet': 0.10, 'pass_rate': 0.05, 'kha_gioi': 0.05,
        'cc_cn': 0.02, 'bang_cap': 0.03,
        'truyendat_dugio': 0.10, 'truyendat_uprank': 0.10,
        'csat': 0.05, 'hoclieu_duyet': 0.05,
        'toiuu_ct': 0.0, 'hieusuat_duoi': 0.0, 'thuonghieu': 0.0, 'sangkien': 0.0
    },
    'QL': { # Rank 5-6 (Quản lý)
        'tuanchu': 0.05, 'kyluat': 0.02, 'hoclieu': 0.03, 'phoihop': 0.05,
        'siso': 0.05, 'el': 0.03, 'bt': 0.02,
        'giaiquyet': 0.10, 'pass_rate': 0.03, 'kha_gioi': 0.02,
        'cc_cn': 0.02, 'bang_cap': 0.03,
        'truyendat_dugio': 0.02, 'truyendat_uprank': 0.05,
        'csat': 0.05, 'hoclieu_duyet': 0.03,
        'toiuu_ct': 0.15, 'hieusuat_duoi': 0.10, 'thuonghieu': 0.10, 'sangkien': 0.05 # Double-checked V2 rows
    },
    'TH': { # Rank 5-6 (GV Thương hiệu)
        'tuanchu': 0.05, 'kyluat': 0.05, 'hoclieu': 0.03, 'phoihop': 0.02,
        'siso': 0.05, 'el': 0.03, 'bt': 0.02,
        'giaiquyet': 0.10, 'pass_rate': 0.05, 'kha_gioi': 0.05,
        'cc_cn': 0.02, 'bang_cap': 0.03,
        'truyendat_dugio': 0.10, 'truyendat_uprank': 0.10,
        'csat': 0.05, 'hoclieu_duyet': 0.05,
        'toiuu_ct': 0.20, 'hieusuat_duoi': 0.0, 'thuonghieu': 0.0, 'sangkien': 0.0
    },
    'Lead': { # Rank 7-8
        'tuanchu': 0.02, 'kyluat': 0.01, 'hoclieu': 0.02, 'phoihop': 0.01,
        'siso': 0.02, 'el': 0.01, 'bt': 0.01,
        'giaiquyet': 0.05, 'pass_rate': 0.02, 'kha_gioi': 0.01,
        'cc_cn': 0.01, 'bang_cap': 0.01,
        'truyendat_dugio': 0.0, 'truyendat_uprank': 0.05,
        'csat': 0.03, 'hoclieu_duyet': 0.02,
        'toiuu_ct': 0.10, 'hieusuat_duoi': 0.25, 'thuonghieu': 0.15, 'sangkien': 0.20
    }
}

# Clean helper
def clean_instructor_name(name):
    if not name:
        return ""
    name_clean = name.strip()
    special_mappings = {
        "lưu hoàng xuân nguyên": "Lưu Xuân Hoàng Nguyên",
        "xuân nguyên": "Lưu Xuân Hoàng Nguyên",
        "lưu xuân hoàng nguyên": "Lưu Xuân Hoàng Nguyên",
        "nguyễn huyền trang": "Nguyễn Thị Huyền Trang",
        "nguyễn thị huyền trang": "Nguyễn Thị Huyền Trang"
    }
    if name_clean.lower() in special_mappings:
        return special_mappings[name_clean.lower()]
    return name_clean

# Helper to normalize class names
def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

def parse_date(d_val):
    from datetime import datetime, date
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

# Load Daily Logs
daily_log_data = {}
if os.path.exists(DAILY_LOG_JSON_PATH):
    with open(DAILY_LOG_JSON_PATH, "r", encoding="utf-8") as f:
        daily_log_data = json.load(f).get("monthly_stats", {})

# Load Predictions
predictions_data = {}
if os.path.exists(PRED_JSON_PATH):
    with open(PRED_JSON_PATH, "r", encoding="utf-8") as f:
        p_data = json.load(f)
        dashboard_data = p_data.get('dashboard_data', {})
        for batch_key, batch_val in dashboard_data.items():
            for c in batch_val.get('cv', []):
                cname = c.get('class_name')
                predictions_data[cname] = c.get('actual_pass', 100.0)
            for c in batch_val.get('curr', []):
                cname = c.get('class_name')
                predictions_data[cname] = c.get('pred_new', 100.0)

# Load Violations
violations_count = defaultdict(int)
if os.path.exists(VIOLATIONS_JSON_PATH):
    with open(VIOLATIONS_JSON_PATH, "r", encoding="utf-8") as f:
        v_list = json.load(f)
        for v in v_list:
            inst = v.get("Instructor")
            if inst:
                violations_count[inst.strip().lower()] += 1

# Read Excel (Correct TA mapping)
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
target_sheets = [
    'KS24-JavaAdvance', 'KS24_JavaWeb', 'KS24_JWS', 'KS24_AI',
    'KS25_Javascript', 'KS25_Database', 'KS25_Python', 'KS25_Python_Web',
    'KS25_QTKD_M103', 'KS25_QTKD_M104', 'KS25_QTKD_DTB201', 'KS25_QTKD_DTB202',
    'KS25_QTKD_PRJ302'
]

excel_metrics = defaultdict(lambda: {
    'classes': set(),
    'cc_vals': [],
    'bt_vals': [],
    'el_vals': []
})

# Helper to normalize search whitelist names
whitelist_names_lower = {k.strip().lower(): v for k, v in WHITELIST_STAFF.items()}

for sheetname in target_sheets:
    if sheetname not in wb.sheetnames:
        continue
    sheet = wb[sheetname]
    max_r = sheet.max_row
    max_c = sheet.max_column
    if max_r < 5 or max_c < 4:
        continue
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]

    col_info = []
    current_date = None
    for c_idx in range(3, max_c):
        if c_idx < len(row3) and row3[c_idx]:
            parsed = parse_date(row3[c_idx])
            if parsed:
                current_date = parsed
        subheader = row4[c_idx] if c_idx < len(row4) else None
        if current_date and subheader in ('Chuyên cần', 'Bài tập', 'Elearning'):
            col_info.append((c_idx, current_date, subheader))

    for r in range(5, max_r + 1):
        cname_raw = sheet.cell(row=r, column=2).value
        teacher_raw = sheet.cell(row=r, column=3).value
        if not cname_raw:
            continue
        norm_class = normalize_class_name(cname_raw)
        if "l01" in norm_class.lower() or "l02" in norm_class.lower():
            continue

        ta_raw = None
        if r + 1 <= max_r:
            next_cname = sheet.cell(row=r + 1, column=2).value
            if not next_cname:
                ta_raw = sheet.cell(row=r + 1, column=3).value

        teacher_name = clean_instructor_name(teacher_raw).strip()
        ta_name = clean_instructor_name(ta_raw).strip() if ta_raw else ""

        # Check in Whitelist
        t_key = teacher_name.lower()
        ta_key = ta_name.lower()

        is_t_valid = t_key in whitelist_names_lower
        is_ta_valid = ta_key in whitelist_names_lower if ta_name else False

        # Read class metrics
        cc, bt, el = [], [], []
        for c_idx, d, sub in col_info:
            val = sheet.cell(row=r, column=c_idx + 1).value
            if isinstance(val, (int, float)):
                val_f = float(val)
                if sub == 'Chuyên cần': cc.append(val_f)
                elif sub == 'Bài tập': bt.append(val_f)
                elif sub == 'Elearning': el.append(val_f)

        if is_t_valid:
            canon_name = whitelist_names_lower[t_key]["name"]
            excel_metrics[canon_name]['classes'].add(f"{norm_class} ({sheetname})")
            excel_metrics[canon_name]['cc_vals'].extend(cc)
            excel_metrics[canon_name]['bt_vals'].extend(bt)
            excel_metrics[canon_name]['el_vals'].extend(el)

        if is_ta_valid:
            canon_name_ta = whitelist_names_lower[ta_key]["name"]
            excel_metrics[canon_name_ta]['classes'].add(f"{norm_class} ({sheetname})")
            # CRITICAL FIX: Assistant (TA) metrics are read from the teacher row 'r', not assistant row 'r+1' which is empty
            excel_metrics[canon_name_ta]['cc_vals'].extend(cc)
            excel_metrics[canon_name_ta]['bt_vals'].extend(bt)
            excel_metrics[canon_name_ta]['el_vals'].extend(el)

wb.close()

# Evaluate everyone in whitelist
evaluation_results = []

for name_lower, info in whitelist_names_lower.items():
    display_name = info["name"]
    group = info["group"]
    role = info["role"]
    rank_str = info["rank"]
    
    log_key = name_lower
    if name_lower == "nguyễn thị huyền trang":
        log_key = "nguyễn huyền trang"
    log_info = daily_log_data.get(log_key, {})
    
    try:
        rank_val = int(rank_str)
    except ValueError:
        rank_val = 3
        
    # Determine rank category weights
    if rank_val in (1, 2):
        w_cat = 'TA'
    elif rank_val in (3, 4):
        w_cat = 'GV'
    elif rank_val in (5, 6):
        # We classify as 'QL' (Quản lý) by default for Rank 5-6, but can customize if needed
        # Let's use 'QL' weights
        w_cat = 'QL'
    else:
        w_cat = 'Lead'
        
    weights = WEIGHTS_V2[w_cat]
    
    # Excel data
    ex_data = excel_metrics.get(display_name, {})
    classes_list = list(ex_data.get('classes', []))
    classes_str = ", ".join(classes_list) if classes_list else "Không phụ trách lớp Excel"
    
    cc_vals = ex_data.get('cc_vals', [])
    bt_vals = ex_data.get('bt_vals', [])
    el_vals = ex_data.get('el_vals', [])
    
    avg_cc = sum(cc_vals) / len(cc_vals) if cc_vals else 0.0
    avg_bt = sum(bt_vals) / len(bt_vals) if bt_vals else 0.0
    avg_el = sum(el_vals) / len(el_vals) if el_vals else 0.0
    
    student_cc_rate = 100.0 - avg_cc
    student_bt_rate = 100.0 - avg_bt
    student_el_rate = 100.0 - avg_el
    
    if not classes_list:
        student_cc_rate = 95.0
        student_bt_rate = 92.0
        student_el_rate = 90.0
        
    # Pass rate
    pass_rates = []
    for cls in classes_list:
        norm_cls = normalize_class_name(cls)
        pr = predictions_data.get(norm_cls)
        if pr is not None:
            pass_rates.append(pr)
    student_pass_rate = sum(pass_rates) / len(pass_rates) if pass_rates else 90.0
    if not classes_list:
        student_pass_rate = 95.0
        
    # -------------------------------------------------------------
    # SCORING RUBRICS
    # -------------------------------------------------------------
    # 1. Tuân thủ (Báo cáo ngày)
    missing_days = log_info.get("missing_days", [])
    time_violations = log_info.get("time_violations", [])
    compliance_violations_count = len(missing_days) + len(time_violations)
    
    if compliance_violations_count == 0:
        score_tuanchu = 10.0
    elif compliance_violations_count <= 2:
        score_tuanchu = 8.0
    elif compliance_violations_count <= 4:
        score_tuanchu = 5.0
    elif compliance_violations_count <= 6:
        score_tuanchu = 3.0
    else:
        score_tuanchu = 1.0
        
    # 2. Kỷ luật tác nghiệp (from vi_pham_gvtg.json)
    op_viols = violations_count.get(display_name.lower(), 0)
    if op_viols <= 1:
        score_kyluat = 10.0
    elif op_viols <= 2:
        score_kyluat = 8.0
    elif op_viols <= 3:
        score_kyluat = 5.0
    elif op_viols <= 4:
        score_kyluat = 3.0
    elif op_viols <= 5:
        score_kyluat = 1.0
    else:
        score_kyluat = 0.0
        
    # 3. Tiến độ học liệu
    uncompleted_tasks = log_info.get("uncompleted_tasks", [])
    material_violations_count = len(uncompleted_tasks)
    
    if material_violations_count == 0:
        score_hoclieu = 10.0
    elif material_violations_count <= 2:
        score_hoclieu = 8.0
    elif material_violations_count <= 4:
        score_hoclieu = 5.0
    elif material_violations_count <= 6:
        score_hoclieu = 3.0
    else:
        score_hoclieu = 1.0
        
    # 4. Phối hợp nội bộ
    score_phoihop = 5.0
    
    # 5. Sĩ số chuyên cần
    if student_cc_rate >= 90.0:
        score_siso = 10.0
    elif student_cc_rate >= 80.0:
        score_siso = 8.0
    elif student_cc_rate >= 70.0:
        score_siso = 5.0
    elif student_cc_rate >= 60.0:
        score_siso = 3.0
    else:
        score_siso = 1.0
        
    # 6. Đôn đốc E-learning
    if student_el_rate >= 90.0:
        score_el = 10.0
    elif student_el_rate >= 80.0:
        score_el = 8.0
    elif student_el_rate >= 70.0:
        score_el = 5.0
    elif student_el_rate >= 60.0:
        score_el = 3.0
    else:
        score_el = 1.0
        
    # 7. Đôn đốc BTVN
    if student_bt_rate >= 90.0:
        score_bt = 10.0
    elif student_bt_rate >= 80.0:
        score_bt = 8.0
    elif student_bt_rate >= 70.0:
        score_bt = 5.0
    elif student_bt_rate >= 60.0:
        score_bt = 3.0
    else:
        score_bt = 1.0
        
    # 8. Giải quyết vấn đề
    score_giaiquyet = 5.0
    
    # 9. Tỷ lệ pass
    if student_pass_rate >= 80.0:
        score_pass_rate = 10.0
    elif student_pass_rate >= 70.0:
        score_pass_rate = 8.0
    elif student_pass_rate >= 60.0:
        score_pass_rate = 5.0
    elif student_pass_rate >= 50.0:
        score_pass_rate = 3.0
    else:
        score_pass_rate = 1.0
        
    # Các tiêu chí khác mặc định 5.0 (cho Phương án B)
    score_kha_gioi = 5.0
    score_cc_cn = 5.0
    score_bang_cap = 5.0
    score_truyendat_dugio = 5.0
    score_truyendat_uprank = 5.0
    score_csat = 5.0
    score_hoclieu_duyet = 5.0
    score_toiuu_ct = 5.0
    score_hieusuat_duoi = 5.0
    score_thuonghieu = 5.0
    score_sangkien = 5.0
    
    # -------------------------------------------------------------
    # CALCULATE SCORES
    # -------------------------------------------------------------
    # Option A: Scale based on criteria with actual data
    # Criteria with data: tuanchu, kyluat, hoclieu, siso, el, bt, pass_rate
    w_tuanchu = weights.get('tuanchu', 0.0)
    w_kyluat = weights.get('kyluat', 0.0)
    w_hoclieu = weights.get('hoclieu', 0.0)
    w_siso = weights.get('siso', 0.0)
    w_el = weights.get('el', 0.0)
    w_bt = weights.get('bt', 0.0)
    w_pass_rate = weights.get('pass_rate', 0.0)
    
    sum_w_real = w_tuanchu + w_kyluat + w_hoclieu + w_siso + w_el + w_bt + w_pass_rate
    weighted_score_real = (
        score_tuanchu * w_tuanchu +
        score_kyluat * w_kyluat +
        score_hoclieu * w_hoclieu +
        score_siso * w_siso +
        score_el * w_el +
        score_bt * w_bt +
        score_pass_rate * w_pass_rate
    )
    overall_score_A = weighted_score_real / sum_w_real if sum_w_real > 0 else 5.0
    
    # Option B: Complete weights with defaults (5.0) inserted
    weighted_score_all = (
        score_tuanchu * weights.get('tuanchu', 0.0) +
        score_kyluat * weights.get('kyluat', 0.0) +
        score_hoclieu * weights.get('hoclieu', 0.0) +
        score_phoihop * weights.get('phoihop', 0.0) +
        score_siso * weights.get('siso', 0.0) +
        score_el * weights.get('el', 0.0) +
        score_bt * weights.get('bt', 0.0) +
        score_giaiquyet * weights.get('giaiquyet', 0.0) +
        score_pass_rate * weights.get('pass_rate', 0.0) +
        score_kha_gioi * weights.get('kha_gioi', 0.0) +
        score_cc_cn * weights.get('cc_cn', 0.0) +
        score_bang_cap * weights.get('bang_cap', 0.0) +
        score_truyendat_dugio * weights.get('truyendat_dugio', 0.0) +
        score_truyendat_uprank * weights.get('truyendat_uprank', 0.0) +
        score_csat * weights.get('csat', 0.0) +
        score_hoclieu_duyet * weights.get('hoclieu_duyet', 0.0) +
        score_toiuu_ct * weights.get('toiuu_ct', 0.0) +
        score_hieusuat_duoi * weights.get('hieusuat_duoi', 0.0) +
        score_thuonghieu * weights.get('thuonghieu', 0.0) +
        score_sangkien * weights.get('sangkien', 0.0)
    )
    overall_score_B = weighted_score_all
    
    # Classification
    if overall_score_A >= 10.0:
        classification = "Vượt mức"
    elif overall_score_A >= 7.5:
        classification = "Đạt"
    elif overall_score_A >= 5.0:
        classification = "Cần cố gắng"
    else:
        classification = "Không đạt kỳ vọng"
        
    evaluation_results.append({
        'name': display_name,
        'group': group,
        'role': role,
        'rank': rank_str,
        'rank_cat': w_cat,
        'classes': classes_str,
        'raw_metrics': {
            'cc_rate': student_cc_rate,
            'bt_rate': student_bt_rate,
            'el_rate': student_el_rate,
            'pass_rate': student_pass_rate,
            'compliance_violations': compliance_violations_count,
            'op_violations': op_viols,
            'material_violations': material_violations_count
        },
        'scores': {
            'tuanchu': score_tuanchu,
            'kyluat': score_kyluat,
            'hoclieu': score_hoclieu,
            'siso': score_siso,
            'el': score_el,
            'bt': score_bt,
            'pass_rate': score_pass_rate
        },
        'overall_score_A': overall_score_A,
        'overall_score_B': overall_score_B,
        'classification': classification
    })

# Write results to Markdown file in UTF-8
report_path = "scratch/target_report.md"
with open(report_path, "w", encoding="utf-8") as md:
    md.write("# BÁO CÁO CHI TIẾT ĐÁNH GIÁ NĂNG LỰC THEO TIÊU CHUẨN MỚI V2\n\n")
    md.write("> [!NOTE]\n")
    md.write("> Báo cáo này áp dụng bảng trọng số và cấu trúc tiêu chuẩn mới nhất của tệp **[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG (2).xlsx**.\n")
    md.write("> Đã hiệu chỉnh logic đọc chỉ số của Trợ giảng (TA) đồng bộ từ dòng Giảng viên chính của lớp học.\n\n")

    md.write("## 1. Chi tiết điểm thành phần của các nhân sự được yêu cầu\n\n")
    
    target_names = [t.strip().lower() for t in TARGETS]
    for r in evaluation_results:
        if r['name'].strip().lower() in target_names:
            md.write(f"### 👤 Họ và tên: {r['name']}\n")
            md.write(f"- **Khối phòng ban**: {r['group']}\n")
            md.write(f"- **Cấp bậc hiện tại**: Rank {r['rank']} ({r['rank_cat']})\n")
            md.write(f"- **Lớp phụ trách**: {r['classes']}\n")
            md.write(f"- **Chỉ số thực tế đạt được**:\n")
            md.write(f"  * Sĩ số lớp trung bình (Chuyên cần): **{r['raw_metrics']['cc_rate']:.2f}%**\n")
            md.write(f"  * Tỷ lệ học viên hoàn thành E-learning: **{r['raw_metrics']['el_rate']:.2f}%**\n")
            md.write(f"  * Tỷ lệ học viên nộp BTVN/Checkpoint: **{r['raw_metrics']['bt_rate']:.2f}%**\n")
            md.write(f"  * Tỷ lệ đỗ môn trung bình (Đầu ra): **{r['raw_metrics']['pass_rate']:.2f}%**\n")
            md.write(f"  * Số lần vi phạm báo cáo ngày: **{r['raw_metrics']['compliance_violations']} lần**\n")
            md.write(f"  * Số lần học liệu trễ kế hoạch: **{r['raw_metrics']['material_violations']} lần**\n")
            md.write(f"  * Số lỗi vi phạm kỷ luật tác nghiệp: **{r['raw_metrics']['op_violations']} lần**\n")
            md.write(f"- **Điểm các tiêu chí định lượng (Thang 10)**:\n")
            md.write(f"  * Tuân thủ quy trình đào tạo (báo cáo ngày): **{r['scores']['tuanchu']} / 10**\n")
            md.write(f"  * Kỷ luật tác nghiệp: **{r['scores']['kyluat']} / 10**\n")
            md.write(f"  * Tiến độ học liệu: **{r['scores']['hoclieu']} / 10**\n")
            md.write(f"  * Quản lý sĩ số: **{r['scores']['siso']} / 10**\n")
            md.write(f"  * Đôn đốc E-learning: **{r['scores']['el']} / 10**\n")
            md.write(f"  * Đôn đốc BTVN: **{r['scores']['bt']} / 10**\n")
            md.write(f"  * Tỷ lệ pass môn: **{r['scores']['pass_rate']} / 10**\n")
            md.write(f"- **Kết quả quá trình tổng hợp**:\n")
            md.write(f"  * **Điểm A (Scale thực tế)**: **{r['overall_score_A']:.2f} / 10**\n")
            md.write(f"  * **Điểm B (Chèn điểm Đạt)**: **{r['overall_score_B']:.2f} / 10**\n")
            md.write(f"  * **Xếp loại năng lực**: **{r['classification']}**\n\n")

    # Lâm Tùng Dương Rank 2 vs Rank 3 comparison
    duong_data = [r for r in evaluation_results if "Lâm Tùng Dương" in r['name']][0]
    md.write("## 2. Phân tích đối chiếu trường hợp Lâm Tùng Dương (Rank 2 vs Rank 3)\n\n")
    md.write("> [!TIP]\n")
    md.write("> Thầy Lâm Tùng Dương vừa được thăng chức từ Trợ giảng (Rank 2) lên Giảng viên (Rank 3). Dưới đây là phân tích sự thay đổi trọng số và điểm số tương ứng.\n\n")
    
    w_r2 = WEIGHTS_V2['TA']
    w_r3 = WEIGHTS_V2['GV']
    
    # Calculate Rank 2 scores
    real_w_r2 = w_r2['tuanchu'] + w_r2['kyluat'] + w_r2['hoclieu'] + w_r2['siso'] + w_r2['el'] + w_r2['bt'] + w_r2['pass_rate']
    weighted_real_r2 = (
        duong_data['scores']['tuanchu'] * w_r2['tuanchu'] +
        duong_data['scores']['kyluat'] * w_r2['kyluat'] +
        duong_data['scores']['hoclieu'] * w_r2['hoclieu'] +
        duong_data['scores']['siso'] * w_r2['siso'] +
        duong_data['scores']['el'] * w_r2['el'] +
        duong_data['scores']['bt'] * w_r2['bt'] +
        duong_data['scores']['pass_rate'] * w_r2['pass_rate']
    )
    score_A_r2 = weighted_real_r2 / real_w_r2
    
    weighted_all_r2 = (
        duong_data['scores']['tuanchu'] * w_r2['tuanchu'] +
        duong_data['scores']['kyluat'] * w_r2['kyluat'] +
        duong_data['scores']['hoclieu'] * w_r2['hoclieu'] +
        5.0 * w_r2['phoihop'] +
        duong_data['scores']['siso'] * w_r2['siso'] +
        duong_data['scores']['el'] * w_r2['el'] +
        duong_data['scores']['bt'] * w_r2['bt'] +
        5.0 * w_r2['giaiquyet'] +
        duong_data['scores']['pass_rate'] * w_r2['pass_rate'] +
        5.0 * w_r2['kha_gioi'] +
        5.0 * w_r2['cc_cn'] +
        5.0 * w_r2['bang_cap'] +
        5.0 * w_r2['csat'] +
        5.0 * w_r2['hoclieu_duyet']
    )
    
    md.write("| Cấp bậc đánh giá | Trọng số áp dụng | Điểm A (Scale thực tế) | Điểm B (Chèn điểm Đạt) | Nhận xét phân tích |\n")
    md.write("| :--- | :---: | :---: | :---: | :--- |\n")
    md.write(f"| **Rank 2 (Trợ giảng)** | Chuẩn TA (V2) | **{score_A_r2:.2f} / 10** | **{weighted_all_r2:.2f} / 10** | Tập trung vào Vận hành (35%), Sĩ số & Đôn đốc (30%). Trọng số sư phạm là 0%. |\n")
    md.write(f"| **Rank 3 (Giảng viên)** | Chuẩn GV (V2) | **{duong_data['overall_score_A']:.2f} / 10** | **{duong_data['overall_score_B']:.2f} / 10** | Bổ sung trọng số sư phạm & truyền đạt (30%), giảm tải trọng số Sĩ số & Đôn đốc xuống 20%. |\n\n")

    md.write("## 3. Bảng xếp hạng năng lực toàn bộ nhân sự khối CNTT & QTKD (Tiêu chuẩn mới V2)\n\n")
    md.write("| Họ và tên | Khối phòng ban | Vai trò / Cấp bậc | Điểm A (Thực tế) | Điểm B (Đạt) | Xếp loại năng lực |\n")
    md.write("| :--- | :--- | :---: | :---: | :---: | :---: |\n")
    
    for r in sorted(evaluation_results, key=lambda x: x['overall_score_A'], reverse=True):
        md.write(f"| **{r['name']}** | {r['group']} | {r['role']} (Rank {r['rank']}) | **{r['overall_score_A']:.2f}** | {r['overall_score_B']:.2f} | **{r['classification']}** |\n")

print(f"Báo cáo chi tiết đã được xuất ra file: {report_path}")
print(f"  - Điểm A (Scale thực tế): {score_A_r2:.2f} / 10")
print(f"  - Điểm B (Mặc định đạt): {weighted_all_r2:.2f} / 10")
print(f"Lâm Tùng Dương ở Rank 3 (Giảng viên) [Hiện tại]:")
print(f"  - Điểm A (Scale thực tế): {duong_data['overall_score_A']:.2f} / 10")
print(f"  - Điểm B (Mặc định đạt): {duong_data['overall_score_B']:.2f} / 10")

print("\n=== FULL LIST OF STAFF IN CNTT & QTKD ===")
for r in sorted(evaluation_results, key=lambda x: x['overall_score_A'], reverse=True):
    print(f"{r['name']} ({r['group']}) | Rank {r['rank']} | Score A: {r['overall_score_A']:.2f} | Score B: {r['overall_score_B']:.2f} | Class: {r['classification']}")
