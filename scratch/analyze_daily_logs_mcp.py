import urllib.request
import urllib.error
import json
import ssl
import sys
import os
import openpyxl
import re

sys.stdout.reconfigure(encoding='utf-8')

# Define targets (39 personnel)
target_groups = {
    "Khối QTKD": [
        "Hoàng Thị Kim Oanh",
        "Hoàng Thị Hậu",
        "Nguyễn Thị Hồng Minh",
        "Đặng Quỳnh Trang",
        "Nguyễn Ngọc Vân Khanh",
        "Triệu Thị Thanh Tâm",
        "Lê Thành Ngọc",
        "Nguyễn Thị Như Quỳnh",
        "Lê Nhựt Mi",
        "Lê Thị Bảo Yến"
    ],
    "Khối CNTT - Cơ sở HN (KS25)": [
        "Trịnh Quốc Hai",
        "Nguyễn Quảng An",
        "Lương Quốc Tuấn",
        "Phạm Ngọc Kiên",
        "Lại Trung Lâm",
        "Lâm Tùng Dương",
        "Ngọ Văn Quý",
        "Trần Minh Cường"
    ],
    "Khối CNTT - Cơ sở HN (KS24)": [
        "Bùi Thanh Hải",
        "Mai Xuân Chinh",
        "Đinh Thành Nam",
        "Nguyễn Công Hưởng",
        "Phạm Tuấn Bình",
        "Nguyễn Bá Minh Đạo"
    ],
    "Khối CNTT - Cơ sở HCM": [
        "Lê Hà Thanh Sang",
        "Phạm Viết Hùng",
        "Trần Quốc Tuấn",
        "Lưu Xuân Hoàng Nguyên",
        "Nguyễn Đức Minh",
        "Nguyễn Ngọc Sơn",
        "Phan Ngọc Tài",
        "Đặng Minh Luân"
    ],
    "Khối Ngoại ngữ - Kỹ năng mềm": [
        "Giáp Thị Minh Hằng",
        "Lò Thị Ngọc Anh",
        "Lê Thị Đỏ",
        "Ngô Quang Huấn"
    ],
    "Khối Quản lý Đào tạo (QLĐT)": [
        "Nguyễn Thị Tươi",
        "Trần Thị Mỹ Phước",
        "Nguyễn Huyền Trang",
        "Nguyễn Xuân Bách"
    ]
}

special_mappings = {
    "lưu hoàng xuân nguyên": "lưu xuân hoàng nguyên",
    "xuân nguyên": "lưu xuân hoàng nguyên"
}

def normalize_name(name):
    norm = name.strip().lower()
    if norm in special_mappings:
        norm = special_mappings[norm]
    return norm

def call_mcp_tool(tool_name, arguments={}):
    url = "https://pm.rikkei.edu.vn/api/mcp"
    payload = {
        "jsonrpc": "2.0",
        "id": 1,
        "method": "tools/call",
        "params": {
            "name": tool_name,
            "arguments": arguments
        }
    }
    headers = {
        "Authorization": "Bearer wl_jtpd1dOgxnUm5n2d7V6dxBT_AZHNrnCK",
        "Content-Type": "application/json",
        "Accept": "application/json, text/event-stream"
    }

    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    req = urllib.request.Request(
        url, 
        data=json.dumps(payload).encode("utf-8"), 
        headers=headers, 
        method="POST"
    )

    try:
        with urllib.request.urlopen(req, context=ctx) as response:
            resp_str = response.read().decode("utf-8")
            for line in resp_str.split("\n"):
                if line.startswith("data:"):
                    json_str = line[5:].strip()
                    data = json.loads(json_str)
                    return data
    except urllib.error.HTTPError as e:
        print(f"HTTP Error calling {tool_name}:", e.code, e.reason)
    except Exception as e:
        print(f"Error calling {tool_name}:", e)
    return None

def load_staff_profiles():
    profiles = {}
    
    # 1. QTKD STAFF
    qtkd_path = r"C:\Users\DELL\Downloads\_Task Management_ QL Khối QTKD.xlsx"
    if os.path.exists(qtkd_path):
        try:
            wb = openpyxl.load_workbook(qtkd_path, data_only=True)
            if "STAFF" in wb.sheetnames:
                sheet = wb["STAFF"]
                headers = [str(sheet.cell(row=1, column=c).value).strip().lower() for c in range(1, sheet.max_column + 1)]
                name_idx = next((i for i, h in enumerate(headers) if "tên" in h or "name" in h), 0) + 1
                role_idx = next((i for i, h in enumerate(headers) if "role" in h or "vị trí" in h), 1) + 1
                rank_idx = next((i for i, h in enumerate(headers) if "rank" in h or "cấp" in h), 2) + 1
                
                for r in range(2, sheet.max_row + 1):
                    name = sheet.cell(row=r, column=name_idx).value
                    role = sheet.cell(row=r, column=role_idx).value
                    rank = sheet.cell(row=r, column=rank_idx).value
                    if name:
                        norm = normalize_name(str(name))
                        profiles[norm] = {
                            "role": str(role).strip() if role else "Giảng viên",
                            "rank": str(rank).strip() if rank else "3"
                        }
            wb.close()
        except Exception as e:
            print("Warning loading QTKD Staff profiles:", e)
            
    # 2. CNTT STAFF
    cntt_path = r"C:\Users\DELL\Downloads\Quản lý hiệu suất đào tạo.xlsx"
    if os.path.exists(cntt_path):
        try:
            wb = openpyxl.load_workbook(cntt_path, data_only=True)
            if "BC hàng ngày" in wb.sheetnames:
                sheet = wb["BC hàng ngày"]
                headers = [str(sheet.cell(row=1, column=c).value).strip().lower() for c in range(1, sheet.max_column + 1)]
                name_idx = next((i for i, h in enumerate(headers) if "họ và tên" in h or "người tạo" in h), 6) + 1
                role_idx = next((i for i, h in enumerate(headers) if "vị trí" in h), 7) + 1
                rank_idx = next((i for i, h in enumerate(headers) if "rank" in h), 8) + 1
                
                for r in range(2, sheet.max_row + 1):
                    name = sheet.cell(row=r, column=name_idx).value
                    role = sheet.cell(row=r, column=role_idx).value
                    rank = sheet.cell(row=r, column=rank_idx).value
                    if name:
                        norm = normalize_name(str(name))
                        if norm not in profiles:
                            profiles[norm] = {
                                "role": str(role).strip() if role else "Giảng viên",
                                "rank": str(rank).strip() if rank else "3"
                            }
            wb.close()
        except Exception as e:
            print("Warning loading CNTT Staff profiles:", e)
            
    return profiles

def load_kpi_masters():
    qtkd_master = {}
    cntt_master = {}
    
    # 1. QTKD KPI Master
    qtkd_path = r"C:\Users\DELL\Downloads\_Task Management_ QL Khối QTKD.xlsx"
    if os.path.exists(qtkd_path):
        try:
            wb = openpyxl.load_workbook(qtkd_path, data_only=True)
            if "KPI_MASTER" in wb.sheetnames:
                sheet = wb["KPI_MASTER"]
                for r in range(2, sheet.max_row + 1):
                    key = sheet.cell(row=r, column=6).value
                    std_time = sheet.cell(row=r, column=5).value
                    if key and std_time is not None:
                        qtkd_master[str(key).strip()] = float(std_time)
            wb.close()
        except Exception as e:
            print("Warning loading QTKD KPI Master:", e)
            
    # 2. CNTT KPI Master
    cntt_path = r"C:\Users\DELL\Downloads\Quản lý hiệu suất đào tạo.xlsx"
    if os.path.exists(cntt_path):
        try:
            wb = openpyxl.load_workbook(cntt_path, data_only=True)
            sheetname = "Cấu trúc KPI công việc GV. TG"
            if sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                for r in range(2, sheet.max_row + 1):
                    key = sheet.cell(row=r, column=7).value
                    std_time = sheet.cell(row=r, column=5).value
                    if key and std_time is not None:
                        cntt_master[str(key).strip()] = float(std_time)
            wb.close()
        except Exception as e:
            print("Warning loading CNTT KPI Master:", e)
            
    return qtkd_master, cntt_master

def match_kpi_standard_time(group, name, role, rank, task_title, kpi_master_qtkd, kpi_master_cntt):
    title_norm = task_title.strip().lower()
    
    is_qtkd = "QTKD" in group
    kpi_db = kpi_master_qtkd if is_qtkd else kpi_master_cntt
    
    role_norm = "giảng viên" if "giảng" in role.lower() or "gv" in role.lower() else "trợ giảng"
    try:
        rank_val = int(rank)
    except:
        rank_val = 3
        
    matched_task_type = None
    if any(k in title_norm for k in ["giảng dạy", "lên lớp", "dạy lý thuyết", "triển khai buổi học"]):
        matched_task_type = "Giảng dạy lý thuyết - Buổi học" if role_norm == "giảng viên" else "Triển khai buổi thực hành - Buổi"
    elif any(k in title_norm for k in ["chuẩn bị", "soạn slide", "soạn bài", "soạn giáo án"]):
        matched_task_type = "Chuẩn bị giảng dạy - Buổi học" if role_norm == "giảng viên" else "Chuẩn bị buổi thực hành - Buổi"
    elif "mindmap" in title_norm or "bản đồ tư duy" in title_norm:
        matched_task_type = "Làm mindmap bài học - Session"
    elif any(k in title_norm for k in ["support", "hỗ trợ", "fix bug", "sửa lỗi", "hướng dẫn"]):
        matched_task_type = "Báo cáo hỗ trợ SV.HV" if not is_qtkd else "Hỗ trợ học viên"
    elif any(k in title_norm for k in ["chấm bài", "chấm thi", "chấm thực hành"]):
        matched_task_type = "Chấm thi thực hành - Bài"
    elif "chấm sản phẩm" in title_norm or "chấm project" in title_norm or "chấm prj" in title_norm:
        matched_task_type = "Chấm thi sản phẩm - Sản phẩm"
    elif "vấn đáp" in title_norm or "chấm vấn đáp" in title_norm:
        matched_task_type = "Chấm thi vấn đáp - Sinh viên"
    elif "trông thi" in title_norm:
        matched_task_type = "Trông thi thực hành - Ca" if "thực hành" in title_norm else "Trông thi lý thuyết - Ca"
        
    if matched_task_type:
        for key, std_time in kpi_db.items():
            key_norm = key.lower()
            if is_qtkd:
                role_part = "giảng viên" if "giảng" in key_norm else "trợ giảng"
                rank_part = re.search(r'-(\d+)-', key_norm)
                rank_num = int(rank_part.group(1)) if rank_part else 3
                task_part = key_norm.split('-')[-1]
                if role_part == role_norm and rank_num == rank_val and (matched_task_type.lower() in key_norm or task_part in title_norm):
                    return std_time, matched_task_type, False
            else:
                role_part = "giảng viên" if "giảng" in key_norm else "trợ giảng"
                rank_part = re.search(r'-(\d+)$', key_norm)
                rank_num = int(rank_part.group(1)) if rank_part else 3
                task_part = key_norm.split('-')[0]
                if role_part == role_norm and rank_num == rank_val and (matched_task_type.lower() in key_norm or task_part in title_norm):
                    return std_time, matched_task_type, False

    for key, std_time in kpi_db.items():
        key_norm = key.lower()
        if role_norm in key_norm and f"-{rank_val}" in key_norm:
            task_name = key_norm.split('-')[0] if not is_qtkd else key_norm.split('-')[-1]
            if task_name in title_norm or title_norm in task_name:
                return std_time, key.split('-')[0] if not is_qtkd else key.split('-')[-1], False
                
    return 30.0, "Đầu việc tự do/chưa định mức", True

def process_stats_for_period(results, staff_profiles, kpi_master_qtkd, kpi_master_cntt, target_dates):
    analysis = {}
    total_days = len(target_dates)
    
    for group, members in results.items():
        for m, m_data in members.items():
            reported_days_list = [d for d in target_dates if m_data["reports"][d] is not None]
            reported_days_count = len(reported_days_list)
            missing_days = [d for d in target_dates if m_data["reports"][d] is None]
            
            norm_name = normalize_name(m)
            profile = staff_profiles.get(norm_name, {"role": "Giảng viên", "rank": "3"})
            role = profile["role"]
            rank = profile["rank"]
            
            total_tasks = 0
            completed_tasks = 0
            declared_hours = 0.0
            uncompleted_reasons = []
            
            time_score = 100.0
            time_violations = []
            warning_flags = []
            
            for d in target_dates:
                r = m_data["reports"][d]
                if r:
                    stats = r.get("stats", {})
                    declared_hours += float(stats.get("hours", 0.0))
                    
                    tasks = r.get("tasks", [])
                    for t in tasks:
                        total_tasks += 1
                        t_title = t.get("title", "")
                        t_hours = float(t.get("hours", 0.0))
                        
                        std_time, matched_cat, is_wildcard = match_kpi_standard_time(
                            group, m, role, rank, t_title, kpi_master_qtkd, kpi_master_cntt
                        )
                        std_hours = std_time / 60.0
                        
                        if is_wildcard:
                            # KHÔNG PHẠT TRỪ ĐIỂM CHO TASK LẠ, chỉ lưu ghi nhận
                            warning_flags.append(
                                f"Task '{t_title}' chưa định dạng (gán tạm {std_time:.0f} phút)"
                            )
                        else:
                            # Có trong KPI Master -> check over-reporting
                            if t_hours > std_hours * 1.5:
                                time_score -= 5.0
                                time_violations.append(
                                    f"{d.split('-')[-1]}/{d.split('-')[-2]}: Task '{t_title}' khai báo {t_hours}h so với định mức tiêu chuẩn {std_hours:.1f}h"
                                )

                        if t.get("done") is True or str(t.get("percent")) == "100":
                            completed_tasks += 1
                        else:
                            uncompleted_reasons.append(f"{t_title} ({t.get('percent', 0)}%)")

            time_score = max(0.0, time_score)
            report_rate = (reported_days_count / float(total_days)) if total_days > 0 else 0.0
            
            completion_rate = 1.0
            if total_tasks > 0:
                completion_rate = completed_tasks / total_tasks

            # Work Score = (Report Rate * 40%) + (Completion Rate * 40%) + (Time Score * 20%)
            work_score = (report_rate * 40.0) + (completion_rate * 40.0) + (time_score * 0.20)
            if reported_days_count == 0:
                work_score = 0.0 

            analysis[norm_name] = {
                "name": m,
                "group": group,
                "role": role,
                "rank": rank,
                "reported_days": reported_days_count,
                "missing_days": missing_days,
                "total_tasks": total_tasks,
                "completed_tasks": completed_tasks,
                "completion_rate": completion_rate * 100.0,
                "time_score": time_score,
                "work_score": round(work_score, 1),
                "declared_hours": declared_hours,
                "time_violations": time_violations[:3],
                "warning_flags": warning_flags[:3],
                "uncompleted_tasks": uncompleted_reasons[:3]
            }
            
    return analysis

def main():
    from datetime import datetime, timedelta
    print("Agent 4: Bắt đầu fetch dữ liệu báo cáo ngày từ Worklane PM...")
    
    # Tính toán ngày động
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)
    
    # Nếu ngày hôm trước là cuối tuần, điều chỉnh về Thứ Sáu gần nhất để kiểm toán báo cáo
    adjusted_yesterday = yesterday
    if yesterday.weekday() == 5: # Thứ 7
        adjusted_yesterday = yesterday - timedelta(days=1)
    elif yesterday.weekday() == 6: # Chủ nhật
        adjusted_yesterday = yesterday - timedelta(days=2)
        
    yesterday_str = adjusted_yesterday.strftime("%Y-%m-%d")
    
    # Tính danh sách ngày làm việc tuần hiện tại (Thứ 2 đến hôm qua)
    start_of_week = today - timedelta(days=today.weekday())
    dates_weekly = []
    curr = start_of_week
    while curr <= yesterday:
        if curr.weekday() < 5: # Chỉ lấy Thứ 2 - Thứ 6
            dates_weekly.append(curr.strftime("%Y-%m-%d"))
        curr += timedelta(days=1)
        
    # Tính danh sách ngày làm việc tháng 7 hiện tại (01/07 đến hôm qua)
    dates_all = []
    curr = datetime(today.year, 7, 1).date()
    while curr <= yesterday:
        if curr.weekday() < 5:
            dates_all.append(curr.strftime("%Y-%m-%d"))
        curr += timedelta(days=1)
        
    # Trường hợp chạy test hoặc chưa có ngày nào trong tuần/tháng, ít nhất phải có ngày hôm qua
    if not dates_weekly and adjusted_yesterday.strftime("%Y-%m-%d") not in dates_weekly:
        # Nếu hôm nay là Thứ Hai, tuần mới chưa có ngày nào trước đó, dates_weekly có thể rỗng.
        pass
    if not dates_all:
        dates_all = [yesterday_str]

    print(f"  Thời gian kiểm toán báo cáo: hôm qua ({yesterday_str})")
    print(f"  Danh sách ngày tuần: {dates_weekly}")
    print(f"  Danh sách ngày tháng: {dates_all}")
    
    print("  Đang nạp thông tin nhân sự và định mức KPI Master từ Excel...")
    staff_profiles = load_staff_profiles()
    kpi_master_qtkd, kpi_master_cntt = load_kpi_masters()
    
    results = {}
    for group, members in target_groups.items():
        results[group] = {}
        for m in members:
            results[group][m] = {
                "name": m,
                "reports": {d: None for d in dates_all}
            }

    # Fetch daily reports cho các ngày
    for d in dates_all:
        print(f"  Tải dữ liệu ngày {d}...")
        res = call_mcp_tool("list_daily_reports", {"date": d, "department": "DT"})
        if res and "result" in res:
            try:
                inner_str = res["result"]["content"][0].get("text", "")
                inner_json = json.loads(inner_str)
                reports = inner_json.get("reports", [])
                for r in reports:
                    user_name = r.get("user")
                    norm_user = normalize_name(user_name)
                    found = False
                    for group, members in target_groups.items():
                        for m in members:
                            if normalize_name(m) == norm_user:
                                results[group][m]["reports"][d] = r
                                found = True
                                break
                        if found:
                            break
            except Exception as e:
                print(f"  Error parsing data for day {d}:", e)

    # Phát hiện nhân sự không báo cáo ngày hôm trước
    missing_yesterday = []
    if yesterday_str in dates_all:
        for group, members in results.items():
            for m, m_data in members.items():
                if m_data["reports"].get(yesterday_str) is None:
                    norm_name = normalize_name(m)
                    profile = staff_profiles.get(norm_name, {"role": "Giảng viên", "rank": "3"})
                    missing_yesterday.append({
                        "name": m,
                        "group": group,
                        "role": profile["role"]
                    })

    print("Agent 4: Tiến hành phân tích riêng biệt theo Tuần và Tháng...")
    weekly_analysis = process_stats_for_period(results, staff_profiles, kpi_master_qtkd, kpi_master_cntt, dates_weekly) if dates_weekly else {}
    monthly_analysis = process_stats_for_period(results, staff_profiles, kpi_master_qtkd, kpi_master_cntt, dates_all)
    
    combined_output = {
        "yesterday": yesterday_str,
        "missing_yesterday": missing_yesterday,
        "dates_weekly": dates_weekly,
        "dates_monthly": dates_all,
        "weekly_stats": weekly_analysis,
        "monthly_stats": monthly_analysis,
        "raw_reports": results
    }

    output_dir = "data"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    output_path = os.path.join(output_dir, "daily_log_analysis.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(combined_output, f, indent=2, ensure_ascii=False)
        
    print(f"Agent 4: Phân tích tuần/tháng thành công! Kết quả lưu tại {output_path}")


if __name__ == "__main__":
    main()
