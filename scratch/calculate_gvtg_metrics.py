import mysql.connector
import openpyxl
import sys
import os
import unicodedata
import json
from collections import defaultdict

# Force UTF-8 output
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"data/PTIT_Chiso.xlsx"
PRED_JSON_PATH = r"scratch/predictions_cv_data.json"

TARGETS = [
    "Trần Quốc Tuấn",
    "Lê Hà Thanh Sang",
    "Phạm Viết Hùng",
    "Lưu Xuân Hoàng Nguyên"
]

# Sheet to Course ID mappings based on courses table in qldt_el
COURSE_MAPPING = {
    'KS25_Javascript': 124,     # [IT103B-K25] Xây dựng ứng dụng web với Javascript
    'KS25_Database': 183,       # [IT202-K25] Cơ sở dữ liệu
    'KS25_Python': 193,         # [IT205-K25] Lập trình ứng dụng với Python
    'KS25_Python_Web': 217,     # [IT-215] Phát triển dịch vụ Web với FastAPI
    'KS24-JavaAdvance': 162,    # [IT203B- K24] - Java Advanced
    'KS24_JavaWeb': 177,        # [IT210 - K24] - Java Web Application
    'KS24_JWS': 194,            # [ IT211 - K24 ] Java Web Service
    'KS24_AI': 214              # [IT-212] AI Application in Action
}

def normalize_vietnamese_name(name):
    if not name:
        return ""
    name = " ".join(name.strip().split())
    name = name.lower()
    name = unicodedata.normalize('NFKD', name)
    name = "".join([c for c in name if not unicodedata.combining(c)])
    name = name.replace("đ", "d")
    return name

def clean_instructor_name(name):
    if not name:
        return ""
    name_clean = name.strip()
    special_mappings = {
        "lưu hoàng xuân nguyên": "Lưu Xuân Hoàng Nguyên",
        "xuân nguyên": "Lưu Xuân Hoàng Nguyên",
        "lưu xuân hoàng nguyên": "Lưu Xuân Hoàng Nguyên",
        "lê hà thanh sang": "Lê Hà Thanh Sang",
        "trần quốc tuấn": "Trần Quốc Tuấn",
        "phạm viết hùng": "Phạm Viết Hùng"
    }
    if name_clean.lower() in special_mappings:
        return special_mappings[name_clean.lower()]
    return name_clean

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

def parse_date(d_val):
    from datetime import datetime, date
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

def main():
    # 1. Load predictions JSON for fallback (with normalized class names as keys)
    predictions_data = {}
    if os.path.exists(PRED_JSON_PATH):
        with open(PRED_JSON_PATH, "r", encoding="utf-8") as f:
            try:
                p_data = json.load(f)
                dashboard_data = p_data.get('dashboard_data', {})
                for batch_key, batch_val in dashboard_data.items():
                    for c in batch_val.get('cv', []):
                        cname = c.get('class_name')
                        norm_cname = normalize_class_name(cname)
                        predictions_data[norm_cname] = {
                            'type': 'actual_cv',
                            'pass_rate': c.get('actual_pass')
                        }
                    for c in batch_val.get('curr', []):
                        cname = c.get('class_name')
                        norm_cname = normalize_class_name(cname)
                        predictions_data[norm_cname] = {
                            'type': 'predicted_curr',
                            'pass_rate': c.get('pred_new')
                        }
            except Exception as e:
                print(f"Error loading predictions JSON: {e}")

    # 2. Connect to MySQL to retrieve database mapping & data
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="qldt_el",
            port=3307
        )
        cursor = conn.cursor(dictionary=True)
        
        # Load all classes from DB for name mapping
        cursor.execute("SELECT id, name FROM classes")
        db_classes = cursor.fetchall()
        
        # Load all courses from DB
        cursor.execute("SELECT id, name FROM courses")
        db_courses = {c['id']: c['name'] for c in cursor.fetchall()}
        
    except Exception as e:
        print("Error connecting to MySQL:", str(e))
        return

    # Helper function to find class ID in DB
    def find_db_class(excel_class_name):
        norm_excel = normalize_class_name(excel_class_name).lower().replace("ks25", "k25").replace("ks24", "k24").strip()
        # Direct matching first
        for dbc in db_classes:
            norm_db = dbc['name'].lower().replace("ks25", "k25").replace("ks24", "k24").strip()
            if norm_excel == norm_db:
                return dbc
            # Remove suffix like _HK2 or _HL from DB class name
            db_base = dbc['name'].split('(')[0].strip()
            for suffix in ['_HK2', '_HL', '-HL', '_GL']:
                if db_base.endswith(suffix):
                    db_base = db_base[:-len(suffix)].strip()
            norm_db_base = db_base.lower().replace("ks25", "k25").replace("ks24", "k24").strip()
            if norm_excel == norm_db_base:
                return dbc
        return None

    # 3. Read Excel
    if not os.path.exists(EXCEL_PATH):
        print(f"Excel file not found at {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # We will accumulate metrics per instructor
    instructor_results = defaultdict(list)

    for sheetname in COURSE_MAPPING.keys():
        if sheetname not in wb.sheetnames:
            continue
        sheet = wb[sheetname]
        max_r = sheet.max_row
        max_c = sheet.max_column
        if max_r < 5 or max_c < 4:
            continue

        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]

        col_info = []
        current_date = None
        for c_idx in range(3, max_c):
            if c_idx < len(row3) and row3[c_idx]:
                parsed = parse_date(row3[c_idx])
                if parsed:
                    current_date = parsed
            subheader = row4[c_idx] if c_idx < len(row4) else None
            if current_date and subheader in ('Chuyên cần', 'Bài tập', 'Elearning'):
                col_info.append((c_idx, current_date, subheader))

        for r in range(5, max_r + 1):
            cname_raw = sheet.cell(row=r, column=2).value
            teacher_raw = sheet.cell(row=r, column=3).value
            if not cname_raw:
                continue
            norm_class = normalize_class_name(cname_raw)
            if "l01" in norm_class.lower() or "l02" in norm_class.lower():
                continue

            # Read TA from r+1
            ta_raw = None
            if r + 1 <= max_r:
                next_cname = sheet.cell(row=r + 1, column=2).value
                if not next_cname:
                    ta_raw = sheet.cell(row=r + 1, column=3).value

            teacher_name = clean_instructor_name(teacher_raw)
            ta_name = clean_instructor_name(ta_raw)

            # Read attendance violations from row r (GV row has the data)
            cc_vals = []
            for c_idx, d, sub in col_info:
                val = sheet.cell(row=r, column=c_idx + 1).value
                if isinstance(val, (int, float)):
                    cc_vals.append(float(val))
                    
            avg_cc = sum(cc_vals)/len(cc_vals) if cc_vals else 0.0

            # Match DB class and course
            course_id = COURSE_MAPPING[sheetname]
            course_name = db_courses.get(course_id, sheetname)
            
            db_class = find_db_class(cname_raw)
            pass_rate = None
            pass_rate_type = "N/A"
            total_students = 0
            passed_students = 0
            graded_students = 0

            pred_key = normalize_class_name(cname_raw)

            if db_class:
                class_id = db_class['id']
                # Query pass rate and graded count from DB final_results
                cursor.execute("""
                    SELECT 
                        COUNT(student_id) as total, 
                        SUM(CASE WHEN pass = 1 THEN 1 ELSE 0 END) as passed,
                        SUM(CASE WHEN pass IS NOT NULL THEN 1 ELSE 0 END) as graded
                    FROM final_results 
                    WHERE class_id = %s AND course_id = %s
                """, (class_id, course_id))
                res = cursor.fetchone()
                if res and res['total'] > 0:
                    total_students = res['total']
                    graded_students = res['graded'] if res['graded'] is not None else 0
                    passed_students = res['passed'] if res['passed'] is not None else 0
                    
                    if graded_students > 0:
                        # Database has graded results
                        pass_rate = (float(passed_students) / float(total_students)) * 100.0
                        pass_rate_type = "Thực tế DB"
                    else:
                        # Database exists but no grades yet (all pass are NULL) -> Fallback to predictions
                        if pred_key in predictions_data:
                            pass_rate = predictions_data[pred_key]['pass_rate']
                            pass_rate_type = "Dự đoán (Model)"
                        else:
                            pass_rate = None
                            pass_rate_type = "Chưa chốt điểm / Không có dự đoán"
                else:
                    # Fallback to predictions JSON
                    if pred_key in predictions_data:
                        pass_rate = predictions_data[pred_key]['pass_rate']
                        pass_rate_type = "Dự đoán (Model)"
            else:
                # Fallback to predictions JSON
                if pred_key in predictions_data:
                    pass_rate = predictions_data[pred_key]['pass_rate']
                    pass_rate_type = "Dự đoán (Model)"

            # Check matching targets
            for t in TARGETS:
                norm_t = normalize_vietnamese_name(t)
                is_gv = normalize_vietnamese_name(teacher_name) == norm_t
                is_ta = ta_name and normalize_vietnamese_name(ta_name) == norm_t
                
                if is_gv or is_ta:
                    role = "Giảng viên" if is_gv else "Trợ giảng"
                    instructor_results[t].append({
                        'sheet': sheetname,
                        'class_excel': cname_raw,
                        'class_norm': norm_class,
                        'class_db': db_class['name'] if db_class else "Không khớp",
                        'role': role,
                        'avg_cc_violation': avg_cc,
                        'cc_sessions': len(cc_vals),
                        'course_id': course_id,
                        'course_name': course_name,
                        'total_students': total_students,
                        'passed_students': passed_students,
                        'pass_rate': pass_rate,
                        'pass_rate_type': pass_rate_type
                    })

    wb.close()
    conn.close()

    # 4. Print & Save Results
    print("=== SUMMARY RESULTS ===")
    
    report_md = []
    report_md.append("# Báo cáo Phân tích Chi tiết Giảng viên & Trợ giảng\n")
    report_md.append(f"**Dữ liệu cập nhật từ Excel:** `{EXCEL_PATH}`\n")
    report_md.append(f"**Dữ liệu kết quả từ MySQL DB:** `qldt_el` (cổng 3307)\n")
    
    for t in TARGETS:
        print(f"\nInstructor: {t}")
        report_md.append(f"## Giảng viên: {t}\n")
        
        classes = instructor_results.get(t, [])
        if not classes:
            print("  No classes found.")
            report_md.append("*Không tìm thấy lớp học nào.*\n")
            continue
            
        print(f"  Total classes: {len(classes)}")
        
        # Markdown table
        report_md.append("| Lớp học (Excel) | Lớp học (DB) | Môn học | Vai trò | Số buổi điểm danh | Vi phạm Chuyên cần TB (%) | Tỷ lệ qua môn | Nguồn tỷ lệ qua môn |")
        report_md.append("| :--- | :--- | :--- | :--- | :---: | :---: | :---: | :--- |")
        
        total_cc_violations = 0.0
        cc_class_count = 0
        
        # Lists to calculate different pass rate averages
        actual_pass_rates = []
        all_pass_rates = []  # Includes actual and predicted
        
        for c in classes:
            pass_str = "Chưa chốt / Đang học"
            if c['pass_rate'] is not None:
                pass_str = f"{c['pass_rate']:.2f}%"
                all_pass_rates.append(c['pass_rate'])
                if c['pass_rate_type'] == "Thực tế DB":
                    actual_pass_rates.append(c['pass_rate'])
                
            cc_str = f"{c['avg_cc_violation']:.2f}%"
            total_cc_violations += c['avg_cc_violation']
            cc_class_count += 1
            
            print(f"  - Class: {c['class_excel']} | Course: {c['course_name']} | Role: {c['role']} | CC Violation: {cc_str} | Pass Rate: {pass_str} ({c['pass_rate_type']})")
            
            report_md.append(f"| {c['class_excel']} | {c['class_db']} | {c['course_name']} | {c['role']} | {c['cc_sessions']} | {cc_str} | {pass_str} | {c['pass_rate_type']} |")
            
        avg_cc_all = total_cc_violations / cc_class_count if cc_class_count > 0 else 0.0
        avg_pass_actual = sum(actual_pass_rates) / len(actual_pass_rates) if actual_pass_rates else 0.0
        avg_pass_all = sum(all_pass_rates) / len(all_pass_rates) if all_pass_rates else 0.0
        
        print(f"  => Average CC Violation: {avg_cc_all:.2f}%")
        print(f"  => Average Actual Pass Rate (DB chốt): {avg_pass_actual:.2f}% (trên {len(actual_pass_rates)} lớp)")
        print(f"  => Average Pass Rate (Gộp cả Dự đoán): {avg_pass_all:.2f}% (trên {len(all_pass_rates)} lớp)")
        
        report_md.append(f"\n**Tóm tắt chỉ số trung bình của {t}:**")
        report_md.append(f"- **Chỉ số vi phạm chuyên cần trung bình:** `{avg_cc_all:.2f}%` (Tỷ lệ chuyên cần đi học: `{100.0 - avg_cc_all:.2f}%`)")
        
        actual_str = f"`{avg_pass_actual:.2f}%` (tính trên `{len(actual_pass_rates)}/{len(classes)}` lớp đã chốt điểm thực tế)" if actual_pass_rates else "`N/A` (chưa có lớp nào chốt điểm)"
        all_str = f"`{avg_pass_all:.2f}%` (tính trên `{len(all_pass_rates)}/{len(classes)}` lớp bao gồm cả dự báo của Model)" if all_pass_rates else "`N/A`"
        
        report_md.append(f"- **Tỷ lệ pass môn thực tế (Đã chốt DB):** {actual_str}")
        report_md.append(f"- **Tỷ lệ pass môn kết hợp (Thực tế + Dự báo Model):** {all_str}\n")
        report_md.append("---")
        
    # Save markdown report to a file
    report_file_path = "output/kpi_gvtg_specific_report.md"
    os.makedirs("output", exist_ok=True)
    with open(report_file_path, "w", encoding="utf-8") as f:
        f.write("\n".join(report_md))
    print(f"\nSaved report to {report_file_path}")

if __name__ == "__main__":
    main()
