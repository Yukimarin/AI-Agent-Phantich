import sys
import os
import openpyxl

# Reconfigure stdout to use UTF-8 encoding
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

files = {
    "old": r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG.xlsx",
    "v1": r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG (1).xlsx",
    "v2": r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG (2).xlsx"
}

for name, path in files.items():
    print(f"=== File: {name} ({os.path.basename(path)}) ===")
    if not os.path.exists(path):
        print("File does not exist!")
        continue
    wb = openpyxl.load_workbook(path, read_only=True)
    print(f"Sheets: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"  Sheet '{sheet_name}': max_row={sheet.max_row}, max_column={sheet.max_column}")
        # Print first 20 rows of each sheet to inspect
        rows = list(sheet.iter_rows(max_row=25, max_col=12, values_only=True))
        for r_idx, r in enumerate(rows):
            val_strs = [str(x) if x is not None else "" for x in r]
            if any(val_strs):
                # Print row elements clearly, clean spacing
                print(f"    Row {r_idx+1}: {val_strs[:10]}")
    wb.close()
    print()
