import openpyxl
import sys
from datetime import datetime, date
from collections import defaultdict
import numpy as np

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'docs/PTIT_Chiso.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

sheets_to_check = {
    'KS24_AI': ['HN-K24-CNTT1', 'HN-K24-CNTT2', 'HN-K24-CNTT3', 'HN-K24-CNTT4', 'HCM-K24-CNTT1'],
    'KS25_Python_Web': ['HN-K25-CNTT1', 'HN-K25-CNTT2', 'HN-K25-CNTT3', 'HN-K25-CNTT4', 'HN-K25-CNTT5', 'HN-K25-CNTT6', 'HN-K25-CNTT8', 'HCM-K25-CNTT5', 'HCM-K25-CNTT6', 'HCM-K25-CNTT7', 'HCM-K25-CNTT8'],
    'KS25_QTKD_PRJ302': ['HN-K25-QTKD1', 'HN-K25-QTKD2', 'HN-K25-QTKD3']
}

start_date = date(2026, 6, 29)
end_date = date(2026, 7, 5)

print("=== PHÂN TÍCH CHI TIẾT VI PHẠM TUẦN 27 (29/06 - 05/07/2026) ===")
for sheetname, classes_target in sheets_to_check.items():
    if sheetname not in wb.sheetnames:
        continue
    sheet = wb[sheetname]
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    dates_list = []
    current_date = None
    for c_idx in range(3, len(row3)):
        val3 = row3[c_idx]
        val4 = row4[c_idx]
        if val3:
            current_date = parse_date(val3)
        if current_date and start_date <= current_date <= end_date:
            dates_list.append((c_idx, current_date, val4))
            
    print(f"\n--- Sheet: {sheetname} (Số ngày có dữ liệu trong tuần: {len(set([d for _, d, _ in dates_list]))}) ---")
    
    current_class = None
    for r in range(5, sheet.max_row + 1):
        cname = sheet.cell(row=r, column=2).value
        teacher_val = sheet.cell(row=r, column=3).value
        
        if cname:
            current_class = normalize_class_name(cname)
            if current_class not in classes_target:
                continue
                
            teacher_name = str(teacher_val).strip() if teacher_val else "N/A"
            
            day_vals = defaultdict(list)
            for c_idx, d, val4 in dates_list:
                val = sheet.cell(row=r, column=c_idx + 1).value
                if val is not None:
                    try:
                        day_vals[val4].append(float(val))
                    except ValueError:
                        pass
            
            print(f"Lớp: {current_class} - GV: {teacher_name}")
            for metric in ['Chuyên cần', 'Bài tập', 'Elearning']:
                vals = day_vals.get(metric, [])
                avg = sum(vals) / len(vals) if vals else 0.0
                max_val = max(vals) if vals else 0.0
                print(f"  + {metric}: Trung bình = {avg:.2f}%, Lớn nhất = {max_val:.2f}% (Số log: {len(vals)})")
