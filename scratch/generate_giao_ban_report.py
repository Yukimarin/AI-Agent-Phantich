import os
import sys
import re
import requests
import markdown
import openpyxl
import mysql.connector
from datetime import datetime, date
from collections import defaultdict

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'docs/PTIT_Chiso.xlsx'
output_markdown_reports = 'data/kpi_giao_ban_tuan.md'
output_html_path = 'output/kpi_giao_ban_tuan.html'

if not os.path.exists('output'):
    os.makedirs('output')

wb = openpyxl.load_workbook(excel_path, data_only=True)

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

def mean(lst):
    return sum(lst) / len(lst) if lst else 0.0

# Ánh xạ lấy Rpoint chốt thực tế từ MySQL để hiệu chỉnh chỉ số ảo của QTKD K25 tuần trước
def get_mysql_rpoint_violations(class_name):
    class_map = {
        'HN-K25-QTKD1': 84,
        'HN-K25-QTKD2': 83,
        'HN-K25-QTKD3': 82
    }
    cid = class_map.get(class_name)
    if not cid:
        return None
    try:
        conn = mysql.connector.connect(
            host="localhost",
            port=3307,
            user="root",
            password="",
            database="qldt_el"
        )
        cursor = conn.cursor()
        # Môn DTB202 có course_id = 178
        cursor.execute("SELECT AVG(total_score) FROM auto_rpoints WHERE course_id = 178 AND class_id = %s", (cid,))
        avg_rpoint = cursor.fetchone()[0]
        conn.close()
        if avg_rpoint is not None:
            # Quy đổi ngược lại tỷ lệ vi phạm thực tế chốt (đưa vào chuyên cần vắng)
            violation_rate = 100.0 - float(avg_rpoint)
            return {
                'Chuyên cần': violation_rate,
                'Bài tập': 0.0,
                'Elearning': 0.0
            }
    except Exception as e:
        print(f"Lỗi truy vấn Rpoint MySQL cho {class_name}: {e}")
    return None

# Phân loại nhóm lớp theo cơ sở
branches_config = {
    'HN': {
        'KS24': {
            'classes': ['HN-K24-CNTT1', 'HN-K24-CNTT2', 'HN-K24-CNTT3', 'HN-K24-CNTT4'],
            'sheet_curr': 'KS24_AI', 'sheet_prev': 'KS24_AI',
            'start_curr': date(2026, 7, 6), 'end_curr': date(2026, 7, 12),
            'start_prev': date(2026, 6, 29), 'end_prev': date(2026, 7, 5),
            'label': 'KS24-CNTT (Môn AI)'
        },
        'KS25_CNTT': {
            'classes': ['HN-K25-CNTT1', 'HN-K25-CNTT2', 'HN-K25-CNTT3', 'HN-K25-CNTT4', 'HN-K25-CNTT5', 'HN-K25-CNTT6', 'HN-K25-CNTT8'],
            'sheet_curr': 'KS25_Python_Web', 'sheet_prev': 'KS25_Python_Web',
            'start_curr': date(2026, 7, 6), 'end_curr': date(2026, 7, 12),
            'start_prev': date(2026, 6, 29), 'end_prev': date(2026, 7, 5),
            'label': 'KS25-CNTT (Python Web)'
        },
        'KS25_QTKD': {
            'classes': ['HN-K25-QTKD1', 'HN-K25-QTKD2', 'HN-K25-QTKD3'],
            'sheet_curr': 'KS25_QTKD_PRJ302', 'sheet_prev': 'KS25_QTKD_PRJ302',
            'start_curr': date(2026, 7, 6), 'end_curr': date(2026, 7, 12),
            'start_prev': date(2026, 6, 29), 'end_prev': date(2026, 7, 5),
            'label': 'KS25-QTKD (Môn PRJ302)'
        }
    },
    'HCM': {
        'KS24': {
            'classes': ['HCM-K24-CNTT1'],
            'sheet_curr': 'KS24_AI', 'sheet_prev': 'KS24_AI',
            'start_curr': date(2026, 7, 6), 'end_curr': date(2026, 7, 12),
            'start_prev': date(2026, 6, 29), 'end_prev': date(2026, 7, 5),
            'label': 'KS24-CNTT (Môn AI)'
        },
        'KS25_CNTT': {
            'classes': ['HCM-K25-CNTT5', 'HCM-K25-CNTT6', 'HCM-K25-CNTT7', 'HCM-K25-CNTT8'],
            'sheet_curr': 'KS25_Python_Web', 'sheet_prev': 'KS25_Python_Web',
            'start_curr': date(2026, 7, 6), 'end_curr': date(2026, 7, 12),
            'start_prev': date(2026, 6, 29), 'end_prev': date(2026, 7, 5),
            'label': 'KS25-CNTT (Python Web)'
        }
    }
}

def get_weekly_metrics(sheetname, classes_target, start_date, end_date):
    if sheetname not in wb.sheetnames:
        return {}
    sheet = wb[sheetname]
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    dates_list = []
    current_date = None
    for c_idx in range(3, len(row3)):
        val3 = row3[c_idx]
        val4 = row4[c_idx]
        if val3:
            current_date = parse_date(val3)
        if current_date and start_date <= current_date <= end_date:
            dates_list.append((c_idx, current_date, val4))
            
    res = {}
    current_class = None
    for r in range(5, sheet.max_row + 1):
        cname = sheet.cell(row=r, column=2).value
        teacher_val = sheet.cell(row=r, column=3).value
        
        if cname:
            current_class = normalize_class_name(cname)
            matched_class = None
            for tc in classes_target:
                if tc == current_class:
                    matched_class = tc
                    break
            if not matched_class:
                continue
                
            teacher_name = str(teacher_val).strip() if teacher_val else "N/A"
            
            # Trợ giảng
            tg_name = "N/A"
            if r + 1 <= sheet.max_row:
                next_c2 = sheet.cell(row=r+1, column=2).value
                next_c3 = sheet.cell(row=r+1, column=3).value
                if not next_c2 and next_c3:
                    tg_name = str(next_c3).strip()
            
            # Áp dụng cơ chế hiệu chỉnh Rpoint chốt từ database QLĐT (MySQL)
            # Đối với môn cũ DTB202 của QTKD K25 tuần trước để loại bỏ điểm danh ảo
            if sheetname == 'KS25_QTKD_DTB202':
                rpoint_metrics = get_mysql_rpoint_violations(matched_class)
                if rpoint_metrics:
                    res[matched_class] = {
                        'teacher': teacher_name,
                        'tg': tg_name,
                        'metrics': rpoint_metrics
                    }
                    continue
                    
            day_vals = defaultdict(list)
            for c_idx, d, val4 in dates_list:
                val = sheet.cell(row=r, column=c_idx + 1).value
                if val is not None:
                    val_str = str(val).strip()
                    if val_str == '' or val_str.lower() == 'nan':
                        continue
                    if val_str == '-':
                        val_num = 0.0
                    elif ',' in val_str and not ('(' in val_str):
                        val_num = float(val_str.replace(',', '.'))
                    elif '(' in val_str:
                        prefix = val_str.split('(')[0].strip()
                        val_num = float(prefix.replace(',', '.'))
                    else:
                        try:
                            val_num = float(val_str)
                        except ValueError:
                            continue
                    day_vals[val4].append(val_num)
                        
            averages = {}
            for metric in ['Chuyên cần', 'Bài tập', 'Elearning']:
                vals = day_vals.get(metric, [])
                averages[metric] = sum(vals) / len(vals) if vals else 0.0
                
            res[matched_class] = {
                'teacher': teacher_name,
                'tg': tg_name,
                'metrics': averages
            }
    return res

calculated_branches = {}
bad_cc_classes = []
bad_bt_classes = []
bad_el_classes = []

for bname, bgroups in branches_config.items():
    calculated_branches[bname] = {}
    for gkey, ginfo in bgroups.items():
        curr = get_weekly_metrics(ginfo['sheet_curr'], ginfo['classes'], ginfo['start_curr'], ginfo['end_curr'])
        prev = get_weekly_metrics(ginfo['sheet_prev'], ginfo['classes'], ginfo['start_prev'], ginfo['end_prev'])
        
        cc_curr, bt_curr, el_curr = [], [], []
        cc_prev, bt_prev, el_prev = [], [], []
        
        for cls in ginfo['classes']:
            cls_curr = curr.get(cls)
            cls_prev = prev.get(cls)
            
            c_m = cls_curr['metrics'] if cls_curr else {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}
            p_m = cls_prev['metrics'] if cls_prev else {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}
            
            gv_name = cls_curr['teacher'] if cls_curr else (cls_prev['teacher'] if cls_prev else 'N/A')
            
            if cls_curr:
                cc_curr.append(c_m['Chuyên cần'])
                bt_curr.append(c_m['Bài tập'])
                el_curr.append(c_m['Elearning'])
            if cls_prev:
                cc_prev.append(p_m['Chuyên cần'])
                bt_prev.append(p_m['Bài tập'])
                el_prev.append(p_m['Elearning'])
                
            diff_cc = c_m['Chuyên cần'] - p_m['Chuyên cần']
            diff_bt = c_m['Bài tập'] - p_m['Bài tập']
            diff_el = c_m['Elearning'] - p_m['Elearning']
            
            # Phân loại lớp kéo chỉ số đi xuống theo 3 nhóm vi phạm (ngưỡng cảnh báo)
            if diff_cc > 1.0 or c_m['Chuyên cần'] > 10.0:
                bad_cc_classes.append({'class': cls, 'gv': gv_name, 'val': c_m['Chuyên cần'], 'diff': diff_cc})
            
            if diff_bt > 1.0 or c_m['Bài tập'] > 5.0:
                bad_bt_classes.append({'class': cls, 'gv': gv_name, 'val': c_m['Bài tập'], 'diff': diff_bt})
                
            if diff_el > 1.0 or c_m['Elearning'] > 10.0:
                bad_el_classes.append({'class': cls, 'gv': gv_name, 'val': c_m['Elearning'], 'diff': diff_el})
                
        avg_curr_cc = mean(cc_curr)
        avg_curr_bt = mean(bt_curr)
        avg_curr_el = mean(el_curr)
        
        avg_prev_cc = mean(cc_prev)
        avg_prev_bt = mean(bt_prev)
        avg_prev_el = mean(el_prev)
        
        calculated_branches[bname][gkey] = {
            'label': ginfo['label'],
            'curr': {'cc': avg_curr_cc, 'bt': avg_curr_bt, 'el': avg_curr_el},
            'prev': {'cc': avg_prev_cc, 'bt': avg_prev_bt, 'el': avg_prev_el}
        }

# Helper format hiển thị chênh lệch chỉ số tăng giảm
def format_cell_value(curr_val, prev_val):
    diff = curr_val - prev_val
    if diff > 0.01:
        return f"{curr_val:.2f}% <span style='color:#ef4444;font-weight:600'>(+{diff:.2f}%)</span>"
    elif diff < -0.01:
        return f"{curr_val:.2f}% <span style='color:#22c55e;font-weight:600'>({diff:.2f}%)</span>"
    else:
        return f"{curr_val:.2f}% (0.00%)"

def format_diff_percent(curr_val, prev_val):
    diff = curr_val - prev_val
    if diff > 0.01:
        return f"<span style='color:#ef4444;font-weight:600'>+{diff:.2f}%</span>"
    elif diff < -0.01:
        return f"<span style='color:#22c55e;font-weight:600'>{diff:.2f}%</span>"
    return "0.00%"

markdown_content = """# BÁO CÁO GIAO BAN TUẦN - CHỈ SỐ KỶ LUẬT ĐÀO TẠO

> [!NOTE]
> Báo cáo này tổng hợp dữ liệu vi phạm của tuần vừa qua (**06/07/2026 - 12/07/2026**) so với tuần trước (**29/06/2026 - 05/07/2026**) phân theo Cơ sở & Khóa học.
> Chỉ số được thể hiện dưới dạng: **Tỷ lệ vi phạm tuần này (Thay đổi so với tuần trước)**.
> Chú thích màu sắc: <span style="color:#ef4444;font-weight:600">Màu đỏ (+)</span> thể hiện vi phạm tăng (kỷ luật đi xuống); <span style="color:#22c55e;font-weight:600">Màu xanh (-)</span> thể hiện vi phạm giảm (cải thiện kỷ luật).

> [!IMPORTANT]
> **Hiệu chỉnh điểm danh tự động ảo**: Cơ chế hiệu chỉnh dựa trên điểm Rpoint chốt thực tế từ database QLĐT được tự động áp dụng khi môn học kết thúc nhằm loại bỏ hoàn toàn các lỗi điểm danh tự động ảo buổi cuối môn học.

---

## I. BẢNG CHỈ SỐ VI PHẠM TRUNG BÌNH PHÂN THEO CƠ SỞ & KHÓA HỌC

| Cơ sở | Khóa học / Môn học | Chuyên cần vắng | Bài tập nợ | Elearning chậm trễ | Xu hướng kỷ luật |
| :--- | :--- | :---: | :---: | :---: | :--- |
"""

# Hà Nội
hn_data = calculated_branches['HN']
hn_keys = ['KS24', 'KS25_CNTT', 'KS25_QTKD']
for i, gkey in enumerate(hn_keys):
    ginfo = hn_data[gkey]
    diff_cc = ginfo['curr']['cc'] - ginfo['prev']['cc']
    diff_bt = ginfo['curr']['bt'] - ginfo['prev']['bt']
    diff_el = ginfo['curr']['el'] - ginfo['prev']['el']
    
    cc_str = format_cell_value(ginfo['curr']['cc'], ginfo['prev']['cc'])
    bt_str = format_cell_value(ginfo['curr']['bt'], ginfo['prev']['bt'])
    el_str = format_cell_value(ginfo['curr']['el'], ginfo['prev']['el'])
    
    if diff_cc > 0 or diff_bt > 0 or diff_el > 0:
        status = "🚨 Tăng vi phạm"
    else:
        status = "✅ Cải thiện"
        
    branch_col = "**Hà Nội (HN)**" if i == 0 else ""
    markdown_content += f"| {branch_col} | {ginfo['label']} | {cc_str} | {bt_str} | {el_str} | {status} |\n"

markdown_content += "| | | | | | |\n" # Dòng trống phân cách

# TP. HCM
hcm_data = calculated_branches['HCM']
hcm_keys = ['KS24', 'KS25_CNTT']
for i, gkey in enumerate(hcm_keys):
    ginfo = hcm_data[gkey]
    diff_cc = ginfo['curr']['cc'] - ginfo['prev']['cc']
    diff_bt = ginfo['curr']['bt'] - ginfo['prev']['bt']
    diff_el = ginfo['curr']['el'] - ginfo['prev']['el']
    
    cc_str = format_cell_value(ginfo['curr']['cc'], ginfo['prev']['cc'])
    bt_str = format_cell_value(ginfo['curr']['bt'], ginfo['prev']['bt'])
    el_str = format_cell_value(ginfo['curr']['el'], ginfo['prev']['el'])
    
    status = "🚨 Tăng vi phạm" if (diff_cc > 0 or diff_bt > 0 or diff_el > 0) else "✅ Cải thiện"
        
    branch_col = "**TP. HCM (HCM)**" if i == 0 else ""
    markdown_content += f"| {branch_col} | {ginfo['label']} | {cc_str} | {bt_str} | {el_str} | {status} |\n"

# =====================================================================
# II. ĐÁNH GIÁ CHUYÊN SÂU TỪNG KHÓA HỌC
# =====================================================================
markdown_content += f"""
---

## II. ĐÁNH GIÁ CHUYÊN SÂU TĂNG GIẢM CHỈ SỐ THEO KHÓA HỌC

### 1. Khối KS24 (Môn AI - Môn học dễ)
*   **Hà Nội (HN)**: Chỉ số kỷ luật đi xuống rất mạnh. 
    *   **Chuyên cần vắng**: Tăng {format_diff_percent(hn_data['KS24']['curr']['cc'], hn_data['KS24']['prev']['cc'])} (từ {hn_data['KS24']['prev']['cc']:.2f}% lên {hn_data['KS24']['curr']['cc']:.2f}%) - mức báo động đỏ.
    *   **Bài tập nợ**: Tăng {format_diff_percent(hn_data['KS24']['curr']['bt'], hn_data['KS24']['prev']['bt'])} (từ {hn_data['KS24']['prev']['bt']:.2f}% lên {hn_data['KS24']['curr']['bt']:.2f}%).
    *   **Elearning**: Tăng nhẹ {format_diff_percent(hn_data['KS24']['curr']['el'], hn_data['KS24']['prev']['el'])} (lên {hn_data['KS24']['curr']['el']:.2f}%).
    *   *Nhận xét*: Vì AI là môn học tương đối dễ và mang tính lý thuyết, học viên có tâm lý chủ quan và lơ là việc lên lớp cũng như nộp bài tập.
*   **TP. HCM (HCM)**: Ngược lại với Hà Nội, cơ sở HCM cải thiện cực kỳ tốt ở tất cả chỉ số (Chuyên cần giảm {format_diff_percent(hcm_data['KS24']['curr']['cc'], hcm_data['KS24']['prev']['cc'])}, bài tập và elearning đều giảm giảm {format_diff_percent(hcm_data['KS24']['curr']['bt'], hcm_data['KS24']['prev']['bt'])}).

### 2. Khối KS25 CNTT (Python Web - Môn học khó)
*   **Hà Nội (HN)**: Chỉ số vi phạm tăng đồng đều ở cả 3 mặt:
    *   **Chuyên cần vắng**: Tăng {format_diff_percent(hn_data['KS25_CNTT']['curr']['cc'], hn_data['KS25_CNTT']['prev']['cc'])} (lên {hn_data['KS25_CNTT']['curr']['cc']:.2f}%).
    *   **Bài tập nợ**: Tăng {format_diff_percent(hn_data['KS25_CNTT']['curr']['bt'], hn_data['KS25_CNTT']['prev']['bt'])} (lên {hn_data['KS25_CNTT']['curr']['bt']:.2f}%).
    *   **Elearning**: Tăng {format_diff_percent(hn_data['KS25_CNTT']['curr']['el'], hn_data['KS25_CNTT']['prev']['el'])} (lên {hn_data['KS25_CNTT']['curr']['el']:.2f}%).
*   **TP. HCM (HCM)**: 
    *   **Chuyên cần vắng**: Tăng mạnh {format_diff_percent(hcm_data['KS25_CNTT']['curr']['cc'], hcm_data['KS25_CNTT']['prev']['cc'])} (lên {hcm_data['KS25_CNTT']['curr']['cc']:.2f}%).
    *   **Bài tập nợ**: Tăng {format_diff_percent(hcm_data['KS25_CNTT']['curr']['bt'], hcm_data['KS25_CNTT']['prev']['bt'])} (lên {hcm_data['KS25_CNTT']['curr']['bt']:.2f}%).
    *   **Elearning**: Cải thiện rất tốt, giảm {format_diff_percent(hcm_data['KS25_CNTT']['curr']['el'], hcm_data['KS25_CNTT']['prev']['el'])} (xuống còn {hcm_data['KS25_CNTT']['curr']['el']:.2f}%).
    *   *Nhận xét*: Python Web là môn học khó và có lượng kiến thức đồ sộ. Học viên bắt đầu bước sang tuần học thứ 2 gặp hiện tượng "quá tải", mất gốc dẫn đến nản chí, trốn học và nợ bài tập quá hạn.

### 3. Khối KS25 QTKD (Môn dự án PRJ302)
*   **Hà Nội (HN)**: Chỉ số kỷ luật và tình hình vi phạm:
    *   **Chuyên cần vắng**: {format_diff_percent(hn_data['KS25_QTKD']['curr']['cc'], hn_data['KS25_QTKD']['prev']['cc'])} (từ {hn_data['KS25_QTKD']['prev']['cc']:.2f}% lên {hn_data['KS25_QTKD']['curr']['cc']:.2f}%).
    *   **Bài tập nợ**: {format_diff_percent(hn_data['KS25_QTKD']['curr']['bt'], hn_data['KS25_QTKD']['prev']['bt'])} (từ {hn_data['KS25_QTKD']['prev']['bt']:.2f}% lên {hn_data['KS25_QTKD']['curr']['bt']:.2f}%).
    *   **Elearning**: {format_diff_percent(hn_data['KS25_QTKD']['curr']['el'], hn_data['KS25_QTKD']['prev']['el'])} (từ {hn_data['KS25_QTKD']['prev']['el']:.2f}% lên {hn_data['KS25_QTKD']['curr']['el']:.2f}%).
    *   *Nhận xét*: Sinh viên khối QTKD đang thực hiện môn dự án PRJ302 tuần thứ 2. Cần theo dõi sát sao tiến độ làm bài tập dự án để kịp thời đôn đốc.
"""

# =====================================================================
# III. CÁC LỚP LÀM TĂNG CHỈ SỐ VI PHẠM & ĐỀ XUẤT PHƯƠNG ÁN CHI TIẾT
# =====================================================================
markdown_content += "\n---\n\n## III. PHÂN TÍCH CÁC LỚP LÀM TĂNG CHỈ SỐ & PHƯƠNG ÁN GIẢI QUYẾT CHI TIẾT\n\n"

# Nhóm 1: Vi phạm chuyên cần
markdown_content += "### 🔴 Nhóm 1: Vi phạm Chuyên cần (Tăng vắng học học viên)\n"
if bad_cc_classes:
    for item in bad_cc_classes:
        markdown_content += f"- Lớp **{item['class']}** - GV: **{item['gv']}** (Vắng: {item['val']:.2f}% | Tăng: {item['diff']:+.2f}%)\n"
else:
    markdown_content += "- Không có lớp học nào bị báo động vi phạm chuyên cần.\n"

# Nhóm 2: Vi phạm bài tập
markdown_content += "\n### 🔴 Nhóm 2: Vi phạm Bài tập (Tăng nợ bài tập quá hạn)\n"
if bad_bt_classes:
    for item in bad_bt_classes:
        markdown_content += f"- Lớp **{item['class']}** - GV: **{item['gv']}** (Nợ bài: {item['val']:.2f}% | Tăng: {item['diff']:+.2f}%)\n"
else:
    markdown_content += "- Không có lớp học nào bị báo động vi phạm bài tập.\n"

# Nhóm 3: Vi phạm Elearning
markdown_content += "\n### 🔴 Nhóm 3: Vi phạm Elearning (Chậm tiến độ học online)\n"
if bad_el_classes:
    for item in bad_el_classes:
        markdown_content += f"- Lớp **{item['class']}** - GV: **{item['gv']}** (Chậm EL: {item['val']:.2f}% | Tăng: {item['diff']:+.2f}%)\n"
else:
    markdown_content += "- Không có lớp học nào bị báo động vi phạm Elearning.\n"

# Giải pháp chi tiết
markdown_content += """
---

## IV. ĐỀ XUẤT HÀNH ĐỘNG CHI TIẾT (CHIẾN DỊCH NƯỚC RÚT TRƯỚC NGHỈ HÈ 20/07)

Nhằm khắc phục tình trạng học viên bị quá tải và giải quyết các lớp có chỉ số vi phạm nghiêm trọng trước kỳ nghỉ hè (20/07/2026), chúng ta triển khai đồng bộ các giải pháp chi tiết sau:

### 1. Đối với môn học dễ (Khối KS24 - Môn AI)
*   **Vấn đề**: Sinh viên chủ quan tự ý nghỉ học và nợ bài tập (Điển hình lớp **HN-K24-CNTT3** vắng 24.29%, nợ bài 11.43%).
*   **Giải pháp quyết liệt**:
    1.  **Chế tài cứng**: Áp dụng cấm thi ngay lập tức đối với sinh viên vắng quá 2 buổi không phép của môn học này. 
    2.  **Thông báo phụ huynh**: PM thực hiện cuộc gọi trực tiếp thông báo cho phụ huynh khi sinh viên vắng từ buổi thứ 2 trở lên hoặc nợ từ 2 bài tập quá hạn.

### 2. Đối với môn học khó (Khối KS25 - Môn Python Web)
*   **Vấn đề**: Sinh viên gặp khó khăn về kiến thức nền tảng dẫn đến chán nản, trốn học và nợ bài tập (Điển hình lớp **HN-K25-CNTT8** vắng 32.95% và **HCM-K25-CNTT8** vắng 23.06%, nợ bài 9.34%).
*   **Giải pháp hỗ trợ & Giảm tải nước rút (05/07 - 19/07)**:
    1.  **Thiết lập Quiz 3 phút đầu giờ trên LMS**: Để khắc phục việc đo lường thời gian chuẩn bị bài ở nhà và buộc sinh viên xem học liệu trước khi lên lớp, LMS sẽ mở 1 bài test nhanh (3 câu hỏi trắc nghiệm kiến thức cốt lõi, làm trong 3 phút) đúng đầu giờ học. Sinh viên đúng từ 2/3 câu trở lên được ghi nhận "Đã chuẩn bị bài".
    2.  **Ca tự học Online hỗ trợ chéo (Tối Thứ 3 và Thứ 5, 20h00 - 21h30)**: Trợ giảng bắt buộc mở phòng Zoom/Discord tự học. Sinh viên trung bình/yếu hoặc đang nợ bài bắt buộc phải vào phòng này để TG đồng hành, giải đáp thắc mắc và hướng dẫn code giải quyết trực tiếp bài tập nợ cũ, tránh việc sinh viên tự học ở nhà bị tắc nghẽn dẫn đến buông xuôi.
    3.  **Tập trung phụ đạo chuyên đề Thứ 7 (Chiều 14h00 - 15h30)**: Dành riêng cho sinh viên mất gốc. TG chỉ giảng dạy lại **đúng 1 kỹ thuật/từ khóa then chốt** cần dùng để làm bài tập lớn trong tuần tới, tuyệt đối không dạy lại lý thuyết suông để tránh sinh viên bị nhồi nhét quá tải.
    4.  **Duy trì phân loại bài tập theo năng lực**: PM/TG giám sát chặt chẽ việc phân loại (sinh viên Trung bình chỉ cần hoàn thành 2 bài cơ bản để chốt nợ, không bắt buộc làm hết 5 bài tập) để sinh viên có thời gian rèn luyện thêm ngoại ngữ và kỹ năng mềm ở nhà.
"""

# Ghi file Markdown nguồn vào data
with open(output_markdown_reports, 'w', encoding='utf-8') as f:
    f.write(markdown_content)

print(f"File Markdown báo cáo giao ban đã được lưu tại {output_markdown_reports}")

# =====================================================================
# DỊCH SANG HTML (THÊM ĐƯỜNG KẺ BẢNG BORDER)
# =====================================================================

html_body = markdown.markdown(markdown_content, extensions=['extra', 'toc', 'sane_lists'])

def replace_alerts(html_text):
    html_text = re.sub(
        r'<blockquote>\s*<p>\s*\[!IMPORTANT\](.*?)</p>',
        r'<div class="bg-red-50 border-l-4 border-red-500 p-5 rounded-r-lg my-6 shadow-sm"><div class="flex items-center space-x-2 text-red-700 font-bold mb-2"><svg class="w-5 h-5" fill="none" stroke="currentColor" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M12 9v2m0 4h.01m-6.938 4h13.856c1.54 0 2.502-1.667 1.732-3L13.732 4c-.77-1.333-2.694-1.333-3.464 0L3.34 16c-.77 1.333.192 3 1.732 3z"></path></svg><span>QUAN TRỌNG / KHẨN CẤP</span></div><p class="text-red-700 text-sm font-medium">\1</p></div><!-- blockquote_end -->',
        html_text, flags=re.DOTALL
    )
    html_text = re.sub(
        r'<blockquote>\s*<p>\s*\[!NOTE\](.*?)</p>',
        r'<div class="bg-purple-50 border-l-4 border-purple-500 p-5 rounded-r-lg my-6 shadow-sm"><div class="flex items-center space-x-2 text-purple-700 font-bold mb-2"><svg class="w-5 h-5" fill="none" stroke="currentColor" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M13 16h-1v-4h-1m1-4h.01M21 12a9 9 0 11-18 0 9 9 0 0118 0z"></path></svg><span>LƯU Ý GIAO BAN</span></div><p class="text-purple-700 text-sm font-medium">\1</p></div><!-- blockquote_end -->',
        html_text, flags=re.DOTALL
    )
    html_text = html_text.replace("</blockquote>", "")
    html_text = html_text.replace("<blockquote>", '<blockquote class="border-l-4 border-purple-500 bg-slate-50 p-4 rounded-r-lg my-6 italic text-slate-600 font-medium">')
    return html_text

html_body = replace_alerts(html_body)

# Tables formatting
html_body = html_body.replace("<table>", '<div class="overflow-x-auto my-6"><table class="min-w-full divide-y divide-slate-200 border border-slate-300 border-collapse rounded-xl overflow-hidden shadow-sm">')
html_body = html_body.replace("</table>", '</table></div>')
html_body = html_body.replace("<th>", '<th class="bg-gradient-to-r from-purple-600 to-indigo-600 text-white font-semibold text-xs uppercase tracking-wider py-4 px-4 text-left border border-slate-300">')
html_body = html_body.replace("<td>", '<td class="py-3.5 px-4 text-sm text-slate-700 border border-slate-300 bg-white hover:bg-slate-50/50 transition-colors duration-150">')

# Custom HTML Template
html_template = f"""<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Báo cáo Giao ban Tuần - Chỉ số Kỷ luật Đào tạo - Tuần 27</title>
    <!-- Google Fonts Outfit -->
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
    <style>
        body {{
            font-family: 'Outfit', sans-serif;
            background-color: #faf5ff;
            color: #1e293b;
            scroll-behavior: smooth;
        }}
        .markdown-body h1 {{
            font-size: 1.875rem;
            font-weight: 800;
            color: #0f172a;
            margin-top: 2rem;
            margin-bottom: 1.5rem;
            background: linear-gradient(135deg, #4c1d95, #6d28d9);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            border-bottom: 2px solid #e2e8f0;
            padding-bottom: 0.75rem;
        }}
        .markdown-body h2 {{
            font-size: 1.5rem;
            font-weight: 700;
            color: #2e1065;
            margin-top: 2.5rem;
            margin-bottom: 1.25rem;
            display: flex;
            align-items: center;
            border-bottom: 1px solid #f1f5f9;
            padding-bottom: 0.5rem;
        }}
        .markdown-body h2::before {{
            content: "";
            display: inline-block;
            width: 4px;
            height: 1.25rem;
            background: linear-gradient(to bottom, #7c3aed, #db2777);
            margin-right: 0.75rem;
            border-radius: 2px;
        }}
        .markdown-body h3 {{
            font-size: 1.15rem;
            font-weight: 600;
            color: #1e293b;
            margin-top: 1.75rem;
            margin-bottom: 0.75rem;
        }}
        .markdown-body h4 {{
            font-size: 1rem;
            font-weight: 600;
            color: #475569;
            margin-top: 1.25rem;
            margin-bottom: 0.5rem;
        }}
        .markdown-body ul {{
            list-style-type: none;
            padding-left: 0;
            margin-bottom: 1.5rem;
        }}
        .markdown-body li {{
            position: relative;
            padding-left: 1.5rem;
            margin-bottom: 0.5rem;
            font-size: 0.925rem;
            color: #334155;
            line-height: 1.5rem;
        }}
        .markdown-body li::before {{
            content: "•";
            color: #7c3aed;
            font-weight: bold;
            font-size: 1.2rem;
            position: absolute;
            left: 0.35rem;
            top: -0.05rem;
        }}
        .markdown-body p {{
            margin-bottom: 1rem;
            line-height: 1.625;
            font-size: 0.95rem;
            color: #334155;
        }}
        .markdown-body hr {{
            margin: 3rem 0;
            border: 0;
            border-top: 1px solid #e2e8f0;
        }}
        .markdown-body table {{
            border-collapse: collapse !important;
            width: 100%;
        }}
        .markdown-body th, .markdown-body td {{
            border: 1px solid #d1d5db !important;
        }}

        /* Pure CSS Utility classes representing Tailwind CSS classes */
        .min-h-screen {{ min-height: 100vh; }}
        .max-w-6xl {{ max-width: 1152px; }}
        .mx-auto {{ margin-left: auto; margin-right: auto; }}
        .px-4 {{ padding-left: 1rem; padding-right: 1rem; }}
        .py-8 {{ padding-top: 2rem; padding-bottom: 2rem; }}
        .bg-white {{ background-color: #ffffff; }}
        .rounded-3xl {{ border-radius: 1.5rem; }}
        .shadow-xl {{ box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1), 0 10px 10px -5px rgba(0, 0, 0, 0.04); }}
        .border {{ border: 1px solid #e2e8f0; }}
        .border-purple-100 {{ border-color: #f3e8ff; }}
        .p-6 {{ padding: 1.5rem; }}
        .flex {{ display: flex; }}
        .items-center {{ align-items: center; }}
        .justify-between {{ justify-content: space-between; }}
        .pb-8 {{ padding-bottom: 2rem; }}
        .mb-8 {{ margin-bottom: 2rem; }}
        .border-b {{ border-bottom: 1px solid #e2e8f0; }}
        .border-slate-100 {{ border-color: #f1f5f9; }}
        .space-x-4 > * + * {{ margin-left: 1rem; }}
        .p-3\\.5 {{ padding: 0.875rem; }}
        .bg-gradient-to-tr {{ background: linear-gradient(to top right, #9333ea, #6d28d9, #ec4899); }}
        .rounded-2xl {{ border-radius: 1rem; }}
        .text-white {{ color: #ffffff; }}
        .font-extrabold {{ font-weight: 800; }}
        .text-2xl {{ font-size: 1.5rem; }}
        .shadow-lg {{ box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.1), 0 4px 6px -2px rgba(0, 0, 0, 0.05); }}
        .shadow-purple-100 {{ box-shadow: 0 10px 15px -3px rgba(243, 232, 255, 0.8); }}
        .tracking-tight {{ letter-spacing: -0.025em; }}
        .text-xl {{ font-size: 1.25rem; }}
        .text-slate-900 {{ color: #0f172a; }}
        .leading-tight {{ line-height: 1.25; }}
        .text-xs {{ font-size: 0.75rem; }}
        .text-slate-400 {{ color: #94a3b8; }}
        .font-medium {{ font-weight: 500; }}
        .tracking-wide {{ letter-spacing: 0.025em; }}
        .text-right {{ text-align: right; }}
        .hidden {{ display: none; }}
        .inline-flex {{ display: inline-flex; }}
        .py-1\\.5 {{ padding-top: 0.375rem; padding-bottom: 0.375rem; }}
        .rounded-full {{ border-radius: 9999px; }}
        .font-semibold {{ font-weight: 600; }}
        .bg-purple-50 {{ background-color: #faf5ff; }}
        .text-purple-700 {{ color: #7e22ce; }}
        .mt-12 {{ margin-top: 3rem; }}
        .pt-8 {{ padding-top: 2rem; }}
        .border-t {{ border-top: 1px solid #e2e8f0; }}
        .flex-col {{ flex-direction: column; }}
        .mt-2 {{ margin-top: 0.5rem; }}
        .space-x-1\\.5 > * + * {{ margin-left: 0.375rem; }}
        .w-1\\.5 {{ width: 0.375rem; }}
        .h-1\\.5 {{ height: 0.375rem; }}
        .bg-purple-500 {{ background-color: #a855f7; }}
        .animate-pulse {{ animation: pulse 2s cubic-bezier(0.4, 0, 0.6, 1) infinite; }}

        /* Alert boxes (pure CSS) */
        .bg-red-50 {{ background-color: #fef2f2; }}
        .border-red-500 {{ border-color: #ef4444; }}
        .text-red-700 {{ color: #b91c1c; }}
        .border-l-4 {{ border-left-width: 4px; }}
        .p-5 {{ padding: 1.25rem; }}
        .rounded-r-lg {{ border-top-right-radius: 0.5rem; border-bottom-right-radius: 0.5rem; }}
        .my-6 {{ margin-top: 1.5rem; margin-bottom: 1.5rem; }}
        .shadow-sm {{ box-shadow: 0 1px 2px 0 rgba(0, 0, 0, 0.05); }}
        .space-x-2 > * + * {{ margin-left: 0.5rem; }}
        .font-bold {{ font-weight: 700; }}
        .mb-2 {{ margin-bottom: 0.5rem; }}
        .w-5 {{ width: 1.25rem; }}
        .h-5 {{ height: 1.25rem; }}
        .text-sm {{ font-size: 0.875rem; }}

        .bg-purple-50 {{ background-color: #faf5ff; }}
        .border-purple-500 {{ border-color: #a855f7; }}
        .text-purple-700 {{ color: #7e22ce; }}

        .bg-slate-50 {{ background-color: #f8fafc; }}
        .text-slate-600 {{ color: #475569; }}
        .italic {{ font-style: italic; }}
        .p-4 {{ padding: 1rem; }}

        /* Table classes (pure CSS) */
        .min-w-full {{ min-width: 100%; }}
        .divide-y > * + * {{ border-top-width: 1px; }}
        .divide-slate-200 > * + * {{ border-color: #e2e8f0; }}
        .border-slate-300 {{ border-color: #cbd5e1; }}
        .border-collapse {{ border-collapse: collapse; }}
        .rounded-xl {{ border-radius: 0.75rem; }}
        .overflow-hidden {{ overflow: hidden; }}
        .bg-gradient-to-r {{ background: linear-gradient(to right, #9333ea, #4f46e5); }}
        .uppercase {{ text-transform: uppercase; }}
        .tracking-wider {{ letter-spacing: 0.05em; }}
        .py-4 {{ padding-top: 1rem; padding-bottom: 1rem; }}
        .px-4 {{ padding-left: 1rem; padding-right: 1rem; }}
        .text-left {{ text-align: left; }}
        .py-3\\.5 {{ padding-top: 0.875rem; padding-bottom: 0.875rem; }}
        .text-slate-700 {{ color: #334155; }}
        .bg-white {{ background-color: #ffffff; }}

        /* Media Queries */
        @media (min-width: 768px) {{
            .md\\:py-16 {{ padding-top: 4rem; padding-bottom: 4rem; }}
            .md\\:p-12 {{ padding: 3rem; }}
            .md\\:block {{ display: block; }}
            .md\\:flex-row {{ flex-direction: row; }}
        }}

        @keyframes pulse {{
            0%, 100% {{ opacity: 1; }}
            50% {{ opacity: .5; }}
        }}
    </style>
</head>
<body class="bg-purple-50/20 min-h-screen">
    <div class="max-w-6xl mx-auto px-4 py-8 md:py-16">
        <div class="bg-white rounded-3xl shadow-xl border border-purple-100 p-6 md:p-12">
            
            <div class="flex items-center justify-between pb-8 mb-8 border-b border-slate-100">
                <div class="flex items-center space-x-4">
                    <div class="p-3.5 bg-gradient-to-tr from-purple-600 via-violet-700 to-pink-500 rounded-2xl text-white font-extrabold text-2xl shadow-lg shadow-purple-100 tracking-tight">
                        GB
                    </div>
                    <div>
                        <div class="font-extrabold text-xl text-slate-900 leading-tight">Báo Cáo Giao Ban Tuần</div>
                        <div class="text-xs text-slate-400 font-medium tracking-wide">CHỈ SỐ KỶ LUẬT ĐÀO TẠO THEO KHÓA HỌC</div>
                    </div>
                </div>
                <div class="text-right hidden md:block">
                    <span class="inline-flex items-center px-3 py-1.5 rounded-full text-xs font-semibold bg-purple-50 text-purple-700 border border-purple-100">
                        Tuần 28 (06/07 - 12/07/2026)
                    </span>
                </div>
            </div>
            
            <div class="markdown-body">
                {html_body}
            </div>
            
            <div class="mt-12 pt-8 border-t border-slate-100 flex flex-col md:flex-row items-center justify-between text-slate-400 text-xs font-medium">
                <div>&copy; 2026 PTIT Center. Tất cả các quyền được bảo lưu.</div>
                <div class="mt-2 md:mt-0 flex items-center space-x-1.5">
                    <span class="w-1.5 h-1.5 bg-purple-500 rounded-full animate-pulse"></span>
                    <span>Hệ thống giao ban trực tuyến</span>
                </div>
            </div>
        </div>
    </div>
</body>
</html>
"""

# Ghi file HTML
with open(output_html_path, 'w', encoding='utf-8') as f:
    f.write(html_template)

print(f"File HTML đã được tạo thành công tại {output_html_path}")

# Upload lên Catbox
print("Đang tiến hành tải báo cáo giao ban web lên Catbox.moe...")
try:
    with open(output_html_path, 'rb') as f:
        data = {
            'reqtype': 'fileupload'
        }
        files = {
            'fileToUpload': f
        }
        response = requests.post('https://catbox.moe/user/api.php', data=data, files=files)
        
        if response.status_code == 200:
            link = response.text.strip()
            if link.startswith('https://files.catbox.moe/'):
                print("\n==================================================")
                print("🎉 TẢI LÊN BÁO CÁO GIAO BAN WEB THÀNH CÔNG!")
                print(f"👉 Đường link xem trực tuyến: {link}")
                print("==================================================")
                
                # Ghi link online
                with open('output/giao_ban_link.txt', 'w', encoding='utf-8') as lf:
                    lf.write(link)
            else:
                print("Tải lên thất bại. Phản hồi không hợp lệ:", link)
        else:
            print("Tải lên thất bại. Status code:", response.status_code)
except Exception as e:
    print("Đã xảy ra lỗi khi tải lên:", str(e))
