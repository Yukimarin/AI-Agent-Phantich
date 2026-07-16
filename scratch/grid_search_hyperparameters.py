import mysql.connector
import sys
import openpyxl
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

def mean(lst):
    return sum(lst) / len(lst) if lst else 0.0

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = (name_str
                .replace("KS25", "K25")
                .replace("KS24", "K24")
                .replace("KS23", "K23"))
    return name_str

course_to_sheet_map = {
    'javascript': 'KS25_Javascript',
    'cơ sở dữ liệu': 'KS25_Database',
    'database': 'KS25_Database',
    'python': 'KS25_Python',
    'java fundamental': 'KS24-JavaAdvance',
    'java advance': 'KS24-JavaAdvance',
    'java web application': 'KS24_JavaWeb',
    'java web service': 'KS24_JWS',
    'agile': 'KS24_JavaWeb',
    'trí tuệ': 'KS24_AI',
    'ai': 'KS24_AI'
}

def get_excel_chot_data(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    class_data = defaultdict(dict)
    for sheetname in wb.sheetnames:
        if sheetname == 'Sheet1' or 'SKL' in sheetname:
            continue
        sheet = wb[sheetname]
        max_r = sheet.max_row
        max_c = sheet.max_column
        if max_r < 5:
            continue
        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
        dates_list = []
        for c_idx in range(3, max_c):
            val3 = row3[c_idx]
            val4 = row4[c_idx]
            if val3:
                dates_list.append((c_idx, val3, val4))
            elif dates_list:
                dates_list.append((c_idx, dates_list[-1][1], val4))
        
        # Find Rpoint col (only after the last daily column)
        rp_col_idx = None
        last_daily_col = max(idx for idx, d, sub in dates_list) if dates_list else 2
        for c_idx in range(max_c - 1, last_daily_col, -1):
            h3 = row3[c_idx]
            h4 = row4[c_idx]
            if h4 in ('Chuyên cần', 'Bài tập', 'Elearning'):
                continue
            vals = []
            for r in range(5, max_r + 1):
                val = sheet.cell(row=r, column=c_idx + 1).value
                if val is not None:
                    try: vals.append(float(val))
                    except: pass
            if len(vals) >= 2:
                avg_val = mean(vals)
                if 30.0 <= avg_val <= 115.0:
                    rp_col_idx = c_idx
                    break
        
        # Read class rows
        for r in range(5, max_r + 1):
            cname = sheet.cell(row=r, column=2).value
            if cname:
                norm_name = normalize_class_name(cname)
                cc_val, bt_val, el_val = None, None, None
                for c_idx, d, sub in reversed(dates_list):
                    val = sheet.cell(row=r, column=c_idx + 1).value
                    if val is not None:
                        try:
                            val_f = float(val)
                            if sub == 'Chuyên cần' and cc_val is None: cc_val = val_f
                            elif sub == 'Bài tập' and bt_val is None: bt_val = val_f
                            elif sub == 'Elearning' and el_val is None: el_val = val_f
                        except: pass
                
                rp_val = None
                if rp_col_idx is not None:
                    val = sheet.cell(row=r, column=rp_col_idx + 1).value
                    if val is not None:
                        try: rp_val = float(val)
                        except: pass
                        
                class_data[norm_name][sheetname] = {
                    'cc': cc_val,
                    'bt': bt_val,
                    'el': el_val,
                    'rp': rp_val
                }
    return class_data

def calibrate_students(students_results, excel_disc):
    att_list = [float(s['attendance']) if s['attendance'] is not None else 0.0 for s in students_results]
    hw_list = [float(s['homework']) if s['homework'] is not None else 100.0 for s in students_results]
    el_list = [float(s['elearning']) if s['elearning'] is not None else 0.0 for s in students_results]
    rp_list = [float(s['rpoints']) if s['rpoints'] is not None else 100.0 for s in students_results]
    
    if not excel_disc:
        calibrated = []
        for i, s in enumerate(students_results):
            calibrated.append({
                'student_id': s['student_id'],
                'attendance': att_list[i],
                'homework': hw_list[i],
                'elearning': el_list[i],
                'rpoints': rp_list[i],
                'hackathon_1': s['hackathon_1'],
                'hackathon_2': s['hackathon_2'],
                'project': s['project'],
                'pass': s['pass']
            })
        return calibrated
        
    db_att_avg = mean(att_list)
    db_hw_avg = mean(hw_list)
    db_el_avg = mean(el_list)
    db_rp_avg = mean(rp_list)
    
    excel_cc = excel_disc['cc'] if excel_disc['cc'] is not None else db_att_avg
    excel_bt_err = excel_disc['bt'] if excel_disc['bt'] is not None else (100.0 - db_hw_avg)
    excel_hw = 100.0 - excel_bt_err
    excel_el = excel_disc['el'] if excel_disc['el'] is not None else db_el_avg
    excel_rp = excel_disc['rp'] if excel_disc['rp'] is not None else None
    if excel_rp is None:
        excel_rp = max(0.0, 100.0 - excel_cc - excel_bt_err - excel_el)
        
    calibrated = []
    for i, s in enumerate(students_results):
        rp_db = rp_list[i]
        rp_cal = rp_db + (excel_rp - db_rp_avg)
        rp_cal = min(120.0, max(0.0, rp_cal))
        
        att_db = att_list[i]
        att_cal = att_db + (excel_cc - db_att_avg)
        att_cal = min(100.0, max(0.0, att_cal))
        
        hw_db = hw_list[i]
        hw_cal = hw_db + (excel_hw - db_hw_avg)
        hw_cal = min(100.0, max(0.0, hw_cal))
        
        el_db = el_list[i]
        el_cal = 0.0 if excel_disc.get('el') == 0.0 else el_db
            
        calibrated.append({
            'student_id': s['student_id'],
            'attendance': att_cal,
            'homework': hw_cal,
            'elearning': el_cal,
            'rpoints': rp_cal,
            'hackathon_1': s['hackathon_1'],
            'hackathon_2': s['hackathon_2'],
            'project': s['project'],
            'pass': s['pass']
        })
    return calibrated

def main():
    conn = mysql.connector.connect(
        host="localhost",
        port=3307,
        user="root",
        password="",
        database="qldt_el"
    )
    cursor = conn.cursor(dictionary=True)
    
    # Load metadata
    cursor.execute("SELECT id, name FROM classes;")
    classes = cursor.fetchall()
    class_id_to_name = {c['id']: c['name'] for c in classes}
    
    cursor.execute("SELECT id, name FROM courses;")
    courses = cursor.fetchall()
    course_id_to_name = {c['id']: c['name'] for c in courses}
    
    cursor.execute("SELECT DISTINCT class_id, course_id FROM final_results;")
    all_class_courses = cursor.fetchall()
    
    # Load excel data
    excel_data = get_excel_chot_data("docs/PTIT_Chiso.xlsx")
    
    # Class course sequence
    class_course_seq = defaultdict(list)
    cursor.execute("""
        SELECT classes_id, courses_id, MIN(date) as first_date
        FROM attendance
        GROUP BY classes_id, courses_id
        ORDER BY classes_id, first_date;
    """)
    for row in cursor.fetchall():
        if row['classes_id'] and row['courses_id']:
            class_course_seq[int(row['classes_id'])].append(int(row['courses_id']))
            
    # Gather database student pass histories
    cursor.execute("SELECT student_id, course_id, pass FROM final_results WHERE pass IS NOT NULL;")
    histories_raw = cursor.fetchall()
    student_pass_history = {}
    for h in histories_raw:
        student_pass_history[(int(h['student_id']), int(h['course_id']))] = int(h['pass'])

    # Build dataset
    dataset = []
    for pair in all_class_courses:
        cid = int(pair['class_id'])
        co_id = int(pair['course_id'])
        cname = class_id_to_name.get(cid, "N/A")
        coname = course_id_to_name.get(co_id, "N/A")
        
        if not any(b in cname for b in ['K24', 'K25', 'KS24', 'KS25']):
            continue
            
        norm_cname = normalize_class_name(cname)
        
        cursor.execute("""
            SELECT student_id, homework, elearning, attendance, hackathon_1, hackathon_2, rpoints, project, pass
            FROM final_results
            WHERE class_id = %s AND course_id = %s AND pass IS NOT NULL;
        """, (cid, co_id))
        students_results = cursor.fetchall()
        
        if not students_results:
            continue
            
        excel_disc = excel_data.get(norm_cname, {}).get(course_to_sheet_map.get(coname.lower()))
        # Try fuzzy match if sheet map is missing
        if not excel_disc:
            low_course = coname.lower()
            target_sheet = None
            for kw, sheet in course_to_sheet_map.items():
                if kw in low_course:
                    target_sheet = sheet
                    break
            excel_disc = excel_data.get(norm_cname, {}).get(target_sheet) if target_sheet else None
            
        calibrated_students = calibrate_students(students_results, excel_disc)
        
        total_students = len(calibrated_students)
        actual_pass_count = sum(1 for s in calibrated_students if s['pass'] == 1)
        actual_pass_rate = (actual_pass_count / total_students) * 100
        
        seq = class_course_seq.get(cid, [])
        prereq_course_id = None
        if co_id in seq:
            idx = seq.index(co_id)
            if idx > 0:
                prereq_course_id = seq[idx - 1]
                
        # Detect soft skills (SKL) and internship (TTRK) courses
        is_soft_skill = any(kw in coname.lower() for kw in ['kỹ năng', 'tin học văn phòng', 'skl', 'thực tập', 'ttrk', 'project'])
        
        students_data = []
        for s in calibrated_students:
            # Check prerequisite history
            has_prev = False
            prev_passed = False
            if prereq_course_id:
                prev_pass = student_pass_history.get((s['student_id'], prereq_course_id))
                if prev_pass is not None:
                    has_prev = True
                    prev_passed = (prev_pass == 1)
            
            # Hackathon
            h_vals = [v for v in [s['hackathon_1'], s['hackathon_2']] if v is not None]
            has_hack = len(h_vals) > 0
            avg_hack_val = mean(h_vals) if has_hack else 65.0
            
            students_data.append({
                'has_prev': has_prev,
                'prev_passed': prev_passed,
                'avg_hack_val': avg_hack_val,
                'has_hack': has_hack
            })
            
        batch = "K25" if ("KS25" in cname or "K25" in cname) else "K24"
        avg_hack_all = mean([s['avg_hack_val'] for s in students_data])
        if actual_pass_rate < 5.0 and avg_hack_all > 60.0:
            continue
            
        dataset.append({
            'class_name': cname,
            'course_name': coname,
            'batch': batch,
            'total': total_students,
            'actual_pass_rate': actual_pass_rate,
            'is_soft_skill': is_soft_skill,
            'students': students_data
        })
        
    print(f"Dataset prepared with {len(dataset)} class-courses.")
    
    # Define ranges for grid search
    w1_range = [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
    p_prereq_pass_range = [0.85, 0.90, 0.95, 0.98]
    p_prereq_fail_range = [0.10, 0.20, 0.30, 0.40]
    p_hack_mult_range = [1.0, 1.1, 1.2, 1.25, 1.3]
    base_scale_range = [0.85, 0.90, 0.95, 1.0]
    
    for target_batch in ['K24', 'K25']:
        print(f"\n--- GRID SEARCH HYPERPARAMETERS FOR KHÓA {target_batch} ---")
        best_params = {}
        min_mae = 999.0
        
        # Grid search loop
        for w1 in w1_range:
            w2 = 1.0 - w1
            for p_prereq_pass in p_prereq_pass_range:
                for p_prereq_fail in p_prereq_fail_range:
                    for p_hack_mult in p_hack_mult_range:
                        for base_scale in base_scale_range:
                            
                            errors = []
                            for cc in dataset:
                                if cc['batch'] != target_batch:
                                    continue
                                    
                                if cc['is_soft_skill']:
                                    pred_pass_rate = 95.0
                                else:
                                    sum_p = 0.0
                                    for s in cc['students']:
                                        # P_prereq
                                        if s['has_prev']:
                                            P_prereq = p_prereq_pass if s['prev_passed'] else p_prereq_fail
                                        else:
                                            P_prereq = 0.8
                                            
                                        # P_hack
                                        p_hack = min(100.0, s['avg_hack_val'] * p_hack_mult)
                                        
                                        if s['has_hack']:
                                            P_learning = w1 * P_prereq + w2 * (p_hack/100.0)
                                        else:
                                            # When no hackathon, fallback weight is fixed or proportional
                                            P_learning = 0.3 * P_prereq + 0.7 * (p_hack/100.0)
                                            
                                        p_eligible = P_learning * base_scale
                                        sum_p += p_eligible
                                        
                                    pred_pass_rate = (sum_p / cc['total']) * 100
                                err = abs(pred_pass_rate - cc['actual_pass_rate'])
                                errors.append(err)
                                
                            mae = mean(errors) if errors else 0.0
                            if errors and mae < min_mae:
                                min_mae = mae
                                best_params = {
                                    'w1': w1,
                                    'w2': w2,
                                    'p_prereq_pass': p_prereq_pass,
                                    'p_prereq_fail': p_prereq_fail,
                                    'p_hack_mult': p_hack_mult,
                                    'base_scale': base_scale
                                }
                                
        print(f"BEST FOR {target_batch}:")
        print(f"  MAE: {min_mae:.2f}%")
        print(f"  w1 (prereq weight): {best_params['w1']:.2f}")
        print(f"  w2 (hackathon weight): {best_params['w2']:.2f}")
        print(f"  Prereq Pass Base: {best_params['p_prereq_pass']:.2f}")
        print(f"  Prereq Fail Base: {best_params['p_prereq_fail']:.2f}")
        print(f"  Hackathon Multiplier: {best_params['p_hack_mult']:.2f}")
        print(f"  Base Scale (Multiplier): {best_params['base_scale']:.2f}")
        
    cursor.close()
    conn.close()

if __name__ == "__main__":
    main()
