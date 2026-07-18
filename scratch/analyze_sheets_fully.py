import sys
import os
import openpyxl

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

files = {
    "old": r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG.xlsx",
    "v1": r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG (1).xlsx",
    "v2": r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG (2).xlsx"
}

def print_weights_sheet(path, label):
    print(f"\n==================== WEIGHTS: {label} ====================")
    if not os.path.exists(path):
        print("Does not exist!")
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb['Trọng số đánh giá']
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        # Print non-empty rows
        val_strs = [str(x) if x is not None else "" for x in row]
        if any(val_strs):
            print(f"Row {r_idx+1}: {val_strs}")
    wb.close()

def print_rubric_sheet(path, label):
    print(f"\n==================== RUBRIC: {label} ====================")
    if not os.path.exists(path):
        print("Does not exist!")
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb['Rubric tính điểm']
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        val_strs = [str(x) if x is not None else "" for x in row]
        if any(val_strs):
            # Print just the criteria name and level summaries
            criteria_name = val_strs[0].replace('\n', ' ')
            m1 = val_strs[1].replace('\n', ' ')[:20]
            m5 = val_strs[5].replace('\n', ' ')[:20]
            print(f"Row {r_idx+1}: {criteria_name[:60]} | M1: {m1}... | M5: {m5}...")
    wb.close()

# Inspect 'Trọng số đánh giá' for all versions
print_weights_sheet(files["old"], "OLD")
print_weights_sheet(files["v1"], "V1")
print_weights_sheet(files["v2"], "V2")

# Inspect 'Rubric tính điểm' for V2
print_rubric_sheet(files["v2"], "V2")
