import sys
import os
import json
import openpyxl
from datetime import datetime, date
from collections import defaultdict

# Reconfigure stdout to use UTF-8 encoding to avoid Windows encoding errors
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

# Constants & Paths
EXCEL_PATH = r"data/PTIT_Chiso.xlsx"
DAILY_LOG_JSON_PATH = r"data/daily_log_analysis.json"
PRED_JSON_PATH = r"scratch/predictions_cv_data.json"
OUTPUT_MD_PATH = r"output/kpi_classification_report.md"
OUTPUT_HTML_PATH = r"output/kpi_classification_report.html"

# Helper function to parse dates from Excel
def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

# Helper to normalize class names
def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

def clean_instructor_name(name):
    if not name:
        return ""
    name_clean = name.strip()
    special_mappings = {
        "lưu hoàng xuân nguyên": "Lưu Xuân Hoàng Nguyên",
        "xuân nguyên": "Lưu Xuân Hoàng Nguyên",
        "lưu xuân hoàng nguyên": "Lưu Xuân Hoàng Nguyên",
        "nguyễn huyền trang": "Nguyễn Thị Huyền Trang",
        "nguyễn thị huyền trang": "Nguyễn Thị Huyền Trang"
    }
    if name_clean.lower() in special_mappings:
        return special_mappings[name_clean.lower()]
    return name_clean

# 1. DEFINE FULL USER WHITELIST (44 Staff members including QLCLĐT, Foreign Languages, and BGĐ)
WHITELIST_STAFF = {
    # A. Khối CNTT-KS24-HN
    "hồ xuân hùng": {"name": "Hồ Xuân Hùng", "group": "Khối CNTT-KS24-HN", "role": "Quản lý", "rank": "5"},
    "bùi thanh hải": {"name": "Bùi Thanh Hải", "group": "Khối CNTT-KS24-HN", "role": "Giảng viên", "rank": "3"},
    "mai xuân chinh": {"name": "Mai Xuân Chinh", "group": "Khối CNTT-KS24-HN", "role": "Giảng viên", "rank": "3"},
    "đinh thành nam": {"name": "Đinh Thành Nam", "group": "Khối CNTT-KS24-HN", "role": "Giảng viên", "rank": "3"},
    "nguyễn công hưởng": {"name": "Nguyễn Công Hưởng", "group": "Khối CNTT-KS24-HN", "role": "Giảng viên", "rank": "3"},
    "phạm tuấn bình": {"name": "Phạm Tuấn Bình", "group": "Khối CNTT-KS24-HN", "role": "Giảng viên", "rank": "3"},

    # B. Khối CNTT-KS25-HN
    "trịnh quốc hai": {"name": "Trịnh Quốc Hai", "group": "Khối CNTT-KS25-HN", "role": "Quản lý", "rank": "5"},
    "lương quốc tuấn": {"name": "Lương Quốc Tuấn", "group": "Khối CNTT-KS25-HN", "role": "Giảng viên", "rank": "3"},
    "nguyễn quảng an": {"name": "Nguyễn Quảng An", "group": "Khối CNTT-KS25-HN", "role": "Giảng viên", "rank": "3"},
    "lại trung lâm": {"name": "Lại Trung Lâm", "group": "Khối CNTT-KS25-HN", "role": "Giảng viên", "rank": "3"},
    "phạm ngọc kiên": {"name": "Phạm Ngọc Kiên", "group": "Khối CNTT-KS25-HN", "role": "Giảng viên", "rank": "3"},
    "ngọ văn quý": {"name": "Ngọ Văn Quý", "group": "Khối CNTT-KS25-HN", "role": "Giảng viên", "rank": "3"},
    "lâm tùng dương": {"name": "Lâm Tùng Dương", "group": "Khối CNTT-KS25-HN", "role": "Giảng viên", "rank": "3"},
    "trần minh cường": {"name": "Trần Minh Cường", "group": "Khối CNTT-KS25-HN", "role": "Quản lý", "rank": "5"},

    # C. CNTT-HCM
    "nguyễn bá minh đạo": {"name": "Nguyễn Bá Minh Đạo", "group": "Khối CNTT-HCM", "role": "Quản lý", "rank": "5"},
    "trần quốc tuấn": {"name": "Trần Quốc Tuấn", "group": "Khối CNTT-HCM", "role": "Giảng viên", "rank": "3"},
    "lê hà thanh sang": {"name": "Lê Hà Thanh Sang", "group": "Khối CNTT-HCM", "role": "Giảng viên", "rank": "3"},
    "phạm viết hùng": {"name": "Phạm Viết Hùng", "group": "Khối CNTT-HCM", "role": "Giảng viên", "rank": "3"},
    "lưu xuân hoàng nguyên": {"name": "Lưu Xuân Hoàng Nguyên", "group": "Khối CNTT-HCM", "role": "Giảng viên", "rank": "3"},
    "nguyễn đức minh": {"name": "Nguyễn Đức Minh", "group": "Khối CNTT-HCM", "role": "Giảng viên", "rank": "3"},
    "phan ngọc tài": {"name": "Phan Ngọc Tài", "group": "Khối CNTT-HCM", "role": "Giảng viên", "rank": "3"},
    "đặng minh luân": {"name": "Đặng Minh Luân", "group": "Khối CNTT-HCM", "role": "Giảng viên", "rank": "3"},
    "nguyễn ngọc sơn": {"name": "Nguyễn Ngọc Sơn", "group": "Khối CNTT-HCM", "role": "Giảng viên", "rank": "3"},

    # D. QTKD - HN
    "hoàng thị kim oanh": {"name": "Hoàng Thị Kim Oanh", "group": "Khối QTKD-HN", "role": "Quản lý", "rank": "5"},
    "lê thành ngọc": {"name": "Lê Thành Ngọc", "group": "Khối QTKD-HN", "role": "Quản lý", "rank": "5"},
    "hoàng thị hậu": {"name": "Hoàng Thị Hậu", "group": "Khối QTKD-HN", "role": "Giảng viên", "rank": "3"},
    "đặng quỳnh trang": {"name": "Đặng Quỳnh Trang", "group": "Khối QTKD-HN", "role": "Giảng viên", "rank": "3"},
    "nguyễn ngọc vân khanh": {"name": "Nguyễn Ngọc Vân Khanh", "group": "Khối QTKD-HN", "role": "Giảng viên", "rank": "3"},
    "nguyễn thị hồng minh": {"name": "Nguyễn Thị Hồng Minh", "group": "Khối QTKD-HN", "role": "Giảng viên", "rank": "3"},
    "triệu thị thanh tâm": {"name": "Triệu Thị Thanh Tâm", "group": "Khối QTKD-HN", "role": "Trợ giảng", "rank": "1"},
    "nguyễn thị như quỳnh": {"name": "Nguyễn Thị Như Quỳnh", "group": "Khối QTKD-HN", "role": "Trợ giảng", "rank": "1"},

    # E. QTKD - HCM
    "lê nhựt mi": {"name": "Lê Nhựt Mi", "group": "Khối QTKD-HCM", "role": "Giảng viên", "rank": "3"},
    "lê thị bảo yến": {"name": "Lê Thị Bảo Yến", "group": "Khối QTKD-HCM", "role": "Giảng viên", "rank": "3"},

    # F. QLCLĐT
    "nguyễn thị tươi": {"name": "Nguyễn Thị Tươi", "group": "Khối QLCLĐT", "role": "Giảng viên", "rank": "3"},
    "nguyễn thị huyền trang": {"name": "Nguyễn Thị Huyền Trang", "group": "Khối QLCLĐT", "role": "Giảng viên", "rank": "3"},
    "nguyễn huyền trang": {"name": "Nguyễn Thị Huyền Trang", "group": "Khối QLCLĐT", "role": "Giảng viên", "rank": "3"},
    "trần thị mỹ phước": {"name": "Trần Thị Mỹ Phước", "group": "Khối QLCLĐT", "role": "Giảng viên", "rank": "3"},
    "nguyễn xuân bách": {"name": "Nguyễn Xuân Bách", "group": "Khối QLCLĐT", "role": "Giảng viên", "rank": "4"},

    # G. Khối Ngoại ngữ và Kỹ năng mềm
    "lò thị ngọc anh": {"name": "Lò Thị Ngọc Anh", "group": "Khối Ngoại ngữ và Kỹ năng mềm", "role": "Quản lý tiếng Anh", "rank": "5"},
    "giáp thị minh hằng": {"name": "Giáp Thị Minh Hằng", "group": "Khối Ngoại ngữ và Kỹ năng mềm", "role": "Quản lý tiếng Nhật", "rank": "5"},
    "lê thị đỏ": {"name": "Lê Thị Đỏ", "group": "Khối Ngoại ngữ và Kỹ năng mềm", "role": "Giảng viên", "rank": "3"},
    "ngô quang huấn": {"name": "Ngô Quang Huấn", "group": "Khối Ngoại ngữ và Kỹ năng mềm", "role": "Quản lý Kỹ năng mềm", "rank": "5"},
    "hoàng phương thảo": {"name": "Hoàng Phương Thảo", "group": "Khối Ngoại ngữ và Kỹ năng mềm", "role": "Giảng viên", "rank": "3"},
    "bùi thị xuân mai": {"name": "Bùi Thị Xuân Mai", "group": "Khối Ngoại ngữ và Kỹ năng mềm", "role": "Giảng viên", "rank": "3"},

    # Ban giám đốc
    "nguyễn duy quang": {"name": "Nguyễn Duy Quang", "group": "Ban Giám Đốc", "role": "Giám đốc đào tạo", "rank": "7"}
}

# 2. LOAD DATA
print("1. Đang nạp dữ liệu từ các nguồn...")

# Load daily log data
daily_log_data = {}
if os.path.exists(DAILY_LOG_JSON_PATH):
    try:
        with open(DAILY_LOG_JSON_PATH, "r", encoding="utf-8") as f:
            daily_log_data = json.load(f).get("monthly_stats", {})
        print(f"Loaded daily log stats for {len(daily_log_data)} instructors.")
    except Exception as e:
        print(f"Error loading daily log json: {e}")
else:
    print(f"File not found: {DAILY_LOG_JSON_PATH}")

# Load academic predictions
predictions_data = {}
if os.path.exists(PRED_JSON_PATH):
    try:
        with open(PRED_JSON_PATH, "r", encoding="utf-8") as f:
            p_data = json.load(f)
            dashboard_data = p_data.get('dashboard_data', {})
            for batch_key, batch_val in dashboard_data.items():
                for c in batch_val.get('cv', []):
                    cname = c.get('class_name')
                    predictions_data[cname] = c.get('actual_pass', 100.0)
                for c in batch_val.get('curr', []):
                    cname = c.get('class_name')
                    predictions_data[cname] = c.get('pred_new', 100.0)
        print(f"Loaded academic pass rates for {len(predictions_data)} classes.")
    except Exception as e:
        print(f"Error loading predictions json: {e}")
else:
    print(f"File not found: {PRED_JSON_PATH}")

# Load operational violations
violations_data = []
violation_path = "data/vi_pham_gvtg.json"
if os.path.exists(violation_path):
    try:
        with open(violation_path, "r", encoding="utf-8") as f:
            violations_data = json.load(f)
        print(f"Loaded {len(violations_data)} operational violations.")
    except Exception as e:
        print(f"Error loading violations json: {e}")

# 3. READ EXCEL DATA (PTIT_Chiso.xlsx)
print("2. Đang đọc dữ liệu chỉ số đào tạo sinh viên từ Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

target_sheets = [
    'KS25_Python_Web',
    'KS25_QTKD_PRJ302'
]

# Map instructor -> { 'classes': set(), 'cc_vals': [], 'bt_vals': [] }
excel_instructor_metrics = defaultdict(lambda: {
    'classes': set(),
    'cc_vals': [],
    'bt_vals': [],
    'el_vals': []
})

for sheetname in target_sheets:
    if sheetname not in wb.sheetnames:
        continue
    sheet = wb[sheetname]
    max_r = sheet.max_row
    max_c = sheet.max_column
    if max_r < 5 or max_c < 4:
        continue

    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]

    # Build column -> (date, subheader) mapping
    col_info = []
    current_date = None
    for c_idx in range(3, max_c):
        if c_idx < len(row3) and row3[c_idx]:
            parsed = parse_date(row3[c_idx])
            if parsed:
                current_date = parsed
        subheader = row4[c_idx] if c_idx < len(row4) else None
        if current_date and subheader in ('Chuyên cần', 'Bài tập', 'Elearning'):
            col_info.append((c_idx, current_date, subheader))

    if not col_info:
        continue

    for r in range(5, max_r + 1):
        cname_raw = sheet.cell(row=r, column=2).value
        teacher_raw = sheet.cell(row=r, column=3).value
        if not cname_raw:
            continue

        norm_class = normalize_class_name(cname_raw)

        # Bỏ qua các lớp học mẫu L01, L02
        if "l01" in norm_class.lower() or "l02" in norm_class.lower():
            continue

        # Read TA on next row (where B column is empty)
        ta_raw = None
        if r + 1 <= max_r:
            next_cname = sheet.cell(row=r + 1, column=2).value
            if not next_cname:
                ta_raw = sheet.cell(row=r + 1, column=3).value

        # Clean names
        teacher_name = clean_instructor_name(teacher_raw)
        ta_name = clean_instructor_name(ta_raw)

        # Check against whitelist
        is_teacher_valid = teacher_name.strip().lower() in WHITELIST_STAFF
        is_ta_valid = ta_name.strip().lower() in WHITELIST_STAFF if ta_name else False

        if not is_teacher_valid and not is_ta_valid:
            continue

        # Parse metrics for teacher
        cc_t, bt_t, el_t = [], [], []
        # Parse metrics for TA
        cc_ta, bt_ta, el_ta = [], [], []

        for c_idx, d, sub in col_info:
            val_t = sheet.cell(row=r, column=c_idx + 1).value
            val_ta = sheet.cell(row=r + 1, column=c_idx + 1).value if ta_name else None

            # For Teacher
            if is_teacher_valid and isinstance(val_t, (int, float)):
                val_float = float(val_t)
                if sub == 'Chuyên cần': cc_t.append(val_float)
                elif sub == 'Bài tập': bt_t.append(val_float)
                elif sub == 'Elearning': el_t.append(val_float)

            # For TA (take class metrics from teacher row r to prevent blank cell bias)
            if is_ta_valid and isinstance(val_t, (int, float)):
                val_float_t = float(val_t)
                if sub == 'Chuyên cần': cc_ta.append(val_float_t)
                elif sub == 'Bài tập': bt_ta.append(val_float_t)
                elif sub == 'Elearning': el_ta.append(val_float_t)

        # Accumulate
        if is_teacher_valid:
            canon_name = WHITELIST_STAFF[teacher_name.strip().lower()]["name"]
            excel_instructor_metrics[canon_name]['classes'].add(f"{norm_class} ({sheetname})")
            excel_instructor_metrics[canon_name]['cc_vals'].extend(cc_t)
            excel_instructor_metrics[canon_name]['bt_vals'].extend(bt_t)
            excel_instructor_metrics[canon_name]['el_vals'].extend(el_t)

        if is_ta_valid:
            canon_name_ta = WHITELIST_STAFF[ta_name.strip().lower()]["name"]
            excel_instructor_metrics[canon_name_ta]['classes'].add(f"{norm_class} ({sheetname})")
            excel_instructor_metrics[canon_name_ta]['cc_vals'].extend(cc_ta)
            excel_instructor_metrics[canon_name_ta]['bt_vals'].extend(bt_ta)
            excel_instructor_metrics[canon_name_ta]['el_vals'].extend(el_ta)

wb.close()
print(f"Finished reading Excel. Found data for {len(excel_instructor_metrics)} whitelisted instructors from class sheets.")

# 4. CALCULATE KPI & CLASSIFICATION
print("3. Đang tính toán điểm KPI theo tiêu chuẩn mới...")

# Define standard weights based on Rank category (V2 Updated Standard)
WEIGHTS = {
    'TA': { # Rank 1-2
        'tuanchu': 0.10, 'kyluat': 0.10, 'hoclieu': 0.10, 'phoihop': 0.05,
        'siso': 0.15, 'el': 0.10, 'bt': 0.05,
        'giaiquyet': 0.05, 'pass_rate': 0.05, 'kha_gioi': 0.05,
        'cc_cn': 0.02, 'bang_cap': 0.03,
        'truyendat_dugio': 0.0, 'truyendat_uprank': 0.0,
        'csat': 0.05, 'hoclieu_duyet': 0.10,
        'toiuu_ct': 0.0, 'hieusuat_duoi': 0.0, 'thuonghieu': 0.0, 'sangkien': 0.0
    },
    'GV': { # Rank 3-4
        'tuanchu': 0.10, 'kyluat': 0.05, 'hoclieu': 0.05, 'phoihop': 0.05,
        'siso': 0.10, 'el': 0.05, 'bt': 0.05,
        'giaiquyet': 0.10, 'pass_rate': 0.05, 'kha_gioi': 0.05,
        'cc_cn': 0.02, 'bang_cap': 0.03,
        'truyendat_dugio': 0.10, 'truyendat_uprank': 0.10,
        'csat': 0.05, 'hoclieu_duyet': 0.05,
        'toiuu_ct': 0.0, 'hieusuat_duoi': 0.0, 'thuonghieu': 0.0, 'sangkien': 0.0
    },
    'QL': { # Rank 5-6 (Quản lý)
        'tuanchu': 0.05, 'kyluat': 0.02, 'hoclieu': 0.03, 'phoihop': 0.05,
        'siso': 0.05, 'el': 0.03, 'bt': 0.02,
        'giaiquyet': 0.10, 'pass_rate': 0.03, 'kha_gioi': 0.02,
        'cc_cn': 0.02, 'bang_cap': 0.03,
        'truyendat_dugio': 0.02, 'truyendat_uprank': 0.05,
        'csat': 0.05, 'hoclieu_duyet': 0.03,
        'toiuu_ct': 0.15, 'hieusuat_duoi': 0.10, 'thuonghieu': 0.10, 'sangkien': 0.05
    },
    'TH': { # Rank 5-6 (GV Thương hiệu)
        'tuanchu': 0.05, 'kyluat': 0.05, 'hoclieu': 0.03, 'phoihop': 0.02,
        'siso': 0.05, 'el': 0.03, 'bt': 0.02,
        'giaiquyet': 0.10, 'pass_rate': 0.05, 'kha_gioi': 0.05,
        'cc_cn': 0.02, 'bang_cap': 0.03,
        'truyendat_dugio': 0.10, 'truyendat_uprank': 0.10,
        'csat': 0.05, 'hoclieu_duyet': 0.05,
        'toiuu_ct': 0.20, 'hieusuat_duoi': 0.0, 'thuonghieu': 0.10, 'sangkien': 0.0
    },
    'Lead': { # Rank 7-8
        'tuanchu': 0.02, 'kyluat': 0.01, 'hoclieu': 0.02, 'phoihop': 0.01,
        'siso': 0.02, 'el': 0.01, 'bt': 0.01,
        'giaiquyet': 0.05, 'pass_rate': 0.02, 'kha_gioi': 0.01,
        'cc_cn': 0.01, 'bang_cap': 0.01,
        'truyendat_dugio': 0.0, 'truyendat_uprank': 0.05,
        'csat': 0.03, 'hoclieu_duyet': 0.02,
        'toiuu_ct': 0.10, 'hieusuat_duoi': 0.25, 'thuonghieu': 0.15, 'sangkien': 0.20
    }
}

evaluation_results = []

# Deduplicate whitelist keys to avoid duplicate entries (e.g. nguyễn thị huyền trang vs nguyễn huyền trang)
unique_whitelist = {}
for k, v in WHITELIST_STAFF.items():
    canon_name = v["name"]
    unique_whitelist[canon_name.lower()] = v

for name_lower, info in sorted(unique_whitelist.items()):
    display_name = info["name"]
    group = info["group"]
    role = info["role"]
    rank_str = info["rank"]
    
    # Try to load log info from daily_log_data
    # Map name variants for matching logs
    log_key = name_lower
    if name_lower == "nguyễn thị huyền trang":
        log_key = "nguyễn huyền trang"
        
    log_info = daily_log_data.get(log_key, {})
    
    try:
        rank_val = int(rank_str)
    except ValueError:
        rank_val = 3
        
    # Determine rank category for weights (V2 Updated)
    if rank_val in (1, 2):
        w_cat = 'TA'
    elif rank_val in (3, 4):
        w_cat = 'GV'
    elif rank_val in (5, 6):
        if "thương hiệu" in role.lower():
            w_cat = 'TH'
        else:
            w_cat = 'QL'
    else:
        w_cat = 'Lead'
        
    weights = WEIGHTS[w_cat]
    
    # Extract raw metrics from Excel (V2 Updated)
    excel_info = excel_instructor_metrics.get(display_name, {})
    classes_list = list(excel_info.get('classes', []))
    classes_str = ", ".join(classes_list) if classes_list else "Không phụ trách lớp Excel"
    
    cc_vals = excel_info.get('cc_vals', [])
    bt_vals = excel_info.get('bt_vals', [])
    el_vals = excel_info.get('el_vals', [])
    
    # Average student violations (from Excel)
    avg_cc_violation = sum(cc_vals) / len(cc_vals) if cc_vals else 0.0
    avg_bt_violation = sum(bt_vals) / len(bt_vals) if bt_vals else 0.0
    avg_el_violation = sum(el_vals) / len(el_vals) if el_vals else 0.0
    
    # Clean percentages (Invert BT violation percentage to completion percentage)
    student_cc_rate = 100.0 - avg_cc_violation
    student_bt_rate = 100.0 - avg_bt_violation
    student_el_rate = 100.0 - avg_el_violation
    
    # If no classes, set to default high values
    if not classes_list:
        student_cc_rate = 95.0
        student_bt_rate = 92.0
        student_el_rate = 90.0
        
    # Quality score (pass rate)
    pass_rates = []
    for cls in classes_list:
        norm_cls = normalize_class_name(cls)
        p_rate = predictions_data.get(norm_cls)
        if p_rate is not None:
            pass_rates.append(p_rate)
    
    student_pass_rate = sum(pass_rates) / len(pass_rates) if pass_rates else 90.0
    if not classes_list:
        student_pass_rate = 95.0
        
    # Count operational violations from JSON
    op_viols = 0
    for v in violations_data:
        v_name = clean_instructor_name(v.get('Họ và tên', ''))
        if v_name.strip().lower() == name_lower:
            op_viols += 1
            
    # Compliance logs violations from daily log data
    missing_days = log_info.get("missing_days", [])
    time_violations = log_info.get("time_violations", [])
    uncompleted_tasks = log_info.get("uncompleted_tasks", [])
    
    compliance_violations_count = len(missing_days) + len(time_violations)
    material_violations_count = len(uncompleted_tasks)
    
    # Apply Rubrics (Thang 10)
    # 1. Tuân thủ (Báo cáo ngày)
    if compliance_violations_count == 0:
        score_tuanchu = 10.0
    elif compliance_violations_count <= 2:
        score_tuanchu = 8.0
    elif compliance_violations_count <= 4:
        score_tuanchu = 5.0
    elif compliance_violations_count <= 6:
        score_tuanchu = 3.0
    else:
        score_tuanchu = 1.0
        
    # 2. Kỷ luật tác nghiệp (Op violations)
    if op_viols == 0:
        score_kyluat = 10.0
    elif op_viols <= 2:
        score_kyluat = 8.0
    elif op_viols <= 4:
        score_kyluat = 5.0
    elif op_viols <= 6:
        score_kyluat = 3.0
    else:
        score_kyluat = 1.0
        
    # 3. Tiến độ học liệu (uncompleted tasks)
    if material_violations_count == 0:
        score_hoclieu = 10.0
    elif material_violations_count <= 2:
        score_hoclieu = 8.0
    elif material_violations_count <= 4:
        score_hoclieu = 5.0
    elif material_violations_count <= 6:
        score_hoclieu = 3.0
    else:
        score_hoclieu = 1.0
        
    # 4. Quản lý sĩ số (chuyên cần trung bình)
    if student_cc_rate >= 90.0:
        score_siso = 10.0
    elif student_cc_rate >= 80.0:
        score_siso = 8.0
    elif student_cc_rate >= 70.0:
        score_siso = 5.0
    elif student_cc_rate >= 60.0:
        score_siso = 3.0
    else:
        score_siso = 1.0
        
    # 5. Đôn đốc E-learning
    if student_el_rate >= 90.0:
        score_el = 10.0
    elif student_el_rate >= 80.0:
        score_el = 8.0
    elif student_el_rate >= 70.0:
        score_el = 5.0
    elif student_el_rate >= 60.0:
        score_el = 3.0
    else:
        score_el = 1.0
        
    # 6. Đôn đốc BTVN
    if student_bt_rate >= 90.0:
        score_bt = 10.0
    elif student_bt_rate >= 80.0:
        score_bt = 8.0
    elif student_bt_rate >= 70.0:
        score_bt = 5.0
    elif student_bt_rate >= 60.0:
        score_bt = 3.0
    else:
        score_bt = 1.0
        
    # 7. Tỷ lệ pass (Đầu ra)
    if student_pass_rate >= 80.0:
        score_pass_rate = 10.0
    elif student_pass_rate >= 70.0:
        score_pass_rate = 8.0
    elif student_pass_rate >= 60.0:
        score_pass_rate = 5.0
    elif student_pass_rate >= 50.0:
        score_pass_rate = 3.0
    else:
        score_pass_rate = 1.0
        
    # Mặc định 5 điểm cho các tiêu chí khác chưa đo lường định lượng
    score_phoihop = 5.0
    score_giaiquyet = 5.0
    score_kha_gioi = 5.0
    score_cc_cn = 5.0
    score_bang_cap = 5.0
    score_truyendat_dugio = 5.0
    score_truyendat_uprank = 5.0
    score_csat = 5.0
    score_hoclieu_duyet = 5.0
    score_toiuu_ct = 5.0
    score_hieusuat_duoi = 5.0
    score_thuonghieu = 5.0
    score_sangkien = 5.0
    
    # 4. CALCULATE OVERALL PROCESS SCORES (V2 Weights matching)
    w_tuanchu = weights.get('tuanchu', 0.0)
    w_kyluat = weights.get('kyluat', 0.0)
    w_hoclieu = weights.get('hoclieu', 0.0)
    w_siso = weights.get('siso', 0.0)
    w_el = weights.get('el', 0.0)
    w_bt = weights.get('bt', 0.0)
    w_pass_rate = weights.get('pass_rate', 0.0)
    
    sum_w_real = w_tuanchu + w_kyluat + w_hoclieu + w_siso + w_el + w_bt + w_pass_rate
    
    weighted_score_real = (
        score_tuanchu * w_tuanchu +
        score_kyluat * w_kyluat +
        score_hoclieu * w_hoclieu +
        score_siso * w_siso +
        score_el * w_el +
        score_bt * w_bt +
        score_pass_rate * w_pass_rate
    )
    overall_score_A = weighted_score_real / sum_w_real if sum_w_real > 0 else 5.0
    
    # Option B: Complete weights with defaults (5.0) inserted
    weighted_score_all = (
        score_tuanchu * weights.get('tuanchu', 0.0) +
        score_kyluat * weights.get('kyluat', 0.0) +
        score_hoclieu * weights.get('hoclieu', 0.0) +
        score_phoihop * weights.get('phoihop', 0.0) +
        score_siso * weights.get('siso', 0.0) +
        score_el * weights.get('el', 0.0) +
        score_bt * weights.get('bt', 0.0) +
        score_giaiquyet * weights.get('giaiquyet', 0.0) +
        score_pass_rate * weights.get('pass_rate', 0.0) +
        score_kha_gioi * weights.get('kha_gioi', 0.0) +
        score_cc_cn * weights.get('cc_cn', 0.0) +
        score_bang_cap * weights.get('bang_cap', 0.0) +
        score_truyendat_dugio * weights.get('truyendat_dugio', 0.0) +
        score_truyendat_uprank * weights.get('truyendat_uprank', 0.0) +
        score_csat * weights.get('csat', 0.0) +
        score_hoclieu_duyet * weights.get('hoclieu_duyet', 0.0) +
        score_toiuu_ct * weights.get('toiuu_ct', 0.0) +
        score_hieusuat_duoi * weights.get('hieusuat_duoi', 0.0) +
        score_thuonghieu * weights.get('thuonghieu', 0.0) +
        score_sangkien * weights.get('sangkien', 0.0)
    )
    overall_score_B = weighted_score_all
    
    # 5. PERFORMANCE CLASSIFICATION
    if overall_score_A >= 10.0:
        classification = "Vượt mức"
        badge_color = "green"
    elif overall_score_A >= 7.5:
        classification = "Đạt"
        badge_color = "blue"
    elif overall_score_A >= 5.0:
        classification = "Cần cố gắng"
        badge_color = "yellow"
    else:
        classification = "Không đạt kỳ vọng"
        badge_color = "red"
        
    # Text Analysis
    strengths_list = []
    weaknesses_list = []
    recs_list = []
    
    if score_tuanchu >= 8.0:
        strengths_list.append("Tuân thủ rất tốt quy trình báo cáo ngày và khai báo thời gian đúng hạn.")
    else:
        weaknesses_list.append(f"Thiếu nộp báo cáo ngày hoặc khai báo vượt định mức nhiều lần (vi phạm {compliance_violations_count} lần trong tháng 7).")
        recs_list.append("Cần cải thiện kỷ luật nộp báo cáo ngày đầy đủ và rà soát định mức thời gian khai báo.")
        
    if score_hoclieu >= 8.0:
        strengths_list.append("Quản lý tiến độ học liệu tốt, không có task bị chậm trễ.")
    else:
        weaknesses_list.append(f"Có task học liệu hoặc nhiệm vụ đào tạo chậm deadline ({material_violations_count} task tồn đọng).")
        recs_list.append("Cần phân bổ thời gian hợp lý để đẩy nhanh tiến độ hoàn thành các task được giao.")
        
    if classes_list:
        if score_siso >= 8.0:
            strengths_list.append(f"Quản lý sĩ số lớp tốt, tỷ lệ chuyên cần trung bình cao ({student_cc_rate:.1f}%).")
        else:
            weaknesses_list.append(f"Tỷ lệ chuyên cần của lớp phụ trách ở mức thấp ({student_cc_rate:.1f}%).")
            recs_list.append("Cần sát sao hơn trong việc điểm danh sinh viên và thông báo cảnh báo chuyên cần sớm.")
            
        if score_bt >= 8.0:
            strengths_list.append(f"Đôn đốc sinh viên hoàn thành BTVN tốt ({student_bt_rate:.1f}% học viên hoàn thành).")
        else:
            weaknesses_list.append(f"Tỷ lệ nợ bài tập của sinh viên lớp phụ trách khá cao ({100.0-student_bt_rate:.1f}% nợ bài).")
            recs_list.append("Cần tăng cường hỗ trợ và đôn đốc sinh viên nộp bài tập đúng hạn để tránh hổng kiến thức.")
            
        if score_pass_rate >= 8.0:
            strengths_list.append(f"Chất lượng đầu ra lớp phụ trách tốt, tỷ lệ sinh viên pass môn cao ({student_pass_rate:.1f}%).")
        else:
            weaknesses_list.append(f"Tỷ lệ pass môn dự kiến của sinh viên ở mức thấp ({student_pass_rate:.1f}%).")
            recs_list.append("Phối hợp với giáo vụ/trợ giảng hỗ trợ học viên yếu kém, tổ chức các buổi bổ trợ kiến thức nền tảng.")
    else:
        strengths_list.append("Duy trì chất lượng công việc hỗ trợ đào tạo chuyên môn ổn định.")
        
    if not strengths_list:
        strengths_list.append("Thực hiện đầy đủ các nhiệm vụ đào tạo cơ bản được giao.")
    if not weaknesses_list:
        weaknesses_list.append("Không ghi nhận vi phạm nghiêm trọng về kỷ luật hoặc chỉ số lớp học.")
    if not recs_list:
        recs_list.append("Tiếp tục phát huy hiệu suất công việc hiện tại và nâng cao năng lực chuyên môn.")

    evaluation_results.append({
        'name': display_name,
        'group': group,
        'role': role,
        'rank': rank_str,
        'rank_cat': w_cat,
        'classes': classes_str,
        'raw_metrics': {
            'cc_rate': student_cc_rate,
            'bt_rate': student_bt_rate,
            'pass_rate': student_pass_rate,
            'compliance_violations': compliance_violations_count,
            'material_violations': material_violations_count,
            'el_rate': student_el_rate,
            'op_violations': op_viols
        },
        'rubric_scores': {
            'tuanchu': score_tuanchu,
            'kyluat': score_kyluat,
            'hoclieu': score_hoclieu,
            'siso': score_siso,
            'dondoc': score_bt, # Map BTVN/dondoc for compatibility
            'phoihop': score_phoihop,
            'giaiquyet': score_giaiquyet,
            'daura': score_pass_rate, # Map pass_rate/daura for compatibility
            'phattrien': score_cc_cn, # Map cc_cn/phattrien for compatibility
            'truyendat_dugio': score_truyendat_dugio,
            'truyendat_uprank': score_truyendat_uprank,
            'csat': score_csat,
            'hoclieu_duyet': score_hoclieu_duyet,
            'toiuu_ct': score_toiuu_ct,
            'hieusuat_duoi': score_hieusuat_duoi,
            'thuonghieu': score_thuonghieu,
            'sangkien': score_sangkien
        },
        'overall_score_A': overall_score_A,
        'overall_score_B': overall_score_B,
        'classification': classification,
        'badge_color': badge_color,
        'strengths': " ".join(strengths_list),
        'weaknesses': " ".join(weaknesses_list),
        'recommendations': " ".join(recs_list)
    })

# 5. WRITE MARKDOWN REPORT
print("4. Đang sinh file báo cáo Markdown...")
with open(OUTPUT_MD_PATH, "w", encoding="utf-8") as md:
    md.write("# Báo cáo Đánh giá & Xếp loại Năng lực GV/TG Học kỳ\n\n")
    md.write("> [!NOTE]\n")
    md.write("> Báo cáo này đánh giá hiệu suất năng lực của giảng viên và trợ giảng dựa theo khung tiêu chuẩn của tài liệu [[data/[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG.xlsx]].\n")
    md.write("> Điểm số được tính toán dựa trên **Phương án A (Scale trọng số theo các tiêu chí thực tế có dữ liệu)** để phản ánh chân thực kết quả giảng dạy và làm việc.\n")
    md.write("> Danh sách nhân sự được cập nhật chính xác theo Whitelist của phòng Đào tạo (44 nhân sự hoạt động chính thức).\n")
    md.write("> *Lưu ý:* Các giảng viên thuộc khối Ngoại ngữ & KNM và QLCLĐT tạm thời sử dụng Rank mặc định hoặc Rank cũ do chưa có khung chia cụ thể, các Quản lý khối vẫn được tính theo Rank 5 (Quản lý).\n\n")
    
    # Summary of stats
    total_staff = len(evaluation_results)
    vuot_muc = sum(1 for r in evaluation_results if r['classification'] == "Vượt mức")
    dat = sum(1 for r in evaluation_results if r['classification'] == "Đạt")
    can_co_gang = sum(1 for r in evaluation_results if r['classification'] == "Cần cố gắng")
    khong_dat = sum(1 for r in evaluation_results if r['classification'] == "Không đạt kỳ vọng")
    
    md.write("## 1. Thống kê Xếp loại Năng lực Chung\n\n")
    md.write(f"- **Tổng số nhân sự đánh giá**: {total_staff} thầy cô\n")
    md.write(f"- **Vượt mức (Điểm >= 10.0)**: {vuot_muc} ({vuot_muc/total_staff*100:.1f}%)\n")
    md.write(f"- **Đạt (7.5 <= Điểm < 10.0)**: {dat} ({dat/total_staff*100:.1f}%)\n")
    md.write(f"- **Cần cố gắng (5.0 <= Điểm < 7.5)**: {can_co_gang} ({can_co_gang/total_staff*100:.1f}%)\n")
    md.write(f"- **Không đạt kỳ vọng (Điểm < 5.0)**: {khong_dat} ({khong_dat/total_staff*100:.1f}%)\n\n")
    
    # Bảng tổng hợp
    md.write("## 2. Bảng tổng hợp xếp loại năng lực theo Khối phòng ban\n\n")
    
    departments = [
        "Khối CNTT-KS24-HN", "Khối CNTT-KS25-HN", "Khối CNTT-HCM",
        "Khối QTKD-HN", "Khối QTKD-HCM", "Khối QLCLĐT",
        "Khối Ngoại ngữ và Kỹ năng mềm", "Ban Giám Đốc"
    ]
    for dept in departments:
        md.write(f"### {dept}\n\n")
        md.write("| Họ và tên | Vai trò / Cấp bậc | Lớp phụ trách | Điểm quá trình A | Điểm quá trình B | Xếp loại năng lực |\n")
        md.write("| :--- | :---: | :--- | :---: | :---: | :---: |\n")
        dept_staff = [r for r in evaluation_results if r['group'] == dept]
        for r in sorted(dept_staff, key=lambda x: x['overall_score_A'], reverse=True):
            md.write(f"| **{r['name']}** | {r['role']} (Rank {r['rank']}) | {r['classes']} | **{r['overall_score_A']:.2f}** | {r['overall_score_B']:.2f} | **{r['classification']}** |\n")
        md.write("\n")
        
    md.write("---\n\n")
    md.write("## 3. Đánh giá chi tiết từng cá nhân\n\n")
    
    for dept in departments:
        md.write(f"### 🔹 Chi tiết nhân sự {dept}\n\n")
        dept_staff = [r for r in evaluation_results if r['group'] == dept]
        for r in sorted(dept_staff, key=lambda x: x['overall_score_A'], reverse=True):
            md.write(f"#### {r['role']}. {r['name']}\n")
            md.write(f"- **Cấp bậc**: Rank {r['rank']} ({r['rank_cat']})\n")
            md.write(f"- **Lớp phụ trách**: {r['classes']}\n")
            md.write(f"- **Điểm Quá Trình (Phương án A - Thực tế)**: **{r['overall_score_A']:.2f} / 10**\n")
            md.write(f"- **Điểm Quá Trình (Phương án B - Đạt)**: **{r['overall_score_B']:.2f} / 10**\n")
            md.write(f"- **Xếp loại năng lực**: **{r['classification']}**\n")
            md.write(f"- **Chi tiết điểm các tiêu chí định lượng (thang 10)**:\n")
            md.write(f"  * Tuân thủ quy trình đào tạo (báo cáo ngày): **{r['rubric_scores']['tuanchu']:.1f}**\n")
            md.write(f"  * Kỷ luật tác nghiệp: **{r['rubric_scores']['kyluat']:.1f}**\n")
            md.write(f"  * Tiến độ học liệu: **{r['rubric_scores']['hoclieu']:.1f}**\n")
            if r['classes'] != "Không phụ trách lớp Excel":
                md.write(f"  * Quản lý sĩ số (chuyên cần): **{r['rubric_scores']['siso']:.1f}** (Sĩ số trung bình: {r['raw_metrics']['cc_rate']:.1f}%)\n")
                md.write(f"  * Đôn đốc học tập: **{r['rubric_scores']['dondoc']:.1f}** (Hoàn thành bài tập: {r['raw_metrics']['bt_rate']:.1f}%)\n")
                md.write(f"  * Chất lượng đầu ra (pass môn): **{r['rubric_scores']['daura']:.1f}** (Tỷ lệ đỗ trung bình: {r['raw_metrics']['pass_rate']:.1f}%)\n")
            md.write(f"- **Điểm mạnh**: {r['strengths']}\n")
            md.write(f"- **Điểm yếu / Vi phạm**: {r['weaknesses']}\n")
            md.write(f"- **Đề xuất cải thiện**: {r['recommendations']}\n\n")

print(f"Successfully generated Markdown report: {OUTPUT_MD_PATH}")

# 6. WRITE HTML REPORT
print("5. Đang sinh file báo cáo HTML Premium...")

# Build JSON strings to embed in HTML
serialized_data = json.dumps(evaluation_results, ensure_ascii=False)

html_template = """<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Báo cáo Xếp loại Năng lực GV/TG Rikkei Education</title>
    <!-- Tailwind CSS CDN -->
    <script src="https://cdn.tailwindcss.com"></script>
    <!-- Fira Sans & Fira Code Fonts -->
    <link href="https://fonts.googleapis.com/css2?family=Fira+Sans:wght@300;400;500;600;700&family=Fira+Code:wght@400;500&display=swap" rel="stylesheet">
    <!-- FontAwesome Icons -->
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    
    <script>
        tailwind.config = {
            darkMode: 'class',
            theme: {
                extend: {
                    fontFamily: {
                        sans: ['Fira Sans', 'sans-serif'],
                        mono: ['Fira Code', 'monospace']
                    }
                }
            }
        }
    </script>
    <style>
        .glass-panel {
            background: rgba(255, 255, 255, 0.7);
            backdrop-filter: blur(12px);
            -webkit-backdrop-filter: blur(12px);
            border: 1px solid rgba(255, 255, 255, 0.25);
        }
        .dark .glass-panel {
            background: rgba(30, 41, 59, 0.7);
            border: 1px solid rgba(255, 255, 255, 0.08);
        }
    </style>
</head>
<body class="bg-slate-50 dark:bg-slate-900 text-slate-800 dark:text-slate-200 min-h-screen transition-colors duration-300">
    <!-- Master Container -->
    <div class="max-w-7xl mx-auto px-4 py-8">
        
        <!-- Header -->
        <header class="mb-8 flex flex-col md:flex-row md:items-center md:justify-between gap-4">
            <div>
                <h1 class="text-3xl font-bold tracking-tight text-indigo-600 dark:text-indigo-400">
                    <i class="fa-solid fa-graduation-cap mr-2"></i>Báo cáo Xếp loại Năng lực GV/TG
                </h1>
                <p class="text-slate-500 dark:text-slate-400 mt-2">
                    Đo lường năng lực và xếp hạng nhân sự khối Đào tạo dựa trên danh sách Whitelist chính thức mới.
                </p>
            </div>
            
            <!-- Dark Mode Switcher & Tools -->
            <div class="flex items-center gap-3">
                <button id="theme-toggle" class="p-3 rounded-full bg-slate-200 dark:bg-slate-800 hover:opacity-80 transition-all shadow-sm">
                    <i id="theme-toggle-icon" class="fa-solid fa-moon text-indigo-600 dark:text-yellow-400 text-lg"></i>
                </button>
                <button onclick="exportCSV()" class="px-4 py-2.5 bg-indigo-600 text-white rounded-xl shadow-md hover:bg-indigo-700 transition-all font-medium flex items-center gap-2">
                    <i class="fa-solid fa-file-csv text-lg"></i> Xuất CSV
                </button>
            </div>
        </header>

        <!-- KPI Summary Cards -->
        <div class="grid grid-cols-1 md:grid-cols-4 gap-6 mb-8">
            <div class="glass-panel p-6 rounded-2xl shadow-sm border border-slate-100 flex items-center justify-between">
                <div>
                    <span class="text-sm font-semibold text-emerald-600 dark:text-emerald-450 tracking-wider uppercase block">Vượt Mức</span>
                    <span id="stat-vuot-muc" class="text-3xl font-bold mt-1 block">0</span>
                    <span class="text-xs text-slate-500 mt-2 block">Điểm quá trình A &ge; 10.0</span>
                </div>
                <div class="w-14 h-14 rounded-2xl bg-emerald-100 dark:bg-emerald-950 flex items-center justify-center text-emerald-600 dark:text-emerald-450 text-2xl shadow-inner">
                    <i class="fa-solid fa-award"></i>
                </div>
            </div>
            
            <div class="glass-panel p-6 rounded-2xl shadow-sm border border-slate-100 flex items-center justify-between">
                <div>
                    <span class="text-sm font-semibold text-blue-600 dark:text-blue-400 tracking-wider uppercase block">Đạt chuẩn</span>
                    <span id="stat-dat" class="text-3xl font-bold mt-1 block">0</span>
                    <span class="text-xs text-slate-500 mt-2 block">7.5 &le; Điểm quá trình A &lt; 10.0</span>
                </div>
                <div class="w-14 h-14 rounded-2xl bg-blue-100 dark:bg-blue-950 flex items-center justify-center text-blue-600 dark:text-blue-400 text-2xl shadow-inner">
                    <i class="fa-solid fa-circle-check"></i>
                </div>
            </div>

            <div class="glass-panel p-6 rounded-2xl shadow-sm border border-slate-100 flex items-center justify-between">
                <div>
                    <span class="text-sm font-semibold text-yellow-600 dark:text-yellow-400 tracking-wider uppercase block">Cần cố gắng</span>
                    <span id="stat-can-co-gang" class="text-3xl font-bold mt-1 block">0</span>
                    <span class="text-xs text-slate-500 mt-2 block">5.0 &le; Điểm quá trình A &lt; 7.5</span>
                </div>
                <div class="w-14 h-14 rounded-2xl bg-yellow-100 dark:bg-yellow-950 flex items-center justify-center text-yellow-600 dark:text-yellow-400 text-2xl shadow-inner">
                    <i class="fa-solid fa-triangle-exclamation"></i>
                </div>
            </div>

            <div class="glass-panel p-6 rounded-2xl shadow-sm border border-slate-100 flex items-center justify-between">
                <div>
                    <span class="text-sm font-semibold text-red-600 dark:text-red-400 tracking-wider uppercase block">Không đạt kỳ vọng</span>
                    <span id="stat-khong-dat" class="text-3xl font-bold mt-1 block">0</span>
                    <span class="text-xs text-slate-500 mt-2 block">Điểm quá trình A &lt; 5.0</span>
                </div>
                <div class="w-14 h-14 rounded-2xl bg-red-100 dark:bg-red-950 flex items-center justify-center text-red-600 dark:text-red-400 text-2xl shadow-inner">
                    <i class="fa-solid fa-circle-xmark"></i>
                </div>
            </div>
        </div>

        <!-- Filter bar -->
        <div class="glass-panel p-6 rounded-2xl shadow-sm mb-8 flex flex-col md:flex-row gap-4 items-center justify-between">
            <div class="flex flex-wrap items-center gap-4 w-full md:w-auto">
                <div class="flex flex-col gap-1">
                    <label class="text-xs font-semibold text-slate-400 uppercase">Khối phòng ban</label>
                    <select id="filter-group" onchange="applyFilters()" class="px-4 py-2 bg-slate-100 dark:bg-slate-800 rounded-xl border border-slate-200 dark:border-slate-700 outline-none text-sm font-medium">
                        <option value="all">Tất cả phòng ban</option>
                        <option value="Khối CNTT-KS24-HN">CNTT-KS24-HN</option>
                        <option value="Khối CNTT-KS25-HN">CNTT-KS25-HN</option>
                        <option value="Khối CNTT-HCM">CNTT-HCM</option>
                        <option value="Khối QTKD-HN">QTKD-HN</option>
                        <option value="Khối QTKD-HCM">QTKD-HCM</option>
                        <option value="Khối QLCLĐT">QLCLĐT</option>
                        <option value="Khối Ngoại ngữ và Kỹ năng mềm">Ngoại ngữ & KNM</option>
                        <option value="Ban Giám Đốc">Ban Giám Đốc</option>
                    </select>
                </div>
                
                <div class="flex flex-col gap-1">
                    <label class="text-xs font-semibold text-slate-400 uppercase">Xếp loại năng lực</label>
                    <select id="filter-class" onchange="applyFilters()" class="px-4 py-2 bg-slate-100 dark:bg-slate-800 rounded-xl border border-slate-200 dark:border-slate-700 outline-none text-sm font-medium">
                        <option value="all">Tất cả xếp loại</option>
                        <option value="Vượt mức">Vượt mức</option>
                        <option value="Đạt">Đạt</option>
                        <option value="Cần cố gắng">Cần cố gắng</option>
                        <option value="Không đạt kỳ vọng">Không đạt kỳ vọng</option>
                    </select>
                </div>
            </div>
            
            <div class="relative w-full md:w-80">
                <i class="fa-solid fa-magnifying-glass absolute left-4 top-3 text-slate-400"></i>
                <input id="search-name" type="text" oninput="applyFilters()" placeholder="Tìm kiếm tên GV / TG..." class="w-full pl-10 pr-4 py-2 bg-slate-100 dark:bg-slate-800 rounded-xl border border-slate-200 dark:border-slate-700 outline-none text-sm placeholder-slate-450">
            </div>
        </div>

        <!-- Leaderboard Table -->
        <div class="glass-panel rounded-2xl shadow-sm overflow-hidden mb-8 border border-slate-100">
            <div class="px-6 py-5 border-b border-slate-150 dark:border-slate-800 flex items-center justify-between bg-white/40 dark:bg-slate-900/40">
                <h3 class="font-bold text-lg text-indigo-600 dark:text-indigo-400">Danh sách xếp hạng hiệu suất năng lực</h3>
                <span id="filtered-count" class="text-sm font-medium text-slate-500">Hiển thị: 0/0 nhân sự</span>
            </div>
            
            <div class="overflow-x-auto">
                <table class="w-full border-collapse text-left">
                    <thead>
                        <tr class="bg-slate-100/50 dark:bg-slate-850/50 text-slate-400 text-xs font-bold uppercase tracking-wider border-b border-slate-200 dark:border-slate-800">
                            <th class="px-6 py-4">Nhân sự</th>
                            <th class="px-6 py-4">Phòng ban</th>
                            <th class="px-6 py-4">Vai trò / Cấp bậc</th>
                            <th class="px-6 py-4">Lớp phụ trách</th>
                            <th class="px-6 py-4 text-center">Điểm A (Thực tế)</th>
                            <th class="px-6 py-4 text-center">Điểm B (Mặc định)</th>
                            <th class="px-6 py-4 text-center">Xếp loại</th>
                            <th class="px-6 py-4"></th>
                        </tr>
                    </thead>
                    <tbody id="table-body" class="divide-y divide-slate-150 dark:divide-slate-800/80">
                        <!-- Dynamic content -->
                    </tbody>
                </table>
            </div>
        </div>
        
    </div>

    <!-- Script Data & Logic -->
    <script>
        const instructors = SERIALIZED_DATA_PLACEHOLDER;
        
        // Theme toggle logic
        const toggleBtn = document.getElementById('theme-toggle');
        const toggleIcon = document.getElementById('theme-toggle-icon');
        
        // Apply theme from localStorage on load
        if (localStorage.getItem('theme') === 'dark' || (!localStorage.getItem('theme') && window.matchMedia('(prefers-color-scheme: dark)').matches)) {
            document.documentElement.classList.add('dark');
            toggleIcon.className = "fa-solid fa-sun text-yellow-400 text-lg";
        } else {
            document.documentElement.classList.remove('dark');
            toggleIcon.className = "fa-solid fa-moon text-indigo-600 text-lg";
        }
        
        toggleBtn.addEventListener('click', () => {
            if (document.documentElement.classList.contains('dark')) {
                document.documentElement.classList.remove('dark');
                toggleIcon.className = "fa-solid fa-moon text-indigo-600 text-lg";
                localStorage.setItem('theme', 'light');
            } else {
                document.documentElement.classList.add('dark');
                toggleIcon.className = "fa-solid fa-sun text-yellow-400 text-lg";
                localStorage.setItem('theme', 'dark');
            }
        });

        // Stats calculation
        function calculateStats() {
            let vuotMuc = 0, dat = 0, canCoGang = 0, khongDat = 0;
            instructors.forEach(ins => {
                if (ins.classification === "Vượt mức") vuotMuc++;
                else if (ins.classification === "Đạt") dat++;
                else if (ins.classification === "Cần cố gắng") canCoGang++;
                else if (ins.classification.includes("Không đạt")) khongDat++;
            });
            document.getElementById('stat-vuot-muc').textContent = vuotMuc;
            document.getElementById('stat-dat').textContent = dat;
            document.getElementById('stat-can-co-gang').textContent = canCoGang;
            document.getElementById('stat-khong-dat').textContent = khongDat;
        }

        // Render Table Rows
        function renderTable(data) {
            const tbody = document.getElementById('table-body');
            tbody.innerHTML = '';
            
            data.forEach((ins, idx) => {
                // Badge classes
                let badgeClass = '';
                if (ins.classification === 'Vượt mức') badgeClass = 'bg-emerald-100 text-emerald-800 dark:bg-emerald-950/50 dark:text-emerald-450';
                else if (ins.classification === 'Đạt') badgeClass = 'bg-blue-100 text-blue-800 dark:bg-blue-950/50 dark:text-blue-450';
                else if (ins.classification === 'Cần cố gắng') badgeClass = 'bg-yellow-100 text-yellow-800 dark:bg-yellow-950/50 dark:text-yellow-450';
                else badgeClass = 'bg-red-100 text-red-800 dark:bg-red-950/50 dark:text-red-450';

                // Truncate classes
                const classesDisplay = ins.classes.length > 50 ? ins.classes.substring(0, 50) + '...' : ins.classes;

                const tr = document.createElement('tr');
                tr.className = 'hover:bg-slate-100/30 dark:hover:bg-slate-800/20 transition-all';
                tr.innerHTML = `
                    <td class="px-6 py-4">
                        <div class="flex items-center gap-3">
                            <div class="w-10 h-10 rounded-full bg-indigo-50 dark:bg-slate-800 text-indigo-500 dark:text-indigo-400 flex items-center justify-center font-bold shadow-inner">
                                ${ins.name.substring(0, 2).toUpperCase()}
                            </div>
                            <div>
                                <span class="font-bold block text-slate-800 dark:text-slate-100 hover:text-indigo-600 dark:hover:text-indigo-400 cursor-pointer">${ins.name}</span>
                                <span class="text-xs text-slate-400 block">${ins.role}</span>
                            </div>
                        </div>
                    </td>
                    <td class="px-6 py-4 text-sm font-medium text-slate-500 dark:text-slate-400">${ins.group}</td>
                    <td class="px-6 py-4 text-sm font-medium">${ins.role} (Rank ${ins.rank})</td>
                    <td class="px-6 py-4 text-sm text-slate-450 dark:text-slate-400" title="${ins.classes}">${classesDisplay}</td>
                    <td class="px-6 py-4 text-center font-bold text-indigo-600 dark:text-indigo-400">${ins.overall_score_A.toFixed(2)}</td>
                    <td class="px-6 py-4 text-center font-medium text-slate-400">${ins.overall_score_B.toFixed(2)}</td>
                    <td class="px-6 py-4 text-center">
                        <span class="px-2.5 py-1 rounded-full text-xs font-semibold ${badgeClass}">${ins.classification}</span>
                    </td>
                    <td class="px-6 py-4">
                        <button onclick="toggleDetails('details-${idx}')" class="p-2 text-slate-400 hover:text-indigo-500 dark:hover:text-indigo-400 transition-all">
                            <i class="fa-solid fa-chevron-down"></i>
                        </button>
                    </td>
                `;
                tbody.appendChild(tr);

                // Add Accordion details row
                const detailsTr = document.createElement('tr');
                detailsTr.id = `details-${idx}`;
                detailsTr.className = 'hidden bg-slate-50/50 dark:bg-slate-950/20';
                
                let excelStatsHTML = '';
                if (ins.classes !== "Không phụ trách lớp Excel") {
                    excelStatsHTML = `
                        <div class="grid grid-cols-3 gap-4 mb-4">
                            <div class="p-3 bg-white dark:bg-slate-850 rounded-xl border border-slate-100 dark:border-slate-800">
                                <span class="text-xs text-slate-400 uppercase font-semibold">Chuyên cần lớp</span>
                                <span class="block text-lg font-bold text-slate-800 dark:text-slate-100">${ins.raw_metrics.cc_rate.toFixed(1)}%</span>
                                <span class="text-xs text-slate-400 block mt-1">Rubric: ${ins.rubric_scores.siso} điểm</span>
                            </div>
                            <div class="p-3 bg-white dark:bg-slate-850 rounded-xl border border-slate-100 dark:border-slate-800">
                                <span class="text-xs text-slate-400 uppercase font-semibold">Hoàn thành bài tập</span>
                                <span class="block text-lg font-bold text-slate-800 dark:text-slate-100">${ins.raw_metrics.bt_rate.toFixed(1)}%</span>
                                <span class="text-xs text-slate-400 block mt-1">Rubric: ${ins.rubric_scores.dondoc} điểm</span>
                            </div>
                            <div class="p-3 bg-white dark:bg-slate-850 rounded-xl border border-slate-100 dark:border-slate-800">
                                <span class="text-xs text-slate-400 uppercase font-semibold">Tỷ lệ pass dự kiến</span>
                                <span class="block text-lg font-bold text-slate-800 dark:text-slate-100">${ins.raw_metrics.pass_rate.toFixed(1)}%</span>
                                <span class="text-xs text-slate-400 block mt-1">Rubric: ${ins.rubric_scores.daura} điểm</span>
                            </div>
                        </div>
                    `;
                }

                detailsTr.innerHTML = `
                    <td colspan="8" class="px-8 py-6 border-b border-slate-200 dark:border-slate-800">
                        <div class="flex flex-col md:flex-row gap-6">
                            <div class="flex-1">
                                <h4 class="font-bold text-sm text-indigo-500 dark:text-indigo-400 uppercase tracking-wider mb-3">Chỉ số thực tế & Rubric chi tiết</h4>
                                <div class="grid grid-cols-2 md:grid-cols-4 gap-4 mb-4">
                                    <div class="p-3 bg-white dark:bg-slate-850 rounded-xl border border-slate-100 dark:border-slate-800">
                                        <span class="text-xs text-slate-400 uppercase font-semibold">Lỗi Báo cáo ngày</span>
                                        <span class="block text-lg font-bold text-slate-800 dark:text-slate-100">${ins.raw_metrics.compliance_violations} lỗi</span>
                                        <span class="text-xs text-slate-400 block mt-1">Rubric: ${ins.rubric_scores.tuanchu} điểm</span>
                                    </div>
                                    <div class="p-3 bg-white dark:bg-slate-850 rounded-xl border border-slate-100 dark:border-slate-800">
                                        <span class="text-xs text-slate-400 uppercase font-semibold">Lỗi học liệu trễ</span>
                                        <span class="block text-lg font-bold text-slate-800 dark:text-slate-100">${ins.raw_metrics.material_violations} lỗi</span>
                                        <span class="text-xs text-slate-400 block mt-1">Rubric: ${ins.rubric_scores.hoclieu} điểm</span>
                                    </div>
                                    <div class="p-3 bg-white dark:bg-slate-850 rounded-xl border border-slate-100 dark:border-slate-800">
                                        <span class="text-xs text-slate-400 uppercase font-semibold">Kỷ luật tác nghiệp</span>
                                        <span class="block text-lg font-bold text-slate-800 dark:text-slate-100">0 lỗi</span>
                                        <span class="text-xs text-slate-400 block mt-1">Rubric: ${ins.rubric_scores.kyluat} điểm</span>
                                    </div>
                                    <div class="p-3 bg-white dark:bg-slate-850 rounded-xl border border-slate-100 dark:border-slate-800">
                                        <span class="text-xs text-slate-400 uppercase font-semibold">CSAT mặc định</span>
                                        <span class="block text-lg font-bold text-slate-800 dark:text-slate-100">4.1 / 5.0</span>
                                        <span class="text-xs text-slate-400 block mt-1">Rubric: ${ins.rubric_scores.csat} điểm</span>
                                    </div>
                                </div>
                                ${excelStatsHTML}
                            </div>
                            <div class="flex-1 border-t md:border-t-0 md:border-l border-slate-200 dark:border-slate-800 pt-4 md:pt-0 md:pl-6">
                                <h4 class="font-bold text-sm text-indigo-500 dark:text-indigo-400 uppercase tracking-wider mb-3">Đánh giá chất lượng</h4>
                                <div class="space-y-3">
                                    <div>
                                        <span class="text-xs font-bold text-emerald-600 dark:text-emerald-450 uppercase block"><i class="fa-solid fa-thumbs-up mr-1"></i> Điểm mạnh</span>
                                        <p class="text-sm text-slate-600 dark:text-slate-300 mt-0.5">${ins.strengths}</p>
                                    </div>
                                    <div>
                                        <span class="text-xs font-bold text-red-500 dark:text-red-400 uppercase block"><i class="fa-solid fa-circle-exclamation mr-1"></i> Điểm yếu / Vi phạm</span>
                                        <p class="text-sm text-slate-600 dark:text-slate-300 mt-0.5">${ins.weaknesses}</p>
                                    </div>
                                    <div>
                                        <span class="text-xs font-bold text-blue-500 dark:text-blue-400 uppercase block"><i class="fa-solid fa-lightbulb mr-1"></i> Đề xuất cải thiện</span>
                                        <p class="text-sm text-slate-600 dark:text-slate-300 mt-0.5">${ins.recommendations}</p>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </td>
                `;
                tbody.appendChild(detailsTr);
            });
        }

        // Accordion toggle
        function toggleDetails(id) {
            const detailsRow = document.getElementById(id);
            if (detailsRow.classList.contains('hidden')) {
                detailsRow.classList.remove('hidden');
            } else {
                detailsRow.classList.add('hidden');
            }
        }

        // Dynamic Filtering
        function applyFilters() {
            const groupFilter = document.getElementById('filter-group').value;
            const classFilter = document.getElementById('filter-class').value;
            const searchQuery = document.getElementById('search-name').value.trim().toLowerCase();
            
            let filtered = instructors;
            
            if (groupFilter !== 'all') {
                filtered = filtered.filter(ins => ins.group === groupFilter);
            }
            if (classFilter !== 'all') {
                filtered = filtered.filter(ins => ins.classification === classFilter);
            }
            if (searchQuery) {
                filtered = filtered.filter(ins => ins.name.toLowerCase().includes(searchQuery));
            }
            
            renderTable(filtered);
            document.getElementById('filtered-count').textContent = `Hiển thị: ${filtered.length} / ${instructors.length} nhân sự`;
        }

        // CSV Export function
        function exportCSV() {
            let csvContent = "data:text/csv;charset=utf-8,\\uFEFF";
            csvContent += "Họ và tên,Vai trò,Khối,Cấp bậc,Lớp phụ trách,Điểm Quá trình A,Điểm Quá trình B,Xếp loại\\r\\n";
            
            instructors.forEach(ins => {
                const classesEscaped = ins.classes.replace(/"/g, '""');
                csvContent += `"${ins.name}","${ins.role}","${ins.group}","Rank ${ins.rank}","${classesEscaped}",${ins.overall_score_A.toFixed(2)},${ins.overall_score_B.toFixed(2)},"${ins.classification}"\\r\\n`;
            });
            
            const encodedUri = encodeURI(csvContent);
            const link = document.createElement("a");
            link.setAttribute("href", encodedUri);
            link.setAttribute("download", "report_xep_loai_nang_luc_gvtg.csv");
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);
        }

        // Initialize table & stats on load
        window.addEventListener('DOMContentLoaded', () => {
            instructors.sort((a, b) => b.overall_score_A - a.overall_score_A);
            calculateStats();
            renderTable(instructors);
            document.getElementById('filtered-count').textContent = `Hiển thị: ${instructors.length} / ${instructors.length} nhân sự`;
        });
    </script>
</body>
</html>
"""

# Replace the placeholder with actual serialized data
html_content = html_template.replace("SERIALIZED_DATA_PLACEHOLDER", serialized_data)

with open(OUTPUT_HTML_PATH, "w", encoding="utf-8") as h:
    h.write(html_content)

print(f"Successfully generated HTML report: {OUTPUT_HTML_PATH}")
print("=== THỰC THI THÀNH CÔNG ===")
