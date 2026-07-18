import sys
import os
import openpyxl
import json
from collections import defaultdict

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

# Constants & Paths
EXCEL_PATH = r"data/PTIT_Chiso.xlsx"
DAILY_LOG_JSON_PATH = r"data/daily_log_analysis.json"
PRED_JSON_PATH = r"scratch/predictions_cv_data.json"
VIOLATIONS_JSON_PATH = r"data/vi_pham_gvtg.json"

# Target Staff names for matching
TARGETS = [
    "Lương Quốc Tuấn",
    "Lâm Tùng Dương",
    "Trần Quốc Tuấn",
    "Phạm Viết Hùng",
    "Nguyễn Ngọc Vân Khanh",
    "Lê Thành Ngọc"
]

def clean_instructor_name(name):
    if not name:
        return ""
    name_clean = name.strip()
    special_mappings = {
        "lưu hoàng xuân nguyên": "Lưu Xuân Hoàng Nguyên",
        "xuân nguyên": "Lưu Xuân Hoàng Nguyên",
        "lưu xuân hoàng nguyên": "Lưu Xuân Hoàng Nguyên",
        "nguyễn huyền trang": "Nguyễn Thị Huyền Trang",
        "nguyễn thị huyền trang": "Nguyễn Thị Huyền Trang"
    }
    if name_clean.lower() in special_mappings:
        return special_mappings[name_clean.lower()]
    return name_clean

# Helper to normalize class names
def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

def parse_date(d_val):
    from datetime import datetime, date
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

# Load Daily Logs
daily_log_data = {}
if os.path.exists(DAILY_LOG_JSON_PATH):
    with open(DAILY_LOG_JSON_PATH, "r", encoding="utf-8") as f:
        daily_log_data = json.load(f).get("monthly_stats", {})

# Load Predictions
predictions_data = {}
if os.path.exists(PRED_JSON_PATH):
    with open(PRED_JSON_PATH, "r", encoding="utf-8") as f:
        p_data = json.load(f)
        dashboard_data = p_data.get('dashboard_data', {})
        for batch_key, batch_val in dashboard_data.items():
            for c in batch_val.get('cv', []):
                cname = c.get('class_name')
                predictions_data[cname] = c.get('actual_pass', 100.0)
            for c in batch_val.get('curr', []):
                cname = c.get('class_name')
                predictions_data[cname] = c.get('pred_new', 100.0)

# Load Violations
violations_count = defaultdict(int)
if os.path.exists(VIOLATIONS_JSON_PATH):
    with open(VIOLATIONS_JSON_PATH, "r", encoding="utf-8") as f:
        v_list = json.load(f)
        for v in v_list:
            inst = v.get("Instructor")
            if inst:
                violations_count[inst.strip()] += 1

# Read Excel
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
target_sheets = [
    'KS24-JavaAdvance', 'KS24_JavaWeb', 'KS24_JWS', 'KS24_AI',
    'KS25_Javascript', 'KS25_Database', 'KS25_Python', 'KS25_Python_Web',
    'KS25_QTKD_M103', 'KS25_QTKD_M104', 'KS25_QTKD_DTB201', 'KS25_QTKD_DTB202',
    'KS25_QTKD_PRJ302'
]

excel_metrics = defaultdict(lambda: {
    'classes': set(),
    'cc_vals': [],
    'bt_vals': [],
    'el_vals': []
})

for sheetname in target_sheets:
    if sheetname not in wb.sheetnames:
        continue
    sheet = wb[sheetname]
    max_r = sheet.max_row
    max_c = sheet.max_column
    if max_r < 5 or max_c < 4:
        continue
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]

    col_info = []
    current_date = None
    for c_idx in range(3, max_c):
        if c_idx < len(row3) and row3[c_idx]:
            parsed = parse_date(row3[c_idx])
            if parsed:
                current_date = parsed
        subheader = row4[c_idx] if c_idx < len(row4) else None
        if current_date and subheader in ('Chuyên cần', 'Bài tập', 'Elearning'):
            col_info.append((c_idx, current_date, subheader))

    for r in range(5, max_r + 1):
        cname_raw = sheet.cell(row=r, column=2).value
        teacher_raw = sheet.cell(row=r, column=3).value
        if not cname_raw:
            continue
        norm_class = normalize_class_name(cname_raw)
        if "l01" in norm_class.lower() or "l02" in norm_class.lower():
            continue

        ta_raw = None
        if r + 1 <= max_r:
            next_cname = sheet.cell(row=r + 1, column=2).value
            if not next_cname:
                ta_raw = sheet.cell(row=r + 1, column=3).value

        teacher_name = clean_instructor_name(teacher_raw)
        ta_name = clean_instructor_name(ta_raw)

        # Check if matched in targets
        for t in TARGETS:
            if teacher_name.strip().lower() == t.strip().lower():
                # Process metrics
                cc, bt, el = [], [], []
                for c_idx, d, sub in col_info:
                    val = sheet.cell(row=r, column=c_idx + 1).value
                    if isinstance(val, (int, float)):
                        val_f = float(val)
                        if sub == 'Chuyên cần': cc.append(val_f)
                        elif sub == 'Bài tập': bt.append(val_f)
                        elif sub == 'Elearning': el.append(val_f)
                excel_metrics[t]['classes'].add(f"{norm_class} ({sheetname})")
                excel_metrics[t]['cc_vals'].extend(cc)
                excel_metrics[t]['bt_vals'].extend(bt)
                excel_metrics[t]['el_vals'].extend(el)
            
            if ta_name and ta_name.strip().lower() == t.strip().lower():
                cc_ta, bt_ta, el_ta = [], [], []
                for c_idx, d, sub in col_info:
                    val_ta = sheet.cell(row=r + 1, column=c_idx + 1).value
                    if isinstance(val_ta, (int, float)):
                        val_f = float(val_ta)
                        if sub == 'Chuyên cần': cc_ta.append(val_f)
                        elif sub == 'Bài tập': bt_ta.append(val_f)
                        elif sub == 'Elearning': el_ta.append(val_f)
                excel_metrics[t]['classes'].add(f"{norm_class} ({sheetname})")
                excel_metrics[t]['cc_vals'].extend(cc_ta)
                excel_metrics[t]['bt_vals'].extend(bt_ta)
                excel_metrics[t]['el_vals'].extend(el_ta)

wb.close()

# Print Results
print("=== RAW METRICS FOR TARGET STAFF ===")
for t in TARGETS:
    print(f"\nStaff: {t}")
    ex_data = excel_metrics.get(t, {})
    classes = list(ex_data.get('classes', []))
    print(f"  Classes ({len(classes)}): {classes}")
    
    cc_vals = ex_data.get('cc_vals', [])
    bt_vals = ex_data.get('bt_vals', [])
    el_vals = ex_data.get('el_vals', [])
    
    avg_cc = sum(cc_vals)/len(cc_vals) if cc_vals else 0.0
    avg_bt = sum(bt_vals)/len(bt_vals) if bt_vals else 0.0
    avg_el = sum(el_vals)/len(el_vals) if el_vals else 0.0
    
    print(f"  Student Chuyên cần violation average: {avg_cc:.2f}% (Rate: {100.0 - avg_cc:.2f}%)")
    print(f"  Student Bài tập violation average (nợ): {avg_bt:.2f}% (Completion Rate: {100.0 - avg_bt:.2f}%)")
    print(f"  Student Elearning violation average: {avg_el:.2f}% (Completion Rate: {100.0 - avg_el:.2f}%)")
    
    # Pass rate from predictions
    pass_rates = []
    for cls in classes:
        norm_cls = normalize_class_name(cls)
        pr = predictions_data.get(norm_cls)
        if pr is not None:
            pass_rates.append(pr)
    avg_pass = sum(pass_rates)/len(pass_rates) if pass_rates else 90.0
    print(f"  Class Average Pass Rate: {avg_pass:.2f}%")
    
    # Daily Log Compliance
    # Match name in daily logs
    log_key = t.strip().lower()
    if log_key == "nguyễn thị huyền trang":
        log_key = "nguyễn huyền trang"
    log_info = daily_log_data.get(log_key, {})
    missing_days = log_info.get("missing_days", [])
    time_violations = log_info.get("time_violations", [])
    uncompleted_tasks = log_info.get("uncompleted_tasks", [])
    print(f"  Báo cáo ngày: {len(missing_days)} missing, {len(time_violations)} time violations (Total: {len(missing_days) + len(time_violations)})")
    print(f"  Task học liệu chậm trễ: {len(uncompleted_tasks)}")
    
    # Operational violations from vi_pham_gvtg.json
    op_viols = violations_count.get(t, 0)
    print(f"  Kỷ luật tác nghiệp vi phạm: {op_viols} times")
