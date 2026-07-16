import openpyxl
from datetime import datetime, date
from collections import defaultdict

SKIP_SHEETS = ('Sheet1', 'Sheet2')
SKL_KEYWORDS = ('SKL', 'skl')


def parse_date(d_val):
    """Convert any date representation to datetime.date."""
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except (ValueError, TypeError):
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except (ValueError, TypeError):
            return None
    return None


def normalize_class_name(name):
    """Normalize class name: strip parentheses, suffixes, and KSxx->Kxx replacements."""
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = (name_str
                .replace("KS25", "K25")
                .replace("KS24", "K24")
                .replace("KS23", "K23"))
    return name_str


def extract_class_size(name):
    if not name or '(' not in str(name):
        return 0
    try:
        content = str(name).split('(')[1].split(')')[0]
        if '-' in content:
            # Lấy số cuối cùng trong chuỗi gạch ngang
            return int(content.split('-')[-1].strip())
        return int(content.strip())
    except (ValueError, IndexError):
        return 0


def is_skill_sheet(sheetname):
    """Return True if this sheet should be excluded (SKL sheets or system sheets)."""
    if sheetname in SKIP_SHEETS:
        return True
    for kw in SKL_KEYWORDS:
        if kw in sheetname:
            return True
    return False


def load_excel_data(filepath):
    """
    Load and parse PTIT_Chiso.xlsx.

    Returns:
        class_course_map: dict[class_name -> list of course_entries sorted by min_date asc]
        Each course_entry is a dict:
        {
            'sheet':    str,
            'teacher':  str,
            'ta':       str,
            'class_size': int,
            'min_date': date,
            'max_date': date,
            'daily':    list of (date, cc_val, bt_val, el_val),  # sorted asc by date
            'cc_all':   list[float],
            'bt_all':   list[float],
            'el_all':   list[float],
        }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    # raw_data[norm_class_name][sheetname] = entry dict
    raw_data = defaultdict(dict)

    for sheetname in wb.sheetnames:
        if is_skill_sheet(sheetname):
            continue
        sheet = wb[sheetname]
        max_r = sheet.max_row
        max_c = sheet.max_column
        if max_r < 5 or max_c < 4:
            continue

        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]

        # Build column -> (date, subheader) mapping
        col_info = []   # (col_idx_0based, date, subheader_str)
        current_date = None
        for c_idx in range(3, max_c):
            if c_idx < len(row3) and row3[c_idx]:
                parsed = parse_date(row3[c_idx])
                if parsed:
                    current_date = parsed
            subheader = row4[c_idx] if c_idx < len(row4) else None
            if current_date and subheader in ('Chuyên cần', 'Bài tập', 'Elearning'):
                col_info.append((c_idx, current_date, subheader))

        if not col_info:
            continue

        all_dates = [d for _, d, _ in col_info]
        sheet_min_date = min(all_dates)
        sheet_max_date = max(all_dates)

        for r in range(5, max_r + 1):
            cname_raw = sheet.cell(row=r, column=2).value
            teacher_raw = sheet.cell(row=r, column=3).value
            if not cname_raw:
                continue

            norm_name = normalize_class_name(cname_raw)
            class_size = extract_class_size(cname_raw)

            # Đọc Trợ giảng (TA) từ dòng r+1
            ta_raw = None
            if r + 1 <= max_r:
                next_cname = sheet.cell(row=r + 1, column=2).value
                if not next_cname: # Nếu dòng dưới là dòng TA (cột tên lớp trống)
                    ta_raw = sheet.cell(row=r + 1, column=3).value

            teacher_str = (str(teacher_raw).strip()
                           if teacher_raw and str(teacher_raw).strip() not in ('None', '')
                           else "Chưa phân công")
            ta_str = (str(ta_raw).strip()
                      if ta_raw and str(ta_raw).strip() not in ('None', '')
                      else "Không có")

            # Accumulate values per date
            daily_map = defaultdict(lambda: {'cc': None, 'bt': None, 'el': None})
            cc_all, bt_all, el_all = [], [], []

            for c_idx, d, sub in col_info:
                val = sheet.cell(row=r, column=c_idx + 1).value
                if not isinstance(val, (int, float)):
                    continue
                if sub == 'Chuyên cần':
                    daily_map[d]['cc'] = float(val)
                    cc_all.append(float(val))
                elif sub == 'Bài tập':
                    daily_map[d]['bt'] = float(val)
                    bt_all.append(float(val))
                elif sub == 'Elearning':
                    daily_map[d]['el'] = float(val)
                    el_all.append(float(val))

            # Only include rows with at least some data
            if not cc_all and not bt_all and not el_all:
                continue

            # Build sorted daily list
            daily = sorted(
                [
                    (d, v['cc'], v['bt'], v['el'])
                    for d, v in daily_map.items()
                    if any(x is not None for x in (v['cc'], v['bt'], v['el']))
                ],
                key=lambda x: x[0]
            )

            raw_data[norm_name][sheetname] = {
                'sheet':    sheetname,
                'teacher':  teacher_str,
                'ta':       ta_str,
                'class_size': class_size,
                'min_date': sheet_min_date,
                'max_date': sheet_max_date,
                'daily':    daily,
                'cc_all':   cc_all,
                'bt_all':   bt_all,
                'el_all':   el_all,
            }

    # Sort each class's courses by min_date ascending → auto course-sequence detection
    class_course_map = {}
    for norm_name, courses in raw_data.items():
        sorted_courses = sorted(courses.values(), key=lambda x: x['min_date'])
        class_course_map[norm_name] = sorted_courses

    return class_course_map
