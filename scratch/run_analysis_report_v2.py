import sys
import openpyxl
from datetime import datetime, date
from collections import defaultdict
import numpy as np

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'docs/PTIT_Chiso.xlsx'
output_path = 'reports/analysis_ptit_chiso.md'

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

wb = openpyxl.load_workbook(excel_path, data_only=True)

# ----------------------------------------------------
# 1. PARSE DATA & DETECT ROLES (PROPAGATING MERGED VALUES)
# ----------------------------------------------------
class_course_data = defaultdict(dict)
teacher_stats = {}

target_sheets = [
    'KS24-JavaAdvance', 'KS24_JavaWeb', 'KS24_JWS', 'KS24_AI',
    'KS25_Javascript', 'KS25_Database', 'KS25_Python', 'KS25_Python_Web',
    'KS25_QTKD_M103', 'KS25_QTKD_M104', 'KS25_QTKD_DTB201', 'KS25_QTKD_DTB202'
]

for sheetname in target_sheets:
    if sheetname not in wb.sheetnames:
        continue
    sheet = wb[sheetname]
    max_r = sheet.max_row
    max_c = sheet.max_column
    if max_r < 5:
        continue
        
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    dates_list = []
    current_date = None
    for c_idx in range(3, max_c):
        val3 = row3[c_idx]
        val4 = row4[c_idx]
        if val3:
            current_date = parse_date(val3)
        if current_date:
            dates_list.append((c_idx, current_date, val4))
            
    current_class = None
    gv_cc_vals = []
    gv_bt_vals = []
    gv_el_vals = []
    
    for r in range(5, max_r + 1):
        cname = sheet.cell(row=r, column=2).value
        teacher = sheet.cell(row=r, column=3).value
        
        if cname:
            current_class = normalize_class_name(cname)
            gv_cc_vals = []
            gv_bt_vals = []
            gv_el_vals = []
            for c_idx, d, sub in dates_list:
                val = sheet.cell(row=r, column=c_idx + 1).value
                if isinstance(val, (int, float)):
                    if sub == 'Chuyên cần':
                        gv_cc_vals.append(val)
                    elif sub == 'Bài tập':
                        gv_bt_vals.append(val)
                    elif sub == 'Elearning':
                        gv_el_vals.append(val)
                        
        if not current_class or not teacher or str(teacher).strip() in ['', 'None', 'Chưa phân công', 'Giảng viên/Trợ giảng']:
            continue
            
        teacher_name = str(teacher).strip()
        role = 'GV' if cname is not None else 'TG'
        dept = 'QTKD' if 'QTKD' in sheetname or 'QTKD' in current_class else 'CNTT'
        
        cc_vals = gv_cc_vals.copy()
        bt_vals = gv_bt_vals.copy()
        el_vals = gv_el_vals.copy()
        
        cc_avg = np.mean(cc_vals) if cc_vals else 0
        bt_avg = np.mean(bt_vals) if bt_vals else 0
        el_avg = np.mean(el_vals) if el_vals else 0
        
        if cc_vals or bt_vals or el_vals:
            class_course_data[(current_class, role)][sheetname] = {
                'cc_avg': cc_avg,
                'bt_avg': bt_avg,
                'el_avg': el_avg,
                'teacher': teacher_name,
                'min_date': min(d for c, d, s in dates_list) if dates_list else date.min
            }
            
        if teacher_name not in teacher_stats:
            teacher_stats[teacher_name] = {
                'role': role,
                'department': dept,
                'cc': [],
                'bt': [],
                'el': [],
                'classes': set()
            }
            
        teacher_stats[teacher_name]['classes'].add(current_class)
        if cc_vals: teacher_stats[teacher_name]['cc'].extend(cc_vals)
        if bt_vals: teacher_stats[teacher_name]['bt'].extend(bt_vals)
        if el_vals: teacher_stats[teacher_name]['el'].extend(el_vals)

# ----------------------------------------------------
# 2. RECONSTRUCT ALL TRANSITIONS (CONSECUTIVE PAIRS)
# ----------------------------------------------------
transitions = []
for (cname, role), courses in sorted(class_course_data.items()):
    if len(courses) < 2:
        continue
    # Sort courses by min_date
    sorted_courses = sorted(courses.items(), key=lambda x: x[1]['min_date'])
    for i in range(1, len(sorted_courses)):
        prev_sheet, prev_data = sorted_courses[i-1]
        curr_sheet, curr_data = sorted_courses[i]
        
        delta_cc = prev_data['cc_avg'] - curr_data['cc_avg']
        delta_bt = prev_data['bt_avg'] - curr_data['bt_avg']
        delta_el = prev_data['el_avg'] - curr_data['el_avg']
        
        transitions.append({
            'class': cname,
            'role': role,
            'prev_sheet': prev_sheet,
            'prev_teacher': prev_data['teacher'],
            'prev_cc': prev_data['cc_avg'],
            'prev_bt': prev_data['bt_avg'],
            'prev_el': prev_data['el_avg'],
            'curr_sheet': curr_sheet,
            'curr_teacher': curr_data['teacher'],
            'curr_cc': curr_data['cc_avg'],
            'curr_bt': curr_data['bt_avg'],
            'curr_el': curr_data['el_avg'],
            'delta_cc': delta_cc,
            'delta_bt': delta_bt,
            'delta_el': delta_el
        })

# ----------------------------------------------------
# 3. CLASSIFY TEACHERS & TAs USING ALL HISTORICAL TRANSITIONS
# ----------------------------------------------------
evaluated_teachers = []
watchlist_teachers = []

for name, stats in teacher_stats.items():
    cc_mean = np.mean(stats['cc']) if stats['cc'] else 0.0
    bt_mean = np.mean(stats['bt']) if stats['bt'] else 0.0
    el_mean = np.mean(stats['el']) if stats['el'] else 0.0
    
    # Get all transition data points where this teacher is the CURRENT teacher
    t_imps = [item for item in transitions if item['curr_teacher'] == name and item['role'] == stats['role']]
    
    if t_imps:
        cmi_values = []
        for class_item in t_imps:
            cmi_cc = 0.5 * class_item['prev_cc'] - class_item['curr_cc'] + 15.0
            cmi_bt = 0.5 * class_item['prev_bt'] - class_item['curr_bt'] + 15.0
            cmi_el = 0.5 * class_item['prev_el'] - class_item['curr_el'] + 15.0
            cmi_values.append((cmi_cc + cmi_bt + cmi_el) / 3.0)
            
        mgmt_score = np.mean(cmi_values)
        
        if mgmt_score > 12.0:
            classification = "Rescuers (Giải cứu xuất sắc)"
        elif mgmt_score < 0.0 or cc_mean > 25.0 or bt_mean > 20.0 or el_mean > 20.0:
            classification = "Cần Hỗ Trợ (Needs Support)"
        else:
            classification = "Duy trì (Maintainers)"
            
        avg_delta_cc = np.mean([item['delta_cc'] for item in t_imps])
        avg_delta_bt = np.mean([item['delta_bt'] for item in t_imps])
        avg_delta_el = np.mean([item['delta_el'] for item in t_imps])
        
        evaluated_teachers.append({
            'name': name,
            'role': stats['role'],
            'department': stats['department'],
            'classes_count': len(stats['classes']),
            'classes_list': ", ".join(list(stats['classes'])),
            'cc': cc_mean,
            'bt': bt_mean,
            'el': el_mean,
            'delta_cc': avg_delta_cc,
            'delta_bt': avg_delta_bt,
            'delta_el': avg_delta_el,
            'mgmt_score': mgmt_score,
            'class': classification
        })
    else:
        watchlist_teachers.append({
            'name': name,
            'role': stats['role'],
            'department': stats['department'],
            'classes_count': len(stats['classes']),
            'classes_list': ", ".join(list(stats['classes'])),
            'cc': cc_mean,
            'bt': bt_mean,
            'el': el_mean
        })

# ----------------------------------------------------
# 4. WRITE THE REFINE REPORT
# ----------------------------------------------------
with open(output_path, 'w', encoding='utf-8') as f:
    f.write("# BÁO CÁO PHÂN TÍCH CHỈ SỐ ĐÀO TẠO & XẾP HẠNG QUẢN TRỊ LỚP (HỆ PTIT)\n\n")
    f.write("> [!IMPORTANT]\n")
    f.write("> Phương pháp đánh giá mới áp dụng **Chỉ số Quản trị Hợp phần (CMI - Composite Management Index)** trên toàn bộ lịch sử chuyển tiếp của lớp học.\n")
    f.write("> Các lớp học được đối chiếu xuyên suốt lộ trình (ví dụ: JS ➔ Database ➔ Python ➔ Python Web) thay vì chỉ so sánh môn học cuối cùng.\n\n")
    
    # Explain CMI
    f.write("## 📝 1. PHƯƠNG PHÁP & CÔNG THỨC ĐÁNH GIÁ MỚI (CMI)\n\n")
    f.write("Thay vì chỉ nhìn vào mức độ cải thiện đơn thuần (Delta Score) dễ dẫn đến sai lệch khi lớp cũ đã quá tốt, chúng ta áp dụng chỉ số CMI:\n\n")
    f.write("$$\\text{CMI} = 0.5 \\times V_{\\text{prev}} - V_{\\text{curr}} + 15\\%$$\n\n")
    f.write("Trong đó:\n")
    f.write("*   $V_{\\text{prev}}$: Tỷ lệ vi phạm ở môn liền trước (Chuyên cần, Bài tập, Elearning).\n")
    f.write("*   $V_{\\text{curr}}$: Tỷ lệ vi phạm ở môn học hiện tại.\n")
    f.write("*   $15\\%$: Hằng số chuẩn hóa điểm nền.\n\n")
    f.write("**💡 Ý nghĩa của CMI:**\n")
    f.write("*   **Cộng điểm dư địa ($0.5 \\times V_{\\text{prev}}$)**: Giảng viên nhận một lớp có kỷ luật ban đầu rất tệ (vi phạm cao) sẽ được cộng điểm bù đắp cho sự vất vả khi nhận lớp khó.\n")
    f.write("*   **Phạt điểm kết quả ($-V_{\\text{curr}}$)**: Yêu cầu tối thượng là phải giữ tỷ lệ vi phạm hiện tại ở mức thấp nhất. Nếu vi phạm hiện tại cao, điểm CMI sẽ bị kéo xuống rất mạnh.\n")
    f.write("*   **Công bằng cho người duy trì**: Một giảng viên nhận lớp ngoan ($V_{\\text{prev}} = 0\\%$) và duy trì hoàn hảo ($V_{\\text{curr}} = 0\\%$) sẽ nhận điểm CMI là $+15.0\\%$ (Xếp nhóm Duy trì tốt/Giải cứu).\n\n")
    f.write("### 📊 Tiêu chuẩn phân hạng năng lực:\n")
    f.write("*   **Rescuers (Giải cứu xuất sắc)**: $\\text{CMI} > 12.0\\%$. Quản lý lớp cực tốt, giúp kéo giảm vi phạm sâu hoặc duy trì xuất sắc lớp ngoan.\n")
    f.write("*   **Duy trì (Maintainers)**: $0.0\\% \\le \\text{CMI} \\le 12.0\\%$. Giữ kỷ luật lớp học ổn định ở mức chấp nhận được.\n")
    f.write("*   **Cần Hỗ Trợ (Needs Support)**: $\\text{CMI} < 0.0\\%$ HOẶC có bất kỳ chỉ số vi phạm hiện tại nào ở mức Báo động (CC vắng > 25.0%, BT nợ > 20.0%, EL chậm > 20.0%).\n\n")
    f.write("---\n\n")
    
    # Part 1: Transitions
    f.write("## 2. PHÂN TÍCH TIẾN BỘ LỚP HỌC (JS ➔ Database ➔ Python ➔ Python Web)\n")
    f.write("Bảng thống kê sự thay đổi tỷ lệ vi phạm chuyên cần (CC), bài tập (BT), và Elearning (EL) giữa hai môn học liền kề:\n\n")
    f.write("| Tên Lớp | Vai trò | Môn Trước (GV/TG) | Môn Hiện Tại (GV/TG) | Delta CC | Delta BT | Delta EL | Đánh giá Tiến bộ |\n")
    f.write("| :--- | :---: | :--- | :--- | :---: | :---: | :---: | :--- |\n")
    for t in sorted(transitions, key=lambda x: (x['delta_cc'] + x['delta_bt'] + x['delta_el'])/3, reverse=True):
        score = (t['delta_cc'] + t['delta_bt'] + t['delta_el']) / 3.0
        if score > 5.0:
            progress_comment = "🌟 Tiến bộ vượt bậc"
        elif score > 0:
            progress_comment = "✅ Tiến bộ ổn định"
        elif score > -5.0:
            progress_comment = "⚠️ Giảm nhẹ kỷ luật"
        else:
            progress_comment = "🚨 Kỷ luật sụt giảm mạnh"
            
        f.write(f"| **{t['class']}** | {t['role']} | {t['prev_sheet']} ({t['prev_teacher']}) | {t['curr_sheet']} ({t['curr_teacher']}) | {t['delta_cc']:+.2f}% | {t['delta_bt']:+.2f}% | {t['delta_el']:+.2f}% | {progress_comment} |\n")
    f.write("\n---\n\n")
    
    # Part 2: Rankings
    f.write("## 3. BẢNG XẾP HẠNG QUẢN TRỊ LỚP THEO CHỈ SỐ CMI\n\n")
    
    # 2.1 CNTT
    f.write("### 💻 Khối Công nghệ Thông tin (CNTT)\n\n")
    
    # GV CNTT
    f.write("#### 👤 Bảng xếp hạng Giảng viên (GV)\n")
    f.write("| Hạng | Giảng viên | Số lớp | CC Vi phạm | BT Vi phạm | EL Vi phạm | Delta CC | Delta BT | Delta EL | Chỉ số CMI | Phân loại |\n")
    f.write("| :---: | :--- | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :--- |\n")
    gv_cntt = [t for t in evaluated_teachers if t['department'] == 'CNTT' and t['role'] == 'GV']
    for idx, t in enumerate(sorted(gv_cntt, key=lambda x: x['mgmt_score'], reverse=True), 1):
        f.write(f"| {idx} | **{t['name']}** | {t['classes_count']} | {t['cc']:.2f}% | {t['bt']:.2f}% | {t['el']:.2f}% | {t['delta_cc']:+.2f}% | {t['delta_bt']:+.2f}% | {t['delta_el']:+.2f}% | **{t['mgmt_score']:+.2f}%** | {t['class']} |\n")
    f.write("\n")
    
    # TG CNTT
    f.write("#### 🎓 Bảng xếp hạng Trợ giảng (TG)\n")
    f.write("| Hạng | Trợ giảng | Số lớp | CC Vi phạm | BT Vi phạm | EL Vi phạm | Delta CC | Delta BT | Delta EL | Chỉ số CMI | Phân loại |\n")
    f.write("| :---: | :--- | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :--- |\n")
    tg_cntt = [t for t in evaluated_teachers if t['department'] == 'CNTT' and t['role'] == 'TG']
    for idx, t in enumerate(sorted(tg_cntt, key=lambda x: x['mgmt_score'], reverse=True), 1):
        f.write(f"| {idx} | **{t['name']}** | {t['classes_count']} | {t['cc']:.2f}% | {t['bt']:.2f}% | {t['el']:.2f}% | {t['delta_cc']:+.2f}% | {t['delta_bt']:+.2f}% | {t['delta_el']:+.2f}% | **{t['mgmt_score']:+.2f}%** | {t['class']} |\n")
    f.write("\n")
    
    # 2.2 QTKD
    f.write("### 📈 Khối Quản trị Kinh doanh (QTKD)\n\n")
    
    # GV QTKD
    f.write("#### 👤 Bảng xếp hạng Giảng viên (GV)\n")
    f.write("| Hạng | Giảng viên | Số lớp | CC Vi phạm | BT Vi phạm | EL Vi phạm | Delta CC | Delta BT | Delta EL | Chỉ số CMI | Phân loại |\n")
    f.write("| :---: | :--- | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :--- |\n")
    gv_qtkd = [t for t in evaluated_teachers if t['department'] == 'QTKD' and t['role'] == 'GV']
    for idx, t in enumerate(sorted(gv_qtkd, key=lambda x: x['mgmt_score'], reverse=True), 1):
        f.write(f"| {idx} | **{t['name']}** | {t['classes_count']} | {t['cc']:.2f}% | {t['bt']:.2f}% | {t['el']:.2f}% | {t['delta_cc']:+.2f}% | {t['delta_bt']:+.2f}% | {t['delta_el']:+.2f}% | **{t['mgmt_score']:+.2f}%** | {t['class']} |\n")
    f.write("\n")
    
    # TG QTKD
    f.write("#### 🎓 Bảng xếp hạng Trợ giảng (TG)\n")
    f.write("| Hạng | Trợ giảng | Số lớp | CC Vi phạm | BT Vi phạm | EL Vi phạm | Delta CC | Delta BT | Delta EL | Chỉ số CMI | Phân loại |\n")
    f.write("| :---: | :--- | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :---: | :--- |\n")
    tg_qtkd = [t for t in evaluated_teachers if t['department'] == 'QTKD' and t['role'] == 'TG']
    if not tg_qtkd:
        f.write("| | *Không có trợ giảng nào đủ điều kiện đánh giá ở khối QTKD* | | | | | | | | | |\n")
    else:
        for idx, t in enumerate(sorted(tg_qtkd, key=lambda x: x['mgmt_score'], reverse=True), 1):
            f.write(f"| {idx} | **{t['name']}** | {t['classes_count']} | {t['cc']:.2f}% | {t['bt']:.2f}% | {t['el']:.2f}% | {t['delta_cc']:+.2f}% | {t['delta_bt']:+.2f}% | {t['delta_el']:+.2f}% | **{t['mgmt_score']:+.2f}%** | {t['class']} |\n")
    f.write("\n---\n\n")
    
    # Part 4: Watchlist
    f.write("## 4. DANH SÁCH THEO DÕI (WATCHLIST - GIẢNG VIÊN/TRỢ GIẢNG MỚI)\n")
    f.write("Áp dụng cho nhân sự chỉ phụ trách 1 lớp hoặc chưa có dữ liệu môn học trước đó để so sánh chéo:\n\n")
    
    # Watchlist CNTT
    f.write("### 💻 Khối Công nghệ Thông tin (CNTT)\n")
    f.write("| Họ và Tên | Vai trò | Lớp đang phụ trách | Vi phạm CC | Vi phạm BT | Vi phạm EL |\n")
    f.write("| :--- | :---: | :--- | :---: | :---: | :---: |\n")
    wl_cntt = [t for t in watchlist_teachers if t['department'] == 'CNTT']
    for t in sorted(wl_cntt, key=lambda x: x['cc']):
        f.write(f"| **{t['name']}** | {t['role']} | {t['classes_list']} | {t['cc']:.2f}% | {t['bt']:.2f}% | {t['el']:.2f}% |\n")
    f.write("\n")
    
    # Watchlist QTKD
    f.write("### 📈 Khối Quản trị Kinh doanh (QTKD)\n")
    f.write("| Họ và Tên | Vai trò | Lớp đang phụ trách | Vi phạm CC | Vi phạm BT | Vi phạm EL |\n")
    f.write("| :--- | :---: | :--- | :---: | :---: | :---: |\n")
    wl_qtkd = [t for t in watchlist_teachers if t['department'] == 'QTKD']
    for t in sorted(wl_qtkd, key=lambda x: x['cc']):
        f.write(f"| **{t['name']}** | {t['role']} | {t['classes_list']} | {t['cc']:.2f}% | {t['bt']:.2f}% | {t['el']:.2f}% |\n")
    f.write("\n")

print("Refined progress and CMI management ranking report (including all historical transitions) generated successfully.")
