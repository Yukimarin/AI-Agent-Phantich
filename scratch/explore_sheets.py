import sys
import openpyxl
from datetime import datetime, date

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'C:/Users/DELL/Desktop/Education-DB-Analytic/docs/PTIT_Chiso.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

sheets = ['KS24_AI', 'KS24_JWS', 'KS25_Python_Web', 'KS25_QTKD_DTB202']
for name in sheets:
    if name not in wb.sheetnames:
        print(f"Sheet {name} not found")
        continue
    sheet = wb[name]
    print(f"\n--- Sheet: {name} (Max row: {sheet.max_row}, Max col: {sheet.max_column}) ---")
    
    # In dong 3 va 4 (ngay thang va loai vi pham)
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    dates = []
    current_date = None
    for c_idx in range(3, len(row3)):
        val3 = row3[c_idx]
        val4 = row4[c_idx]
        if val3:
            current_date = parse_date(val3)
        if current_date:
            dates.append((c_idx + 1, current_date, val4))
            
    print("Dates found in columns:")
    # In 15 cot cuoi cung co ngay thang
    for col_idx, d, val4 in dates[-15:]:
        print(f"  Col {col_idx}: Date={d}, Type={val4}")
        
    # In cac lop o cot 2 (tu dong 5 tro di)
    classes = []
    for r in range(5, sheet.max_row + 1):
        cname = sheet.cell(row=r, column=2).value
        teacher = sheet.cell(row=r, column=3).value
        if cname:
            classes.append((r, cname, teacher))
    print(f"Classes found ({len(classes)}):")
    for r, cname, teacher in classes:
        print(f"  Row {r}: Class={cname}, Teacher={teacher}")
