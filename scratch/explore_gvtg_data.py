import mysql.connector
import openpyxl
import sys
import os
import unicodedata
from collections import defaultdict

# Force UTF-8 output
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"data/PTIT_Chiso.xlsx"

# 4 target instructors
TARGETS = [
    "Trần Quốc Tuấn",
    "Lê Hà Thanh Sang",
    "Phạm Viết Hùng",
    "Lưu Xuân Hoàng Nguyên"
]

def normalize_vietnamese_name(name):
    if not name:
        return ""
    name = " ".join(name.strip().split())
    name = name.lower()
    name = unicodedata.normalize('NFKD', name)
    name = "".join([c for c in name if not unicodedata.combining(c)])
    name = name.replace("đ", "d")
    return name

def clean_instructor_name(name):
    if not name:
        return ""
    name_clean = name.strip()
    special_mappings = {
        "lưu hoàng xuân nguyên": "Lưu Xuân Hoàng Nguyên",
        "xuân nguyên": "Lưu Xuân Hoàng Nguyên",
        "lưu xuân hoàng nguyên": "Lưu Xuân Hoàng Nguyên",
        "lê hà thanh sang": "Lê Hà Thanh Sang",
        "trần quốc tuấn": "Trần Quốc Tuấn",
        "phạm viết hùng": "Phạm Viết Hùng"
    }
    if name_clean.lower() in special_mappings:
        return special_mappings[name_clean.lower()]
    return name_clean

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

def parse_date(d_val):
    from datetime import datetime, date
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

def check_excel():
    print("=== READING EXCEL DATA ===")
    if not os.path.exists(EXCEL_PATH):
        print(f"Excel file not found at {EXCEL_PATH}")
        return {}
        
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # We want to scan all sheets to see if there are other sheets not in target_sheets
    print("Sheets available:", wb.sheetnames)
    
    excel_classes = defaultdict(list)
    
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        max_r = sheet.max_row
        max_c = sheet.max_column
        if max_r < 5 or max_c < 4:
            continue
            
        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]

        col_info = []
        current_date = None
        for c_idx in range(3, max_c):
            if c_idx < len(row3) and row3[c_idx]:
                parsed = parse_date(row3[c_idx])
                if parsed:
                    current_date = parsed
            subheader = row4[c_idx] if c_idx < len(row4) else None
            if current_date and subheader in ('Chuyên cần', 'Bài tập', 'Elearning'):
                col_info.append((c_idx, current_date, subheader))

        for r in range(5, max_r + 1):
            cname_raw = sheet.cell(row=r, column=2).value
            teacher_raw = sheet.cell(row=r, column=3).value
            if not cname_raw:
                continue
            norm_class = normalize_class_name(cname_raw)
            if "l01" in norm_class.lower() or "l02" in norm_class.lower():
                continue

            # Read TA from r+1
            ta_raw = None
            if r + 1 <= max_r:
                next_cname = sheet.cell(row=r + 1, column=2).value
                if not next_cname:
                    ta_raw = sheet.cell(row=r + 1, column=3).value

            teacher_name = clean_instructor_name(teacher_raw)
            ta_name = clean_instructor_name(ta_raw)

            # Check if matched in targets
            for t in TARGETS:
                norm_t = normalize_vietnamese_name(t)
                is_gv = normalize_vietnamese_name(teacher_name) == norm_t
                is_ta = ta_name and normalize_vietnamese_name(ta_name) == norm_t
                
                if is_gv or is_ta:
                    role = "GV" if is_gv else "TG"
                    # Read values for this class from row r (always from GV row to avoid empty rows)
                    cc = []
                    for c_idx, d, sub in col_info:
                        val = sheet.cell(row=r, column=c_idx + 1).value
                        if isinstance(val, (int, float)):
                            val_f = float(val)
                            if sub == 'Chuyên cần': 
                                cc.append(val_f)
                    
                    avg_cc = sum(cc)/len(cc) if cc else 0.0
                    excel_classes[t].append({
                        "class_name": norm_class,
                        "raw_class_name": cname_raw,
                        "role": role,
                        "sheet": sheetname,
                        "avg_cc_violation": avg_cc,
                        "cc_count": len(cc)
                    })
                    
    wb.close()
    return excel_classes

def check_mysql():
    print("\n=== CONNECTING TO MYSQL ===")
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="qldt_el",
            port=3307
        )
        cursor = conn.cursor(dictionary=True)
        
        # Look for users in DB
        print("--- Finding Users in DB ---")
        for t in TARGETS:
            norm_t = normalize_vietnamese_name(t)
            cursor.execute("SELECT id, full_name, email FROM user")
            all_users = cursor.fetchall()
            matched_users = []
            for u in all_users:
                if u['full_name'] and normalize_vietnamese_name(u['full_name']) == norm_t:
                    matched_users.append(u)
            print(f"Target: {t} -> Matched in DB: {matched_users}")
            
            # If user found, find classes from classes_users
            if matched_users:
                u_ids = [u['id'] for u in matched_users]
                u_ids_str = ",".join([str(uid) for uid in u_ids])
                query = f"""
                    SELECT c.id as class_id, c.name as class_name, cu.user_id 
                    FROM classes c
                    JOIN classes_users cu ON c.id = cu.classes_id
                    WHERE cu.user_id IN ({u_ids_str})
                """
                cursor.execute(query)
                assigned_classes = cursor.fetchall()
                print(f"  Assigned classes in DB ({len(assigned_classes)}):")
                for ac in assigned_classes:
                    print(f"    Class ID: {ac['class_id']} | Class Name: {ac['class_name']} | User ID: {ac['user_id']}")
                    
        conn.close()
    except Exception as e:
        print("Error connecting to MySQL:", str(e))

if __name__ == "__main__":
    excel_data = check_excel()
    for t in TARGETS:
        print(f"\nGiảng viên: {t}")
        classes = excel_data.get(t, [])
        print(f"Số lớp trong Excel: {len(classes)}")
        for c in classes:
            print(f"  - Lớp: {c['class_name']} ({c['sheet']}) | Vai trò: {c['role']} | Vi phạm CC TB: {c['avg_cc_violation']:.2f}% (Số buổi: {c['cc_count']})")
            
    check_mysql()
