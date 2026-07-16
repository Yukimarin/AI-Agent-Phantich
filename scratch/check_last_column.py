import openpyxl
import sys
from datetime import datetime, date

sys.stdout.reconfigure(encoding='utf-8')

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except:
            return None
    return None

def main():
    wb = openpyxl.load_workbook("docs/PTIT_Chiso.xlsx", data_only=True)
    
    for sheetname in wb.sheetnames[:5]: # Check first 5 sheets
        if sheetname == 'Sheet1':
            continue
        sheet = wb[sheetname]
        max_r = sheet.max_row
        max_c = sheet.max_column
        
        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
        
        dates_list = []
        current_date = None
        for c_idx in range(3, max_c):
            val3 = row3[c_idx]
            val4 = row4[c_idx]
            if val3:
                current_date = parse_date(val3)
            if current_date:
                dates_list.append((c_idx, current_date, val4))
                
        if not dates_list:
            continue
            
        # Group by date to find the last date
        unique_dates = sorted(list(set(d for idx, d, sub in dates_list)))
        last_date = unique_dates[-1] if unique_dates else None
        
        print(f"\nSheet: {sheetname} | Last Date: {last_date}")
        
        # Columns corresponding to the last date
        last_cols = [idx for idx, d, sub in dates_list if d == last_date]
        
        # Print data for first 3 rows (classes)
        for r in range(5, min(max_r + 1, 10)):
            cname = sheet.cell(row=r, column=2).value
            teacher = sheet.cell(row=r, column=3).value
            if cname:
                print(f"  Class: {cname:<20} | Teacher: {str(teacher):<15}")
                for c_idx in last_cols:
                    sub = row4[c_idx]
                    val = sheet.cell(row=r, column=c_idx + 1).value
                    print(f"    {sub}: {val}")

if __name__ == "__main__":
    main()
