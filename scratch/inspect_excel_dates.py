import openpyxl
import sys
from datetime import datetime, date

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'docs/PTIT_Chiso.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

print("Sheets in workbook:", wb.sheetnames)

for sheetname in wb.sheetnames:
    if 'KS' not in sheetname and 'SKL' not in sheetname:
        continue
    sheet = wb[sheetname]
    print(f"\n--- Sheet: {sheetname} ---")
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))
    if not row3:
        print("No row 3")
        continue
    row3 = row3[0]
    
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))
    row4 = row4[0] if row4 else []
    
    dates_found = []
    for c_idx in range(3, len(row3)):
        val3 = row3[c_idx]
        val4 = row4[c_idx] if c_idx < len(row4) else None
        if val3:
            p_date = parse_date(val3)
            if p_date:
                # Đổi tên metric sang tiếng Anh/không dấu khi in để tránh lỗi console
                metric_name = "unknown"
                if val4:
                    val4_str = str(val4).strip().lower()
                    if 'chuyên' in val4_str or 'chuyen' in val4_str:
                        metric_name = "ChuyenCan"
                    elif 'bài' in val4_str or 'bai' in val4_str:
                        metric_name = "BaiTap"
                    elif 'elearning' in val4_str or 'el' in val4_str:
                        metric_name = "Elearning"
                    else:
                        metric_name = val4_str
                dates_found.append((c_idx, p_date.strftime("%Y-%m-%d"), metric_name))
                
    if dates_found:
        print(f"Total dates found: {len(dates_found)}")
        print(f"First 5 dates: {dates_found[:5]}")
        print(f"Last 5 dates: {dates_found[-5:]}")
    else:
        print("No valid dates found in row 3")
