import mysql.connector
import os
import sys
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
import json
import re
import openpyxl
import numpy as np
from datetime import datetime, date
from collections import defaultdict

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip().upper()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    name_str = name_str.replace('KS24', 'K24').replace('KS25', 'K25')
    name_str = name_str.lower()
    for word in ['hk2', 'hk1', 'hl', 'cu', 'retake', 'old']:
        name_str = name_str.replace(word, '')
    name_str = re.sub(r'[^a-z0-9]', '', name_str)
    return name_str

def run_query(cursor, query, params=None):
    cursor.execute(query, params or ())
    columns = [desc[0] for desc in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def save_course_difficulty(course_name, difficulty):
    metadata_dir = "data"
    os.makedirs(metadata_dir, exist_ok=True)
    metadata_path = os.path.join(metadata_dir, "course_metadata.json")
    meta = {"course_difficulty": {}, "default_difficulty": 1.2}
    if os.path.exists(metadata_path):
        try:
            with open(metadata_path, 'r', encoding='utf-8') as jf:
                meta = json.load(jf)
        except Exception:
            pass
            
    if "course_difficulty" not in meta:
        meta["course_difficulty"] = {}
        
    meta["course_difficulty"][course_name] = {
        "difficulty_coefficient": difficulty,
        "notes": "Tự động gán và lưu bởi hệ thống"
    }
    
    try:
        with open(metadata_path, 'w', encoding='utf-8') as jf:
            json.dump(meta, jf, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving metadata: {e}")

def get_course_difficulty(course_name, cursor):
    metadata_path = "data/course_metadata.json"
    difficulty = None
    if os.path.exists(metadata_path):
        try:
            with open(metadata_path, 'r', encoding='utf-8') as jf:
                meta = json.load(jf)
                course_diff_map = meta.get("course_difficulty", {})
                if course_name in course_diff_map:
                    difficulty = float(course_diff_map[course_name].get("difficulty_coefficient"))
        except Exception as e:
            print(f"Error reading course_metadata.json: {e}")
            
    if difficulty is not None:
        return difficulty
        
    # 2. Fallback to historical fail rate in final_results
    try:
        cursor.execute("SELECT id FROM courses WHERE name = %s", (course_name,))
        row = cursor.fetchone()
        if row:
            co_id = row[0]
            cursor.execute("""
                SELECT 
                    SUM(CASE WHEN pass = '0' THEN 1 ELSE 0 END) as fail_cnt,
                    COUNT(*) as total_cnt
                FROM final_results 
                WHERE course_id = %s;
            """, (co_id,))
            res = cursor.fetchone()
            if res and res[1] and res[1] > 0:
                fail_rate = (float(res[0]) / float(res[1])) * 100.0
                difficulty = round(max(1.0, min(2.0, 1.0 + fail_rate / 50.0)), 2)
                print(f"Calculated CDC for '{course_name}' from DB historical fail rate: {difficulty}")
    except Exception as e:
        print(f"Error calculating CDC from DB for '{course_name}': {e}")
        
    if difficulty is not None:
        save_course_difficulty(course_name, difficulty)
        return difficulty
        
    # 3. Fallback to heuristics based on course name
    coname_lower = course_name.lower()
    if "ai" in coname_lower or "trí tuệ nhân tạo" in coname_lower or "machine learning" in coname_lower:
        difficulty = 1.8
    elif "database" in coname_lower or "cơ sở dữ liệu" in coname_lower or "cấu trúc dữ liệu" in coname_lower:
        difficulty = 1.5
    elif "python" in coname_lower or "java" in coname_lower or "web" in coname_lower:
        difficulty = 1.3
    else:
        difficulty = 1.2
        
    print(f"Calculated CDC for '{course_name}' via heuristics: {difficulty}")
    save_course_difficulty(course_name, difficulty)
    return difficulty

def main():
    conn = mysql.connector.connect(
        host="localhost",
        port=3307,
        user="root",
        password="",
        database="qldt_el"
    )
    cursor = conn.cursor()
    
    # Load Excel data for teacher names fallback
    excel_path = "docs/PTIT_Chiso.xlsx"
    if not os.path.exists(excel_path):
        excel_path = "data/PTIT_Chiso.xlsx"
    excel_data = load_excel_data(excel_path) if os.path.exists(excel_path) else {}
    # Target batches & classes configuration
    ks24_classes = [
        {'id': 48, 'name': 'HN-KS24-CNTT1', 'course_id': 194, 'curr_course_name': 'AI (KS24)', 'prev_course_id': 194},
        {'id': 49, 'name': 'HN-KS24-CNTT2', 'course_id': 194, 'curr_course_name': 'AI (KS24)', 'prev_course_id': 194},
        {'id': 156, 'name': 'HN-KS24-CNTT3', 'course_id': 194, 'curr_course_name': 'AI (KS24)', 'prev_course_id': 194},
        {'id': 51, 'name': 'HN-KS24-CNTT4', 'course_id': 194, 'curr_course_name': 'AI (KS24)', 'prev_course_id': 194},
        {'id': 63, 'name': 'HCM-KS24-CNTT1', 'course_id': 194, 'curr_course_name': 'AI (KS24)', 'prev_course_id': 194},
        {'id': 64, 'name': 'HCM-KS24-CNTT2', 'course_id': 162, 'curr_course_name': 'AI (KS24)', 'prev_course_id': 162},
    ]
    
    ks25_classes = [
        {'id': 77, 'name': 'HN-KS25-CNTT1', 'course_id': 193, 'curr_course_name': '[IT205-K25] Lập trình ứng dụng với Python', 'prev_course_id': 124},
        {'id': 76, 'name': 'HN-KS25-CNTT2', 'course_id': 193, 'curr_course_name': '[IT205-K25] Lập trình ứng dụng với Python', 'prev_course_id': 124},
        {'id': 75, 'name': 'HN-KS25-CNTT3', 'course_id': 193, 'curr_course_name': '[IT205-K25] Lập trình ứng dụng với Python', 'prev_course_id': 124},
        {'id': 74, 'name': 'HN-KS25-CNTT4', 'course_id': 193, 'curr_course_name': '[IT205-K25] Lập trình ứng dụng với Python', 'prev_course_id': 124},
        {'id': 73, 'name': 'HN-KS25-CNTT5', 'course_id': 193, 'curr_course_name': '[IT205-K25] Lập trình ứng dụng với Python', 'prev_course_id': 124},
        {'id': 72, 'name': 'HN-KS25-CNTT6', 'course_id': 193, 'curr_course_name': '[IT205-K25] Lập trình ứng dụng với Python', 'prev_course_id': 124},
    ]
    
    qtkd_classes = [
        {'id': 84, 'name': 'HN-K25-QTKD1', 'course_id': 178, 'curr_course_name': '[DTB202] Chuyển đổi số trong doanh nghiệp', 'prev_course_id': 188},
        {'id': 83, 'name': 'HN-K25-QTKD2', 'course_id': 178, 'curr_course_name': '[DTB202] Chuyển đổi số trong doanh nghiệp', 'prev_course_id': 188},
        {'id': 82, 'name': 'HN-K25-QTKD3', 'course_id': 178, 'curr_course_name': '[DTB202] Chuyển đổi số trong doanh nghiệp', 'prev_course_id': 188},
    ]
    
    all_target_classes = ks24_classes + ks25_classes + qtkd_classes
    
    # Load student pass history (to calculate prerequisite GPA and test scores)
    student_pass_history = {}
    pass_results = run_query(cursor, "SELECT class_id, course_id, student_id, homework, rpoints, pass, project, mutiple_choice_1, essay_1, hackathon_1, hackathon_2, attendance FROM qldt_el.final_results;")
    for r in pass_results:
        sid = int(r['student_id'])
        co_id = int(r['course_id'])
        student_pass_history[(sid, co_id)] = {
            'homework': float(r['homework']) if r['homework'] is not None else 100.0,
            'rpoints': float(r['rpoints']) if r['rpoints'] is not None else 100.0,
            'pass': int(r['pass']) if r['pass'] is not None else None,
            'project': float(r['project']) if r['project'] is not None else None,
            'mutiple_choice_1': float(r['mutiple_choice_1']) if r['mutiple_choice_1'] is not None else None,
            'essay_1': float(r['essay_1']) if r['essay_1'] is not None else None,
            'hackathon_1': float(r['hackathon_1']) if r['hackathon_1'] is not None else None,
            'hackathon_2': float(r['hackathon_2']) if r['hackathon_2'] is not None else None,
            'attendance': float(r['attendance']) if r['attendance'] is not None else 0.0
        }

    # Load class course sequence to determine historical prior courses
    course_seq_raw = run_query(cursor, """
        SELECT classes_id, courses_id, MIN(date) as first_date
        FROM qldt_el.attendance
        GROUP BY classes_id, courses_id
        ORDER BY classes_id, first_date ASC;
    """)
    class_course_seq = defaultdict(list)
    for row in course_seq_raw:
        if row['classes_id'] and row['courses_id']:
            class_course_seq[int(row['classes_id'])].append(int(row['courses_id']))

    # Query Hackathon scores
    hackathon_map = {}
    hackathon_raw = run_query(cursor, """
        SELECT r.student_id, ts.course_id, AVG(r.point) as avg_point
        FROM qldt_el.result_test r
        JOIN qldt_el.test_schedule ts ON r.test_schedule_id = ts.id
        WHERE ts.type = 'THI HACKATHON' AND r.point IS NOT NULL
        GROUP BY r.student_id, ts.course_id;
    """)
    for r in hackathon_raw:
        hackathon_map[(int(r['student_id']), int(r['course_id']))] = float(r['avg_point'])

    risk_results = {}
    
    for c_info in all_target_classes:
        cid = c_info['id']
        cname = c_info['name']
        co_id = c_info['course_id']
        prev_co_id = c_info['prev_course_id']
        
        # 1. Query all active students in class
        students = run_query(cursor, """
            SELECT sc.student_id, s.full_name as student_name, s.student_code
            FROM qldt_el.student_class sc
            JOIN qldt_el.students s ON sc.student_id = s.id
            WHERE sc.class_id = %s AND s.status = 'ĐANG HỌC' AND sc.is_active = 1;
        """, (cid,))
        
        if cid == 156 and not students:
            # Gộp lớp 156
            students = run_query(cursor, """
                SELECT sc.student_id, s.full_name as student_name, s.student_code
                FROM qldt_el.student_class sc
                JOIN qldt_el.students s ON sc.student_id = s.id
                WHERE sc.class_id IN (156, 50, 52) AND s.status = 'ĐANG HỌC' AND sc.is_active = 1;
            """)
            
        if not students:
            continue
            
        # 2. Query attendance details for this class and course (ordered by date/id)
        att_raw = run_query(cursor, """
            SELECT ad.student_id, ad.status, a.date, a.id as attendance_id
            FROM qldt_el.attendance_detail ad
            JOIN qldt_el.attendance a ON ad.attendance_id = a.id
            WHERE a.classes_id = %s AND a.courses_id = %s
            ORDER BY a.date ASC, a.id ASC;
        """, (cid, co_id))
        if cid == 156 and not att_raw:
            att_raw = run_query(cursor, """
                SELECT ad.student_id, ad.status, a.date, a.id as attendance_id
                FROM qldt_el.attendance_detail ad
                JOIN qldt_el.attendance a ON ad.attendance_id = a.id
                WHERE a.classes_id IN (156, 50, 52) AND a.courses_id = %s
                ORDER BY a.date ASC, a.id ASC;
            """, (co_id,))
            
        student_att = defaultdict(list)
        for row in att_raw:
            student_att[int(row['student_id'])].append({
                'status': str(row['status']),
                'id': int(row['attendance_id'])
            })
            
        # 3. Query homeworks (exercises) for this class and course (ordered by id)
        hw_raw = run_query(cursor, """
            SELECT e.student_id, e.check, e.link_git, e.id as exercise_id
            FROM qldt_el.exercise e
            WHERE e.class_id = %s AND e.course_id = %s
            ORDER BY e.id ASC;
        """, (cid, co_id))
        if cid == 156 and not hw_raw:
            hw_raw = run_query(cursor, """
                SELECT e.student_id, e.check, e.link_git, e.id as exercise_id
                FROM qldt_el.exercise e
                WHERE e.class_id IN (156, 50, 52) AND e.course_id = %s
                ORDER BY e.id ASC;
            """, (co_id,))
            
        student_hw_details = defaultdict(list)
        student_hw_total = defaultdict(int)
        student_hw_debt = defaultdict(int)
        for row in hw_raw:
            sid = int(row['student_id'])
            check = int(row['check'] or 0)
            git = row['link_git']
            student_hw_total[sid] += 1
            is_git_empty = not git or git.strip() == "" or "placeholder" in str(git).lower()
            is_debt = (check == 2 or (check == 0 and is_git_empty))
            if is_debt:
                student_hw_debt[sid] += 1
            student_hw_details[sid].append({
                'is_debt': is_debt,
                'id': int(row['exercise_id'])
            })
                
        # 4. Query Elearning late submissions
        el_raw = run_query(cursor, """
            SELECT el.student_id, COUNT(*) as late_count
            FROM qldt_el.elearning_late el
            JOIN qldt_el.sessions s ON el.session_id = s.id
            WHERE s.course_id = %s
            GROUP BY el.student_id;
        """, (co_id,))
        student_el_late = {int(r['student_id']): int(r['late_count']) for r in el_raw}
        
        # 5. Query Rpoints actual scores
        rp_raw = run_query(cursor, """
            SELECT r.student_id, AVG(r.actual_score) as avg_score
            FROM qldt_el.rpoints r
            WHERE r.course_id = %s
            GROUP BY r.student_id;
        """, (co_id,))
        student_rp = {int(r['student_id']): float(r['avg_score']) for r in rp_raw}
        
        # Fallback Rpoint/Project/Homework/Attendance/Elearning from final_results for current course (if present)
        fr_raw = run_query(cursor, """
            SELECT student_id, rpoints, project, homework, attendance, elearning
            FROM qldt_el.final_results
            WHERE class_id = %s AND course_id = %s;
        """, (cid, co_id))
        if cid == 156 and not fr_raw:
            fr_raw = run_query(cursor, """
                SELECT student_id, rpoints, project, homework, attendance, elearning
                FROM qldt_el.final_results
                WHERE class_id IN (156, 50, 52) AND course_id = %s;
            """, (co_id,))
        student_fr_rp = {int(r['student_id']): float(r['rpoints']) if r['rpoints'] is not None else None for r in fr_raw}
        student_fr_proj = {int(r['student_id']): float(r['project']) if r['project'] is not None else None for r in fr_raw}
        student_fr_hw = {int(r['student_id']): float(r['homework']) if r['homework'] is not None else None for r in fr_raw}
        student_fr_att = {int(r['student_id']): float(r['attendance']) if r['attendance'] is not None else None for r in fr_raw}
        student_fr_el = {int(r['student_id']): float(r['elearning']) if r['elearning'] is not None else None for r in fr_raw}
        
        # 6. Determine 2 prior courses of this class
        seq = class_course_seq.get(cid, [])
        prev_courses = []
        if co_id in seq:
            idx = seq.index(co_id)
            if idx >= 1:
                prev_courses.append(seq[idx - 1])
            if idx >= 2:
                prev_courses.append(seq[idx - 2])
        if prev_co_id and prev_co_id not in prev_courses:
            prev_courses.append(prev_co_id)
        prev_courses = prev_courses[:2]
        
        # Load auto_rpoints for prior courses
        auto_rp_map = defaultdict(lambda: defaultdict(float))
        for prev_c in prev_courses:
            auto_rp_raw = run_query(cursor, """
                SELECT student_id, total_score, recorded_date
                FROM qldt_el.auto_rpoints
                WHERE course_id = %s;
            """, (prev_c,))
            latest_dates = {}
            for r in auto_rp_raw:
                sid_raw = int(r['student_id'])
                score = float(r['total_score'])
                rdate = r['recorded_date']
                if sid_raw not in latest_dates or rdate > latest_dates[sid_raw]:
                    latest_dates[sid_raw] = rdate
                    auto_rp_map[prev_c][sid_raw] = score
                    
        is_ks24 = "KS24" in cname
        is_qtkd = c_info in qtkd_classes
        if is_ks24:
            base_scale = 1.10
        elif is_qtkd:
            base_scale = 0.80
        else:
            base_scale = 0.85

        class_risk_list = []
        
        for s in students:
            sid = int(s['student_id'])
            sname = s['student_name']
            scode = s['student_code']
            # --- CALCULATE INDICATORS ---
            
            # A. Attendance (absent rate %)
            fr_att = student_fr_att.get(sid)
            att_list = student_att.get(sid, [])
            total_sessions = len(att_list)
            if fr_att is not None:
                att_val = fr_att
            else:
                absences = sum(1 for item in att_list if item['status'] in ('0', '2'))
                att_val = (absences / total_sessions) * 100 if total_sessions > 0 else 0.0
            
            # B. Homework % complete
            fr_hw = student_fr_hw.get(sid)
            if fr_hw is not None:
                hw_val = fr_hw
                hw_total = 10
            else:
                hw_total = student_hw_total.get(sid, 0)
                hw_debt = student_hw_debt.get(sid, 0)
                hw_val = ((hw_total - hw_debt) / hw_total) * 100 if hw_total > 0 else 100.0
            
            # C. Elearning late count
            fr_el = student_fr_el.get(sid)
            if fr_el is not None:
                el_val = fr_el
            else:
                el_val = float(student_el_late.get(sid, 0))
                
            # D. Check for Resumed/Suspended student (no history in both 2 prior courses)
            is_resumed_student = False
            if len(prev_courses) >= 1:
                has_history = False
                for prev_c in prev_courses:
                    prev_fr = student_pass_history.get((sid, prev_c))
                    if prev_fr:
                        test_scores = []
                        for score_key in ['project', 'mutiple_choice_1', 'essay_1', 'hackathon_1', 'hackathon_2']:
                            val = prev_fr.get(score_key)
                            if val is not None:
                                test_scores.append(float(val))
                        if test_scores or prev_fr.get('pass') is not None:
                            has_history = True
                            break
                if not has_history:
                    is_resumed_student = True
            
            # E. Discipline Score (Rpoint 1-2 prior courses + Current course)
            discipline_prev = None
            if not is_resumed_student:
                prev_rps = []
                for prev_c in prev_courses:
                    rpoint_val = None
                    prev_fr = student_pass_history.get((sid, prev_c))
                    if prev_fr and prev_fr.get('rpoints') is not None:
                        rpoint_val = prev_fr['rpoints']
                    if rpoint_val is None:
                        rpoint_val = auto_rp_map[prev_c].get(sid)
                    if rpoint_val is not None:
                        prev_rps.append(float(rpoint_val))
                if prev_rps:
                    discipline_prev = np.mean(prev_rps)
            
            if is_resumed_student:
                discipline_prev = 70.0
            elif discipline_prev is None:
                discipline_prev = 100.0
                
            discipline_curr = max(0.0, 100.0 - att_val)
            discipline_val = 0.5 * discipline_prev + 0.5 * discipline_curr
            
            # F. Study Performance: P_prereq (Test score of prerequisite course)
            penalty_resumption = 1.0
            if is_resumed_student:
                P_prereq = 50.0
                penalty_resumption = 0.85
            elif len(prev_courses) >= 1:
                prev_c_main = prev_courses[0]
                prev_fr = student_pass_history.get((sid, prev_c_main))
                if prev_fr:
                    prev_att = prev_fr.get('attendance', 0.0)
                    test_scores = []
                    for score_key in ['project', 'mutiple_choice_1', 'essay_1', 'hackathon_1', 'hackathon_2']:
                        val = prev_fr.get(score_key)
                        if val is not None:
                            test_scores.append(float(val))
                    prev_score = np.mean(test_scores) if test_scores else None
                    
                    if prev_score is not None:
                        P_prereq_base = prev_score
                    else:
                        fallback_grades = [v for v in [prev_fr.get('homework'), prev_fr.get('rpoints')] if v is not None]
                        P_prereq_base = np.mean(fallback_grades) if fallback_grades else 75.0
                        
                    is_prev_failed_hard_test = (prev_score is not None and prev_score < 40.0)
                    if prev_att > 30.0 or is_prev_failed_hard_test:
                        P_prereq = P_prereq_base * (0.90 if is_ks24 else 0.80)
                    else:
                        P_prereq = P_prereq_base
                else:
                    P_prereq = 75.0
            else:
                P_prereq = 75.0
                
            # G. Current test grade (P_hack) & Heuristic estimation if not tested
            shack = hackathon_map.get((sid, co_id))
            if shack is not None:
                shack_val = float(shack)
                P_hack = min(100.0, shack_val * 1.25)
            else:
                # Optimal Heuristic: 65% Prereq + 35% Current Discipline
                P_hack_est = 0.65 * P_prereq + 0.35 * discipline_curr
                P_hack = min(100.0, P_hack_est * 1.25)
            
            is_qtkd = c_info in qtkd_classes
            if not is_qtkd:
                P_learning = 0.42 * P_prereq + 0.58 * P_hack
            else:
                P_learning = P_prereq
            P_learning = min(100.0, max(0.0, P_learning))
            
            # H. Project grade
            proj_val = student_fr_proj.get(sid)
            
            # --- CONSECUTIVE PENALTIES WITH BATCH ADAPTATION ---
            consecutive_abs = 0
            for item in reversed(att_list):
                if item['status'] in ('0', '2'):
                    consecutive_abs += 1
                else:
                    break
            
            if is_ks24:
                penalty_abs = 0.90 if consecutive_abs >= 2 else 1.0
            else:
                penalty_abs = 0.75 if consecutive_abs >= 2 else 1.0
            
            consecutive_hw = 0
            hw_details = student_hw_details.get(sid, [])
            for item in reversed(hw_details):
                if item['is_debt']:
                    consecutive_hw += 1
                else:
                    break
            
            if is_ks24:
                penalty_hw = 0.92 if consecutive_hw >= 2 else 1.0
            else:
                penalty_hw = 0.78 if consecutive_hw >= 2 else 1.0
            
            # --- CRITERIA CHECKS ---
            reasons = []
            # --- CRITERIA CHECKS WITH SOFT RELAXATION ---
            is_ks24 = "KS24" in cname
            is_failed = False
            soft_penalty_factor = 1.0
            
            if total_sessions > 3:
                if is_ks24:
                    # KHÓA CŨ (KS24): Không cấm thi hay phạt điểm kỷ luật, chỉ lưu cảnh báo nhắc nhở
                    if att_val > 20.0:
                        reasons.append(f"Cảnh báo vắng học ({att_val:.1f}%)")
                    if hw_total >= 2 and hw_val < 80.0:
                        reasons.append(f"Cảnh báo nợ bài tập ({hw_val:.1f}% xong)")
                    if el_val > 3.0:
                        reasons.append(f"Cảnh báo Elearning ({el_val:.0f} bài)")
                    is_course_finished = any(v is not None for v in student_fr_rp.values())
                    if is_course_finished and discipline_val < 80.0:
                        reasons.append(f"Cảnh báo Rpoint thấp ({discipline_val:.1f}/80)")
                else:
                    # KHÓA MỚI (KS25 & QTKD): Cấm thi cứng khi vi phạm quá nặng, phạt điểm vừa phải khi vi phạm nhẹ
                    # 1. Chuyên cần (vắng học)
                    if att_val > 30.0:
                        is_failed = True
                        reasons.append(f"Cấm thi: Vắng học quá nặng ({att_val:.1f}%)")
                    elif att_val > 20.0:
                        soft_penalty_factor *= 0.65
                        reasons.append(f"Cảnh báo vắng học ({att_val:.1f}%)")
                    
                    # 2. Bài tập nợ
                    if hw_total >= 2:
                        if hw_val < 50.0:
                            is_failed = True
                            reasons.append(f"Cấm thi: Nợ bài tập quá nặng ({hw_val:.1f}% xong)")
                        elif hw_val < 80.0:
                            soft_penalty_factor *= 0.70
                            reasons.append(f"Cảnh báo nợ bài tập ({hw_val:.1f}% xong)")
                    
                    # 3. Elearning vi phạm
                    if el_val > 5.0:
                        is_failed = True
                        reasons.append(f"Cấm thi: Vi phạm Elearning quá nặng ({el_val:.0f} bài)")
                    elif el_val > 3.0:
                        soft_penalty_factor *= 0.75
                        reasons.append(f"Cảnh báo Elearning ({el_val:.0f} bài)")

                    # 4. Rpoint chốt
                    is_course_finished = any(v is not None for v in student_fr_rp.values())
                    if is_course_finished:
                        if discipline_val < 65.0:
                            is_failed = True
                            reasons.append(f"Cấm thi: Rpoint quá thấp ({discipline_val:.1f}/80)")
                        elif discipline_val < 80.0:
                            soft_penalty_factor *= 0.70
                            reasons.append(f"Cảnh báo Rpoint thấp ({discipline_val:.1f}/80)")
                        
            if is_qtkd and proj_val is not None and proj_val < 50.0:
                is_failed = True
                reasons.append("Trượt Project")
            
            # --- CALCULATE FINAL PREDICTED SCORE WITH CDC ---
            cdc_val = get_course_difficulty(c_info['curr_course_name'], cursor)
            P_learning_adj = P_learning / cdc_val
            p_eligible = P_learning_adj * 0.6 + discipline_val * 0.4
            
            # Áp dụng các hình phạt chuyên cần liên tiếp, nợ bài tập, bảo lưu, vi phạm kỷ luật nhẹ và base_scale tối ưu
            p_eligible = p_eligible * penalty_abs * penalty_hw * penalty_resumption * soft_penalty_factor * base_scale
            p_eligible = min(100.0, max(0.0, p_eligible))
            
            if is_failed:
                p_eligible = 0.0
                
            is_risk = is_failed or (p_eligible < 50.0)
            if not is_failed and p_eligible < 50.0:
                reason_detail = f"Xác suất đỗ thấp ({p_eligible:.1f}%)"
                if is_resumed_student:
                    reason_detail += " [Cảnh báo] Học viên mới / quay lại sau bảo lưu"
                else:
                    prev_c_main = prev_courses[0] if prev_courses else None
                    prev_fr_c = student_pass_history.get((sid, prev_c_main)) if prev_c_main else None
                    t_sc = []
                    if prev_fr_c:
                        for sk in ['project', 'mutiple_choice_1', 'essay_1', 'hackathon_1', 'hackathon_2']:
                            v = prev_fr_c.get(sk)
                            if v is not None: t_sc.append(float(v))
                    pr_sc = np.mean(t_sc) if t_sc else None
                    if prev_fr_c and (prev_fr_c.get('pass') == 0 or (pr_sc is not None and pr_sc < 50.0)):
                        reason_detail += " [Học lực] Yếu kiến thức nền tảng (Trượt môn trước)"
                
                if consecutive_abs >= 2:
                    reason_detail += f" [Cảnh báo] Nghỉ liên tiếp ({consecutive_abs} buổi)"
                if consecutive_hw >= 2:
                    reason_detail += f" [Cảnh báo] Nợ bài tập liên tiếp ({consecutive_hw} bài)"
                    
                reasons.append(reason_detail)
                
            if is_risk:
                class_risk_list.append({
                    'code': scode,
                    'name': sname,
                    'att': att_val,
                    'hw': hw_val,
                    'el': el_val,
                    'rp': discipline_val,
                    'p_eligible': p_eligible,
                    'is_failed': is_failed,
                    'reasons': ", ".join(reasons)
                })
                
        # Get class teacher
        cursor.execute("SELECT name FROM classes WHERE id = %s", (cid,))
        raw_cname = cursor.fetchone()[0]
        norm_cname = normalize_class_name(raw_cname)
        
        # Try to find teacher in Excel data fallback
        teacher_name = "Chưa rõ"
        for sheet_key, sheet_dict in excel_data.items():
            if normalize_class_name(sheet_key) == norm_cname:
                # Find first teacher name
                for k, entry in sheet_dict.items():
                    if 'teacher' in entry:
                        teacher_name = entry['teacher']
                        break
        
        risk_results[cname] = {
            'total_students': len(students),
            'risk_count': len(class_risk_list),
            'risk_rate': (len(class_risk_list) / len(students)) * 100 if students else 0.0,
            'risk_students': class_risk_list,
            'teacher': teacher_name
        }

    # Save to JSON
    with open('scratch/student_risk_data.json', 'w', encoding='utf-8') as jf:
        json.dump(risk_results, jf, ensure_ascii=False, indent=4)
        
    # Standalone Markdown report generator
    with open('data/student_risk_report.md', 'w', encoding='utf-8') as mf:
        mf.write("# Báo cáo Chi tiết Sinh viên có nguy cơ trượt môn (Thực tế 100% từ DB)\n\n")
        mf.write("Báo cáo này truy vấn trực tiếp dữ liệu vi phạm kỷ luật thực tế của từng học viên trong Cơ sở dữ liệu MySQL (Không cào bằng hiệu chỉnh Excel).\n\n")
        
        mf.write("## I. BẢNG TỔNG HỢP TỶ LỆ NGUY CƠ TRƯỢT THEO LỚP\n\n")
        mf.write("| Tên lớp | Giảng viên | Sĩ số | Số SV nguy cơ | Tỷ lệ nguy cơ |\n")
        mf.write("| :--- | :--- | :---: | :---: | :---: |\n")
        for cname, info in risk_results.items():
            mf.write(f"| {cname} | {info['teacher']} | {info['total_students']} | {info['risk_count']} | **{info['risk_rate']:.2f}%** |\n")
            
        mf.write("\n## II. CHI TIẾT DANH SÁCH SINH VIÊN THEO TỪNG LỚP\n\n")
        for cname, info in risk_results.items():
            mf.write(f"### Lớp: {cname}\n")
            mf.write(f"*   **Giảng viên**: {info['teacher']} | Sĩ số: {info['total_students']} SV | Số SV nguy cơ: {info['risk_count']} SV ({info['risk_rate']:.2f}%)\n\n")
            
            discipline_risks = [s for s in info['risk_students'] if s['is_failed']]
            academic_risks = [s for s in info['risk_students'] if not s['is_failed']]
            
            if discipline_risks:
                mf.write("#### ⚠️ Nhóm 1: Cảnh báo Kỷ luật (Nguy cơ cấm thi môn hiện tại)\n")
                mf.write("| MSSV | Họ và tên | Chuyên cần (vắng) | Bài tập (% xong) | Elearning (lỗi) | Rpoint | Điểm dự báo | Chi tiết vi phạm |\n")
                mf.write("| :--- | :--- | :---: | :---: | :---: | :---: | :---: | :--- |\n")
                for s in discipline_risks:
                    mf.write(f"| {s['code']} | {s['name']} | {s['att']:.1f}% | {s['hw']:.1f}% | {s['el']:.0f} | {s['rp']:.1f} | {s['p_eligible']:.1f}% | {s['reasons']} |\n")
                mf.write("\n")
                
            if academic_risks:
                mf.write("#### 📉 Nhóm 2: Cảnh báo Học lực (Nguy cơ trượt môn hiện tại)\n")
                mf.write("| MSSV | Họ và tên | Chuyên cần (vắng) | Bài tập (% xong) | Elearning (lỗi) | Rpoint | Điểm dự báo | Chi tiết vi phạm |\n")
                mf.write("| :--- | :--- | :---: | :---: | :---: | :---: | :---: | :--- |\n")
                for s in academic_risks:
                    mf.write(f"| {s['code']} | {s['name']} | {s['att']:.1f}% | {s['hw']:.1f}% | {s['el']:.0f} | {s['rp']:.1f} | {s['p_eligible']:.1f}% | {s['reasons']} |\n")
                mf.write("\n")
                
            if not info['risk_students']:
                mf.write("> 🎉 Lớp không có sinh viên nào có nguy cơ trượt môn.\n\n")
            mf.write("\n" + "-"*50 + "\n\n")
            
    print("Real risk analysis completed.")
    cursor.close()
    conn.close()

# Helper for load_excel_data to retrieve teacher name
def load_excel_data(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    class_data = defaultdict(dict)
    for sheetname in wb.sheetnames:
        if sheetname == 'Sheet1' or 'SKL' in sheetname:
            continue
        sheet = wb[sheetname]
        max_r = sheet.max_row
        for r in range(5, max_r + 1):
            cname = sheet.cell(row=r, column=2).value
            teacher = sheet.cell(row=r, column=3).value
            if cname:
                norm_name = normalize_class_name(cname)
                class_data[norm_name][sheetname] = {
                    'teacher': str(teacher).strip() if teacher else "Ẩn danh"
                }
    return class_data

if __name__ == '__main__':
    main()
