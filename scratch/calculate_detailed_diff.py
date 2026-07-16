import openpyxl
import sys
from datetime import datetime, date
from collections import defaultdict
import numpy as np

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'docs/PTIT_Chiso.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

branches_config = {
    'HN': {
        'KS24': {
            'classes': ['HN-K24-CNTT1', 'HN-K24-CNTT2', 'HN-K24-CNTT3', 'HN-K24-CNTT4'],
            'sheet_curr': 'KS24_AI', 'sheet_prev': 'KS24_AI',
            'start_curr': date(2026, 6, 29), 'end_curr': date(2026, 7, 5),
            'start_prev': date(2026, 6, 22), 'end_prev': date(2026, 6, 28),
            'label': 'KS24-CNTT (Môn AI)'
        },
        'KS25_CNTT': {
            'classes': ['HN-K25-CNTT1', 'HN-K25-CNTT2', 'HN-K25-CNTT3', 'HN-K25-CNTT4', 'HN-K25-CNTT5', 'HN-K25-CNTT6', 'HN-K25-CNTT8'],
            'sheet_curr': 'KS25_Python_Web', 'sheet_prev': 'KS25_Python_Web',
            'start_curr': date(2026, 6, 29), 'end_curr': date(2026, 7, 5),
            'start_prev': date(2026, 6, 22), 'end_prev': date(2026, 6, 28),
            'label': 'KS25-CNTT (Python Web)'
        },
        'KS25_QTKD': {
            'classes': ['HN-K25-QTKD1', 'HN-K25-QTKD2', 'HN-K25-QTKD3'],
            'sheet_curr': 'KS25_QTKD_PRJ302', 'sheet_prev': 'KS25_QTKD_DTB202',
            'start_curr': date(2026, 6, 29), 'end_curr': date(2026, 7, 5),
            'start_prev': date(2026, 6, 22), 'end_prev': date(2026, 6, 28),
            'label': 'KS25-QTKD (Môn PRJ302)'
        }
    },
    'HCM': {
        'KS24': {
            'classes': ['HCM-K24-CNTT1'],
            'sheet_curr': 'KS24_AI', 'sheet_prev': 'KS24_AI',
            'start_curr': date(2026, 6, 29), 'end_curr': date(2026, 7, 5),
            'start_prev': date(2026, 6, 22), 'end_prev': date(2026, 6, 28),
            'label': 'KS24-CNTT (Môn AI)'
        },
        'KS25_CNTT': {
            'classes': ['HCM-K25-CNTT5', 'HCM-K25-CNTT6', 'HCM-K25-CNTT7', 'HCM-K25-CNTT8'],
            'sheet_curr': 'KS25_Python_Web', 'sheet_prev': 'KS25_Python_Web',
            'start_curr': date(2026, 6, 29), 'end_curr': date(2026, 7, 5),
            'start_prev': date(2026, 6, 22), 'end_prev': date(2026, 6, 28),
            'label': 'KS25-CNTT (Python Web)'
        }
    }
}

def get_weekly_metrics(sheetname, classes_target, start_date, end_date):
    if sheetname not in wb.sheetnames:
        return {}
    sheet = wb[sheetname]
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    dates_list = []
    current_date = None
    for c_idx in range(3, len(row3)):
        val3 = row3[c_idx]
        val4 = row4[c_idx]
        if val3:
            current_date = parse_date(val3)
        if current_date and start_date <= current_date <= end_date:
            dates_list.append((c_idx, current_date, val4))
            
    res = {}
    current_class = None
    for r in range(5, sheet.max_row + 1):
        cname = sheet.cell(row=r, column=2).value
        teacher_val = sheet.cell(row=r, column=3).value
        
        if cname:
            current_class = normalize_class_name(cname)
            matched_class = None
            for tc in classes_target:
                if tc == current_class:
                    matched_class = tc
                    break
            if not matched_class:
                continue
                
            teacher_name = str(teacher_val).strip() if teacher_val else "N/A"
            
            day_vals = defaultdict(list)
            for c_idx, d, val4 in dates_list:
                val = sheet.cell(row=r, column=c_idx + 1).value
                if val is not None:
                    # Clean/parse values just like clean_excel_data.py
                    val_str = str(val).strip()
                    if val_str == '' or val_str.lower() == 'nan':
                        continue
                    if val_str == '-':
                        val_num = 0.0
                    elif ',' in val_str and not ('(' in val_str):
                        val_num = float(val_str.replace(',', '.'))
                    elif '(' in val_str:
                        prefix = val_str.split('(')[0].strip()
                        val_num = float(prefix.replace(',', '.'))
                    else:
                        try:
                            val_num = float(val_str)
                        except ValueError:
                            continue
                    day_vals[val4].append(val_num)
                    
            averages = {}
            for metric in ['Chuyên cần', 'Bài tập', 'Elearning']:
                vals = day_vals.get(metric, [])
                averages[metric] = sum(vals) / len(vals) if vals else 0.0
                
            res[matched_class] = {
                'teacher': teacher_name,
                'metrics': averages
            }
    return res

print("=== SO SÁNH CHI TIẾT TỪNG CHỈ SỐ THEO KHÓA HỌC ===")
for bname, bgroups in branches_config.items():
    print(f"\nCƠ SỞ: {bname}")
    for gkey, ginfo in bgroups.items():
        curr = get_weekly_metrics(ginfo['sheet_curr'], ginfo['classes'], ginfo['start_curr'], ginfo['end_curr'])
        prev = get_weekly_metrics(ginfo['sheet_prev'], ginfo['classes'], ginfo['start_prev'], ginfo['end_prev'])
        
        cc_curr, bt_curr, el_curr = [], [], []
        cc_prev, bt_prev, el_prev = [], [], []
        
        for cls in ginfo['classes']:
            cls_curr = curr.get(cls)
            cls_prev = prev.get(cls)
            
            c_m = cls_curr['metrics'] if cls_curr else {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}
            p_m = cls_prev['metrics'] if cls_prev else {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}
            
            if cls_curr:
                cc_curr.append(c_m['Chuyên cần'])
                bt_curr.append(c_m['Bài tập'])
                el_curr.append(c_m['Elearning'])
            if cls_prev:
                cc_prev.append(p_m['Chuyên cần'])
                bt_prev.append(p_m['Bài tập'])
                el_prev.append(p_m['Elearning'])
                
        avg_curr_cc = np.mean(cc_curr) if cc_curr else 0.0
        avg_curr_bt = np.mean(bt_curr) if bt_curr else 0.0
        avg_curr_el = np.mean(el_curr) if el_curr else 0.0
        
        avg_prev_cc = np.mean(cc_prev) if cc_prev else 0.0
        avg_prev_bt = np.mean(bt_prev) if bt_prev else 0.0
        avg_prev_el = np.mean(el_prev) if el_prev else 0.0
        
        print(f"  + {ginfo['label']}:")
        print(f"    - Chuyên cần: {avg_prev_cc:.2f}% -> {avg_curr_cc:.2f}% | Chênh lệch: {avg_curr_cc - avg_prev_cc:+.2f}%")
        print(f"    - Bài tập: {avg_prev_bt:.2f}% -> {avg_curr_bt:.2f}% | Chênh lệch: {avg_curr_bt - avg_prev_bt:+.2f}%")
        print(f"    - Elearning: {avg_prev_el:.2f}% -> {avg_curr_el:.2f}% | Chênh lệch: {avg_curr_el - avg_prev_el:+.2f}%")
