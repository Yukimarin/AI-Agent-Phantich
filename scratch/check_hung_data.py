import sys
import os
import openpyxl

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

path = r"data/PTIT_Chiso.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

target_sheets = [
    'KS24-JavaAdvance', 'KS24_JavaWeb', 'KS24_JWS', 'KS24_AI',
    'KS25_Javascript', 'KS25_Database', 'KS25_Python', 'KS25_Python_Web',
    'KS25_QTKD_M103', 'KS25_QTKD_M104', 'KS25_QTKD_DTB201', 'KS25_QTKD_DTB202',
    'KS25_QTKD_PRJ302'
]

for name in target_sheets:
    if name not in wb.sheetnames:
        continue
    ws = wb[name]
    for r in range(1, ws.max_row+1):
        for c in range(1, 6):
            val = ws.cell(row=r, column=c).value
            if val and "Hùng" in str(val):
                print(f"Sheet: {name} | Row {r} Col {c}: {val}")
                # Print row details
                row_vals = [ws.cell(row=r, column=x).value for x in range(1, 15)]
                print(f"  Row {r}: {row_vals}")
                # If there's an assistant role (row r or r+1 or r-1), check
                prev_row = [ws.cell(row=r-1, column=x).value for x in range(1, 15)] if r > 1 else []
                next_row = [ws.cell(row=r+1, column=x).value for x in range(1, 15)] if r < ws.max_row else []
                print(f"  Row {r-1} (Prev): {prev_row}")
                print(f"  Row {r+1} (Next): {next_row}")
wb.close()
