import shutil
import os
import openpyxl
from datetime import datetime, date

src = r"C:\Users\DELL\Desktop\Backup\PTIT\PTIT_Chiso.xlsx"
dest_docs = r"docs/PTIT_Chiso.xlsx"
dest_data = r"data/PTIT_Chiso.xlsx"

# Copy files
print(f"Copying {src} to {dest_docs}...")
try:
    shutil.copy(src, dest_docs)
    print("Copied to docs successfully.")
except Exception as e:
    print(f"Error copying to docs: {e}")

try:
    shutil.copy(src, dest_data)
    print("Copied to data successfully.")
except Exception as e:
    print(f"Error copying to data: {e}")

# Inspect dates
if os.path.exists(dest_docs):
    wb = openpyxl.load_workbook(dest_docs, data_only=True)
    print("Sheets:", wb.sheetnames)
    
    def parse_date(d_val):
        if not d_val:
            return None
        if isinstance(d_val, datetime):
            return d_val.date()
        if isinstance(d_val, date):
            return d_val
        d_str = str(d_val).strip()
        parts = d_str.split('/')
        if len(parts) == 2:
            try:
                return date(2026, int(parts[1]), int(parts[0]))
            except ValueError:
                return None
        elif len(parts) == 3:
            try:
                year = int(parts[2])
                if year < 100:
                    year += 2000
                return date(year, int(parts[1]), int(parts[0]))
            except ValueError:
                return None
        return None

    for sheetname in wb.sheetnames:
        if 'KS' not in sheetname and 'SKL' not in sheetname:
            continue
        sheet = wb[sheetname]
        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))
        if not row3 or not row3[0]:
            continue
        row3 = row3[0]
        
        dates = []
        for c_idx in range(3, len(row3)):
            val = row3[c_idx]
            if val:
                p_date = parse_date(val)
                if p_date:
                    dates.append(p_date)
        if dates:
            print(f"Sheet '{sheetname}': min date = {min(dates)}, max date = {max(dates)}")
else:
    print("Dest file does not exist.")
