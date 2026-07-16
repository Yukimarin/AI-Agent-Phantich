import sys
import openpyxl
from datetime import datetime, date
from collections import defaultdict

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'C:/Users/DELL/Desktop/Education-DB-Analytic/docs/PTIT_Chiso.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

# Kiem tra cac ngay co trong sheet KS25_Python de xem co phai tuan truoc do khong
def check_python_sheet():
    if 'KS25_Python' not in wb.sheetnames:
        print("Sheet KS25_Python not found")
        return
    sheet = wb['KS25_Python']
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    dates_list = []
    current_date = None
    for c_idx in range(3, len(row3)):
        val3 = row3[c_idx]
        val4 = row4[c_idx]
        if val3:
            current_date = parse_date(val3)
        if current_date:
            dates_list.append((c_idx, current_date, val4))
    
    print("\n--- Sheet KS25_Python Dates ---")
    print(f"Total dates: {len(dates_list)}")
    unique_dates = sorted(list(set([d for _, d, _ in dates_list])))
    print(f"Unique dates: {unique_dates[-10:]}")
    
    # In ra vi pham cua cac lop trong 5 ngay cuoi cung cua sheet Python
    classes = ['HN-K25-CNTT1', 'HN-K25-CNTT2', 'HN-K25-CNTT3', 'HN-K25-CNTT4', 'HN-K25-CNTT5', 'HN-K25-CNTT6', 'HCM-K25-CNTT8', 'HCM-K25-CNTT7', 'HCM-K25-CNTT6', 'HCM-K25-CNTT5']
    for r in range(5, sheet.max_row + 1):
        cname = sheet.cell(row=r, column=2).value
        if not cname:
            continue
        cname_norm = normalize_class_name(cname)
        if cname_norm in classes:
            print(f"\nClass {cname_norm} in KS25_Python:")
            # nhom theo ngay
            day_data = defaultdict(dict)
            for c_idx, d, val4 in dates_list:
                val = sheet.cell(row=r, column=c_idx + 1).value
                day_data[d][val4] = val
            for d in sorted(day_data.keys())[-5:]:
                print(f"  Date {d} -> {dict(day_data[d])}")

check_python_sheet()
