import sys
import openpyxl
from datetime import datetime, date
from collections import defaultdict

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'C:/Users/DELL/Desktop/Education-DB-Analytic/docs/PTIT_Chiso.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

# Phân nhóm lớp
# KS24:
# - HN: HN-K24-CNTT1, HN-K24-CNTT2, HN-K24-CNTT3, HN-K24-CNTT4
# - HCM: HCM-K24-CNTT1
# KS25 CNTT:
# - HN: HN-K25-CNTT1, HN-K25-CNTT2, HN-K25-CNTT3, HN-K25-CNTT4, HN-K25-CNTT5, HN-K25-CNTT6
# - HCM: HCM-K25-CNTT5, HCM-K25-CNTT6, HCM-K25-CNTT7, HCM-K25-CNTT8
# KS25 QTKD:
# - HN: HN-K25-QTKD1, HN-K25-QTKD2, HN-K25-QTKD3

groups = {
    'KS24_HN': {
        'sheet_curr': 'KS24_AI',
        'sheet_prev': 'KS24_AI', # Môn AI bắt đầu từ 17/6, nên tuần trước cũng lấy AI (17-19/6)
        'classes': ['HN-K24-CNTT1', 'HN-K24-CNTT2', 'HN-K24-CNTT3', 'HN-K24-CNTT4']
    },
    'KS24_HCM': {
        'sheet_curr': 'KS24_AI', # Người dùng nói HCM đang học JWS, nhưng ta thấy dữ liệu tuần vừa qua của nó nằm trong sheet KS24_AI (ngày 22, 23/6) hay KS24_JWS?
        # Hãy kiểm tra cả 2 sheet để xem sheet nào có dữ liệu tuần này.
        'sheet_prev': 'KS24_AI',
        'classes': ['HCM-K24-CNTT1']
    },
    'KS25_CNTT_HN': {
        'sheet_curr': 'KS25_Python_Web',
        'sheet_prev': 'KS25_Python',
        'classes': ['HN-K25-CNTT1', 'HN-K25-CNTT2', 'HN-K25-CNTT3', 'HN-K25-CNTT4', 'HN-K25-CNTT5', 'HN-K25-CNTT6']
    },
    'KS25_CNTT_HCM': {
        'sheet_curr': 'KS25_Python_Web',
        'sheet_prev': 'KS25_Python',
        'classes': ['HCM-K25-CNTT5', 'HCM-K25-CNTT6', 'HCM-K25-CNTT7', 'HCM-K25-CNTT8']
    },
    'KS25_QTKD_HN': {
        'sheet_curr': 'KS25_QTKD_DTB202',
        'sheet_prev': 'KS25_QTKD_DTB202',
        'classes': ['HN-K25-QTKD1', 'HN-K25-QTKD2', 'HN-K25-QTKD3']
    }
}

# Khoảng thời gian
start_prev = date(2026, 6, 15)
end_prev = date(2026, 6, 21)
start_curr = date(2026, 6, 22)
end_curr = date(2026, 6, 28)

def get_sheet_data(sheetname, classes_target, start_date, end_date):
    if sheetname not in wb.sheetnames:
        return {}
    sheet = wb[sheetname]
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    dates_list = []
    current_date = None
    for c_idx in range(3, len(row3)):
        val3 = row3[c_idx]
        val4 = row4[c_idx]
        if val3:
            current_date = parse_date(val3)
        if current_date and start_date <= current_date <= end_date:
            dates_list.append((c_idx, current_date, val4))
            
    # Đọc dữ liệu từng lớp
    result = {}
    for r in range(5, sheet.max_row + 1):
        cname = sheet.cell(row=r, column=2).value
        if not cname:
            continue
        cname_norm = normalize_class_name(cname)
        
        # Check xem có thuộc classes_target không
        matched_class = None
        for tc in classes_target:
            if tc == cname_norm:
                matched_class = tc
                break
        if not matched_class:
            continue
            
        teacher = sheet.cell(row=r, column=3).value
        tg = sheet.cell(row=r, column=4).value
        
        # Nhóm điểm theo ngày và loại vi phạm
        day_vals = defaultdict(list)
        for c_idx, d, val4 in dates_list:
            val = sheet.cell(row=r, column=c_idx + 1).value
            if val is not None:
                try:
                    val_float = float(val)
                    day_vals[val4].append(val_float)
                except ValueError:
                    pass
        
        # Tính trung bình cho từng loại vi phạm trong khoảng thời gian này
        averages = {}
        for metric in ['Chuyên cần', 'Bài tập', 'Elearning']:
            vals = day_vals.get(metric, [])
            averages[metric] = sum(vals) / len(vals) if vals else 0.0
            
        result[matched_class] = {
            'teacher': teacher,
            'tg': tg,
            'metrics': averages,
            'raw_count': {m: len(day_vals.get(m, [])) for m in ['Chuyên cần', 'Bài tập', 'Elearning']}
        }
    return result

print("=== PHÂN TÍCH CHỈ SỐ TUẦN VÀ XU HƯỚNG ===")
for gname, ginfo in groups.items():
    print(f"\nGroup: {gname}")
    classes = ginfo['classes']
    
    # Tuần trước
    data_prev = get_sheet_data(ginfo['sheet_prev'], classes, start_prev, end_prev)
    # Tuần này
    data_curr = get_sheet_data(ginfo['sheet_curr'], classes, start_curr, end_curr)
    
    # In chi tiết từng lớp
    print("  --- CHI TIẾT TỪNG LỚP ---")
    for cls in classes:
        p_info = data_prev.get(cls, {'metrics': {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}})
        c_info = data_curr.get(cls, {'metrics': {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}})
        
        p_m = p_info['metrics']
        c_m = c_info['metrics']
        
        print(f"  Class: {cls} | GV: {c_info.get('teacher', 'N/A')} | TG: {c_info.get('tg', 'N/A')}")
        print(f"    Tuần trước -> CC: {p_m['Chuyên cần']:.2f}%, BT: {p_m['Bài tập']:.2f}%, EL: {p_m['Elearning']:.2f}%")
        print(f"    Tuần này   -> CC: {c_m['Chuyên cần']:.2f}%, BT: {c_m['Bài tập']:.2f}%, EL: {c_m['Elearning']:.2f}%")
        
    # Tính trung bình nhóm
    avg_prev = {'Chuyên cần': [], 'Bài tập': [], 'Elearning': []}
    avg_curr = {'Chuyên cần': [], 'Bài tập': [], 'Elearning': []}
    
    for cls in classes:
        if cls in data_prev:
            for m in ['Chuyên cần', 'Bài tập', 'Elearning']:
                avg_prev[m].append(data_prev[cls]['metrics'][m])
        if cls in data_curr:
            for m in ['Chuyên cần', 'Bài tập', 'Elearning']:
                avg_curr[m].append(data_curr[cls]['metrics'][m])
                
    group_prev = {m: sum(avg_prev[m])/len(avg_prev[m]) if avg_prev[m] else 0.0 for m in ['Chuyên cần', 'Bài tập', 'Elearning']}
    group_curr = {m: sum(avg_curr[m])/len(avg_curr[m]) if avg_curr[m] else 0.0 for m in ['Chuyên cần', 'Bài tập', 'Elearning']}
    
    print("  --- TRUNG BÌNH CẢ NHÓM ---")
    print(f"    Tuần trước -> CC: {group_prev['Chuyên cần']:.2f}%, BT: {group_prev['Bài tập']:.2f}%, EL: {group_prev['Elearning']:.2f}%")
    print(f"    Tuần này   -> CC: {group_curr['Chuyên cần']:.2f}%, BT: {group_curr['Bài tập']:.2f}%, EL: {group_curr['Elearning']:.2f}%")
    
    diff = {m: group_curr[m] - group_prev[m] for m in ['Chuyên cần', 'Bài tập', 'Elearning']}
    diff_str = ", ".join([f"{m}: {diff[m]:+.2f}%" for m in ['Chuyên cần', 'Bài tập', 'Elearning']])
    print(f"    Thay đổi   -> {diff_str}")
