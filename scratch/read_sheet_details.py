import sys
import openpyxl
from datetime import datetime, date

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'C:/Users/DELL/Desktop/Education-DB-Analytic/docs/PTIT_Chiso.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

sheets_to_read = {
    'KS24_AI': ['HN-K24-CNTT1(38)', 'HN-K24-CNTT2(39)', 'HN-K24-CNTT3(42)', 'HN-K24-CNTT4(34-32)', 'HCM-K24-CNTT1(44)'],
    'KS24_JWS': ['HCM-K24-CNTT1(44)'],
    'KS25_Python_Web': ['HN-K25-CNTT1(42)', 'HN-K25-CNTT2(43)', 'HN-K25-CNTT3(37)', 'HN-K25-CNTT4(42)', 'HN-K25-CNTT5(42)', 'HN-K25-CNTT6(33)', 'HCM-K25-CNTT8(38)', 'HCM-K25-CNTT7(42-41)', 'HCM-K25-CNTT6(40)', 'HCM-K25-CNTT5(39)'],
    'KS25_QTKD_DTB202': ['HN-K25-QTKD1(37-33)', 'HN-K25-QTKD2(40)', 'HN-K25-QTKD3(27)']
}

for sheetname, class_list in sheets_to_read.items():
    sheet = wb[sheetname]
    print(f"\n==================== SHEET: {sheetname} ====================")
    
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    dates_list = []
    current_date = None
    for c_idx in range(3, len(row3)):
        val3 = row3[c_idx]
        val4 = row4[c_idx]
        if val3:
            current_date = parse_date(val3)
        if current_date:
            dates_list.append((c_idx, current_date, val4))
            
    print(f"Total date columns: {len(dates_list)}")
    
    # Lay tat ca cac dong du lieu cua cac lop target
    for r in range(5, sheet.max_row + 1):
        cname = sheet.cell(row=r, column=2).value
        if not cname:
            continue
        cname_str = str(cname).strip()
        # check xem co trong list target ko
        matched = False
        for target in class_list:
            if target in cname_str or cname_str in target:
                matched = True
                break
        if not matched:
            continue
            
        teacher = sheet.cell(row=r, column=3).value
        tg = sheet.cell(row=r, column=4).value
        print(f"\n  Class: {cname_str} | GV: {teacher} | TG: {tg}")
        
        # In gia tri cua tung ngay
        # Nhom theo ngay
        day_data = {}
        for c_idx, d, val4 in dates_list:
            val = sheet.cell(row=r, column=c_idx + 1).value
            if d not in day_data:
                day_data[d] = {}
            day_data[d][val4] = val
            
        for d in sorted(day_data.keys()):
            metrics = day_data[d]
            metrics_str = ", ".join([f"{k}: {v}" for k, v in metrics.items()])
            print(f"    Date {d} -> {metrics_str}")
