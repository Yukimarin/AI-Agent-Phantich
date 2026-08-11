import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("data/inputs/PTIT_Chiso.xlsx", data_only=True)
for name in wb.sheetnames:
    if "KS25" in name or "K25" in name:
        sheet = wb[name]
        print(f"\nSheet: {name}, max_row={sheet.max_row}, max_col={sheet.max_column}")
        header_idx = None
        for r in range(1, 15):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
            if any(isinstance(x, str) and 'Lớp' in x for x in row_vals if x):
                header_idx = r
                break
        if header_idx:
            row_header = [sheet.cell(row=header_idx, column=c).value for c in range(1, sheet.max_column + 1)]
            print("  Header row index:", header_idx)
            print("  Header row (first 15 cols):", row_header[:15])
            print("  Header row (last 10 cols):", [x for x in row_header if x][-10:])
wb.close()
