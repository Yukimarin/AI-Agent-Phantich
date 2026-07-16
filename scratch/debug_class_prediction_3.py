import mysql.connector
import sys
import os
import openpyxl
import numpy as np

sys.stdout.reconfigure(encoding='utf-8')

# Import normalized name from excel_loader if exists
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from excel_loader import normalize_class_name

def main():
    conn = mysql.connector.connect(
        host="localhost",
        port=3306,
        user="root",
        password="",
        database="qldt_el"
    )
    cursor = conn.cursor()
    
    # Check HCM-KS24-CNTT1 and IT203A-K24 (Java Fundamental)
    class_raw = run_query(cursor, "SELECT id, name FROM qldt_el.classes WHERE name = 'HCM-KS24-CNTT1';")
    course_raw = run_query(cursor, "SELECT id, name FROM qldt_el.courses WHERE name LIKE '%Java Fundamental%' AND name LIKE '%K24%';")
    
    print("Class ID raw:", class_raw)
    print("Course ID raw:", course_raw)
    
    if not class_raw or not course_raw:
        return
        
    cid = class_raw[0]['id']
    co_id = course_raw[0]['id']
    
    students = run_query(cursor, """
        SELECT student_id, homework, elearning, attendance, hackathon_1, hackathon_2, rpoints, project, pass
        FROM qldt_el.final_results
        WHERE class_id = %s AND course_id = %s AND pass IS NOT NULL;
    """, (cid, co_id))
    
    print(f"\nDebug student details for class {cid} and course {co_id}:")
    print(f"Total students: {len(students)}")
    
    # Calculate averages
    att_list = [s['attendance'] if s['attendance'] is not None else 0.0 for s in students]
    hw_list = [s['homework'] if s['homework'] is not None else 100.0 for s in students]
    el_list = [s['elearning'] if s['elearning'] is not None else 0.0 for s in students]
    rp_list = [s['rpoints'] if s['rpoints'] is not None else 100.0 for s in students]
    
    db_att_avg = np.mean(att_list)
    db_hw_avg = np.mean(hw_list)
    db_el_avg = np.mean(el_list)
    db_rp_avg = np.mean(rp_list)
    
    print(f"DB averages: att_avg={db_att_avg:.2f}, hw_avg={db_hw_avg:.2f}, el_avg={db_el_avg:.2f}, rp_avg={db_rp_avg:.2f}")
    
    # Load Excel chot
    excel_path = "docs/PTIT_Chiso.xlsx"
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    norm_cname = normalize_class_name(class_raw[0]['name'])
    
    # Search in KS24-JavaAdvance sheet
    sheet = wb['KS24-JavaAdvance']
    excel_disc = None
    for r in range(5, sheet.max_row+1):
        if sheet.cell(row=r, column=2).value and normalize_class_name(sheet.cell(row=r, column=2).value) == norm_cname:
            excel_disc = {
                'cc': sheet.cell(row=r, column=37).value, # col 36 is chuyên cần
                'bt': sheet.cell(row=r, column=38).value, # col 37 is bài tập
                'el': sheet.cell(row=r, column=39).value, # col 38 is elearning
                'rp': sheet.cell(row=r, column=41).value  # col 40 is rpoint chot
            }
            break
            
    print("Excel chot found:", excel_disc)
    
    # Run calibration and rules check
    excel_cc = excel_disc['cc'] if (excel_disc and excel_disc['cc'] is not None) else db_att_avg
    excel_bt_err = excel_disc['bt'] if (excel_disc and excel_disc['bt'] is not None) else (100.0 - db_hw_avg)
    excel_hw = 100.0 - excel_bt_err
    excel_el = excel_disc['el'] if (excel_disc and excel_disc['el'] is not None) else db_el_avg
    excel_rp = excel_disc['rp'] if (excel_disc and excel_disc['rp'] is not None) else max(0.0, 100.0 - excel_cc - excel_bt_err - excel_el)
    
    print(f"Calibrating to: excel_cc={excel_cc}, excel_hw={excel_hw}, excel_el={excel_el}, excel_rp={excel_rp}")
    
    has_project = any(s['project'] is not None and s['project'] > 0.0 for s in students)
    print(f"Course has_project: {has_project}")
    
    # Test for first 5 students
    for i, s in enumerate(students[:10]):
        rp_db = rp_list[i]
        rp_cal = rp_db + (excel_rp - db_rp_avg)
        rp_cal = min(120.0, max(0.0, rp_cal))
        
        att_db = att_list[i]
        att_cal = att_db + (excel_cc - db_att_avg)
        att_cal = min(100.0, max(0.0, att_cal))
        
        hw_db = hw_list[i]
        hw_cal = hw_db + (excel_hw - db_hw_avg)
        hw_cal = min(100.0, max(0.0, hw_cal))
        
        el_db = el_list[i]
        el_cal = 0.0 if (excel_disc and excel_disc.get('el') == 0.0) else el_db
        
        proj = s['project']
        
        is_cc_old = att_cal <= 20.0
        is_rp_old = rp_cal >= 80.0
        is_proj_old_ok = True
        if has_project and proj is not None and proj < 50.0:
            is_proj_old_ok = False
            
        print(f"Student {s['student_id']}: rp_db={rp_db}, rp_cal={rp_cal:.2f}, att_cal={att_cal:.1f}%, hw_cal={hw_cal:.1f}%, proj={proj}, pass={s['pass']}")
        print(f"  Old rules check: is_cc_old={is_cc_old}, is_rp_old={is_rp_old}, is_proj_old_ok={is_proj_old_ok}")

def run_query(cursor, query, params=None):
    cursor.execute(query, params or ())
    columns = [desc[0] for desc in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

if __name__ == "__main__":
    main()
