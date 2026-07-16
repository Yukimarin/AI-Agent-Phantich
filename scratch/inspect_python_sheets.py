import openpyxl

excel_path = 'docs/PTIT_Chiso.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

for sheetname in ['KS25_Python', 'KS25_Python_Web']:
    if sheetname not in wb.sheetnames:
        print(f"Sheet {sheetname} not found")
        continue
    sheet = wb[sheetname]
    classes = set()
    for r in range(5, sheet.max_row + 1):
        cval = sheet.cell(row=r, column=2).value
        if cval:
            classes.add(str(cval).strip())
    print(f"Sheet {sheetname} classes: {sorted(list(classes))}")
