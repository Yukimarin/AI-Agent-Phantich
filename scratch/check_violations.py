import openpyxl

excel_path = 'C:/Users/DELL/Desktop/Education-DB-Analytic/docs/PTIT_Chiso.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

for sname in ['KS24_AI', 'KS25_Python_Web', 'KS25_Python', 'KS25_QTKD_DTB202']:
    sheet = wb[sname]
    print(f"\n--- Checking TG points in {sname} ---")
    tg_rows = 0
    tg_rows_with_data = 0
    
    for r in range(5, sheet.max_row + 1):
        c2 = sheet.cell(row=r, column=2).value
        c3 = sheet.cell(row=r, column=3).value
        if not c2 and c3: # Day la dong TG
            tg_rows += 1
            has_data = False
            for c in range(4, sheet.max_column + 1):
                val = sheet.cell(row=r, column=c).value
                if val is not None:
                    has_data = True
                    break
            if has_data:
                tg_rows_with_data += 1
                
    print(f"Total TG rows: {tg_rows}, TG rows with data: {tg_rows_with_data}")
