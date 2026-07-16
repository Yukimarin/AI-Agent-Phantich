import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = (name_str
                .replace("KS25", "K25")
                .replace("KS24", "K24")
                .replace("KS23", "K23"))
    return name_str

def main():
    wb = openpyxl.load_workbook("docs/PTIT_Chiso.xlsx", data_only=True)
    
    classes_to_check = [
        'HCM-KS25-CNTT5_HK2',
        'HCM-KS25-CNTT6_HK2',
        'HCM-KS25-CNTT7_HK2',
        'HCM-KS25-CNTT8_HK2'
    ]
    
    sheets_to_check = ['KS25_Database', 'KS25_Python', 'KS25_Python_Web']
    
    print("=== EXCEL DATA FOR KS25 HCM CLASSES ===")
    for cname_db in classes_to_check:
        norm_db = normalize_class_name(cname_db)
        print(f"\nDB Class: {cname_db} (Normalized: {norm_db})")
        
        for sheetname in sheets_to_check:
            sheet = wb[sheetname]
            # Find class row
            found = False
            for r in range(5, sheet.max_row + 1):
                val = sheet.cell(row=r, column=2).value
                if val:
                    norm_excel = normalize_class_name(val)
                    if norm_excel == norm_db:
                        # Print some last columns values to see what is fetched
                        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
                        row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
                        
                        dates_list = []
                        for c_idx in range(3, sheet.max_column):
                            val3 = row3[c_idx]
                            val4 = row4[c_idx]
                            if val3:
                                dates_list.append((c_idx, val3, val4))
                            elif dates_list:
                                dates_list.append((c_idx, dates_list[-1][1], val4))
                                
                        # Last date values
                        cc_val, bt_val, el_val = None, None, None
                        for c_idx, d, sub in reversed(dates_list):
                            v = sheet.cell(row=r, column=c_idx + 1).value
                            if v is not None:
                                try:
                                    v_f = float(v)
                                    if sub == 'Chuyên cần' and cc_val is None: cc_val = v_f
                                    elif sub == 'Bài tập' and bt_val is None: bt_val = v_f
                                    elif sub == 'Elearning' and el_val is None: el_val = v_f
                                except: pass
                        
                        # Find Rpoint col
                        rp_col_idx = None
                        for c_idx in range(sheet.max_column - 1, 2, -1):
                            vals = []
                            for row_idx in range(5, sheet.max_row + 1):
                                cell_val = sheet.cell(row=row_idx, column=c_idx + 1).value
                                if cell_val is not None:
                                    try: vals.append(float(cell_val))
                                    except: pass
                            if len(vals) >= 2:
                                avg_val = sum(vals)/len(vals)
                                if 30.0 <= avg_val <= 115.0:
                                    if not row3[c_idx]:
                                        rp_col_idx = c_idx
                                        break
                        
                        rp_val = None
                        if rp_col_idx is not None:
                            v = sheet.cell(row=r, column=rp_col_idx + 1).value
                            if v is not None:
                                try: rp_val = float(v)
                                except: pass
                                
                        print(f"  Sheet: {sheetname} | Found row in Excel: {val} | CC (Vắng): {cc_val}% | BT (Nợ): {bt_val}% | EL (Muộn): {el_val} | RP (Chốt): {rp_val}")
                        found = True
                        break
            if not found:
                print(f"  Sheet: {sheetname} | NOT FOUND in Excel!")

if __name__ == "__main__":
    main()
