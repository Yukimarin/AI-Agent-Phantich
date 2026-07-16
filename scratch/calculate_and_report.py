import sys
import openpyxl
from datetime import datetime, date
from collections import defaultdict
import numpy as np

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'C:/Users/DELL/Desktop/Education-DB-Analytic/docs/PTIT_Chiso.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

# Danh sach cac lop dich va thong tin mon hoc hien tai + truoc do
target_groups = {
    'KS24_HN': {
        'classes': ['HN-K24-CNTT1', 'HN-K24-CNTT2', 'HN-K24-CNTT3', 'HN-K24-CNTT4'],
        'sheet_curr': 'KS24_AI',
        'sheet_prev': 'KS24_AI',
        'start_curr': date(2026, 6, 22), 'end_curr': date(2026, 6, 28),
        'start_prev': date(2026, 6, 15), 'end_prev': date(2026, 6, 21),
    },
    'KS24_HCM': {
        'classes': ['HCM-K24-CNTT1'],
        'sheet_curr': 'KS24_AI', # trong sheet KS24_AI co du lieu tuan nay, con sheet JWS chi co den 15/6
        'sheet_prev': 'KS24_AI',
        'start_curr': date(2026, 6, 22), 'end_curr': date(2026, 6, 28),
        'start_prev': date(2026, 6, 15), 'end_prev': date(2026, 6, 21),
    },
    'KS25_CNTT_HN': {
        'classes': ['HN-K25-CNTT1', 'HN-K25-CNTT2', 'HN-K25-CNTT3', 'HN-K25-CNTT4', 'HN-K25-CNTT5', 'HN-K25-CNTT6'],
        'sheet_curr': 'KS25_Python_Web',
        'sheet_prev': 'KS25_Python',
        'start_curr': date(2026, 6, 22), 'end_curr': date(2026, 6, 28),
        'start_prev': date(2026, 6, 15), 'end_prev': date(2026, 6, 21),
    },
    'KS25_CNTT_HCM': {
        'classes': ['HCM-K25-CNTT5', 'HCM-K25-CNTT6', 'HCM-K25-CNTT7', 'HCM-K25-CNTT8'],
        'sheet_curr': 'KS25_Python_Web',
        'sheet_prev': 'KS25_Python',
        'start_curr': date(2026, 6, 22), 'end_curr': date(2026, 6, 28),
        'start_prev': date(2026, 6, 15), 'end_prev': date(2026, 6, 21),
    },
    'KS25_QTKD_HN': {
        'classes': ['HN-K25-QTKD1', 'HN-K25-QTKD2', 'HN-K25-QTKD3'],
        'sheet_curr': 'KS25_QTKD_DTB202',
        'sheet_prev': 'KS25_QTKD_DTB202',
        'start_curr': date(2026, 6, 22), 'end_curr': date(2026, 6, 28),
        'start_prev': date(2026, 6, 15), 'end_prev': date(2026, 6, 21),
    }
}

# Ham lay du lieu cua cac lop trong sheet theo thoi gian, phan ra GV va TG
def parse_sheet_data(sheetname, classes_target, start_date, end_date):
    if sheetname not in wb.sheetnames:
        return {}
    sheet = wb[sheetname]
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    dates_list = []
    current_date = None
    for c_idx in range(3, len(row3)):
        val3 = row3[c_idx]
        val4 = row4[c_idx]
        if val3:
            current_date = parse_date(val3)
        if current_date and start_date <= current_date <= end_date:
            dates_list.append((c_idx, current_date, val4))
            
    res = {}
    current_class = None
    current_gv_name = None
    
    for r in range(5, sheet.max_row + 1):
        cname = sheet.cell(row=r, column=2).value
        teacher_or_tg = sheet.cell(row=r, column=3).value
        teacher_or_tg_name = str(teacher_or_tg).strip() if teacher_or_tg else ""
        
        if cname:
            current_class = normalize_class_name(cname)
            role = 'GV'
            current_gv_name = teacher_or_tg_name
        else:
            role = 'TG'
            
        if not current_class:
            continue
            
        # Check xem co trong lop target khong
        matched_class = None
        for tc in classes_target:
            if tc == current_class:
                matched_class = tc
                break
        if not matched_class:
            continue
            
        if teacher_or_tg_name in ['', 'None', 'Chưa phân công', 'Giảng viên/Trợ giảng']:
            continue
            
        # Nhom diem
        day_vals = defaultdict(list)
        for c_idx, d, val4 in dates_list:
            val = sheet.cell(row=r, column=c_idx + 1).value
            if val is not None:
                try:
                    val_float = float(val)
                    day_vals[val4].append(val_float)
                except ValueError:
                    pass
                    
        averages = {}
        for metric in ['Chuyên cần', 'Bài tập', 'Elearning']:
            vals = day_vals.get(metric, [])
            averages[metric] = sum(vals) / len(vals) if vals else 0.0
            
        key = (matched_class, role)
        res[key] = {
            'name': teacher_or_tg_name,
            'gv_name': current_gv_name, # Giup TG biet ho di cung GV nao
            'metrics': averages,
            'raw_vals': {m: day_vals.get(m, []) for m in ['Chuyên cần', 'Bài tập', 'Elearning']}
        }
    return res

# Chay phan tich cho tung group
all_results = {}
for gname, ginfo in target_groups.items():
    classes = ginfo['classes']
    curr_data = parse_sheet_data(ginfo['sheet_curr'], classes, ginfo['start_curr'], ginfo['end_curr'])
    prev_data = parse_sheet_data(ginfo['sheet_prev'], classes, ginfo['start_prev'], ginfo['end_prev'])
    
    all_results[gname] = {
        'curr': curr_data,
        'prev': prev_data,
        'classes': classes
    }

# Xuat ra ket qua thong ke de viet bao cao
for gname, gdata in all_results.items():
    print(f"\n==================== NHÓM: {gname} ====================")
    classes = gdata['classes']
    curr = gdata['curr']
    prev = gdata['prev']
    
    # 1. Thong ke trung binh tuan nay cua GV
    gv_curr_cc, gv_curr_bt, gv_curr_el = [], [], []
    gv_prev_cc, gv_prev_bt, gv_prev_el = [], [], []
    
    print("\n  --- CHI TIẾT TỪNG LỚP (GIẢNG VIÊN) ---")
    for cls in classes:
        cls_curr_gv = curr.get((cls, 'GV'))
        cls_prev_gv = prev.get((cls, 'GV'))
        
        c_m = cls_curr_gv['metrics'] if cls_curr_gv else {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}
        p_m = cls_prev_gv['metrics'] if cls_prev_gv else {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}
        
        gv_name = cls_curr_gv['name'] if cls_curr_gv else (cls_prev_gv['name'] if cls_prev_gv else 'N/A')
        
        print(f"  Lớp: {cls} | GV: {gv_name}")
        print(f"    Tuần trước -> CC: {p_m['Chuyên cần']:.2f}%, BT: {p_m['Bài tập']:.2f}%, EL: {p_m['Elearning']:.2f}%")
        print(f"    Tuần này   -> CC: {c_m['Chuyên cần']:.2f}%, BT: {c_m['Bài tập']:.2f}%, EL: {c_m['Elearning']:.2f}%")
        
        if cls_curr_gv:
            gv_curr_cc.append(c_m['Chuyên cần'])
            gv_curr_bt.append(c_m['Bài tập'])
            gv_curr_el.append(c_m['Elearning'])
        if cls_prev_gv:
            gv_prev_cc.append(p_m['Chuyên cần'])
            gv_prev_bt.append(p_m['Bài tập'])
            gv_prev_el.append(p_m['Elearning'])
            
    print("\n  --- CHI TIẾT TỪNG LỚP (TRỢ GIẢNG) ---")
    tg_curr_cc, tg_curr_bt, tg_curr_el = [], [], []
    tg_prev_cc, tg_prev_bt, tg_prev_el = [], [], []
    for cls in classes:
        cls_curr_tg = curr.get((cls, 'TG'))
        cls_prev_tg = prev.get((cls, 'TG'))
        
        c_m = cls_curr_tg['metrics'] if cls_curr_tg else {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}
        p_m = cls_prev_tg['metrics'] if cls_prev_tg else {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}
        
        tg_name = cls_curr_tg['name'] if cls_curr_tg else (cls_prev_tg['name'] if cls_prev_tg else 'N/A')
        
        if tg_name != 'N/A':
            print(f"  Lớp: {cls} | TG: {tg_name}")
            print(f"    Tuần trước -> CC: {p_m['Chuyên cần']:.2f}%, BT: {p_m['Bài tập']:.2f}%, EL: {p_m['Elearning']:.2f}%")
            print(f"    Tuần này   -> CC: {c_m['Chuyên cần']:.2f}%, BT: {c_m['Bài tập']:.2f}%, EL: {c_m['Elearning']:.2f}%")
            
            if cls_curr_tg:
                tg_curr_cc.append(c_m['Chuyên cần'])
                tg_curr_bt.append(c_m['Bài tập'])
                tg_curr_el.append(c_m['Elearning'])
            if cls_prev_tg:
                tg_prev_cc.append(p_m['Chuyên cần'])
                tg_prev_bt.append(p_m['Bài tập'])
                tg_prev_el.append(p_m['Elearning'])
                
    # Trung binh nhom GV (Dai dien cho chi so vi pham cua lop)
    avg_curr_cc = sum(gv_curr_cc) / len(gv_curr_cc) if gv_curr_cc else 0.0
    avg_curr_bt = sum(gv_curr_bt) / len(gv_curr_bt) if gv_curr_bt else 0.0
    avg_curr_el = sum(gv_curr_el) / len(gv_curr_el) if gv_curr_el else 0.0
    
    avg_prev_cc = sum(gv_prev_cc) / len(gv_prev_cc) if gv_prev_cc else 0.0
    avg_prev_bt = sum(gv_prev_bt) / len(gv_prev_bt) if gv_prev_bt else 0.0
    avg_prev_el = sum(gv_prev_el) / len(gv_prev_el) if gv_prev_el else 0.0
    
    print("\n  --- CHỈ SỐ VI PHẠM TRUNG BÌNH CỦA NHÓM (Dựa trên lớp/GV) ---")
    print(f"    Tuần trước -> CC: {avg_prev_cc:.2f}%, BT: {avg_prev_bt:.2f}%, EL: {avg_prev_el:.2f}%")
    print(f"    Tuần này   -> CC: {avg_curr_cc:.2f}%, BT: {avg_curr_bt:.2f}%, EL: {avg_curr_el:.2f}%")
    print(f"    Thay đổi   -> CC: {avg_curr_cc - avg_prev_cc:+.2f}%, BT: {avg_curr_bt - avg_prev_bt:+.2f}%, EL: {avg_curr_el - avg_prev_el:+.2f}%")

# Thong ke nang luc GV/TG trong khoi CNTT & QTKD dua tren chi so tuan nay
# (Hoac so sanh su tien bo, hoac dua tren chi so CMI)
print("\n==================== THỐNG KÊ ĐÁNH GIÁ NĂNG LỰC GV/TG TUẦN NÀY ====================")
# Gom tat ca GV
teachers_eval = defaultdict(lambda: {'cc': [], 'bt': [], 'el': [], 'classes': []})
tgs_eval = defaultdict(lambda: {'cc': [], 'bt': [], 'el': [], 'classes': []})

for gname, gdata in all_results.items():
    curr = gdata['curr']
    for (cls, role), info in curr.items():
        name = info['name']
        m = info['metrics']
        if role == 'GV':
            teachers_eval[name]['cc'].append(m['Chuyên cần'])
            teachers_eval[name]['bt'].append(m['Bài tập'])
            teachers_eval[name]['el'].append(m['Elearning'])
            teachers_eval[name]['classes'].append(cls)
        else:
            tgs_eval[name]['cc'].append(m['Chuyên cần'])
            tgs_eval[name]['bt'].append(m['Bài tập'])
            tgs_eval[name]['el'].append(m['Elearning'])
            tgs_eval[name]['classes'].append(cls)

print("\n--- GIẢNG VIÊN ---")
for name, data in teachers_eval.items():
    cc = np.mean(data['cc'])
    bt = np.mean(data['bt'])
    el = np.mean(data['el'])
    print(f"  GV: {name} | Lớp dạy: {data['classes']} | CC vi phạm: {cc:.2f}%, BT nợ: {bt:.2f}%, EL trễ: {el:.2f}%")

print("\n--- TRỢ GIẢNG ---")
for name, data in tgs_eval.items():
    cc = np.mean(data['cc'])
    bt = np.mean(data['bt'])
    el = np.mean(data['el'])
    print(f"  TG: {name} | Lớp phụ trách: {data['classes']} | CC vi phạm: {cc:.2f}%, BT nợ: {bt:.2f}%, EL trễ: {el:.2f}%")
