import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

def main():
    wb = openpyxl.load_workbook("docs/PTIT_Chiso.xlsx", data_only=True)
    
    for sheetname in wb.sheetnames:
        if sheetname == 'Sheet1':
            continue
        sheet = wb[sheetname]
        max_r = sheet.max_row
        max_c = sheet.max_column
        
        # Get row 3 and 4 headers
        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
        
        print(f"\nSheet: {sheetname} (Columns: {max_c})")
        # Print the last 6 columns headers
        for col_idx in range(max(0, max_c - 6), max_c):
            h3 = row3[col_idx]
            h4 = row4[col_idx]
            print(f"  Col {col_idx}: h3={h3}, h4={h4}")
            
        # Print first class's values for the last 6 columns
        for r in range(5, max_r + 1):
            cname = sheet.cell(row=r, column=2).value
            if cname:
                vals = []
                for col_idx in range(max(0, max_c - 6), max_c):
                    val = sheet.cell(row=r, column=col_idx + 1).value
                    vals.append(f"{col_idx}:{val}")
                print(f"    Class {cname}: {vals}")
                break

if __name__ == "__main__":
    main()
