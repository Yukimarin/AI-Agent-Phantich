import sys
import os
import openpyxl

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

path = r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG (2).xlsx"

wb = openpyxl.load_workbook(path, data_only=True)
ws_demo = wb['Lương Quốc Tuấn - DEMO']
ws_weights = wb['Trọng số đánh giá']

# Read weights from demo sheet
demo_weights = []
for r_idx in range(6, 45):
    name = ws_demo.cell(row=r_idx, column=1).value
    weight = ws_demo.cell(row=r_idx, column=8).value
    score = ws_demo.cell(row=r_idx, column=7).value
    if name and isinstance(weight, (int, float)) and not str(name).startswith('NHÓM'):
        demo_weights.append((r_idx, str(name).strip().replace('\n', ' '), weight, score))

print("=== WEIGHTS AND SCORES IN DEMO SHEET ===")
total_weight = 0
for idx, (r, name, w, s) in enumerate(demo_weights):
    print(f"Row {r}: {name[:55]} | Weight: {w} | Score: {s}")
    total_weight += w
print(f"Total Weight in Demo: {total_weight}\n")

# Read weights for Rank 3-4 (Giảng viên) in 'Trọng số đánh giá' sheet
print("=== WEIGHTS FOR RANK 3-4 (GIẢNG VIÊN) IN OFFICIAL WEIGHTS ===")
total_weights_official = 0
for r_idx in range(5, 31):
    name = ws_weights.cell(row=r_idx, column=2).value
    w_r34 = ws_weights.cell(row=r_idx, column=4).value # Column D is Rank 3-4
    if name and isinstance(w_r34, (int, float)) and not str(name).startswith('TỔNG'):
        print(f"Row {r_idx}: {str(name).strip().replace('\n', ' ')[:55]} | Weight: {w_r34}")
        total_weights_official += w_r34
print(f"Total Official Weight for Rank 3-4: {total_weights_official}")

wb.close()
