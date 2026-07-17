import openpyxl
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\DELL\Downloads\[RE] Đào tạo - Tiêu chuẩn xếp loại năng lực GV_TG (1).xlsx"

if not os.path.exists(file_path):
    print(f"Error: File not found at {file_path}")
    sys.exit(1)

wb = openpyxl.load_workbook(file_path, data_only=True)
print("=== SHEETS ===")
print(wb.sheetnames)

for sheetname in wb.sheetnames:
    print(f"\n=========================================")
    print(f"=== SHEET: {sheetname} ===")
    print(f"=========================================")
    sheet = wb[sheetname]
    for r in range(1, sheet.max_row + 1):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
        # Filter empty rows but keep rows with at least one cell populated
        if any(v is not None for v in row_vals):
            print(f"Row {r:02d}: {row_vals}")
wb.close()
