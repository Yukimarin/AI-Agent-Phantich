import os
import sys
import openpyxl
import shutil
import json
import re
from datetime import datetime, date
from collections import defaultdict
from openpyxl.utils import get_column_letter

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date().isoformat()
    if isinstance(d_val, date):
        return d_val.isoformat()
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0])).isoformat()
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0])).isoformat()
        except ValueError:
            return None
    return None

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

def extract_class_size(class_name):
    match = re.search(r'\((\d+)\)', str(class_name))
    if match:
        return int(match.group(1))
    return 30

def generate_classes_metrics_cache(excel_path, output_json_path="data/processed/classes_metrics_cache.json"):
    """
    Đọc file Excel 1 lần duy nhất và tạo cache JSON chuẩn hóa cho toàn bộ các Agent.
    Giúp tăng tốc độ xử lý của pipeline lên gấp 50-100 lần.
    """
    print(f"DataSanitizer: Tạo Cache JSON chỉ số đào tạo từ {excel_path}...")
    if not os.path.exists(excel_path):
        print(f"Warning: Không tìm thấy {excel_path}")
        return False
        
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        cache_data = {
            "generated_at": datetime.now().isoformat(),
            "sheets": {},
            "classes": {}
        }
        
        active_sheets = [s for s in wb.sheetnames if s.lower() != 'sheet1' and any(k in s for k in ['KS24', 'KS25', 'SKL'])]
        
        for sheetname in active_sheets:
            sheet = wb[sheetname]
            row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
            
            # Map cột ngày học
            dates_list = []
            current_date = None
            for c_idx in range(3, len(row3)):
                col_letter = get_column_letter(c_idx + 1)
                dim = sheet.column_dimensions.get(col_letter)
                if dim and dim.hidden:
                    continue
                val3 = row3[c_idx]
                if val3:
                    current_date = parse_date(val3)
                if current_date:
                    dates_list.append((c_idx, current_date))
                    
            # Đọc từng dòng lớp học
            for r in range(5, sheet.max_row + 1):
                cname = sheet.cell(row=r, column=2).value
                gv_name = sheet.cell(row=r, column=3).value
                
                if not cname:
                    continue
                    
                cname_raw = str(cname).strip()
                cname_norm = normalize_class_name(cname_raw)
                c_size = extract_class_size(cname_raw)
                gv_str = str(gv_name).strip() if gv_name else ""
                
                if cname_norm not in cache_data["classes"]:
                    cache_data["classes"][cname_norm] = {
                        "raw_name": cname_raw,
                        "size": c_size,
                        "sheets": {}
                    }
                    
                metrics_by_date = {}
                idx = 0
                while idx < len(dates_list):
                    c_idx, d_str = dates_list[idx]
                    # CC, BT, EL là 3 cột liên tiếp
                    cc_val = sheet.cell(row=r, column=c_idx + 1).value
                    bt_val = sheet.cell(row=r, column=c_idx + 2).value if (c_idx + 1 < len(row3)) else None
                    el_val = sheet.cell(row=r, column=c_idx + 3).value if (c_idx + 2 < len(row3)) else None
                    
                    if cc_val is not None or bt_val is not None or el_val is not None:
                        try:
                            cc = float(cc_val) if cc_val is not None else 0.0
                            bt = float(bt_val) if bt_val is not None else 0.0
                            el = float(el_val) if el_val is not None else 0.0
                            metrics_by_date[d_str] = {
                                "cc": cc,
                                "bt": bt,
                                "el": el
                            }
                        except (ValueError, TypeError):
                            pass
                    idx += 3
                    
                cache_data["classes"][cname_norm]["sheets"][sheetname] = {
                    "instructor": gv_str,
                    "metrics": metrics_by_date
                }
                
        wb.close()
        
        os.makedirs(os.path.dirname(output_json_path), exist_ok=True)
        with open(output_json_path, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
            
        print(f"✓ DataSanitizer: Xuất Cache thành công tại {output_json_path} ({len(cache_data['classes'])} lớp)")
        return True
    except Exception as e:
        print(f"Lỗi tạo Cache: {e}")
        return False

def sanitize_excel(file_path):
    print(f"DataSanitizer: Bắt đầu làm sạch dữ liệu tại {file_path}")
    
    if not os.path.exists(file_path):
        print(f"Lỗi: Không tìm thấy file {file_path}")
        return False
        
    backup_path = file_path.replace(".xlsx", "_raw.xlsx")
    if not os.path.exists(backup_path):
        shutil.copy2(file_path, backup_path)
        print(f"DataSanitizer: Đã backup file gốc sang {backup_path}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    if isinstance(cell.value, str):
                        cell.value = cell.value.strip()
                        
        wb.save(file_path)
        wb.close()
        print(f"DataSanitizer: Làm sạch thành công file {file_path}")
        return True
    except Exception as e:
        print(f"DataSanitizer Lỗi: {str(e)}")
        return False

def sync_from_backup():
    backup_path = r"C:\Users\DELL\Desktop\Backup\PTIT\PTIT_Chiso.xlsx"
    target_path = "data/inputs/PTIT_Chiso.xlsx"
    sync_meta_path = "data/processed/last_sync.json"
    
    if os.path.exists(backup_path):
        print(f"DataSanitizer: Tìm thấy file backup tại {backup_path}")
        backup_mtime = os.path.getmtime(backup_path)
        
        should_copy = False
        if not os.path.exists(target_path):
            should_copy = True
        else:
            last_sync_time = 0.0
            if os.path.exists(sync_meta_path):
                try:
                    with open(sync_meta_path, "r", encoding="utf-8") as f:
                        meta = json.load(f)
                        last_sync_time = meta.get("backup_mtime", 0.0)
                except Exception:
                    pass
            
            if abs(backup_mtime - last_sync_time) > 0.01:
                should_copy = True
                
        if should_copy:
            try:
                os.makedirs(os.path.dirname(target_path), exist_ok=True)
                shutil.copy2(backup_path, target_path)
                print(f"✓ DataSanitizer: Đã tự động đồng bộ file Excel mới từ Backup vào dự án.")
                os.makedirs(os.path.dirname(sync_meta_path), exist_ok=True)
                with open(sync_meta_path, "w", encoding="utf-8") as f:
                    json.dump({"backup_mtime": backup_mtime}, f)
            except Exception as e:
                print(f"Warning: Không thể copy file từ Backup: {e}")
        else:
            print("DataSanitizer: File Excel trong dự án đã đồng bộ và là mới nhất.")
    else:
        print(f"DataSanitizer: Không tìm thấy file backup tại {backup_path}. Sử dụng file hiện tại trong dự án.")

def main():
    print("=========================================")
    print("KHỞI CHẠY DATA SANITIZER (HARNESS LAYER)")
    print("=========================================")
    
    # 0. Làm sạch trực tiếp file nguồn ngoài Desktop Backup trước (nếu có)
    backup_path = r"C:\Users\DELL\Desktop\Backup\PTIT\PTIT_Chiso.xlsx"
    if os.path.exists(backup_path):
        print(f"DataSanitizer: Tiến hành làm sạch file nguồn ngoài Desktop: {backup_path}")
        sanitize_excel(backup_path)
        
    # 1. Đồng bộ hóa dữ liệu từ thư mục Backup ngoài dự án
    sync_from_backup()
    
    target_file = "data/inputs/PTIT_Chiso.xlsx"
    success = sanitize_excel(target_file)
    
    # 2. Tạo Single Source of Truth Cache JSON
    if success:
        generate_classes_metrics_cache(target_file)
        print("DataSanitizer: Đã sẵn sàng môi trường dữ liệu sạch và Cache JSON!")
        sys.exit(0)
    else:
        print("DataSanitizer: Gặp sự cố làm sạch dữ liệu.")
        sys.exit(1)

if __name__ == "__main__":
    main()
