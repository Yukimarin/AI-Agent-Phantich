import openpyxl
import mysql.connector
import sys
import os
import re
import json
import socket
import time
import subprocess
from datetime import datetime, timedelta

sys.stdout.reconfigure(encoding='utf-8')

def remove_accents(input_str):
    if not input_str:
        return ""
    input_str = str(input_str)
    s1 = u'ÀÁÂÃÈÉÊÌÍÒÓÔÕÙÚÝàáâãèéêìíòóôõùúýĂăĐđĨĩŨũƠơƯưẠạẢảẤấẦầẨẩẪẫẬậẮắẰằẲẳẴẵẬặẸẹẺẻẼẽẾếỀềỂểỄễỆệỊịỎỏỐốỒồỔổỖỗỘộỚớỜờỞởỠỡỢợỤụỦủỨứỪừỬửỮữỰựỲỳỶỷỸỹỸỳ'
    s2 = u'AAAAEEEIIOOOOUUYaaaaeeeiioooouuyAaDdIiUuOoUuAaAaAaAaAaAaAaAaAaAaAaAaEeEeEeEeEeEeEeEeIiOoOoOoOoOoOoOoOoOoOoOoOoUuUuUuUuUuUuUuYyYyYyYy'
    s = ""
    for c in input_str:
        if c in s1:
            s += s2[s1.index(c)]
        else:
            s += c
    s = re.sub(r'\s+', ' ', s)
    return s.lower().strip()

def normalize_class_name(name):
    name = str(name).strip()
    name = re.sub(r'KS(\d+)', r'K\1', name)
    return name

def parse_session_number(session_str):
    match = re.search(r'Buổi\s*(\d+)', session_str, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return None

def generate_heuristics_email(name):
    name_clean = remove_accents(name)
    words = name_clean.split(" ")
    if len(words) >= 2:
        ten = words[-1]
        ho_lot = "".join([w[0] for w in words[:-1]])
        return f"{ten}{ho_lot}@rikkei.edu.vn"
    return f"{name_clean.replace(' ', '')}@rikkei.edu.vn"

def load_staff_emails(json_path="data/processed/daily_log_analysis.json"):
    emails = {}
    if os.path.exists(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            def recurse_extract(obj):
                if isinstance(obj, dict):
                    name = obj.get("name") or obj.get("user")
                    email = obj.get("email")
                    if name and email and ("@rikkei" in email or "@rikkeieducation" in email or "@rikkeiacademy" in email or "@rikkeiedu" in email):
                        # Chuẩn hóa tên không dấu làm key
                        norm_k = remove_accents(str(name))
                        emails[norm_k] = str(email).strip()
                    for k, v in obj.items():
                        recurse_extract(v)
                elif isinstance(obj, list):
                    for item in obj:
                        recurse_extract(item)
            recurse_extract(data)
        except Exception as e:
            print("Error loading staff emails from Worklane:", e)
            
    # Dự phòng leaders & nhân viên chính thức
    default_emails = {
        'Trần Minh Cường': 'cuongtm@rikkei.edu.vn',
        'Nguyễn Thị Tươi': 'tuoint@rikkei.edu.vn',
        'Giáp Thị Minh Hằng': 'hanggtm@rikkeieducation.com',
        'Lò Thị Ngọc Anh': 'anhltn1@rikkeiacademy.net',
        'Nguyễn Bá Minh Đạo': 'daonbm@rikkeiacademy.net',
        'Lương Quốc Tuấn': 'tuanlq@rikkeiacademy.com',
        'Trịnh Quốc Hai': 'haitq@rikkeiacademy.com',
        'Nguyễn Quảng An': 'annq@rikkeiacademy.com',
        'Lại Trung Lâm': 'lamlt@rikkei.edu.vn',
        'Phạm Ngọc Kiên': 'kienpn@rikkeiedu.io.vn',
        'Hồ Xuân Hùng': 'hunghx@rikkei.edu.vn'
    }
    for k, v in default_emails.items():
        norm_k = remove_accents(k)
        if norm_k not in emails:
            emails[norm_k] = v
            
    return emails

def load_all_staffs(md_path="data/inputs/staff_roles_ranks.md", emails_db=None):
    staffs = {}
    if not emails_db:
        emails_db = {}
        
    current_group = "Khối CNTT"
    if os.path.exists(md_path):
        try:
            with open(md_path, "r", encoding="utf-8") as f:
                for line in f:
                    line_strip = line.strip()
                    if line_strip.startswith("## "):
                        group_title = line_strip.replace("## ", "").strip()
                        if "CNTT" in group_title:
                            current_group = "Khối CNTT"
                        elif "Kinh doanh" in group_title or "QTKD" in group_title:
                            current_group = "Khối QTKD"
                        elif "Ngoại ngữ" in group_title:
                            current_group = "Khối Ngoại ngữ"
                        elif "chất lượng" in group_title or "QLCLĐT" in group_title:
                            current_group = "Khối QLCDT"
                        continue
                        
                    if "|" in line:
                        parts = line.split("|")
                        name = parts[0].replace("-", "").strip()
                        role_part = parts[1].strip()
                        rank_part = parts[-1].strip()
                        
                        role = "GV"
                        if any(k in role_part.lower() for k in ["trợ giảng", "tg"]):
                            role = "TG"
                        elif any(k in role_part.lower() for k in ["giảng viên", "gv", "leader", "co-founder", "giáo vụ"]):
                            role = "GV"
                            
                        rank = "N/A"
                        match = re.search(r'rank:\s*(\d+|N/A)|rank\s*(\d+|N/A)', rank_part, re.IGNORECASE)
                        if match:
                            rank = match.group(1) or match.group(2)
                            
                        # Phân chia địa lý cho khối CNTT theo 3 cơ sở chính xác
                        hcm_cntt_list = ["Nguyễn Bá Minh Đạo", "Lê Hà Thanh Sang", "Trần Quốc Tuấn", "Nguyễn Đức Minh", "Đặng Minh Luân", "Lưu Hoàng Xuân Nguyên", "Phan Ngọc Tài", "Nguyễn Ngọc Sơn", "Phạm Viết Hùng"]
                        hn_nt_cntt_list = ["Hồ Xuân Hùng", "Lâm Tùng Dương", "Lương Quốc Tuấn", "Ngọ Văn Quý", "Nguyễn Quảng An", "Lại Trung Lâm", "Phạm Ngọc Kiên"]
                        hn_hpc_cntt_list = ["Trịnh Quốc Hai", "Bùi Thanh Hải", "Nguyễn Xuân Bách", "Phạm Tuấn Bình", "Nguyễn Công Hưởng", "Đinh Thành Nam", "Mai Xuân Chinh"]
                        
                        group_label = current_group
                        if current_group == "Khối CNTT":
                            if name in hcm_cntt_list:
                                group_label = "Khối HCM-CNTT"
                            elif name in hn_nt_cntt_list:
                                group_label = "Khối HN-CNTT Ngọc Trục"
                            elif name in hn_hpc_cntt_list:
                                group_label = "Khối HN-CNTT HPC"
                        
                        # Phân chia khối Ngoại ngữ: tiếng Nhật vs tiếng Anh
                        japanese_list = ["Giáp Thị Minh Hằng", "Lê Thị Đỏ"]
                        if current_group == "Khối Ngoại ngữ":
                            if name in japanese_list:
                                group_label = "Khối Ngoại ngữ tiếng Nhật"
                            else:
                                group_label = "Khối Ngoại ngữ tiếng Anh"
                                
                        norm_name = remove_accents(name)
                        staffs[name] = {
                            "name": name,
                            "role": role,
                            "rank": rank,
                            "group": group_label,
                            "email": emails_db.get(norm_name) or generate_heuristics_email(name),
                            "norm_name": norm_name
                        }
        except Exception as e:
            print("Error parsing staff ranks:", e)
            
    # Dự phòng các giảng viên chính
    default_ranks = {
        'Lương Quốc Tuấn': {'role': 'GV', 'rank': '3', 'group': 'Khối HN-CNTT'}, 
        'Trịnh Quốc Hai': {'role': 'GV', 'rank': '4', 'group': 'Khối HN-CNTT'}, 
        'Nguyễn Quảng An': {'role': 'GV', 'rank': '4', 'group': 'Khối HN-CNTT'}, 
        'Lại Trung Lâm': {'role': 'TG', 'rank': '2', 'group': 'Khối HN-CNTT'}, 
        'Phạm Ngọc Kiên': {'role': 'TG', 'rank': '2', 'group': 'Khối HN-CNTT'}
    }
    for k, v in default_ranks.items():
        if k not in staffs:
            norm_k = remove_accents(k)
            staffs[k] = {
                "name": k,
                "role": v['role'],
                "rank": v['rank'],
                "group": v['group'],
                "email": emails_db.get(norm_k) or generate_heuristics_email(k),
                "norm_name": norm_k
            }
            
    # Sửa ngoại lệ email cho cô Lê Thị Đỏ
    if "Lê Thị Đỏ" in staffs:
        staffs["Lê Thị Đỏ"]["email"] = "dolt@rikkeieducation.top"
        
    return staffs

def match_instructor(excel_name, all_staffs):
    if not excel_name:
        return None
    excel_name_clean = str(excel_name).strip()
    excel_norm = remove_accents(excel_name_clean)
    
    # 1. Khớp chính xác tên không dấu
    for staff_name, info in all_staffs.items():
        if info['norm_name'] == excel_norm or excel_norm in info['norm_name'] or info['norm_name'] in excel_norm:
            return info
            
    # 2. Khớp theo họ + tên
    for staff_name, info in all_staffs.items():
        staff_words = info['norm_name'].split(" ")
        excel_words = excel_norm.split(" ")
        if len(excel_words) >= 2 and len(staff_words) >= 2:
            if excel_words[0] == staff_words[0] and excel_words[-1] == staff_words[-1]:
                return info
                
    # 3. Tạo bản ghi ảo nếu không khớp
    return {
        "name": excel_name_clean,
        "role": "GV",
        "rank": "N/A",
        "group": "Khối HN-CNTT",
        "email": generate_heuristics_email(excel_name_clean)
    }

def find_course_by_name(tkb_course_name, courses_db):
    tkb_clean = remove_accents(tkb_course_name)
    tkb_clean = re.sub(r'\[.*?\]', '', tkb_clean).strip()
    tkb_clean = tkb_clean.replace("mon hoc", "").replace("tieng", "").strip()
    
    if not tkb_clean:
        return None
        
    for co in courses_db:
        co_name_clean = remove_accents(co['name'])
        co_name_clean = re.sub(r'\[.*?\]', '', co_name_clean).strip()
        if tkb_clean in co_name_clean or co_name_clean in tkb_clean:
            return co['id']
            
    tkb_words = [w for w in tkb_clean.split(" ") if len(w) > 2]
    if not tkb_words:
        return None
        
    for co in courses_db:
        co_name_clean = remove_accents(co['name'])
        if all(w in co_name_clean for w in tkb_words):
            return co['id']
            
    return None

def check_mysql_port(port=3307):
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.settimeout(1.0)
    try:
        s.connect(('127.0.0.1', port))
        s.close()
        return True
    except socket.error:
        s.close()
        return False

def ensure_mysql_started():
    if check_mysql_port(3307):
        return True

    mysql_bin = r"C:\Program Files\MySQL\MySQL Server 9.7\bin\mysqld.exe"
    data_dir = os.path.abspath("data/mysql_data_97")
    
    cmd = [
        mysql_bin,
        "--no-defaults",
        f"--datadir={data_dir}",
        "--port=3307",
        "--shared-memory"
    ]
    try:
        subprocess.Popen(cmd, creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0)
        for i in range(15):
            time.sleep(1.0)
            if check_mysql_port(3307):
                return True
        return False
    except:
        return False

def main():
    tkb_path = "data/inputs/1. Thời khóa biểu tổng .xlsx"
    if not os.path.exists(tkb_path):
        print(f"Error: File {tkb_path} not found.")
        sys.exit(1)
        
    if not ensure_mysql_started():
        print("Cannot start MySQL")
        sys.exit(1)
        
    # Tải emails thực tế từ Worklane
    emails_db = load_staff_emails()
    # Tải toàn bộ nhân sự và gán group chi tiết
    all_staffs = load_all_staffs(emails_db=emails_db)
    print(f"Loaded {len(all_staffs)} staff members from configuration.")
        
    try:
        conn = mysql.connector.connect(
            host="127.0.0.1",
            user="root",
            password="",
            database="qldt_el",
            port=3307
        )
        cursor = conn.cursor(dictionary=True)
        print("Connected to MySQL database on port 3307.")
    except Exception as e:
        print(f"Error connecting to MySQL: {e}")
        sys.exit(1)

    # Tải toàn bộ lớp học trong DB
    cursor.execute("SELECT id, name FROM classes")
    classes_db = cursor.fetchall()
    class_map = {}
    for c in classes_db:
        db_name = c['name']
        norm_name = normalize_class_name(db_name)
        class_map[norm_name] = c['id']
        class_map[db_name] = c['id']
        
    # Tải toàn bộ môn học trong DB
    cursor.execute("SELECT id, name FROM courses")
    courses_db = cursor.fetchall()
    
    # 3. Đọc dữ liệu Thời khóa biểu Excel
    print("Đang đọc dữ liệu thời khóa biểu Excel...")
    try:
        wb = openpyxl.load_workbook(tkb_path, data_only=True)
    except Exception as e:
        print("Error opening Excel workbook:", str(e))
        conn.close()
        sys.exit(1)
        
    rows = []
    target_tkb_sheets = ['1.1. TKB Hà Nội tổng', '1.2. TKB Hồ Chí Minh tổng']
    for sheetname in target_tkb_sheets:
        if sheetname not in wb.sheetnames:
            continue
        sheet = wb[sheetname]
        headers = [str(sheet.cell(row=1, column=c).value).strip() for c in range(1, sheet.max_column + 1)]
        
        class_col = 'Lớp đào tạo'
        subject_col = 'Môn học'
        gv_lt_col = 'Giảng viên LT'
        gv_th_col = 'Giảng viên TH'
        date_col = 'Ngày đào tạo'
        session_col = 'Tiến độ đào tạo'
        ca_col = 'Ca đào tạo'
        
        try:
            class_idx = headers.index(class_col) + 1
            subject_idx = headers.index(subject_col) + 1
            gv_lt_idx = headers.index(gv_lt_col) + 1
            gv_th_idx = headers.index(gv_th_col) + 1
            date_idx = headers.index(date_col) + 1
            session_idx = headers.index(session_col) + 1
            ca_idx = headers.index(ca_col) + 1
        except ValueError as e:
            continue
            
        for r in range(2, sheet.max_row + 1):
            class_val = sheet.cell(row=r, column=class_idx).value
            subject_val = sheet.cell(row=r, column=subject_idx).value
            gv_lt_val = sheet.cell(row=r, column=gv_lt_idx).value
            gv_th_val = sheet.cell(row=r, column=gv_th_idx).value
            date_val = sheet.cell(row=r, column=date_idx).value
            session_val = sheet.cell(row=r, column=session_idx).value
            ca_val = sheet.cell(row=r, column=ca_idx).value
            
            if class_val is None and subject_val is None:
                continue
                
            rows.append({
                class_col: class_val,
                subject_col: subject_val,
                gv_lt_col: gv_lt_val,
                gv_th_col: gv_th_val,
                date_col: date_val,
                session_col: session_val,
                ca_col: ca_val
            })
    wb.close()
    
    seen = set()
    unique_rows = []
    today_str = datetime.now().strftime('%Y-%m-%d')
    now = datetime.now()
    
    for row in rows:
        raw_date = row[date_col]
        if isinstance(raw_date, datetime):
            date_str = raw_date.strftime('%Y-%m-%d')
        elif raw_date:
            try:
                if isinstance(raw_date, str):
                    date_str = raw_date.split(" ")[0]
                else:
                    date_str = str(raw_date)
            except:
                continue
        else:
            continue
            
        tkb_class = str(row[class_col] or "").strip()
        subject_val = str(row[subject_col] or "").strip()
        ca_hoc = str(row[ca_col] or "").strip()
        session_str = str(row[session_col] or "").strip()
        
        norm_class = normalize_class_name(tkb_class)
        if (norm_class not in class_map) and (tkb_class not in class_map):
            continue
            
        key = (date_str, tkb_class, subject_val, ca_hoc, session_str)
        if key not in seen:
            seen.add(key)
            row['parsed_date_str'] = date_str
            unique_rows.append(row)
            
    print(f"Tìm thấy {len(unique_rows)} ca học trong TKB khớp với lớp trong DB.")
    
    violations = []
    
    def get_ca_start_time(ca_str):
        match = re.search(r'(\d{{1,2}}:\d{{2}})', ca_str)
        if match:
            return match.group(1) + ":00"
        return "07:00:00"

    course_sessions_cache = {}

    # 4. DUYỆT CÁC LỚP QUÉT LỖI QLĐT
    for row in unique_rows:
        date_str = row['parsed_date_str']
        tkb_class = str(row[class_col] or "").strip()
        norm_class = normalize_class_name(tkb_class)
        cid = class_map.get(norm_class) or class_map.get(tkb_class)
        
        gv_lt_tkb = str(row[gv_lt_col] or "").strip()
        gv_th_tkb = str(row[gv_th_col] or "").strip()
        session_str = str(row[session_col] or "").strip()
        ca_hoc = str(row[ca_col] or "").strip()
        tkb_course = str(row[subject_col] or "").strip()
        
        is_th = bool(gv_th_tkb and gv_th_tkb.lower() != 'nan')
        excel_teacher = gv_th_tkb if is_th else gv_lt_tkb
        
        if not excel_teacher or excel_teacher.lower() in ['nan', 'none', '']:
            continue
            
        staff_info = match_instructor(excel_teacher, all_staffs)
        if not staff_info:
            continue
            
        responsible_person = staff_info['name']
        role = "TG" if is_th else "GV"
        rank = staff_info['rank']
        email = staff_info['email']
        group = staff_info['group']
        
        course_id = find_course_by_name(tkb_course, courses_db)
        if not course_id:
            continue
            
        if course_id not in course_sessions_cache:
            cursor.execute("SELECT id, position, name FROM sessions WHERE course_id = %s", (course_id,))
            sessions_db = cursor.fetchall()
            course_sessions_cache[course_id] = {s['position']: {'id': s['id'], 'name': s['name']} for s in sessions_db}
            
        sess_map = course_sessions_cache[course_id]
        pos = parse_session_number(session_str)
        if not pos or pos not in sess_map:
            continue
            
        sess_id = sess_map[pos]['id']
        sess_name = sess_map[pos]['name']
        
        start_time_str = get_ca_start_time(ca_hoc)
        attendance_time = datetime.strptime(f"{date_str} {start_time_str}", "%Y-%m-%d %H:%M:%S")
        deadline = attendance_time + timedelta(hours=24)
        
        if deadline > now:
            continue
            
        # ─────────────────────────────────────────────────────────────
        # TIÊU CHÍ 2.2a: Tài nguyên học tập (document)
        # ─────────────────────────────────────────────────────────────
        cursor.execute("SELECT created_at FROM documents WHERE class_id = %s AND session_id = %s", (cid, sess_id))
        docs = cursor.fetchall()
        
        if not docs:
            violations.append({
                'Date': date_str, 'Class': norm_class, 'Session': session_str, 'Ca': ca_hoc,
                'Instructor': responsible_person, 'Role': role, 'Rank': rank, 'Email': email, 'Group': group,
                'Category': 'QLDT', 'Error': 'QLDT-DOC-MISSING',
                'Course': tkb_course, 'SessionName': sess_name,
                'Details': f"Chưa upload tài nguyên học tập (Lark link + Source code) quá 24h sau ca học",
                'TKB_Teachers': f"LT: {gv_lt_tkb}, TH: {gv_th_tkb}"
            })
        else:
            doc_created_at = docs[0]['created_at']
            if doc_created_at > deadline:
                violations.append({
                    'Date': date_str, 'Class': norm_class, 'Session': session_str, 'Ca': ca_hoc,
                    'Instructor': responsible_person, 'Role': role, 'Rank': rank, 'Email': email, 'Group': group,
                    'Category': 'QLDT', 'Error': 'QLDT-DOC-LATE',
                    'Course': tkb_course, 'SessionName': sess_name,
                    'Details': f"Upload tài nguyên học tập chậm trễ lúc {doc_created_at.strftime('%Y-%m-%d %H:%M:%S')} (Hạn: {deadline.strftime('%Y-%m-%d %H:%M:%S')})",
                    'TKB_Teachers': f"LT: {gv_lt_tkb}, TH: {gv_th_tkb}"
                })
                
        # ─────────────────────────────────────────────────────────────
        # TIÊU CHÍ 2.2b: Cập nhật trạng thái BTVN (exercise)
        # ─────────────────────────────────────────────────────────────
        cursor.execute("SELECT id FROM homework WHERE session_id = %s", (sess_id,))
        hws = cursor.fetchall()
        
        if not hws:
            violations.append({
                'Date': date_str, 'Class': norm_class, 'Session': session_str, 'Ca': ca_hoc,
                'Instructor': responsible_person, 'Role': role, 'Rank': rank, 'Email': email, 'Group': group,
                'Category': 'QLDT', 'Error': 'QLDT-EX-MISSING',
                'Course': tkb_course, 'SessionName': sess_name,
                'Details': f"Chưa khởi tạo/giao BTVN cho học viên quá 24h sau ca học",
                'TKB_Teachers': f"LT: {gv_lt_tkb}, TH: {gv_th_tkb}"
            })
        else:
            hw_id = hws[0]['id']
            cursor.execute("SELECT updated_at FROM exercise WHERE class_id = %s AND homework_id = %s", (cid, hw_id))
            exercises = cursor.fetchall()
            
            if not exercises:
                violations.append({
                    'Date': date_str, 'Class': norm_class, 'Session': session_str, 'Ca': ca_hoc,
                    'Instructor': responsible_person, 'Role': role, 'Rank': rank, 'Email': email, 'Group': group,
                    'Category': 'QLDT', 'Error': 'QLDT-EX-MISSING',
                    'Course': tkb_course, 'SessionName': sess_name,
                    'Details': f"Chưa khởi tạo trạng thái BTVN (nợ bài/chấm bài) cho học viên quá 24h",
                    'TKB_Teachers': f"LT: {gv_lt_tkb}, TH: {gv_th_tkb}"
                })
            else:
                valid_updates = [ex['updated_at'] for ex in exercises if ex['updated_at']]
                if valid_updates:
                    max_updated_at = max(valid_updates)
                    if max_updated_at > deadline:
                        violations.append({
                            'Date': date_str, 'Class': norm_class, 'Session': session_str, 'Ca': ca_hoc,
                            'Instructor': responsible_person, 'Role': role, 'Rank': rank, 'Email': email, 'Group': group,
                            'Category': 'QLDT', 'Error': 'QLDT-EX-LATE',
                            'Course': tkb_course, 'SessionName': sess_name,
                            'Details': f"Cập nhật trạng thái BTVN chậm trễ lúc {max_updated_at.strftime('%Y-%m-%d %H:%M:%S')} (Hạn: {deadline.strftime('%Y-%m-%d %H:%M:%S')})",
                            'TKB_Teachers': f"LT: {gv_lt_tkb}, TH: {gv_th_tkb}"
                        })

    # ─────────────────────────────────────────────────────────────
    # PHẦN 2.1: GHI NHẬN VI PHẠM WORKLANE (DAILY LOGS & OVERDUE TASKS)
    # ─────────────────────────────────────────────────────────────
    
    # 1. Báo cáo ngày
    daily_log_path = "data/processed/daily_log_analysis.json"
    if os.path.exists(daily_log_path):
        print("Đang đọc dữ liệu báo cáo ngày từ daily_log_analysis.json...")
        with open(daily_log_path, "r", encoding="utf-8") as f:
            daily_data = json.load(f)
            
        weekly_stats = daily_data.get("weekly_stats", {})
        monthly_stats = daily_data.get("monthly_stats", {})
        
        for name, info in all_staffs.items():
            name_key = name.lower().strip()
            
            w_stat = weekly_stats.get(name_key) or {}
            w_miss_days = w_stat.get("missing_days", [])
            m_stat = monthly_stats.get(name_key) or {}
            m_miss_days = m_stat.get("missing_days", [])
            
            all_miss_days = sorted(list(set(w_miss_days + m_miss_days)))
            for m_day in all_miss_days:
                violations.append({
                    'Date': m_day, 'Class': 'Worklane', 'Session': 'Nhật ký công việc', 'Ca': 'N/A',
                    'Instructor': name, 'Role': info['role'], 'Rank': info['rank'], 'Email': info['email'], 'Group': info['group'],
                    'Category': 'WORKLANE', 'Error': 'WL-DAILY-LOG-MISSING',
                    'Course': 'Worklane', 'SessionName': 'Báo cáo ngày',
                    'Details': f"Không nộp báo cáo ngày làm việc {m_day} trên hệ thống Worklane PM",
                    'TKB_Teachers': 'Worklane PM'
                })
                
    # 2. Task trễ hạn
    project_issues_path = "data/processed/project_issues_worklane.json"
    if os.path.exists(project_issues_path):
        print("Đang đọc dữ liệu công việc từ project_issues_worklane.json...")
        with open(project_issues_path, "r", encoding="utf-8") as f:
            project_data = json.load(f)
            
        for proj_key, proj_data in project_data.items():
            p_info = proj_data.get('project_info', {})
            p_name = p_info.get('name', proj_key)
            status = p_info.get('status', 'ACTIVE')
            
            if status.upper() in ['CANCEL', 'CANCELLED', 'HỦY', 'HUY']:
                continue
                
            issues = proj_data.get('issues', {}).get('issues', [])
            for iss in issues:
                iss_state = str(iss.get('state', '')).lower().strip()
                if iss_state in ['hủy', 'huy', 'cancel', 'cancelled', 'hoàn thành', 'done', 'completed', 'chờ duyệt', 'cho duyet', 'pending review', 'review']:
                    continue
                    
                iss_due = iss.get('dueDate')
                if iss_due:
                    due_date_str = iss_due[:10]
                    if due_date_str <= today_str:
                        assignee = iss.get('assignee') or ""
                        
                        matched_staff = match_instructor(assignee, all_staffs)
                        if matched_staff and matched_staff['rank'] != 'N/A':
                            violations.append({
                                'Date': due_date_str, 'Class': p_name, 'Session': f"Task {iss['code']}", 'Ca': 'N/A',
                                'Instructor': matched_staff['name'], 'Role': matched_staff['role'], 'Rank': matched_staff['rank'], 'Email': matched_staff['email'], 'Group': matched_staff['group'],
                                'Category': 'WORKLANE', 'Error': 'WL-TASK-OVERDUE',
                                'Course': p_name,
                                'SessionName': f"{iss['code']} - {iss['title']}",
                                'TaskCode': iss['code'],
                                'TaskTitle': iss['title'],
                                'ProjectName': p_name,
                                'Details': f"Task trễ hạn bản thân thực hiện: '{iss['title']}' (Hạn chót: {due_date_str})",
                                'TKB_Teachers': 'Worklane PM'
                            })

    # ─────────────────────────────────────────────────────────────
    # PHẦN 2.3: GHI NHẬN THÊM CÁC VI PHẠM THỦ CÔNG THEO YÊU CẦU CỦA GIÁM ĐỐC
    # ─────────────────────────────────────────────────────────────
    # 1. Mai Xuân Chinh lỗi về QLĐT chậm trễ chấm BTVN 2 lần
    matched_chinh = match_instructor("Mai Xuân Chinh", all_staffs)
    if matched_chinh:
        violations.append({
            'Date': '2026-08-20', 'Class': 'HN-K24-CNTT3', 'Session': 'Buổi 15', 'Ca': 'Ca 1',
            'Instructor': matched_chinh['name'], 'Role': matched_chinh['role'], 'Rank': matched_chinh['rank'],
            'Email': matched_chinh['email'], 'Group': matched_chinh['group'],
            'Category': 'QLDT', 'Error': 'QLDT-EX-LATE',
            'Course': 'Java Web Service', 'SessionName': 'Project Java Web Service',
            'Details': 'Chậm trễ chấm BTVN cho sinh viên (Lần 1)',
            'TKB_Teachers': 'TH: Mai Xuân Chinh'
        })
        violations.append({
            'Date': '2026-08-21', 'Class': 'HN-K24-CNTT3', 'Session': 'Buổi 16', 'Ca': 'Ca 1',
            'Instructor': matched_chinh['name'], 'Role': matched_chinh['role'], 'Rank': matched_chinh['rank'],
            'Email': matched_chinh['email'], 'Group': matched_chinh['group'],
            'Category': 'QLDT', 'Error': 'QLDT-EX-LATE',
            'Course': 'Java Web Service', 'SessionName': 'Project Java Web Service',
            'Details': 'Chậm trễ chấm BTVN cho sinh viên (Lần 2)',
            'TKB_Teachers': 'TH: Mai Xuân Chinh'
        })

    # 2. Nguyễn Quảng An lỗi về QLĐT chậm trễ việc chốt Rpoint cho lớp CNTT4 và CNTT6
    matched_an = match_instructor("Nguyễn Quảng An", all_staffs)
    if matched_an:
        violations.append({
            'Date': '2026-08-21', 'Class': 'HN-K25-CNTT4', 'Session': 'Chốt Rpoint', 'Ca': 'N/A',
            'Instructor': matched_an['name'], 'Role': matched_an['role'], 'Rank': matched_an['rank'],
            'Email': matched_an['email'], 'Group': matched_an['group'],
            'Category': 'QLDT', 'Error': 'QLDT-RPOINT-LATE',
            'Course': 'Python Web', 'SessionName': 'Chốt điểm Rpoint môn học',
            'Details': 'Chậm trễ việc chốt điểm Rpoint môn học cho lớp HN-K25-CNTT4',
            'TKB_Teachers': 'LT: Nguyễn Quảng An'
        })
        violations.append({
            'Date': '2026-08-21', 'Class': 'HN-K25-CNTT6', 'Session': 'Chốt Rpoint', 'Ca': 'N/A',
            'Instructor': matched_an['name'], 'Role': matched_an['role'], 'Rank': matched_an['rank'],
            'Email': matched_an['email'], 'Group': matched_an['group'],
            'Category': 'QLDT', 'Error': 'QLDT-RPOINT-LATE',
            'Course': 'Python Web', 'SessionName': 'Chốt điểm Rpoint môn học',
            'Details': 'Chậm trễ việc chốt điểm Rpoint môn học cho lớp HN-K25-CNTT6',
            'TKB_Teachers': 'LT: Nguyễn Quảng An'
        })

    # Gán tuần/tháng
    for v in violations:
        try:
            dt = datetime.strptime(v['Date'], '%Y-%m-%d')
            monday = dt - timedelta(days=dt.weekday())
            week_label = f"Tuần {monday.strftime('%d/%m')} - {(monday + timedelta(days=6)).strftime('%d/%m/%Y')}"
            month_label = f"Tháng {dt.strftime('%m/%Y')}"
            v['week_monday'] = monday.strftime('%Y-%m-%d')
            v['week_label'] = week_label
            v['month_label'] = month_label
        except Exception as e:
            v['week_monday'] = 'N/A'
            v['week_label'] = 'Không xác định'
            v['month_label'] = 'Không xác định'

    # Xuất báo cáo vi phạm ra JSON
    output_json_path = "data/processed/agent3_output.json"
    output_report_path = "output/reports/core/agent_3_ops_discipline.md"
    # Đọc cờ kiểm toán Anti-Tampering từ Agent 1 (nếu có hành vi xóa sửa số liệu trái quy định)
    agent1_json_path = "data/processed/agent1_output.json"
    if os.path.exists(agent1_json_path):
        try:
            with open(agent1_json_path, "r", encoding="utf-8") as f_a1:
                a1_data = json.load(f_a1)
                classes_analysis = a1_data.get("classes_analysis", {})
                for c_k, c_info in classes_analysis.items():
                    if c_info.get("tampering_suspect", False):
                        for person in [c_info.get("instructor"), c_info.get("assistant")]:
                            if person and person in all_staffs:
                                s_meta = all_staffs[person]
                                violations.append({
                                    'Date': datetime.now().strftime('%Y-%m-%d'),
                                    'Course': c_info.get('sheet', 'Môn học'),
                                    'Class': c_info.get('class_name', c_k),
                                    'Session': 'Audit Kiểm toán',
                                    'SessionName': 'Kiểm toán tính toàn vẹn dữ liệu điểm',
                                    'Instructor': person,
                                    'Role': s_meta['role'],
                                    'Rank': s_meta['rank'],
                                    'Group': s_meta['group'],
                                    'Email': s_meta['email'],
                                    'Error': 'TAMPERING-DATA-VIOLATION',
                                    'Details': f"Tự ý xóa bớt số ca vi phạm của ngày học cũ ở lớp {c_k} sai quy định quản lý đào tạo.",
                                    'Category': 'QLDT',
                                    'week_label': 'Tuần hiện tại',
                                    'month_label': 'Tháng 08/2026'
                                })
        except Exception as e:
            print("Error checking Anti-tampering in Agent 3:", e)

    # Ghi JSON
    os.makedirs(os.path.dirname(output_json_path), exist_ok=True)
    with open(output_json_path, "w", encoding="utf-8") as f:
        json.dump(violations, f, ensure_ascii=False, indent=2)
    print(f"Đã lưu kết quả JSON tại: {output_json_path}")
    
    # Ghi Markdown
    os.makedirs(os.path.dirname(output_report_path), exist_ok=True)
    with open(output_report_path, "w", encoding="utf-8") as f:
        f.write("# Báo cáo Vi phạm Kỷ luật tác nghiệp GV/TG (Agent 3)\n\n")
        f.write(f"*Thời gian đối chiếu: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}*\n\n")
        
        f.write("> [!IMPORTANT]\n")
        f.write("> Báo cáo này quét và đối chiếu tự động toàn bộ vi phạm trên hệ thống Worklane PM và QLĐT.\n\n")
        
        f.write("## 📊 Tổng hợp số lỗi vi phạm theo từng nhân sự:\n\n")
        
        summary = {}
        for v in violations:
            name = v['Instructor']
            if name not in summary:
                summary[name] = {'worklane': 0, 'qldt': 0, 'rank': v['Rank'], 'role': v['Role'], 'email': v['Email'], 'group': v['Group']}
            cat = v.get('Category', 'QLDT').lower()
            summary[name][cat] += 1
            
        f.write("| Giảng viên/Trợ giảng | Vai trò | Rank | Khối | Email | Lỗi Worklane (2.1) | Lỗi QLĐT (2.2) | Tổng số vi phạm |\n")
        f.write("| :--- | :---: | :---: | :--- | :--- | :---: | :---: | :---: |\n")
        for name, data in sorted(summary.items(), key=lambda x: (x[1]['worklane']+x[1]['qldt']), reverse=True):
            total = data['worklane'] + data['qldt']
            f.write(f"| **{name}** | {data['role']} | {data['rank']} | {data['group']} | `{data['email']}` | {data['worklane']} | {data['qldt']} | **{total}** |\n")
            
        f.write("\n## 📋 PHẦN 2.1: CÁC VI PHẠM TRÊN WORKLANE\n\n")
        worklane_v = [v for v in violations if v.get('Category') == 'WORKLANE']
        if not worklane_v:
            f.write("✅ Không ghi nhận vi phạm Worklane nào.\n")
        else:
            f.write("| Ngày ghi nhận | Dự án / Nguồn | Đầu việc / Task | Giảng viên/Trợ giảng | Email | Vai trò | Rank | Khối | Mã lỗi | Chi tiết vi phạm |\n")
            f.write("| :--- | :--- | :--- | :--- | :--- | :---: | :---: | :--- | :---: | :--- |\n")
            for v in sorted(worklane_v, key=lambda x: (x['Instructor'], x['Date'])):
                f.write(f"| {v['Date']} | {v['Class']} | {v['SessionName']} | **{v['Instructor']}** | `{v['Email']}` | {v['Role']} | {v['Rank']} | {v['Group']} | `{v['Error']}` | {v['Details']} |\n")

        f.write("\n## 📋 PHẦN 2.2: CÁC VI PHẠM TRÊN QLĐT (CHỦ CHỐT TÀI NGUYÊN & BTVN)\n\n")
        qldt_v = [v for v in violations if v.get('Category') == 'QLDT']
        if not qldt_v:
            f.write("✅ Không ghi nhận vi phạm QLĐT nào.\n")
        else:
            f.write("| Ngày dạy | Môn học | Lớp học | Buổi học | Tên bài học | Giảng viên/Trợ giảng | Email | Vai trò | Rank | Khối | Mã lỗi | Chi tiết vi phạm |\n")
            f.write("| :--- | :--- | :--- | :--- | :--- | :--- | :--- | :---: | :---: | :--- | :---: | :--- |\n")
            for v in sorted(qldt_v, key=lambda x: (x['Instructor'], x['Date'])):
                f.write(f"| {v['Date']} | {v['Course']} | {v['Class']} | {v['Session']} | {v['SessionName']} | **{v['Instructor']}** | `{v['Email']}` | {v['Role']} | {v['Rank']} | {v['Group']} | `{v['Error']}` | {v['Details']} |\n")
                
    print(f"Đã xuất báo cáo Markdown tại: {output_report_path}")
    conn.close()

if __name__ == "__main__":
    main()
