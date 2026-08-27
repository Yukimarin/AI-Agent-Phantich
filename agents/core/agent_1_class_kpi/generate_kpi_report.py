# -*- coding: utf-8 -*-
import sys
import openpyxl
from datetime import datetime, date, timedelta
from collections import defaultdict
import numpy as np
import re
from openpyxl.utils import get_column_letter

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

class_sizes = {}
def extract_class_size(class_name):
    match = re.search(r'\((\d+)\)', str(class_name))
    if match:
        return int(match.group(1))
    return 30

excel_path = 'data/inputs/PTIT_Chiso.xlsx'
output_path = 'output/reports/core/agent_1_student_discipline.md'

wb = openpyxl.load_workbook(excel_path, data_only=True)

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

def build_class_timelines(workbook):
    timelines = defaultdict(lambda: defaultdict(list))
    active_sheets = []
    for s_name in workbook.sheetnames:
        if s_name.lower() == 'sheet1':
            continue
        if any(k in s_name for k in ['KS24', 'KS25', 'SKL']):
            active_sheets.append(s_name)
            
    for sheetname in active_sheets:
        sheet = workbook[sheetname]
        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        
        dates_list = []
        current_date = None
        for c_idx in range(3, len(row3)):
            col_letter = get_column_letter(c_idx + 1)
            dim = sheet.column_dimensions.get(col_letter)
            if dim and dim.hidden:
                continue
            val3 = row3[c_idx]
            if val3:
                current_date = parse_date(val3)
            if current_date:
                dates_list.append((c_idx, current_date))
                
        for r in range(5, sheet.max_row + 1):
            cname = sheet.cell(row=r, column=2).value
            if cname:
                cname_str = str(cname).strip()
                current_class = normalize_class_name(cname_str)
                for c_idx, d in dates_list:
                    val = sheet.cell(row=r, column=c_idx + 1).value
                    if val is not None:
                        if d not in timelines[current_class][sheetname]:
                            timelines[current_class][sheetname].append(d)
                            
    sorted_timelines = {}
    for cls, sheets_dict in timelines.items():
        sorted_timelines[cls] = {}
        for s_name, date_list in sheets_dict.items():
            sorted_timelines[cls][s_name] = sorted(list(set(date_list)))
            
    return sorted_timelines

def test_build_class_timelines(workbook):
    timelines = build_class_timelines(workbook)
    assert isinstance(timelines, dict), "Timeline map must be a dictionary"
    assert len(timelines) > 0, "Timelines should not be empty"
    print("test_build_class_timelines: PASS")

test_build_class_timelines(wb)

def get_compare_date_range(class_name, sheet_curr, monday_curr, timelines):
    subject_dates = timelines.get(class_name, {}).get(sheet_curr, [])
    past_dates = [d for d in subject_dates if d < monday_curr]
    
    if past_dates:
        last_date = max(past_dates)
        monday_prev = last_date - timedelta(days=last_date.weekday())
        sunday_prev = monday_prev + timedelta(days=6)
        return monday_prev, sunday_prev, sheet_curr, False
        
    other_subjects = timelines.get(class_name, {})
    candidates = []
    for s_name, date_list in other_subjects.items():
        if s_name == sheet_curr:
            continue
        past_subj_dates = [d for d in date_list if d < monday_curr]
        if past_subj_dates:
            candidates.append((max(past_subj_dates), s_name))
            
    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        last_date, sheet_prev = candidates[0]
        monday_prev = last_date - timedelta(days=last_date.weekday())
        sunday_prev = monday_prev + timedelta(days=6)
        return monday_prev, sunday_prev, sheet_prev, True
        
    return None, None, None, True

def test_get_compare_date_range():
    mock_timelines = {
        'TEST-CLASS': {
            'SubjectA': [date(2026, 7, 10), date(2026, 7, 15)],
            'SubjectB': [date(2026, 8, 11), date(2026, 8, 14)]
        }
    }
    m_p, s_p, s_prev, is_new = get_compare_date_range('TEST-CLASS', 'SubjectA', date(2026, 7, 16), mock_timelines)
    assert s_prev == 'SubjectA'
    assert not is_new
    assert m_p == date(2026, 7, 13)
    
    m_p, s_p, s_prev, is_new = get_compare_date_range('TEST-CLASS', 'SubjectB', date(2026, 8, 10), mock_timelines)
    assert s_prev == 'SubjectA'
    assert is_new
    assert m_p == date(2026, 7, 13)
    
    print("test_get_compare_date_range: PASS")

test_get_compare_date_range()

def get_max_excel_date(wb, sheets):
    all_dates = []
    for sheetname in sheets:
        if sheetname not in wb.sheetnames:
            continue
        sheet = wb[sheetname]
        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        for val in row3:
            parsed = parse_date(val)
            if parsed:
                all_dates.append(parsed)
    return max(all_dates) if all_dates else date(2026, 7, 17)

active_sheets = []
for s_name in wb.sheetnames:
    if s_name.lower() == 'sheet1':
        continue
    if any(k in s_name for k in ['KS24', 'KS25', 'SKL']):
        active_sheets.append(s_name)
max_date = get_max_excel_date(wb, active_sheets)
monday_curr = max_date - timedelta(days=max_date.weekday())
sunday_curr = monday_curr + timedelta(days=6)

monday_prev = monday_curr - timedelta(days=7)
sunday_prev = monday_prev + timedelta(days=6)

start_prev = monday_prev
end_prev = sunday_prev
start_curr = monday_curr
end_curr = sunday_curr

print(f"Tự động phát hiện tuần báo cáo: {start_curr.strftime('%d/%m')} - {end_curr.strftime('%d/%m/%Y')}")
print(f"Tuần đối chiếu: {start_prev.strftime('%d/%m')} - {end_prev.strftime('%d/%m/%Y')}")

weekly_groups = {
    'KS25_CNTT_HN': {
        'classes': ['HN-K25-CNTT1', 'HN-K25-CNTT2', 'HN-K25-CNTT3', 'HN-K25-CNTT4', 'HN-K25-CNTT5', 'HN-K25-CNTT6'],
        'sheet_curr': 'KS25_Python_Web',
        'sheet_prev': 'KS25_Python_Web',
        'label': 'Khóa KS25 CNTT Hà Nội (Python Web)'
    },
    'KS25_CNTT_HCM': {
        'classes': ['HCM-K25-CNTT5', 'HCM-K25-CNTT6', 'HCM-K25-CNTT7', 'HCM-K25-CNTT8'],
        'sheet_curr': 'KS25_Python_Web',
        'sheet_prev': 'KS25_Python_Web',
        'label': 'Khóa KS25 CNTT TP. HCM (Python Web)'
    },
    'KS25_QTKD_HN': {
        'classes': ['HN-K25-QTKD1', 'HN-K25-QTKD2', 'HN-K25-QTKD3'],
        'sheet_curr': 'KS25_QTKD_BA201',
        'sheet_prev': 'KS25_QTKD_PRJ302',
        'label': 'Khóa KS25 QTKD Hà Nội (BA201 / PRJ302)'
    },
    # Sau khi gộp lớp: HN-K24-CNTT5 và HCM-K24-CNTT2 đã giải thể.
    # HCM-K24-CNTT1 được chuyển vào cùng bảng cơ sở HN để quản lý thống nhất.
    'KS24_CNTT_HN': {
        'classes': ['HN-K24-CNTT1', 'HN-K24-CNTT2', 'HN-K24-CNTT3', 'HN-K24-CNTT4', 'HCM-K24-CNTT1'],
        'sheet_curr': 'KS24_AI_Intergration',
        'sheet_prev': 'KS24_AI',
        'label': 'Khóa KS24 CNTT (AI Integration / AI) — Hà Nội & HCM-CNTT1'
    }
}

def format_weekly_cell(curr_val, prev_val):
    diff = curr_val - prev_val
    if diff > 0.05:
        return f"{curr_val:.2f}% <span style='color:#ef4444; font-weight:600;'>(▲ +{diff:.2f}%)</span>"
    elif diff < -0.05:
        return f"{curr_val:.2f}% <span style='color:#10b981; font-weight:600;'>(▼ {diff:.2f}%)</span>"
    else:
        return f"{curr_val:.2f}% <span style='color:#64748b;'>(--)</span>"

def get_weekly_metrics(sheetname, classes_target, start_date, end_date):
    if sheetname not in wb.sheetnames:
        return {}
    sheet = wb[sheetname]
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    dates_list = []
    current_date = None
    for c_idx in range(3, len(row3)):
        col_letter = get_column_letter(c_idx + 1)
        dim = sheet.column_dimensions.get(col_letter)
        if dim and dim.hidden:
            continue
        val3 = row3[c_idx]
        val4 = row4[c_idx]
        if val3:
            current_date = parse_date(val3)
        if current_date and start_date <= current_date <= end_date:
            dates_list.append((c_idx, current_date, val4))
            
    res = {}
    current_class = None
    for r in range(5, sheet.max_row + 1):
        cname = sheet.cell(row=r, column=2).value
        teacher_val = sheet.cell(row=r, column=3).value
        
        if cname:
            cname_str = str(cname).strip()
            current_class = normalize_class_name(cname_str)
            size = extract_class_size(cname_str)
            class_sizes[current_class] = size
            
            matched_class = None
            for tc in classes_target:
                if tc == current_class:
                    matched_class = tc
                    break
            if not matched_class:
                continue
                
            teacher_name = str(teacher_val).strip() if teacher_val else "N/A"
            
            tg_name = "N/A"
            if r + 1 <= sheet.max_row:
                next_c2 = sheet.cell(row=r+1, column=2).value
                next_c3 = sheet.cell(row=r+1, column=3).value
                if not next_c2 and next_c3:
                    tg_name = str(next_c3).strip()
                    
            # Gom giá trị theo ngày học và loại của ngày đó
            date_metrics = defaultdict(dict)
            for c_idx, d, val4 in dates_list:
                val = sheet.cell(row=r, column=c_idx + 1).value
                if val is not None:
                    try:
                        date_metrics[d][val4] = float(val)
                    except ValueError:
                        pass
            
            # Lọc bỏ các ngày trống (tất cả chỉ số vắng/nợ/vi phạm bằng 0 hoặc None)
            active_dates_vals = defaultdict(list)
            for d, metrics in date_metrics.items():
                is_empty_day = True
                for val in metrics.values():
                    if val is not None and val != 0.0:
                        is_empty_day = False
                        break
                
                if not is_empty_day:
                    for val4, val in metrics.items():
                        active_dates_vals[val4].append(val)
            
            averages = {}
            for metric in ['Chuyên cần', 'Bài tập', 'Elearning']:
                vals = active_dates_vals.get(metric, [])
                averages[metric] = sum(vals) / len(vals) if vals else 0.0
                
            res[matched_class] = {
                'teacher': teacher_name,
                'tg': tg_name,
                'metrics': averages
            }
    return res

timelines = build_class_timelines(wb)

def read_class_metrics_for_date(sheet_obj, class_name, target_date):
    row3 = list(sheet_obj.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet_obj.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    cols = []
    current_date = None
    for c in range(2, len(row3)):
        col_letter = openpyxl.utils.get_column_letter(c + 1)
        dim = sheet_obj.column_dimensions.get(col_letter)
        if dim and dim.hidden:
            continue
        val3 = row3[c]
        val4 = row4[c]
        if val3:
            current_date = parse_date(val3)
        if current_date == target_date and val4 in ['Chuyên cần', 'Bài tập', 'Elearning']:
            cols.append((c + 1, val4))
            
    if not cols:
        return None
        
    header_row_idx = None
    for r in range(1, 20):
        row_vals = [str(sheet_obj.cell(row=r, column=c).value or "").strip() for c in range(1, sheet_obj.max_column + 1)]
        if 'Lớp' in row_vals:
            header_row_idx = r
            break
            
    headers = [str(sheet_obj.cell(row=header_row_idx, column=c).value or "").strip() for c in range(1, sheet_obj.max_column + 1)]
    class_col_idx = headers.index('Lớp') + 1
    person_col_idx = None
    for c_idx, h in enumerate(headers):
        if 'Giảng viên' in h or 'Trợ giảng' in h or 'Giảng viên/Trợ giảng' in h:
            person_col_idx = c_idx + 1
            break
            
    current_class = None
    gv_name = "N/A"
    tg_name = "N/A"
    gv_metrics = {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}
    tg_metrics = {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}
    
    has_gv = False
    has_tg = False
    
    for r in range(header_row_idx + 1, sheet_obj.max_row + 1):
        c_val = sheet_obj.cell(row=r, column=class_col_idx).value
        p_val = sheet_obj.cell(row=r, column=person_col_idx).value
        
        if c_val:
            current_class = normalize_class_name(str(c_val).strip())
            
        if current_class == class_name and p_val and str(p_val).strip() not in ['', 'nan', 'Giảng viên/Trợ giảng']:
            name = str(p_val).strip()
            is_tg = (c_val is None or str(c_val).strip() == "")
            
            # Đọc điểm
            metrics = {}
            for col_idx, metric in cols:
                val = sheet_obj.cell(row=r, column=col_idx).value
                if val is not None:
                    metrics[metric] = float(val)
                    
            if not is_tg:
                gv_name = name
                gv_metrics = metrics
                has_gv = True
            else:
                tg_name = name
                tg_metrics = metrics
                has_tg = True
                
    if not has_gv and not has_tg:
        return None
        
    for m in ['Chuyên cần', 'Bài tập', 'Elearning']:
        if m not in gv_metrics or gv_metrics[m] is None:
            gv_metrics[m] = 0.0
            
    if has_tg:
        for m in ['Chuyên cần', 'Bài tập', 'Elearning']:
            if m not in tg_metrics or tg_metrics[m] is None:
                tg_metrics[m] = gv_metrics[m]
                
    return {
        'teacher': gv_name,
        'tg': tg_name,
        'metrics': gv_metrics,
        'date': target_date
    }

def get_class_latest_and_prev_metrics(workbook, sheetname, class_name, ref_date, timelines):
    if sheetname not in workbook.sheetnames:
        return None, None
    sheet = workbook[sheetname]
    
    subject_dates = timelines.get(class_name, {}).get(sheetname, [])
    valid_dates = [d for d in subject_dates if d <= ref_date]
    if not valid_dates:
        return None, None
        
    valid_dates.sort()
    latest_date = valid_dates[-1]
    
    curr_metrics = read_class_metrics_for_date(sheet, class_name, latest_date)
    
    prev_metrics = None
    if len(valid_dates) >= 2:
        prev_date = valid_dates[-2]
        prev_metrics = read_class_metrics_for_date(sheet, class_name, prev_date)
    else:
        other_subjects = timelines.get(class_name, {})
        candidates = []
        for s_name, date_list in other_subjects.items():
            if s_name == sheetname:
                continue
            past_dates = [d for d in date_list if d < latest_date]
            if past_dates:
                candidates.append((max(past_dates), s_name))
        if candidates:
            candidates.sort(key=lambda x: x[0], reverse=True)
            prev_date, prev_sheet = candidates[0]
            prev_metrics = read_class_metrics_for_date(workbook[prev_sheet], class_name, prev_date)
            
    return curr_metrics, prev_metrics

weekly_stats = {}
for gkey, ginfo in weekly_groups.items():
    classes = ginfo['classes']
    curr_data = {}
    prev_data = {}
    
    for cls in classes:
        curr_metrics, prev_metrics = get_class_latest_and_prev_metrics(wb, ginfo['sheet_curr'], cls, max_date, timelines)
        if curr_metrics:
            curr_data[cls] = curr_metrics
            if prev_metrics:
                prev_data[cls] = prev_metrics
            else:
                prev_data[cls] = {
                    'teacher': curr_metrics['teacher'],
                    'tg': curr_metrics['tg'],
                    'metrics': {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}
                }
        else:
            # Fallback
            subject_dates = timelines.get(cls, {}).get(ginfo['sheet_curr'], [])
            if subject_dates:
                last_date = max(subject_dates)
                curr_m, prev_m = get_class_latest_and_prev_metrics(wb, ginfo['sheet_curr'], cls, last_date, timelines)
                if curr_m:
                    curr_data[cls] = curr_m
                    prev_data[cls] = prev_m if prev_m else {
                        'teacher': curr_m['teacher'],
                        'tg': curr_m['tg'],
                        'metrics': {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}
                    }
            else:
                all_dates = []
                for s_n, d_l in timelines.get(cls, {}).items():
                    all_dates.extend([(d, s_n) for d in d_l])
                if all_dates:
                    all_dates.sort(key=lambda x: x[0])
                    last_date, s_n = all_dates[-1]
                    curr_m, prev_m = get_class_latest_and_prev_metrics(wb, s_n, cls, last_date, timelines)
                    if curr_m:
                        curr_data[cls] = curr_m
                        prev_data[cls] = prev_m if prev_m else {
                            'teacher': curr_m['teacher'],
                            'tg': curr_m['tg'],
                            'metrics': {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}
                        }
                        
    weekly_stats[gkey] = {
        'label': ginfo['label'],
        'classes': classes,
        'curr': curr_data,
        'prev': prev_data
    }

all_sheets = []
for s_name in wb.sheetnames:
    if s_name.lower() == 'sheet1':
        continue
    if any(k in s_name for k in ['KS24', 'KS25', 'SKL']):
        all_sheets.append(s_name)

class_course_data = defaultdict(dict)
teacher_stats = {}

for sheetname in all_sheets:
    if sheetname not in wb.sheetnames:
        continue
    sheet = wb[sheetname]
    row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    dates_list = []
    current_date = None
    for c_idx in range(3, len(row3)):
        col_letter = get_column_letter(c_idx + 1)
        dim = sheet.column_dimensions.get(col_letter)
        if dim and dim.hidden:
            continue
        val3 = row3[c_idx]
        val4 = row4[c_idx]
        if val3:
            current_date = parse_date(val3)
        if current_date:
            dates_list.append((c_idx, current_date, val4))
            
    current_class = None
    gv_cc_vals, gv_bt_vals, gv_el_vals = [], [], []
    tg_cc_vals, tg_bt_vals, tg_el_vals = [], [], []
    current_gv_name = None
    current_tg_name = None
    
    for r in range(5, sheet.max_row + 1):
        cname = sheet.cell(row=r, column=2).value
        teacher_tg_val = sheet.cell(row=r, column=3).value
        teacher_tg_name = str(teacher_tg_val).strip() if teacher_tg_val else ""
        
        if cname:
            cname_str = str(cname).strip()
            current_class = normalize_class_name(cname_str)
            size = extract_class_size(cname_str)
            class_sizes[current_class] = size
            role = 'GV'
            current_gv_name = teacher_tg_name
            gv_cc_vals, gv_bt_vals, gv_el_vals = [], [], []
            
            for c_idx, d, val4 in dates_list:
                val = sheet.cell(row=r, column=c_idx + 1).value
                if val is not None:
                    try:
                        val_f = float(val)
                        if val4 == 'Chuyên cần': gv_cc_vals.append(val_f)
                        elif val4 == 'Bài tập': gv_bt_vals.append(val_f)
                        elif val4 == 'Elearning': gv_el_vals.append(val_f)
                    except ValueError:
                        pass
        else:
            role = 'TG'
            current_tg_name = teacher_tg_name
            tg_cc_vals = gv_cc_vals.copy()
            tg_bt_vals = gv_bt_vals.copy()
            tg_el_vals = gv_el_vals.copy()
            
        if not current_class or teacher_tg_name in ['', 'None', 'Chưa phân công', 'Giảng viên/Trợ giảng']:
            continue
            
        name = teacher_tg_name
        dept = 'QTKD' if 'QTKD' in sheetname or 'QTKD' in current_class else 'CNTT'
        
        cc_vals = gv_cc_vals.copy() if role == 'GV' else tg_cc_vals.copy()
        bt_vals = gv_bt_vals.copy() if role == 'GV' else tg_bt_vals.copy()
        el_vals = gv_el_vals.copy() if role == 'GV' else tg_el_vals.copy()
        
        cc_avg = np.mean(cc_vals) if cc_vals else 0.0
        bt_avg = np.mean(bt_vals) if bt_vals else 0.0
        el_avg = np.mean(el_vals) if el_vals else 0.0
        
        class_course_data[(current_class, role)][sheetname] = {
            'cc_avg': cc_avg,
            'bt_avg': bt_avg,
            'el_avg': el_avg,
            'teacher': name,
            'min_date': min(d for c, d, s in dates_list) if dates_list else date.min
        }
        
        if name not in teacher_stats:
            teacher_stats[name] = {
                'role': role,
                'department': dept,
                'cc': [],
                'bt': [],
                'el': [],
                'classes': set(),
                'sheets': set()
            }
        teacher_stats[name]['classes'].add(current_class)
        teacher_stats[name]['sheets'].add(sheetname)
        if cc_vals: teacher_stats[name]['cc'].extend(cc_vals)
        if bt_vals: teacher_stats[name]['bt'].extend(bt_vals)
        if el_vals: teacher_stats[name]['el'].extend(el_vals)

transitions = []
for (cname, role), courses in sorted(class_course_data.items()):
    if len(courses) < 2:
        continue
    sorted_courses = sorted(courses.items(), key=lambda x: x[1]['min_date'])
    for i in range(1, len(sorted_courses)):
        prev_sheet, prev_data = sorted_courses[i-1]
        curr_sheet, curr_data = sorted_courses[i]
        
        delta_cc = prev_data['cc_avg'] - curr_data['cc_avg']
        delta_bt = prev_data['bt_avg'] - curr_data['bt_avg']
        delta_el = prev_data['el_avg'] - curr_data['el_avg']
        
        transitions.append({
            'class': cname,
            'role': role,
            'prev_sheet': prev_sheet,
            'prev_teacher': prev_data['teacher'],
            'prev_cc': prev_data['cc_avg'],
            'prev_bt': prev_data['bt_avg'],
            'prev_el': prev_data['el_avg'],
            'curr_sheet': curr_sheet,
            'curr_teacher': curr_data['teacher'],
            'curr_cc': curr_data['cc_avg'],
            'curr_bt': curr_data['bt_avg'],
            'curr_el': curr_data['el_avg'],
            'delta_cc': delta_cc,
            'delta_bt': delta_bt,
            'delta_el': delta_el
        })

evaluated_staff = []
watchlist_staff = []

for name, stats in teacher_stats.items():
    cc_mean = np.mean(stats['cc']) if stats['cc'] else 0.0
    bt_mean = np.mean(stats['bt']) if stats['bt'] else 0.0
    el_mean = np.mean(stats['el']) if stats['el'] else 0.0
    
    t_imps = [item for item in transitions if item['curr_teacher'] == name and item['role'] == stats['role']]
    
    if t_imps:
        cmi_values = []
        for class_item in t_imps:
            cmi_cc = 0.5 * class_item['prev_cc'] - class_item['curr_cc'] + 15.0
            cmi_bt = 0.5 * class_item['prev_bt'] - class_item['curr_bt'] + 15.0
            cmi_el = 0.5 * class_item['prev_el'] - class_item['curr_el'] + 15.0
            cmi_values.append((cmi_cc + cmi_bt + cmi_el) / 3.0)
            
        mgmt_score = np.mean(cmi_values)
        
        if mgmt_score > 12.0:
            classification = "Rescuers (Giải cứu xuất sắc)"
        elif mgmt_score < 0.0 or cc_mean > 25.0 or bt_mean > 20.0 or el_mean > 20.0:
            classification = "Cần Hỗ Trợ (Needs Support)"
        else:
            classification = "Duy trì (Maintainers)"
            
        avg_delta_cc = np.mean([item['delta_cc'] for item in t_imps])
        avg_delta_bt = np.mean([item['delta_bt'] for item in t_imps])
        avg_delta_el = np.mean([item['delta_el'] for item in t_imps])
        
        evaluated_staff.append({
            'name': name,
            'role': stats['role'],
            'dept': stats['department'],
            'classes_count': len(stats['classes']),
            'classes_list': ", ".join(list(stats['classes'])),
            'sheets_list': ", ".join(list(stats['sheets'])),
            'cc': cc_mean,
            'bt': bt_mean,
            'el': el_mean,
            'delta_cc': avg_delta_cc,
            'delta_bt': avg_delta_bt,
            'delta_el': avg_delta_el,
            'cmi': mgmt_score,
            'class': classification
        })
    else:
        watchlist_staff.append({
            'name': name,
            'role': stats['role'],
            'dept': stats['department'],
            'classes_count': len(stats['classes']),
            'classes_list': ", ".join(list(stats['classes'])),
            'sheets_list': ", ".join(list(stats['sheets'])),
            'cc': cc_mean,
            'bt': bt_mean,
        })

import json
trends_data = {
    'KS24_HN': {'courses': [], 'cc': [], 'bt': [], 'el': []},
    'KS24_HCM': {'courses': [], 'cc': [], 'bt': [], 'el': []},
    'HN': {'courses': [], 'cc': [], 'bt': [], 'el': []},
    'HCM': {'courses': [], 'cc': [], 'bt': [], 'el': []},
    'QTKD': {'courses': [], 'cc': [], 'bt': [], 'el': []}
}
cohort_sheets = {
    # KS24_HN: gộp cả HCM-K24-CNTT1 vào cùng nhóm HN để phản ánh cấu trúc lớp mới sau khi gộp
    'KS24_HN': ['KS24-JavaAdvance', 'KS24_JavaWeb', 'KS24_JWS', 'KS24_AI', 'KS24_AI_Intergration'],
    'HN': ['KS25_Javascript', 'KS25_Database', 'KS25_Python', 'KS25_Python_Web'],
    'HCM': ['KS25_Javascript', 'KS25_Database', 'KS25_Python', 'KS25_Python_Web'],
    'QTKD': ['KS25_QTKD_M103', 'KS25_QTKD_M104', 'KS25_QTKD_DTB201', 'KS25_QTKD_DTB202', 'KS25_QTKD_PRJ302', 'KS25_QTKD_BA201']
}
for cohort, sheets in cohort_sheets.items():
    for sheetname in sheets:
        if sheetname not in wb.sheetnames:
            continue
        cc_weighted_sum, bt_weighted_sum, el_weighted_sum = 0.0, 0.0, 0.0
        total_students_cohort = 0
        
        for (cname, role), courses in class_course_data.items():
            if role != 'GV' or sheetname not in courses:
                continue
            if cohort == 'KS24_HN' and not ('CNTT' in cname and 'K24' in cname): continue
            if cohort == 'HN' and not ('HN' in cname and 'CNTT' in cname and 'K25' in cname): continue
            if cohort == 'HCM' and not ('HCM' in cname and 'CNTT' in cname and 'K25' in cname): continue
            if cohort == 'QTKD' and not 'QTKD' in cname: continue
            
            size = class_sizes.get(cname, 30)
            total_students_cohort += size
            cc_weighted_sum += courses[sheetname]['cc_avg'] * size
            bt_weighted_sum += courses[sheetname]['bt_avg'] * size
            el_weighted_sum += courses[sheetname]['el_avg'] * size
            
        course_name = sheetname.replace('KS24-', '').replace('KS24_', '').replace('KS25_', '').split('_')[-1]
        trends_data[cohort]['courses'].append(course_name)
        
        cc_avg = cc_weighted_sum / total_students_cohort if total_students_cohort > 0 else 0.0
        bt_avg = bt_weighted_sum / total_students_cohort if total_students_cohort > 0 else 0.0
        el_avg = el_weighted_sum / total_students_cohort if total_students_cohort > 0 else 0.0
        
        trends_data[cohort]['cc'].append(round(float(cc_avg), 2))
        trends_data[cohort]['bt'].append(round(float(bt_avg), 2))
        trends_data[cohort]['el'].append(round(float(el_avg), 2))

trends_data['compare'] = {}
for gkey, dict_key in [
    ('KS24_CNTT_HN', 'KS24_HN'),
    ('KS25_CNTT_HN', 'HN'),
    ('KS25_CNTT_HCM', 'HCM'),
    ('KS25_QTKD_HN', 'QTKD')
]:
    stats = weekly_stats.get(gkey, {'curr': {}, 'prev': {}})
    
    total_size_curr = 0
    weighted_cc_curr = 0.0
    weighted_bt_curr = 0.0
    weighted_el_curr = 0.0
    for cls, x in stats['curr'].items():
        size = class_sizes.get(cls, 30)
        total_size_curr += size
        weighted_cc_curr += x['metrics']['Chuyên cần'] * size
        weighted_bt_curr += x['metrics']['Bài tập'] * size
        weighted_el_curr += x['metrics']['Elearning'] * size
        
    curr_cc = weighted_cc_curr / total_size_curr if total_size_curr > 0 else 0.0
    curr_bt = weighted_bt_curr / total_size_curr if total_size_curr > 0 else 0.0
    curr_el = weighted_el_curr / total_size_curr if total_size_curr > 0 else 0.0
    
    total_size_prev = 0
    weighted_cc_prev = 0.0
    weighted_bt_prev = 0.0
    weighted_el_prev = 0.0
    for cls, x in stats['prev'].items():
        size = class_sizes.get(cls, 30)
        total_size_prev += size
        weighted_cc_prev += x['metrics']['Chuyên cần'] * size
        weighted_bt_prev += x['metrics']['Bài tập'] * size
        weighted_el_prev += x['metrics']['Elearning'] * size
        
    prev_cc = weighted_cc_prev / total_size_prev if total_size_prev > 0 else 0.0
    prev_bt = weighted_bt_prev / total_size_prev if total_size_prev > 0 else 0.0
    prev_el = weighted_el_prev / total_size_prev if total_size_prev > 0 else 0.0
    
    trends_data['compare'][dict_key] = {
        'curr': [round(float(curr_cc), 2), round(float(curr_bt), 2), round(float(curr_el), 2)],
        'prev': [round(float(prev_cc), 2), round(float(prev_bt), 2), round(float(prev_el), 2)]
    }

with open('data/processed/historical_trends.json', 'w', encoding='utf-8') as f:
    json.dump(trends_data, f, ensure_ascii=False, indent=4)

markdown_content = f"""# BÁO CÁO THỐNG KÊ CHỈ SỐ VI PHẠM HÀNG NGÀY & NĂNG LỰC QUẢN TRỊ LỚP CỦA GV/TG

<div style="background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%); color: white; padding: 12px 20px; border-radius: 8px; display: inline-block; font-weight: 600; margin-bottom: 20px; box-shadow: 0 4px 15px rgba(59, 130, 246, 0.3);">
  <i class="fas fa-calendar-check" style="margin-right: 8px;"></i> Báo cáo Ngày: {max_date.strftime('%d/%m/%Y')}
</div>
"""

markdown_content += "\n"

def generate_kpi_card(title, curr_val, prev_val):
    diff = curr_val - prev_val
    trend_class = "trend-bad" if diff > 0 else "trend-good"
    arrow = "▲" if diff > 0 else "▼" if diff < 0 else "-"
    diff_str = f"{diff:+.2f}%" if diff != 0 else "0.00%"
    return f"""
    <div class="kpi-card">
        <div class="kpi-title">{title}</div>
        <div class="kpi-value">{curr_val:.2f}%</div>
        <div class="kpi-trend {trend_class}">{arrow} {diff_str} so với hôm qua</div>
    </div>
    """

def generate_cohort_section(cohort_id, weekly_stats_group):
    # Calculate weighted averages
    curr_vals = {}
    prev_vals = {}
    for m in ['Chuyên cần', 'Bài tập', 'Elearning']:
        total_size_curr = 0
        weighted_sum_curr = 0.0
        for cls, x in weekly_stats_group['curr'].items():
            size = class_sizes.get(cls, 30)
            total_size_curr += size
            weighted_sum_curr += x['metrics'][m] * size
        curr_vals[m] = weighted_sum_curr / total_size_curr if total_size_curr > 0 else 0.0
        
        total_size_prev = 0
        weighted_sum_prev = 0.0
        for cls, x in weekly_stats_group['prev'].items():
            size = class_sizes.get(cls, 30)
            total_size_prev += size
            weighted_sum_prev += x['metrics'][m] * size
        prev_vals[m] = weighted_sum_prev / total_size_prev if total_size_prev > 0 else 0.0
    
    cc_diff_overall = curr_vals['Chuyên cần'] - prev_vals['Chuyên cần']
    bt_diff_overall = curr_vals['Bài tập'] - prev_vals['Bài tập']
    el_diff_overall = curr_vals['Elearning'] - prev_vals['Elearning']
    
    issues = []
    gv_actions_cc = []
    gv_actions_bt = []
    gv_actions_el = []
    cv_actions_cc = []
    cv_actions_bt = []
    cv_actions_el = []

    curr_cc = curr_vals['Chuyên cần']
    curr_bt = curr_vals['Bài tập']
    curr_el = curr_vals['Elearning']

    cc_violated = (cc_diff_overall > 1.5 or curr_cc > 5.0)
    bt_violated = (bt_diff_overall > 1.5 or curr_bt > 10.0)
    el_violated = (el_diff_overall > 1.5 or curr_el > 10.0)

    # 1. Phát hiện vấn đề & Ghi nhận cảnh báo
    if cc_violated:
        if cc_diff_overall > 1.5 and curr_cc > 5.0:
            issues.append(f"Tỷ lệ vắng mặt chuyên cần ở mức nguy hiểm ({curr_cc:.2f}%) và tăng nhanh (+{cc_diff_overall:.2f}%) so với hôm qua.")
        elif curr_cc > 5.0:
            issues.append(f"Tỷ lệ vắng mặt chuyên cần duy trì ở mức cao nguy hiểm ({curr_cc:.2f}%).")
        else:
            issues.append(f"Tỷ lệ vắng mặt chuyên cần có xu hướng tăng nhanh (+{cc_diff_overall:.2f}%).")

    if bt_violated:
        if bt_diff_overall > 1.5 and curr_bt > 10.0:
            issues.append(f"Tỷ lệ nợ bài tập ở mức báo động ({curr_bt:.2f}%) và tăng nhanh (+{bt_diff_overall:.2f}%) so với hôm qua.")
        elif curr_bt > 10.0:
            issues.append(f"Tỷ lệ nợ bài tập duy trì ở mức cao báo động ({curr_bt:.2f}%).")
        else:
            issues.append(f"Tỷ lệ nợ bài tập có xu hướng tăng (+{bt_diff_overall:.2f}%).")

    if el_violated:
        if el_diff_overall > 1.5 and curr_el > 10.0:
            issues.append(f"Tỷ lệ vi phạm Elearning ở mức nghiêm trọng ({curr_el:.2f}%) và tăng nhanh (+{el_diff_overall:.2f}%) so với hôm qua.")
        elif curr_el > 10.0:
            issues.append(f"Tỷ lệ vi phạm Elearning duy trì ở mức cao nghiêm trọng ({curr_el:.2f}%).")
        else:
            issues.append(f"Tỷ lệ vi phạm Elearning có xu hướng gia tăng (+{el_diff_overall:.2f}%).")

    # 2. Sinh Kế hoạch Hành động Tích hợp Tinh gọn (Tối đa 2-3 giải pháp đi sâu cho GV và CVHT)
    gv_final = []
    cv_final = []
    
    n_violations = sum([1 for v in [cc_violated, bt_violated, el_violated] if v])

    if n_violations >= 2:
        # Trường hợp vi phạm hỗn hợp (từ 2 lỗi trở lên): Dùng các giải pháp tích hợp đa chiều
        gv_final = [
            "Kiểm soát nghiêm đầu giờ: Điểm danh và kiểm tra nhanh 5 phút đầu buổi dựa trên nội dung Elearning; từ chối cho SV vào lớp nếu chưa hoàn thành lý thuyết Elearning.",
            "Kèm cặp trực tiếp tại lớp: Dành 15-20 phút cuối ca học trực tiếp kèm cặp nhóm sinh viên chưa nộp bài hoàn thành các bài tập trọng tâm.",
            "Cảnh báo & Báo cáo nóng: Gửi tin nhắn Zalo cá nhân cho SV vắng/nợ ngay sau ca học và báo cáo danh sách SV vắng liên tiếp 2 buổi cho CVHT trước 12h ngày hôm sau."
        ]
        cv_final = [
            "Hotline khẩn cấp: Thực hiện cuộc gọi Hotline trực tiếp cho phụ huynh/người giám hộ báo cáo tình trạng vắng học và nợ bài của sinh viên trong vòng 24h.",
            "Triệu tập 3 bên: Triệu tập cuộc gặp trực tiếp bắt buộc (GV - CVHT - SV) đối với các trường hợp vi phạm từ 3 buổi trở lên để ký biên bản cam kết học tập.",
            "Kiểm toán điều kiện học vụ: Rà soát lập danh sách SV có tỷ lệ chuyên cần < 80% hoặc Elearning < 50% để đưa vào diện xem xét cấm thi cuối kỳ sớm."
        ]
    elif cc_violated:
        # Chỉ vi phạm chuyên cần
        gv_final = [
            "Gửi tin nhắn cảnh cáo Zalo cá nhân cho từng sinh viên vắng mặt ngay sau khi kết thúc ca học.",
            "Yêu cầu SV vắng nộp giải trình bằng văn bản trong 24h và áp dụng trừ điểm chuyên cần trực tiếp nếu không phép.",
            "Báo cáo khẩn danh sách SV vắng liên tiếp 2 buổi cho Cố vấn học tập (CVHT) trước 12h ngày hôm sau để can thiệp nóng."
        ]
        cv_final = [
            "Thực hiện cuộc gọi khẩn cấp (Hotline) trực tiếp cho phụ huynh/người giám hộ để cảnh báo tình trạng vắng học.",
            "Triệu tập cuộc gặp 3 bên bắt buộc (GV - CVHT - SV) đối với các trường hợp vắng từ 3 buổi để ký cam kết học tập.",
            "Kích hoạt quy trình xem xét cấm thi cuối kỳ sớm đối với sinh viên có tỷ lệ chuyên cần môn học dưới 80%."
        ]
    elif bt_violated:
        # Chỉ vi phạm bài tập
        gv_final = [
            "Thiết lập thời gian hoàn thành bù bắt buộc trong 24h đối với sinh viên nợ bài và cập nhật tỷ lệ nợ lên nhóm lớp hàng ngày.",
            "Dành 15-20 phút cuối ca học trực tiếp kèm cặp nhóm sinh viên chưa nộp bài hoàn thành các bài tập trọng tâm."
        ]
        cv_final = [
            "Tổ chức các buổi học bù / phòng Lab tự học bắt buộc trong tuần và phân công Trợ giảng giám sát SV hoàn thành bài nợ.",
            "Liên hệ Hotline phụ huynh đối với sinh viên nợ tích lũy từ 3 bài tập trở lên để phối hợp đôn đốc."
        ]
    elif el_violated:
        # Chỉ vi phạm Elearning
        gv_final = [
            "Nhắn tin gửi link bài tập Elearning trực tiếp trước 24h deadline; kiểm tra đột xuất 5 phút đầu giờ dựa trên nội dung Elearning.",
            "Áp dụng quy định từ chối cho SV vào lớp hoặc ghi nhận vắng nếu SV hoàn toàn không học lý thuyết Elearning trước ca học."
        ]
        cv_final = [
            "Rà soát kỹ danh sách sinh viên có tỷ lệ hoàn thành Elearning dưới 50% để đưa vào diện xem xét cấm thi theo Quy chế.",
            "Phối hợp với Ban cán sự lớp tổ chức các nhóm học tập hỗ trợ SV hoàn thành Elearning đúng hạn."
        ]

    ai_insights = ""
    if issues:
        issues_html = "".join(f"<li>{i}</li>" for i in issues)
        gv_html = "".join(f"<li style='margin-bottom:5px;'>{a}</li>" for a in gv_final)
        cv_html = "".join(f"<li style='margin-bottom:5px;'>{a}</li>" for a in cv_final)
        n_issues = len(issues)
        ai_insights = f"""
<div style="background:rgba(15,23,42,0.7);border:1px solid rgba(255,255,255,0.1);border-radius:14px;padding:18px 20px;margin:16px 0;backdrop-filter:blur(10px);">
    <div style="display:flex;align-items:center;gap:8px;margin-bottom:12px;">
        <span style="font-size:1.1rem;">📌</span>
        <span style="font-weight:700;font-size:0.93rem;color:#e2e8f0;">Kế hoạch Hành động Khắc phục Vi phạm Gia tăng</span>
    </div>
    <div style="background:rgba(239,68,68,0.12);border-left:4px solid #ef4444;padding:10px 14px;border-radius:6px;margin-bottom:10px;">
        <div style="color:#fca5a5;font-weight:700;font-size:0.78rem;text-transform:uppercase;letter-spacing:0.06em;margin-bottom:5px;">🚨 {n_issues} vấn đề phát hiện hôm nay</div>
        <ul style="margin:0;padding-left:18px;color:#e2e8f0;font-size:0.88rem;line-height:1.6;">{issues_html}</ul>
    </div>
    <div style="background:rgba(245,158,11,0.08);border-left:4px solid #f59e0b;padding:11px 15px;border-radius:6px;margin-bottom:10px;">
        <div style="color:#fcd34d;font-weight:700;font-size:0.78rem;text-transform:uppercase;letter-spacing:0.05em;margin-bottom:6px;">👨‍🏫 Giảng Viên Lớp &mdash; Thực hiện 24&ndash;48h</div>
        <ul style="margin:0;padding-left:18px;color:#e2e8f0;font-size:0.88rem;line-height:1.65;">{gv_html}</ul>
    </div>
    <div style="background:rgba(59,130,246,0.08);border-left:4px solid #3b82f6;padding:11px 15px;border-radius:6px;">
        <div style="color:#93c5fd;font-weight:700;font-size:0.78rem;text-transform:uppercase;letter-spacing:0.05em;margin-bottom:6px;">🧑‍💼 Cố Vấn Học Tập &mdash; Thực hiện trong hôm nay</div>
        <ul style="margin:0;padding-left:18px;color:#e2e8f0;font-size:0.88rem;line-height:1.65;">{cv_html}</ul>
    </div>
</div>
"""
    else:
        ai_insights = f"""
<div style="background:rgba(16,185,129,0.1);border:1px solid rgba(16,185,129,0.2);border-radius:12px;padding:13px 17px;margin:16px 0;display:flex;align-items:center;gap:11px;">
    <span style="font-size:1.35rem;">✅</span>
    <div><strong style="color:#6ee7b7;">Kỷ luật ổn định:</strong> Khóa {cohort_id} không có chỉ số vi phạm nào tăng đột biến hôm nay. Không cần can thiệp khẩn cấp.</div>
</div>
"""
    
    html = f"""
---
### Khóa {cohort_id}

{ai_insights}

<div class="data-grid-container" markdown="1">

| Tên Lớp | Giảng viên | Trợ giảng | Chuyên cần | Bài tập | Elearning | Xu hướng |
| :--- | :--- | :--- | :---: | :---: | :---: | :--- |
"""
    for cls in weekly_stats_group['classes']:
        curr_info = weekly_stats_group['curr'].get(cls, {'metrics': {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}, 'teacher': 'N/A', 'tg': 'N/A'})
        prev_info = weekly_stats_group['prev'].get(cls, {'metrics': {'Chuyên cần': 0.0, 'Bài tập': 0.0, 'Elearning': 0.0}})
        c_m, p_m = curr_info['metrics'], prev_info['metrics']
        cc_cell = format_weekly_cell(c_m['Chuyên cần'], p_m['Chuyên cần'])
        bt_cell = format_weekly_cell(c_m['Bài tập'], p_m['Bài tập'])
        el_cell = format_weekly_cell(c_m['Elearning'], p_m['Elearning'])
        cc_diff = c_m['Chuyên cần'] - p_m['Chuyên cần']
        bt_diff = c_m['Bài tập'] - p_m['Bài tập']
        el_diff = c_m['Elearning'] - p_m['Elearning']
        
        status_html = "<span class='badge badge-danger'>🚨 Tăng</span>" if (cc_diff > 1.0 or bt_diff > 1.0 or el_diff > 1.0) else "<span class='badge badge-success'>✅ Ổn định</span>"
        html += f"| **{cls}** | {curr_info['teacher']} | {curr_info['tg']} | {cc_cell} | {bt_cell} | {el_cell} | {status_html} |\n"
    
    html += """
</div>
"""
    return html

# KS24: Hà Nội + HCM-CNTT1 gộp chung sau tái cơ cấu lớp
markdown_content += generate_cohort_section('KS24-CNTT (HN & HCM-CNTT1)', weekly_stats['KS24_CNTT_HN'])
markdown_content += generate_cohort_section('HN-KS25-CNTT', weekly_stats['KS25_CNTT_HN'])
markdown_content += generate_cohort_section('HCM-KS25-CNTT', weekly_stats['KS25_CNTT_HCM'])
markdown_content += generate_cohort_section('HN-KS25-QTKD', weekly_stats['KS25_QTKD_HN'])


markdown_content += r"""
## II. ĐÁNH GIÁ NĂNG LỰC QUẢN TRỊ LỚP TÍCH LŨY CỦA GIẢNG VIÊN VÀ TRỢ GIẢNG (CMI HISTORICAL RANKING)

<div class="quote-box" style="margin-bottom: 30px;" markdown="1">
<h3 style="margin-top: 0; color: #60a5fa;"><i class="fas fa-calculator"></i> 3 Bước Đánh Giá CMI</h3>
<ul style="margin-bottom: 0;">
<li style="margin-bottom: 8px;"><b>1. Trọng số:</b> Tính theo công thức <code>50% Tỷ lệ lỗi cũ - 100% Tỷ lệ lỗi mới + Điểm nền (15)</code>.</li>
<li style="margin-bottom: 8px;"><b>2. Ý nghĩa:</b> Điểm càng cao (>12) ➔ Khắc phục lỗi càng tốt (Giải cứu xuất sắc).</li>
<li><b>3. Phân loại:</b> Nếu điểm dưới 0 hoặc vi phạm hiện tại quá cao (CC > 25%, BT > 20%, EL > 20%) ➔ <b>Cần Hỗ Trợ</b>.</li>
</ul>
</div>

### 🏆 Bảng xếp hạng Top 5 Giảng viên xuất sắc (Giảng viên phụ trách)

| Hạng | Họ và Tên | Tổng số lớp đã dạy | Chỉ số CMI tích lũy | Phân loại |
| :---: | :--- | :---: | :---: | :--- |
"""

all_gv = [x for x in evaluated_staff if x['role'] == 'GV']
all_gv = sorted(all_gv, key=lambda x: x['cmi'], reverse=True)[:5]

rank = 1
for item in all_gv:
    markdown_content += f"| {rank} | **{item['name']}** | {item['classes_count']} lớp | **{item['cmi']:+.2f}%** | {item['class']} |\n"
    rank += 1

markdown_content += """
### 🏆 Bảng xếp hạng Top 5 Trợ giảng xuất sắc (Trợ giảng phụ trách)

| Hạng | Họ và Tên | Tổng số lớp đã dạy | Chỉ số CMI tích lũy | Phân loại |
| :---: | :--- | :---: | :---: | :--- |
"""

all_tg = [x for x in evaluated_staff if x['role'] == 'TG']
all_tg = sorted(all_tg, key=lambda x: x['cmi'], reverse=True)[:5]

rank = 1
for item in all_tg:
    markdown_content += f"| {rank} | **{item['name']}** | {item['classes_count']} lớp | **{item['cmi']:+.2f}%** | {item['class']} |\n"
    rank += 1

def generate_staff_card(item):
    badge_color = "#10b981" if "Giải cứu" in item['class'] else "#3b82f6" if "Duy trì" in item['class'] else "#ef4444"
    icon = "fa-star" if "Giải cứu" in item['class'] else "fa-check-circle" if "Duy trì" in item['class'] else "fa-exclamation-triangle"
    
    weakness = "N/A"
    solution = "N/A"
    if item['delta_cc'] < -2.0: 
        weakness = "Tỷ lệ vắng mặt tăng"
        solution = "Gắt gao điểm danh và liên hệ học viên"
    elif item['delta_bt'] < -2.0:
        weakness = "Học viên lười làm bài tập"
        solution = "Chữa bài tập kỹ hơn, giao task nhỏ gọn hơn"
    elif item['delta_el'] < -2.0:
        weakness = "Chậm tiến độ Elearning"
        solution = "Nhắc nhở học viên làm Elearning ngay trên lớp"
    elif "Giải cứu" in item['class']:
        weakness = "Không có (Kỷ luật cực tốt)"
        solution = "Duy trì phong độ và chia sẻ kinh nghiệm cho nhóm"
    else:
        weakness = "Kỷ luật ở mức trung bình ổn"
        solution = "Cải thiện động lực học của sinh viên yếu kém"

    card = f"""<div class="staff-card" style="background: rgba(30, 41, 59, 0.5); border: 1px solid rgba(255, 255, 255, 0.1); border-radius: 12px; padding: 16px;">
<div style="display: flex; align-items: center; justify-content: space-between; border-bottom: 1px solid rgba(255, 255, 255, 0.05); padding-bottom: 12px; margin-bottom: 12px;">
<div>
<h4 style="margin: 0; color: #f8fafc; font-size: 1.1rem;">{item['name']}</h4>
<div style="font-size: 0.8rem; color: #94a3b8; margin-top: 4px;">Vai trò: {item['role']} - {item['classes_count']} Lớp ({item['classes_list']})</div>
</div>
<div style="background: {badge_color}20; color: {badge_color}; padding: 4px 10px; border-radius: 20px; font-size: 0.75rem; font-weight: 600;">
<i class="fas {icon}"></i> {item['class']}
</div>
</div>
<div style="display: flex; gap: 12px; font-size: 0.85rem;">
<div style="flex: 1;">
<div style="color: #ef4444; font-weight: 600; margin-bottom: 4px;"><i class="fas fa-arrow-down"></i> Điểm Yếu</div>
<div style="color: #cbd5e1;">{weakness}</div>
</div>
<div style="flex: 1;">
<div style="color: #10b981; font-weight: 600; margin-bottom: 4px;"><i class="fas fa-lightbulb"></i> Đề Xuất</div>
<div style="color: #cbd5e1;">{solution}</div>
</div>
</div>
</div>
"""
    return card

evaluated_cntt = []
evaluated_qtkd = []
for item in evaluated_staff:
    classes_str = item['classes_list'].upper()
    if 'QTKD' in classes_str:
        evaluated_qtkd.append(item)
    else:
        evaluated_cntt.append(item)

markdown_content += """
---

## III. ĐÁNH GIÁ CHI TIẾT TỪNG NHÂN SỰ VÀ ĐỀ XUẤT CẢI THIỆN
"""

# Khối CNTT
markdown_content += "\n### 💻 Khối Công nghệ Thông tin (CNTT)\n"
html_cards_cntt = "<div class='staff-grid' style='display: grid; grid-template-columns: repeat(auto-fill, minmax(320px, 1fr)); gap: 20px; margin-top: 15px; margin-bottom: 30px;' markdown='1'>\n"
for item in sorted(evaluated_cntt, key=lambda x: x['cmi'], reverse=True):
    html_cards_cntt += generate_staff_card(item)
html_cards_cntt += "</div>\n"
markdown_content += html_cards_cntt

# Khối QTKD
markdown_content += "\n### 📊 Khối Quản trị Kinh doanh (QTKD)\n"
html_cards_qtkd = "<div class='staff-grid' style='display: grid; grid-template-columns: repeat(auto-fill, minmax(320px, 1fr)); gap: 20px; margin-top: 15px; margin-bottom: 30px;' markdown='1'>\n"
for item in sorted(evaluated_qtkd, key=lambda x: x['cmi'], reverse=True):
    html_cards_qtkd += generate_staff_card(item)
html_cards_qtkd += "</div>\n"
markdown_content += html_cards_qtkd


# Lưu báo cáo
with open(output_path, 'w', encoding='utf-8') as f:
    f.write(markdown_content)

print(f"Agent 1: Báo cáo đã được ghi đè thành công tại {output_path}")
wb.close()
