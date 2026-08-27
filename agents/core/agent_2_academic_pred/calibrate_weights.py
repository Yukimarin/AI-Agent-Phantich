import os
import sys
import json
import sqlite3
import mysql.connector
from collections import defaultdict
from statistics import mean

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from run import get_excel_chot_data, calibrate_students, get_course_difficulty_combined, normalize_class_name

sys.stdout.reconfigure(encoding='utf-8')

# Hàm map động tên môn học sang sheet Excel thông minh (hỗ trợ cả QTKD)
def find_target_sheet(coname, cname, sheetnames):
    low_course = coname.lower()
    low_class = cname.lower()
    
    # 1. Xử lý khối QTKD
    if "qtkd" in low_class:
        # Tìm mã môn học trong tên môn (ví dụ DTB202, PRJ302, BA201, M103, M104...)
        for code in ["dtb201", "dtb202", "prj302", "ba201", "m103", "m104", "skl"]:
            if code in low_course:
                for sheet in sheetnames:
                    if "qtkd" in sheet.lower() and code in sheet.lower():
                        return sheet
        # Fallback tìm kiếm tương đối cho QTKD
        for sheet in sheetnames:
            if "qtkd" in sheet.lower():
                # Tách từ khóa và so khớp
                words = [w for w in low_course.split() if len(w) > 2]
                if any(w in sheet.lower() for w in words):
                    return sheet
                    
    # 2. Xử lý khối CNTT
    course_to_sheet_map = {
        'javascript': 'KS25_Javascript',
        'cơ sở dữ liệu': 'KS25_Database',
        'database': 'KS25_Database',
        'python': 'KS25_Python',
        'java fundamental': 'KS24-JavaAdvance',
        'java advance': 'KS24-JavaAdvance',
        'java web application': 'KS24_JavaWeb',
        'java web service': 'KS24_JWS',
        'agile': 'KS24_JavaWeb',
        'trí tuệ': 'KS24_AI',
        'ai': 'KS24_AI',
        'python web': 'KS25_Python_Web',
        'fastapi': 'KS25_Python_Web',
        'dịch vụ web': 'KS25_Python_Web',
        'kỹ năng': 'SKL'
    }
    
    for kw, sheet in course_to_sheet_map.items():
        if kw in low_course:
            # Tìm sheet chính xác nhất trong Excel
            for s in sheetnames:
                if sheet.lower() in s.lower():
                    return s
                    
    # 3. Quét tương đối cuối cùng
    for s in sheetnames:
        if s.lower() in low_course or low_course in s.lower():
            return s
            
    return None

def run_query(cursor, query, params=None):
    cursor.execute(query, params or ())
    columns = [desc[0] for desc in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def predict_in_memory(preload, params):
    w1 = params['w1']
    w2 = params['w2']
    p_hack_mult = params['p_hack_mult']
    base_scale = params['base_scale']
    env_threshold = params['env_threshold']
    
    excel_disc = preload['excel_disc']
    students_db_data = preload['students_db_data']
    hackathon_map = preload['hackathon_map']
    student_att = preload['student_att']
    total_sessions = preload['total_sessions']
    has_hackathon = preload['has_hackathon']
    has_project = preload['has_project']
    prev_courses = preload['prev_courses']
    student_pass_history = preload['student_pass_history']
    auto_rp_map = preload['auto_rp_map']
    batch = preload['batch']
    
    # 1. Hiệu chỉnh học sinh
    calibrated_students = calibrate_students(students_db_data, excel_disc)
    
    # Peer Pressure Penalty
    mult_env = 1.0
    v_class = 0.0
    if excel_disc:
        v_class = (excel_disc['cc'] + excel_disc['bt'] + excel_disc['el']) / 3.0
        if v_class > env_threshold:
            mult_env = max(0.90, 1.0 - 0.5 * (v_class - env_threshold)/100.0)
            
    students_output = []
    
    for s in calibrated_students:
        sid = s['student_id']
        att_val = s['attendance']
        hw_val = s['homework']
        el_val = s['elearning']
        rp = s['rpoints']
        proj = s['project']
        actual_pass = s['pass']
        
        # Consecutive Penalties
        att_list = student_att.get(sid, [])
        consecutive_abs = 0
        for status in reversed(att_list):
            if status in ('0', '2'): consecutive_abs += 1
            else: break
        penalty_abs = 0.5 if consecutive_abs >= 2 else 1.0
        
        is_resumed_student = False
        if len(prev_courses) >= 1:
            has_hist = False
            for prev_c in prev_courses:
                prev_fr = student_pass_history.get((sid, prev_c))
                if prev_fr and prev_fr.get('pass') is not None:
                    has_hist = True
                    break
            if not has_hist:
                is_resumed_student = True
                
        # Discipline prev score
        discipline_prev = None
        if not is_resumed_student:
            prev_rps = []
            for prev_c in prev_courses:
                rpoint_val = None
                if batch == 'KS25':
                    rpoint_val = auto_rp_map[prev_c].get(sid)
                else:
                    prev_fr = student_pass_history.get((sid, prev_c))
                    if prev_fr and prev_fr.get('rpoints') is not None:
                        rpoint_val = prev_fr['rpoints']
                if rpoint_val is not None:
                    prev_rps.append(rpoint_val)
            if prev_rps:
                discipline_prev = mean(prev_rps)
                
        if is_resumed_student:
            discipline_prev = 70.0
        elif discipline_prev is None:
            discipline_prev = 100.0
            
        discipline_curr = max(0.0, 100.0 - att_val)
        discipline_val = 0.5 * discipline_prev + 0.5 * discipline_curr
        
        # Prior pass rate
        P_prereq = 75.0
        penalty_resumption = 1.0
        if is_resumed_student:
            P_prereq = 50.0
            penalty_resumption = 0.85
        elif len(prev_courses) >= 1:
            prev_c_main = prev_courses[0]
            prev_pass_status = student_pass_history.get((sid, prev_c_main))
            if prev_pass_status:
                P_prereq = 90.0 if prev_pass_status['pass'] == 1 else 10.0
                
        shack_val = s['hackathon_1'] if s['hackathon_1'] is not None else 65.0
        p_hack = min(100.0, shack_val * p_hack_mult)
        
        if has_hackathon:
            P_learning = w1 * P_prereq + w2 * (p_hack / 100.0)
        else:
            P_learning = 0.3 * P_prereq/100.0 + 0.7 * (p_hack / 100.0)
        P_learning = min(1.0, max(0.0, P_learning))
        
        p_eligible = P_learning * 0.6 + (discipline_val / 100.0) * 0.4
        p_eligible = p_eligible * penalty_abs * penalty_resumption * base_scale
        p_eligible = min(1.0, max(0.0, p_eligible)) * 100.0
        
        p_final = p_eligible * mult_env
        p_final = min(100.0, max(0.0, p_final))
        
        # Luật cấm thi mới
        is_cc_new = att_val <= 20.0
        is_bt_new = hw_val >= 80.0
        is_el_new = el_val <= 3.0
        is_rp_new = True if (excel_disc is None or excel_disc.get('rp') is None) else (rp >= 80.0)
        is_proj_new_ok = proj >= 50.0 if (has_project and proj is not None) else True
        
        if total_sessions <= 3:
            is_failed_new = False
        else:
            is_failed_new = not (is_cc_new and is_bt_new and is_el_new and is_rp_new and is_proj_new_ok)
            
        p_pred = 0.0 if is_failed_new else p_final
        pred_pass = 1 if p_pred >= 50.0 else 0
        
        # Nhãn thực tế giả lập theo quy chế mới (bị cấm thi theo quy chế mới -> chắc chắn trượt)
        actual_pass_sim = 0 if is_failed_new else actual_pass
        
        students_output.append({
            'student_id': sid,
            'pred_pass': pred_pass,
            'p_pred': p_pred,
            'actual_pass': actual_pass_sim
        })
        
    pred_pass_count = sum(1 for s in students_output if s['pred_pass'] == 1)
    actual_pass_count = sum(1 for s in students_output if s['actual_pass'] == 1)
    pred_rate = (pred_pass_count / len(students_output)) * 100.0 if students_output else 0.0
    actual_rate = (actual_pass_count / len(students_output)) * 100.0 if students_output else 0.0
    
    return {
        'students': students_output,
        'pred_rate': pred_rate,
        'actual_rate': actual_rate
    }

def main():
    print("================================================================================")
    print("KHỞI CHẠY TÌM BỘ THAM SỐ TỐI ƯU HÓA (KẾ THỪA NATIVE LOGIC - AGENT 2)")
    print("================================================================================")
    
    try:
        conn = mysql.connector.connect(
            host="localhost",
            port=3307,
            user="root",
            password="",
            database="qldt_el"
        )
        cursor = conn.cursor()
        print("✓ Connected to MySQL database on port 3307.")
    except Exception as e:
        print(f"✗ Lỗi kết nối MySQL: {e}")
        sys.exit(1)
        
    excel_path = "data/inputs/PTIT_Chiso.xlsx"
    print("Đang nạp dữ liệu Excel sử dụng native logic từ run.py...")
    excel_data = get_excel_chot_data(excel_path)
    
    # Lấy danh sách sheetnames của Excel để map động
    import openpyxl
    wb_temp = openpyxl.load_workbook(excel_path, read_only=True)
    sheetnames = wb_temp.sheetnames
    wb_temp.close()
    
    class_course_seq = defaultdict(list)
    seq_raw = run_query(cursor, "SELECT DISTINCT class_id, course_id FROM final_results ORDER BY class_id, course_id ASC;")
    for r in seq_raw:
        class_course_seq[int(r['class_id'])].append(int(r['course_id']))
        
    classes_raw = run_query(cursor, "SELECT id, name FROM classes;")
    classes_map = {int(c['id']): c['name'] for c in classes_raw}
    
    courses_raw = run_query(cursor, "SELECT id, name FROM courses;")
    courses_map = {int(c['id']): c['name'] for c in courses_raw}
    
    completed_runs = {
        'KS24': [],
        'KS25': [],
        'QTKD': []
    }
    
    for cid, cname in classes_map.items():
        is_qtkd = "QTKD" in cname
        is_ks25 = ("KS25" in cname or "K25" in cname) and not is_qtkd
        is_ks24 = ("KS24" in cname or "K24" in cname)
        
        if not (is_ks25 or is_ks24 or is_qtkd):
            continue
            
        batch = 'KS24' if is_ks24 else ('KS25' if is_ks25 else 'QTKD')
        
        for co_id in class_course_seq[cid]:
            coname = courses_map.get(co_id, "")
            check_raw = run_query(cursor, "SELECT COUNT(*) as cnt FROM final_results WHERE class_id = %s AND course_id = %s AND pass IS NOT NULL;", (cid, co_id))
            if check_raw and check_raw[0]['cnt'] > 0:
                completed_runs[batch].append({
                    'cid': cid,
                    'co_id': co_id,
                    'cname': cname,
                    'coname': coname
                })
                
    print("\nBắt đầu nạp trước (Preload) dữ liệu các lớp vào RAM...")
    preloaded_runs = {
        'KS24': [],
        'KS25': [],
        'QTKD': []
    }
    
    for batch in ['KS24', 'KS25', 'QTKD']:
        for run_info in completed_runs[batch]:
            cid = run_info['cid']
            co_id = run_info['co_id']
            cname = run_info['cname']
            coname = run_info['coname']
            
            # SELECT thêm cột hackathon_1 để lấy điểm thi thực tế lịch sử
            raw_st = run_query(cursor, """
                SELECT f.student_id, f.attendance, f.homework, f.elearning, f.rpoints, f.project, f.pass, f.hackathon_1, s.full_name
                FROM final_results f
                JOIN students s ON f.student_id = s.id
                WHERE f.class_id = %s AND f.course_id = %s;
            """, (cid, co_id))
            
            # Đọc Hackathon từ result_test làm fallback
            hackathon_map = {}
            hack_raw = run_query(cursor, """
                SELECT r.student_id, AVG(r.point) as avg_point
                FROM result_test r
                JOIN test_schedule ts ON r.test_schedule_id = ts.id
                WHERE ts.type = 'THI HACKATHON' AND ts.course_id = %s
                GROUP BY r.student_id;
            """, (co_id,))
            for h in hack_raw:
                hackathon_map[int(h['student_id'])] = float(h['avg_point']) if h['avg_point'] is not None else 65.0
                
            students_db_data = []
            for s in raw_st:
                sid = int(s['student_id'])
                db_hack = float(s['hackathon_1']) if s['hackathon_1'] is not None else None
                students_db_data.append({
                    'student_id': sid,
                    'attendance': float(s['attendance']) if s['attendance'] is not None else 0.0,
                    'homework': float(s['homework']) if s['homework'] is not None else 100.0,
                    'elearning': float(s['elearning']) if s['elearning'] is not None else 0.0,
                    'rpoints': float(s['rpoints']) if s['rpoints'] is not None else 100.0,
                    'project': float(s['project']) if s['project'] is not None else None,
                    'pass': int(s['pass']) if s['pass'] is not None else None,
                    'full_name': s['full_name'],
                    'hackathon_1': db_hack if db_hack is not None else hackathon_map.get(sid),
                    'hackathon_2': None
                })
                
            att_raw = run_query(cursor, """
                SELECT ad.student_id, ad.status
                FROM attendance_detail ad
                JOIN attendance a ON ad.attendance_id = a.id
                WHERE a.classes_id = %s AND a.courses_id = %s
                ORDER BY a.date ASC, a.id ASC;
            """, (cid, co_id))
            student_att = defaultdict(list)
            for row in att_raw:
                student_att[int(row['student_id'])].append(str(row['status']))
                
            num_sessions_db = 0
            sessions_raw = run_query(cursor, "SELECT COUNT(DISTINCT id) as cnt FROM attendance WHERE classes_id = %s AND courses_id = %s;", (cid, co_id))
            if sessions_raw:
                num_sessions_db = sessions_raw[0]['cnt']
                
            target_sheet = find_target_sheet(coname, cname, sheetnames)
            excel_disc = excel_data.get(normalize_class_name(cname), {}).get(target_sheet) if target_sheet else None
            
            num_sessions_excel = excel_disc.get('num_sessions', 0) if excel_disc else 0
            total_sessions = max(num_sessions_db, num_sessions_excel)
            
            has_hackathon = len(hackathon_map) > 0 or any(s['hackathon_1'] is not None for s in students_db_data)
            has_project = any(s['project'] is not None and s['project'] > 5.0 for s in students_db_data)
            
            seq = class_course_seq.get(cid, [])
            prev_courses = []
            if co_id in seq:
                idx = seq.index(co_id)
                if idx >= 1: prev_courses.append(seq[idx - 1])
                if idx >= 2: prev_courses.append(seq[idx - 2])
            prev_courses = prev_courses[:2]
            
            auto_rp_map = defaultdict(lambda: defaultdict(float))
            student_pass_history = {}
            sids = [s['student_id'] for s in students_db_data]
            
            if sids:
                placeholders = ", ".join(["%s"] * len(sids))
                for prev_c in prev_courses:
                    hist_raw = run_query(cursor, f"SELECT student_id, pass, rpoints FROM final_results WHERE course_id = %s AND student_id IN ({placeholders});", [prev_c] + sids)
                    for h in hist_raw:
                        student_pass_history[(int(h['student_id']), prev_c)] = {
                            'pass': int(h['pass']) if h['pass'] is not None else None,
                            'rpoints': float(h['rpoints']) if h['rpoints'] is not None else None
                        }
                    auto_rp_raw = run_query(cursor, f"SELECT student_id, total_score, recorded_date FROM auto_rpoints WHERE course_id = %s AND student_id IN ({placeholders});", [prev_c] + sids)
                    latest_dates = {}
                    for r in auto_rp_raw:
                        sid_raw = int(r['student_id'])
                        score = float(r['total_score'])
                        rdate = r['recorded_date']
                        if sid_raw not in latest_dates or rdate > latest_dates[sid_raw]:
                            latest_dates[sid_raw] = rdate
                            auto_rp_map[prev_c][sid_raw] = score
                            
            preloaded_runs[batch].append({
                'cid': cid,
                'co_id': co_id,
                'cname': cname,
                'coname': coname,
                'excel_disc': excel_disc,
                'students_db_data': students_db_data,
                'hackathon_map': hackathon_map,
                'student_att': student_att,
                'total_sessions': total_sessions,
                'has_hackathon': has_hackathon,
                'has_project': has_project,
                'prev_courses': prev_courses,
                'student_pass_history': student_pass_history,
                'auto_rp_map': auto_rp_map,
                'batch': batch
            })
            
    print("✓ Đã preload xong dữ liệu lịch sử vào RAM!")
    
    # 5. Chạy Grid Search In-Memory
    w1_grid = [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6]
    p_hack_mult_grid = [1.0, 1.1, 1.2, 1.25, 1.3, 1.4, 1.5]
    base_scale_grid = [0.8, 0.85, 0.9, 0.95, 1.0, 1.05, 1.1]
    env_threshold_grid = [5.0, 7.5, 10.0, 12.5, 15.0]
    
    best_params = {}
    
    for batch in ['KS24', 'KS25', 'QTKD']:
        preloads = preloaded_runs[batch]
        if not preloads:
            print(f"\nKhối {batch} không có dữ liệu huấn luyện mẫu.")
            best_params[batch] = {
                'w1': 0.40 if batch == 'KS24' else 0.0,
                'w2': 0.60 if batch == 'KS24' else 1.0,
                'p_hack_mult': 1.25,
                'base_scale': 1.0,
                'env_threshold': 10.0
            }
            continue
            
        print(f"\nĐang chạy Grid Search tối ưu hóa (RAM Mode) cho Khối {batch}...")
        
        min_loss = float('inf')
        optimal_set = None
        
        for w1 in w1_grid:
            w2 = 1.0 - w1
            for p_hack_mult in p_hack_mult_grid:
                for base_scale in base_scale_grid:
                    for env_threshold in env_threshold_grid:
                        params = {
                            'w1': w1,
                            'w2': w2,
                            'p_hack_mult': p_hack_mult,
                            'base_scale': base_scale,
                            'env_threshold': env_threshold
                        }
                        
                        total_students = 0
                        incorrect_predictions = 0
                        abs_errors = []
                        
                        for preload in preloads:
                            res = predict_in_memory(preload, params)
                            incorrect_predictions += sum(1 for s in res['students'] if s['pred_pass'] != s['actual_pass'])
                            total_students += len(res['students'])
                            abs_errors.append(abs(res['pred_rate'] - res['actual_rate']))
                            
                        if total_students == 0:
                            continue
                            
                        misclassification_rate = (incorrect_predictions / total_students) * 100.0
                        class_mae = mean(abs_errors) if abs_errors else 0.0
                        
                        loss = 0.6 * misclassification_rate + 0.4 * class_mae
                        
                        if loss < min_loss:
                            min_loss = loss
                            optimal_set = {
                                'params': params,
                                'loss': loss,
                                'misclass_rate': misclassification_rate,
                                'class_mae': class_mae
                            }
                            
        if optimal_set:
            best_params[batch] = optimal_set['params']
            print(f"✓ BỘ THAM SỐ TỐI ƯU KHỐI {batch}:")
            print(f"  - Tham số: {optimal_set['params']}")
            print(f"  - Sai lệch Loss tổng hợp: {optimal_set['loss']:.2f}%")
            print(f"  - Tỷ lệ dự báo sai sinh viên: {optimal_set['misclass_rate']:.2f}%")
            print(f"  - Sai số tỷ lệ đỗ lớp (MAE): {optimal_set['class_mae']:.2f}%")
        else:
            print(f"✗ Không tìm thấy bộ tham số tối ưu cho Khối {batch}. Dùng mặc định.")
            
    metadata_path = "data/inputs/course_metadata.json"
    metadata = {}
    if os.path.exists(metadata_path):
        try:
            with open(metadata_path, 'r', encoding='utf-8') as f:
                metadata = json.load(f)
        except Exception as e:
            print(f"Warning: Lỗi đọc file metadata hiện tại: {e}")
            
    metadata['calibration_params'] = best_params
    
    try:
        with open(metadata_path, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)
        print(f"\n✓ Hoàn tất! Đã cập nhật và đồng bộ tham số tối ưu mới vào file: {metadata_path}")
    except Exception as e:
        print(f"✗ Lỗi ghi file metadata: {e}")
        
    conn.close()

if __name__ == "__main__":
    main()
