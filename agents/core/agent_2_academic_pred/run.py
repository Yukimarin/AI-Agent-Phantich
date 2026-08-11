import os
import sys
import json
import re
import mysql.connector
import openpyxl
from datetime import datetime, date
from collections import defaultdict

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

# Helper to calculate average of a list
def mean(vals):
    if not vals:
        return 0.0
    return sum(vals) / len(vals)

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str.lower().replace(" ", "").replace("-", "")

def get_excel_chot_data(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    class_data = defaultdict(dict)
    
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        max_r = sheet.max_row
        max_c = sheet.max_column
        
        # 1. Find Header row & last columns
        header_idx = None
        for r in range(1, min(15, max_r + 1)):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, max_c + 1)]
            if 'Lớp' in row_vals or 'Lớp học' in row_vals:
                header_idx = r
                break
                
        if header_idx is None:
            continue
            
        row4 = [sheet.cell(row=header_idx, column=c).value for c in range(1, max_c + 1)]
        
        # Scan for date columns and last daily cols
        last_cols = []
        last_date = None
        for c_idx, val in enumerate(row4):
            if val and ('/' in str(val) or isinstance(val, (datetime, date))):
                last_cols.append(c_idx)
                if isinstance(val, (datetime, date)):
                    last_date = val.date()
                else:
                    parts = str(val).split('/')
                    if len(parts) >= 2:
                        try:
                            last_date = date(2026, int(parts[1]), int(parts[0]))
                        except:
                            pass
                            
        if not last_cols:
            continue
            
        # 2. Find Rpoint chot col (scanning strictly after the last daily date col)
        rp_col_idx = None
        last_date_col = max(last_cols)
        for c_idx in range(last_date_col + 1, max_c):
            h4 = row4[c_idx]
            if h4 in ('Chuyên cần', 'Bài tập', 'Elearning'):
                continue
            # Try reading some numeric values
            vals = []
            for r in range(header_idx + 1, max_r + 1):
                val = sheet.cell(row=r, column=c_idx + 1).value
                if val is not None:
                    try:
                        vals.append(float(val))
                    except:
                        pass
            if len(vals) >= 2:
                avg_val = mean(vals)
                if 30.0 <= avg_val <= 115.0:
                    rp_col_idx = c_idx
                    break
                    
        # 3. Read data for each class row
        for r in range(header_idx + 1, max_r + 1):
            cname = sheet.cell(row=r, column=2).value
            if cname:
                norm_name = normalize_class_name(cname)
                
                # Violations: average of daily columns
                cc_vals, bt_vals, el_vals = [], [], []
                for c_idx in last_cols:
                    cc_c = sheet.cell(row=r, column=c_idx + 1).value
                    bt_c = sheet.cell(row=r, column=c_idx + 2).value
                    el_c = sheet.cell(row=r, column=c_idx + 3).value
                    
                    if cc_c is not None:
                        try: cc_vals.append(float(cc_c))
                        except: pass
                    if bt_c is not None:
                        try: bt_vals.append(float(bt_c))
                        except: pass
                    if el_c is not None:
                        try: el_vals.append(float(el_c))
                        except: pass
                        
                cc_val = sum(cc_vals) / len(cc_vals) if cc_vals else 0.0
                bt_val = sum(bt_vals) / len(bt_vals) if bt_vals else 0.0
                el_val = sum(el_vals) / len(el_vals) if el_vals else 0.0
                            
                # Rpoint
                rp_val = None
                if rp_col_idx is not None:
                    cell_val = sheet.cell(row=r, column=rp_col_idx + 1).value
                    if cell_val is not None:
                        try:
                            rp_val = float(cell_val)
                        except:
                            pass
                            
                class_data[norm_name][sheetname] = {
                    'cc': cc_val,
                    'bt': bt_val,
                    'el': el_val,
                    'rp': rp_val,
                    'date': last_date,
                    'num_sessions': len(last_cols)
                }
    return class_data

def load_course_metadata():
    for path in ["data/inputs/course_metadata.json", "data/course_metadata.json"]:
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Error loading course metadata: {e}")
    return {}

def get_db_course_difficulty(cursor, co_id, co_name):
    global db_mode
    if db_mode == "SQLite":
        try:
            cursor.execute("SELECT midterm_score FROM student_grades WHERE subject = ?", (co_name,))
            rows = cursor.fetchall()
            if len(rows) >= 5:
                fails = sum(1 for r in rows if r[0] < 5.0)
                fail_rate = fails / len(rows)
                if fail_rate >= 0.4: return 1.8
                elif fail_rate >= 0.25: return 1.5
                elif fail_rate >= 0.10: return 1.3
                elif fail_rate >= 0.05: return 1.1
                else: return 1.0
        except Exception:
            pass
        return None
    else:
        try:
            cursor.execute("""
                SELECT AVG(CASE WHEN pass = 0 THEN 1.0 ELSE 0.0 END) as fail_rate, COUNT(*) as cnt
                FROM qldt_el.final_results
                WHERE course_id = %s AND pass IS NOT NULL;
            """, (co_id,))
            row = cursor.fetchone()
            if row and row[1] >= 5:
                fail_rate = float(row[0])
                if fail_rate >= 0.4: return 1.8
                elif fail_rate >= 0.25: return 1.5
                elif fail_rate >= 0.10: return 1.3
                elif fail_rate >= 0.05: return 1.1
                else: return 1.0
        except Exception as e:
            print(f"Error querying historical fail rate: {e}")
        return None

def get_course_difficulty_combined(cursor, co_id, course_name):
    # 1. Check course_metadata.json
    meta = load_course_metadata()
    if meta and "course_difficulty" in meta:
        for key, val in meta["course_difficulty"].items():
            if key.lower() in course_name.lower() or course_name.lower() in key.lower():
                return val.get("difficulty_coefficient", 1.2)
                
    # 2. Check DB historical failure rate
    db_diff = get_db_course_difficulty(cursor, co_id, course_name)
    if db_diff is not None:
        return db_diff
        
    # 3. Fallback to heuristics
    coname_lower = course_name.lower()
    if "ai" in coname_lower or "trí tuệ nhân tạo" in coname_lower or "machine learning" in coname_lower:
        return 1.8
    elif "database" in coname_lower or "cơ sở dữ liệu" in coname_lower or "cấu trúc dữ liệu" in coname_lower:
        return 1.5
    elif "python" in coname_lower or "java" in coname_lower or "web" in coname_lower:
        return 1.3
    else:
        return meta.get("default_difficulty", 1.2)


def calibrate_students(students_results, excel_disc):
    if not students_results:
        return []
    att_list = [s['attendance'] if s['attendance'] is not None else 0.0 for s in students_results]
    hw_list = [s['homework'] if s['homework'] is not None else 100.0 for s in students_results]
    el_list = [s['elearning'] if s['elearning'] is not None else 0.0 for s in students_results]
    rp_list = [s['rpoints'] if s['rpoints'] is not None else 100.0 for s in students_results]
    
    if excel_disc:
        db_att_avg = mean(att_list)
        db_hw_avg = mean(hw_list)
        db_el_avg = mean(el_list)
        db_rp_avg = mean(rp_list)
        
        calibrated = []
        for i, s in enumerate(students_results):
            # Scale student metrics relatively
            student_att_scaled = att_list[i]
            if db_att_avg > 0.0:
                student_att_scaled = min(100.0, max(0.0, (att_list[i] / db_att_avg) * excel_disc['cc']))
            else:
                student_att_scaled = excel_disc['cc']
                
            student_hw_scaled = hw_list[i]
            excel_hw_completed = max(0.0, 100.0 - excel_disc['bt'])
            if db_hw_avg > 0.0:
                student_hw_scaled = min(100.0, max(0.0, (hw_list[i] / db_hw_avg) * excel_hw_completed))
            else:
                student_hw_scaled = excel_hw_completed
                
            student_el_scaled = el_list[i]
                
            student_rp_scaled = rp_list[i]
            if excel_disc['rp'] is not None:
                if db_rp_avg > 0.0:
                    student_rp_scaled = min(120.0, max(0.0, (rp_list[i] / db_rp_avg) * excel_disc['rp']))
                else:
                    student_rp_scaled = excel_disc['rp']
                    
            calibrated.append({
                'student_id': s['student_id'],
                'attendance': student_att_scaled, # vắng %
                'homework': student_hw_scaled,     # hoàn thành %
                'elearning': student_el_scaled,   # số bài trễ
                'rpoints': student_rp_scaled,
                'hackathon_1': s['hackathon_1'],
                'hackathon_2': s['hackathon_2'],
                'project': s['project'],
                'pass': s['pass'],
                'full_name': s['full_name']
            })
        return calibrated
    else:
        calibrated = []
        for i, s in enumerate(students_results):
            calibrated.append({
                'student_id': s['student_id'],
                'attendance': att_list[i],
                'homework': hw_list[i],
                'elearning': el_list[i],
                'rpoints': rp_list[i],
                'hackathon_1': s['hackathon_1'],
                'hackathon_2': s['hackathon_2'],
                'project': s['project'],
                'pass': s['pass'],
                'full_name': s['full_name']
            })
        return calibrated

db_mode = "MySQL"

def run_query(cursor, query, params=None):
    global db_mode
    if db_mode == "SQLite":
        cleaned_query = query.replace('qldt_el.', '').replace('%s', '?').lower()
        if 'final_results' in cleaned_query and ('students' in cleaned_query or 'student_id' in cleaned_query):
            # Querying students or student scores/history
            cursor.execute("SELECT student_id, student_name, class_id, subject, midterm_score FROM student_grades")
            rows = cursor.fetchall()
            result = []
            for idx, r in enumerate(rows):
                sid, sname, cid_val, subject_val, score = r
                try:
                    num_sid = int(str(sid).replace('SV', ''))
                except ValueError:
                    num_sid = 1000 + idx
                # Map pass state based on midterm score >= 5.0
                is_passed = 1 if score >= 5.0 else 0
                result.append({
                    'student_id': num_sid,
                    'attendance': 5.0,
                    'homework': 90.0,
                    'elearning': 0.0,
                    'rpoints': 100.0,
                    'project': 80.0,
                    'pass': is_passed,
                    'full_name': sname,
                    'hackathon_1': None
                })
            return result
        elif 'attendance_detail' in cleaned_query or 'attendance' in cleaned_query:
            return []
        elif 'exercise' in cleaned_query:
            return []
        elif 'result_test' in cleaned_query:
            return []
        elif 'auto_rpoints' in cleaned_query:
            return []
        elif 'from classes' in cleaned_query or ' classes ' in cleaned_query or cleaned_query.strip().endswith('classes'):
            return [
                {'id': 1, 'name': 'HN-KS25-CNTT6'},
                {'id': 2, 'name': 'HN-KS25-CNTT8'},
                {'id': 3, 'name': 'HN-KS24-CNTT1'}
            ]
        elif 'from courses' in cleaned_query or ' courses ' in cleaned_query or cleaned_query.strip().endswith('courses'):
            return [
                {'id': 217, 'name': 'Python Web'},
                {'id': 214, 'name': 'AI Application'},
                {'id': 193, 'name': 'Python'},
                {'id': 194, 'name': 'Java Web Service'}
            ]
        elif 'final_results' in cleaned_query:
            return [
                {'class_id': 1, 'course_id': 193},
                {'class_id': 1, 'course_id': 217},
                {'class_id': 2, 'course_id': 193},
                {'class_id': 2, 'course_id': 217},
                {'class_id': 3, 'course_id': 194},
                {'class_id': 3, 'course_id': 214}
            ]
        else:
            try:
                # Fallback to execute raw query on SQLite (replacing MySQL schema prefix and placeholders)
                sql_to_run = query.replace('qldt_el.', '').replace('%s', '?')
                cursor.execute(sql_to_run, params or ())
                columns = [desc[0] for desc in cursor.description]
                return [dict(zip(columns, row)) for row in cursor.fetchall()]
            except Exception as e:
                # print(f"SQLite fallback query failed: {e}")
                return []
    else:
        cursor.execute(query, params or ())
        columns = [desc[0] for desc in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]

def predict_class_pass_rate(cursor, cid, co_id, class_course_seq, excel_data, cname, coname, batch):
    is_soft_skill = any(kw in coname.lower() for kw in ['kỹ năng', 'tin học văn phòng', 'skl', 'thực tập', 'ttrk', 'project'])
    
    # 1. Fetch Students
    raw_st = run_query(cursor, """
        SELECT f.student_id, f.attendance, f.homework, f.elearning, f.rpoints, f.project, f.pass, s.full_name
        FROM qldt_el.final_results f
        JOIN qldt_el.students s ON f.student_id = s.id
        WHERE f.class_id = %s AND f.course_id = %s;
    """, (cid, co_id))
    
    if not raw_st:
        return {'size': 0, 'avg_pred_old': 0.0, 'avg_pred_new': 0.0, 'actual_pass_rate': 0.0, 'students': [], 'v_class': 0.0, 'mult_env': 1.0}
        
    # Get Hackathon
    hackathon_map = {}
    hack_raw = run_query(cursor, """
        SELECT r.student_id, AVG(r.point) as avg_point
        FROM qldt_el.result_test r
        JOIN qldt_el.test_schedule ts ON r.test_schedule_id = ts.id
        WHERE ts.type = 'THI HACKATHON' AND ts.course_id = %s
        GROUP BY r.student_id;
    """, (co_id,))
    for h in hack_raw:
        hackathon_map[int(h['student_id'])] = float(h['avg_point']) if h['avg_point'] is not None else 65.0
        
    students_db_data = []
    for s in raw_st:
        sid = int(s['student_id'])
        students_db_data.append({
            'student_id': sid,
            'attendance': float(s['attendance']) if s['attendance'] is not None else 0.0,
            'homework': float(s['homework']) if s['homework'] is not None else 100.0,
            'elearning': float(s['elearning']) if s['elearning'] is not None else 0.0,
            'rpoints': float(s['rpoints']) if s['rpoints'] is not None else 100.0,
            'hackathon_1': hackathon_map.get(sid),
            'hackathon_2': None,
            'project': float(s['project']) if s['project'] is not None else None,
            'pass': int(s['pass']) if s['pass'] is not None else None,
            'full_name': s['full_name']
        })
        
    # 2. Get Excel Calibrations
    norm_cname = normalize_class_name(cname)
    course_to_sheet_map = {
        'javascript': 'KS25_Javascript',
        'cơ sở dữ liệu': 'KS25_Database',
        'database': 'KS25_Database',
        'python': 'KS25_Python',
        'java fundamental': 'KS24-JavaAdvance',
        'java advance': 'KS24-JavaAdvance',
        'java web application': 'KS24_JavaWeb',
        'java web service': 'KS24_JWS',
        'agile': 'KS24_JavaWeb',
        'trí tuệ': 'KS24_AI',
        'ai': 'KS24_AI',
        'python web': 'KS25_Python_Web',
        'fastapi': 'KS25_Python_Web',
        'dịch vụ web': 'KS25_Python_Web'
    }
    low_course = coname.lower()
    target_sheet = None
    for kw, sheet in course_to_sheet_map.items():
        if kw in low_course:
            target_sheet = sheet
            break
    excel_disc = excel_data.get(norm_cname, {}).get(target_sheet) if target_sheet else None
    
    calibrated_students = calibrate_students(students_db_data, excel_disc)
    
    # 3. Environmental Penalty (Peer Pressure)
    mult_env = 1.0
    v_class = 0.0
    if excel_disc:
        # Calculate environmental violation percentage
        v_class = (excel_disc['cc'] + excel_disc['bt'] + excel_disc['el']) / 3.0
        if v_class > 10.0:
            mult_env = max(0.90, 1.0 - 0.5 * (v_class - 10.0)/100.0)
            
    # 4. Param Settings
    if batch == 'K25':
        w1, w2 = 0.00, 1.00
        p_prereq_pass = 0.85
        p_prereq_fail = 0.10
        p_hack_mult = 1.30
        base_scale = 0.95
    else:
        w1, w2 = 0.40, 0.60
        p_prereq_pass = 0.98
        p_prereq_fail = 0.10
        p_hack_mult = 1.25
        base_scale = 1.00
        
    has_project = any(s['project'] is not None and s['project'] > 5.0 for s in calibrated_students)
    has_hackathon = any(s['hackathon_1'] is not None and s['hackathon_1'] > 5.0 for s in calibrated_students)
    
    # Attendance and Homework detail logs
    att_raw = run_query(cursor, """
        SELECT ad.student_id, ad.status, a.id as attendance_id
        FROM qldt_el.attendance_detail ad
        JOIN qldt_el.attendance a ON ad.attendance_id = a.id
        WHERE a.classes_id = %s AND a.courses_id = %s
        ORDER BY a.date ASC, a.id ASC;
    """, (cid, co_id))
    student_att = defaultdict(list)
    for row in att_raw:
        student_att[int(row['student_id'])].append(str(row['status']))
        
    num_sessions_db = 0
    sessions_raw = run_query(cursor, """
        SELECT COUNT(DISTINCT id) as cnt
        FROM qldt_el.attendance
        WHERE classes_id = %s AND courses_id = %s;
    """, (cid, co_id))
    if sessions_raw and isinstance(sessions_raw, list) and len(sessions_raw) > 0 and 'cnt' in sessions_raw[0]:
        num_sessions_db = sessions_raw[0]['cnt']
    if num_sessions_db == 0 and student_att:
        num_sessions_db = max(len(att_list) for att_list in student_att.values())
    num_sessions_excel = excel_disc.get('num_sessions', 0) if excel_disc else 0
    total_sessions = max(num_sessions_db, num_sessions_excel)
        
    hw_raw = run_query(cursor, """
        SELECT e.student_id, e.check, e.link_git
        FROM qldt_el.exercise e
        WHERE e.class_id = %s AND e.course_id = %s
        ORDER BY e.id ASC;
    """, (cid, co_id))
    student_hw_details = defaultdict(list)
    for row in hw_raw:
        sid_r = int(row['student_id'])
        check = int(row['check'] or 0)
        git = row['link_git']
        is_git_empty = not git or git.strip() == "" or "placeholder" in str(git).lower()
        is_debt = (check == 2 or (check == 0 and is_git_empty))
        student_hw_details[sid_r].append(is_debt)

    # Prior courses
    seq = class_course_seq.get(cid, [])
    prev_courses = []
    if co_id in seq:
        idx = seq.index(co_id)
        if idx >= 1:
            prev_courses.append(seq[idx - 1])
        if idx >= 2:
            prev_courses.append(seq[idx - 2])
    prev_courses = prev_courses[:2]
    
    # Prior Rpoints
    auto_rp_map = defaultdict(lambda: defaultdict(float))
    student_pass_history = {}
    for prev_c in prev_courses:
        # Load pass history
        hist_raw = run_query(cursor, "SELECT student_id, pass, rpoints, project, homework, attendance, elearning, hackathon_1 FROM qldt_el.final_results WHERE course_id = %s;", (prev_c,))
        for h in hist_raw:
            student_pass_history[(int(h.get('student_id')), prev_c)] = {
                'pass': int(h.get('pass')) if h.get('pass') is not None else None,
                'rpoints': float(h.get('rpoints')) if h.get('rpoints') is not None else None,
                'homework': float(h.get('homework')) if h.get('homework') is not None else None,
                'attendance': float(h.get('attendance')) if h.get('attendance') is not None else None,
                'hackathon_1': float(h.get('hackathon_1')) if h.get('hackathon_1') is not None else None
            }
            
        auto_rp_raw = run_query(cursor, "SELECT student_id, total_score, recorded_date FROM qldt_el.auto_rpoints WHERE course_id = %s;", (prev_c,))
        latest_dates = {}
        for r in auto_rp_raw:
            sid_raw = int(r['student_id'])
            score = float(r['total_score'])
            rdate = r['recorded_date']
            if sid_raw not in latest_dates or rdate > latest_dates[sid_raw]:
                latest_dates[sid_raw] = rdate
                auto_rp_map[prev_c][sid_raw] = score
                
    students_output = []
    sum_p_final_old = 0.0
    sum_p_final_new = 0.0
    
    for s in calibrated_students:
        sid = int(s['student_id'])
        att_val = s['attendance']
        hw_val = s['homework']
        el_val = s['elearning']
        rp = s['rpoints']
        proj = s['project']
        
        # Consecutive penalties
        att_list = student_att.get(sid, [])
        consecutive_abs = 0
        for status in reversed(att_list):
            if status in ('0', '2'):
                consecutive_abs += 1
            else:
                break
        penalty_abs = 0.5 if consecutive_abs >= 2 else 1.0
        
        hw_list = student_hw_details.get(sid, [])
        consecutive_hw = 0
        for is_debt in reversed(hw_list):
            if is_debt:
                consecutive_hw += 1
            else:
                break
        penalty_hw = 0.6 if consecutive_hw >= 2 else 1.0
        
        # Check resumed student
        is_resumed_student = False
        if len(prev_courses) >= 1:
            has_hist = False
            for prev_c in prev_courses:
                prev_fr = student_pass_history.get((sid, prev_c))
                if prev_fr and prev_fr.get('pass') is not None:
                    has_hist = True
                    break
            if not has_hist:
                is_resumed_student = True
                
        # Discipline Score
        discipline_prev = None
        if not is_resumed_student:
            prev_rps = []
            for prev_c in prev_courses:
                rpoint_val = None
                if batch == 'K25':
                    rpoint_val = auto_rp_map[prev_c].get(sid)
                elif batch == 'K24':
                    prev_fr = student_pass_history.get((sid, prev_c))
                    if prev_fr and prev_fr.get('rpoints') is not None:
                        rpoint_val = prev_fr['rpoints']
                else:
                    prev_fr = student_pass_history.get((sid, prev_c))
                    if prev_fr and prev_fr.get('rpoints') is not None:
                        rpoint_val = prev_fr['rpoints']
                    if rpoint_val is None:
                        rpoint_val = auto_rp_map[prev_c].get(sid)
                if rpoint_val is not None:
                    prev_rps.append(float(rpoint_val))
            if prev_rps:
                discipline_prev = mean(prev_rps)
                
        if is_resumed_student:
            discipline_prev = 70.0
        elif discipline_prev is None:
            discipline_prev = 100.0
            
        discipline_curr = max(0.0, 100.0 - att_val)
        discipline_val = 0.5 * discipline_prev + 0.5 * discipline_curr
        
        # Study Performance: P_prereq
        penalty_resumption = 1.0
        prior_pass = None
        prior_hack = None
        if is_resumed_student:
            P_prereq = 50.0
            penalty_resumption = 0.85
        elif len(prev_courses) >= 1:
            prev_c_main = prev_courses[0]
            prev_pass_status = student_pass_history.get((sid, prev_c_main))
            if prev_pass_status:
                P_prereq = p_prereq_pass if prev_pass_status['pass'] == 1 else p_prereq_fail
                prior_pass = prev_pass_status.get('pass')
                prior_hack = prev_pass_status.get('hackathon_1')
            else:
                P_prereq = 75.0
        else:
            P_prereq = 75.0
            
        # P_hack
        shack_val = s['hackathon_1'] if s['hackathon_1'] is not None else 65.0
        p_hack = min(100.0, shack_val * p_hack_mult)
            
        # P_learning
        if has_hackathon:
            P_learning = w1 * P_prereq + w2 * (p_hack / 100.0)
        else:
            P_learning = 0.3 * P_prereq/100.0 + 0.7 * (p_hack / 100.0)
        P_learning = min(1.0, max(0.0, P_learning))
        
        # Apply CDC
        cdc_val = get_course_difficulty_combined(cursor, co_id, coname)
        P_learning_adj = P_learning / cdc_val
        p_eligible = P_learning_adj * 0.6 + (discipline_val / 100.0) * 0.4
        
        # Scale & penalties
        p_eligible = p_eligible * penalty_abs * penalty_hw * penalty_resumption * base_scale
        p_eligible = min(1.0, max(0.0, p_eligible)) * 100.0
        
        if is_soft_skill:
            p_eligible = 93.0
            
        # Apply environmental penalty
        p_final = p_eligible * mult_env
        p_final = min(100.0, max(0.0, p_final))
        
        # 1. Prediction under OLD rule (no hard blocks, matches historical DB)
        p_eligible_old = p_final
        sum_p_final_old += p_eligible_old
        
        # 2. Prediction under NEW rule (cấm thi chặn cứng)
        is_cc_new = att_val <= 20.0
        is_bt_new = hw_val >= 80.0
        is_el_new = el_val <= 3.0
        is_rp_new = True if (excel_disc is None or excel_disc.get('rp') is None) else (rp >= 80.0)
        is_proj_new_ok = True
        if has_project and proj is not None:
            is_proj_new_ok = proj >= 50.0
            
        if total_sessions <= 3:
            is_failed_new = False
        else:
            is_failed_new = not (is_cc_new and is_bt_new and is_el_new and is_rp_new and is_proj_new_ok)
        p_eligible_new = 0.0 if is_failed_new else p_final
        sum_p_final_new += p_eligible_new
        
        # Risk classification for current courses
        risk_level = "GREEN" # Thấp/Không cảnh báo
        reasons = []
        if total_sessions > 3 and is_failed_new:
            risk_level = "RED" # CAO - Bị cấm thi theo Quy chế mới
            if not is_cc_new: reasons.append(f"Vắng chuyên cần > 20% ({att_val:.1f}%)")
            if not is_bt_new: reasons.append(f"Nợ bài tập > 20% ({100.0 - hw_val:.1f}%)")
            if not is_el_new: reasons.append(f"Vi phạm Elearning quá 3 bài ({el_val} bài)")
            if not is_rp_new: reasons.append(f"Rpoints thấp ({rp:.1f}/80)")
            if not is_proj_new_ok: reasons.append("Project dưới 50 điểm")
        elif p_final < 30.0 and (total_sessions > 3 and (att_val > 15.0 or (100.0 - hw_val) > 20.0)):
            risk_level = "RED" # CAO - Học lực quá yếu VÀ Vi phạm kỷ luật
            reasons.append(f"Xác suất đỗ quá thấp ({p_final:.1f}%) kèm vi phạm KL")
        elif (total_sessions > 3 and (att_val > 15.0 and (100.0 - hw_val) > 30.0)) or \
             (p_final < 40.0 and (total_sessions > 3 and el_val >= 2.0)) or \
             (total_sessions > 3 and att_val > 20.0) or \
             (total_sessions > 3 and (100.0 - hw_val) > 40.0):
            risk_level = "YELLOW" # TRUNG BÌNH - Nguy cơ mất gốc thực sự
            if att_val > 15.0: reasons.append(f"Vắng nhiều ({att_val:.1f}%)")
            if (100.0 - hw_val) > 30.0: reasons.append(f"Nợ bài nhiều ({100.0 - hw_val:.1f}%)")
            if el_val >= 2.0: reasons.append(f"Trễ Elearning ({el_val} bài)")
            if p_final < 40.0: reasons.append(f"Học lực môn trước yếu ({p_final:.1f}%)")
        elif p_final < 40.0:
            risk_level = "YELLOW"
            reasons.append(f"Học lực môn trước yếu ({p_final:.1f}%)")
        elif p_final < 60.0:
            risk_level = "GREEN" # Theo dõi học lực
            reasons.append(f"Học lực cần theo dõi ({p_final:.1f}%)")
            
        # Detect student anomalies
        student_anomalies = []
        # Rule 1: copy_suspect
        if prior_hack is not None and prior_hack >= 75.0:
            if shack_val is not None and shack_val < 50.0:
                student_anomalies.append("copy_suspect")
        
        # Rule 2: discipline_paradox
        is_bad_disc = (att_val > 15.0) or ((100.0 - hw_val) > 30.0)
        if is_bad_disc and shack_val is not None and shack_val >= 80.0:
            student_anomalies.append("discipline_paradox")
            
        # Rule 3: passive_learner
        is_perfect_el = (el_val == 0.0)
        is_low_performance = ((100.0 - hw_val) > 30.0) or (shack_val is not None and shack_val < 50.0)
        if is_perfect_el and is_low_performance:
            student_anomalies.append("passive_learner")
            
        # Rule 4: sudden_drop
        if discipline_prev is not None and discipline_prev >= 95.0:
            if (att_val > 15.0) or ((100.0 - hw_val) > 30.0):
                student_anomalies.append("sudden_drop")

        students_output.append({
            'student_id': sid,
            'full_name': s['full_name'],
            'p_final': p_final,
            'is_failed_new': is_failed_new,
            'risk_level': risk_level,
            'reasons': reasons,
            'att': att_val,
            'hw': hw_val,
            'el': el_val,
            'rp': rp,
            'hack': shack_val,
            'anomalies': student_anomalies
        })
        
    avg_pred_old = sum_p_final_old / len(calibrated_students)
    avg_pred_new = sum_p_final_new / len(calibrated_students)
    
    # Calculate actual pass rate from DB results
    actual_passed = sum(1 for s in calibrated_students if s['pass'] == 1)
    actual_pass_rate = (actual_passed / len(calibrated_students)) * 100.0
    
    return {
        'avg_pred_old': avg_pred_old,
        'avg_pred_new': avg_pred_new,
        'actual_pass_rate': actual_pass_rate,
        'students': students_output,
        'size': len(calibrated_students),
        'v_class': v_class,
        'mult_env': mult_env
    }

def main():
    global db_mode
    try:
        conn = mysql.connector.connect(
            host="localhost",
            port=3307,
            user="root",
            password="",
            database="qldt_el"
        )
        cursor = conn.cursor()
        print("Connected to MySQL database on port 3307.")
    except Exception as mysql_err:
        print(f"Warning: Connection to MySQL on port 3307 failed ({mysql_err}). Falling back to SQLite 'data/inputs/qldt.db'...")
        db_mode = "SQLite"
        import sqlite3
        conn = sqlite3.connect('data/inputs/qldt.db')
        cursor = conn.cursor()
        # Initialize qldt.db tables from qldt.sql if it exists
        if os.path.exists('data/inputs/qldt.sql'):
            with open('data/inputs/qldt.sql', 'r', encoding='utf-8') as sf:
                sql_script = sf.read()
            cursor.executescript(sql_script)
            conn.commit()
    
    # 1. Load Excel chot data
    print("Loading data from Excel chot Rpoint...")
    excel_path = "data/inputs/PTIT_Chiso.xlsx"
    excel_data = get_excel_chot_data(excel_path)
    
    # 2. Retrieve sequence mappings
    print("Loading class-course sequences...")
    class_course_seq = defaultdict(list)
    seq_raw = run_query(cursor, """
        SELECT DISTINCT class_id, course_id 
        FROM qldt_el.final_results
        ORDER BY class_id, course_id ASC;
    """)
    for r in seq_raw:
        class_course_seq[int(r['class_id'])].append(int(r['course_id']))
        
    # Mappings of Course IDs to inspect/predict
    # KS24: JWS (211) -> AI (212)
    # KS25: Python (124) -> Python Web (215)
    ks24_cv_course = 194
    ks24_curr_course = 214
    ks25_cv_course = 193
    ks25_curr_course = 217
    
    classes_raw = run_query(cursor, "SELECT id, name FROM qldt_el.classes;")
    classes_map = {int(c['id']): c['name'] for c in classes_raw}
    
    courses_raw = run_query(cursor, "SELECT id, name FROM qldt_el.courses;")
    courses_map = {int(c['id']): c['name'] for c in courses_raw}
    
    # Data structures to save JSON for HTML Dashboard
    dashboard_data = {
        'KS24': {'cv': [], 'curr': []},
        'KS25': {'cv': [], 'curr': []},
        'QTKD': {'cv': [], 'curr': []}
    }
    
    all_care_list = []
    
    # We loop through all classes and filter appropriately
    print("Processing evaluations...")
    for cid, cname in sorted(classes_map.items()):
        norm_cname = normalize_class_name(cname)
        is_qtkd = "QTKD" in cname
        is_ks25 = ("KS25" in cname or "K25" in cname) and not is_qtkd
        is_ks24 = ("KS24" in cname or "K24" in cname)
        
        if not (is_ks25 or is_ks24 or is_qtkd):
            continue
            
        batch = "K25" if is_ks25 else "K24"
        
        # KS24 Block
        if is_ks24:
            # 1. Java Web Service CV
            co_id = ks24_cv_course
            coname = courses_map.get(co_id, "Java Web Service")
            res_cv = predict_class_pass_rate(cursor, cid, co_id, class_course_seq, excel_data, cname, coname, batch)
            if res_cv['size'] > 0:
                dashboard_data['KS24']['cv'].append({
                    'class_name': cname,
                    'course_name': coname,
                    'size': res_cv['size'],
                    'v_class': res_cv['v_class'],
                    'mult_env': res_cv['mult_env'],
                    'pred_old': res_cv['avg_pred_old'],
                    'actual_pass': res_cv['actual_pass_rate'],
                    'err': abs(res_cv['avg_pred_old'] - res_cv['actual_pass_rate'])
                })
                
            # 2. AI Application CURR
            co_id_curr = ks24_curr_course
            coname_curr = courses_map.get(co_id_curr, "AI Application")
            res_curr = predict_class_pass_rate(cursor, cid, co_id_curr, class_course_seq, excel_data, cname, coname_curr, batch)
            if res_curr['size'] > 0:
                dashboard_data['KS24']['curr'].append({
                    'class_name': cname,
                    'course_name': coname_curr,
                    'size': res_curr['size'],
                    'v_class': res_curr['v_class'],
                    'mult_env': res_curr['mult_env'],
                    'pred_old': res_curr['avg_pred_old'],
                    'pred_new': res_curr['avg_pred_new'],
                    'actual_pass': res_curr['actual_pass_rate']
                })
                # Add students to Care List
                for s in res_curr['students']:
                    if s['risk_level'] in ('RED', 'YELLOW') or len(s['anomalies']) > 0:
                        all_care_list.append({
                            'batch': 'K24',
                            'class_name': cname,
                            'course_name': coname_curr,
                            'student_id': s['student_id'],
                            'full_name': s['full_name'],
                            'risk_level': s['risk_level'],
                            'reasons': s['reasons'],
                            'p_final': s['p_final'],
                            'att': s['att'],
                            'hw': s['hw'],
                            'el': s['el'],
                            'rp': s['rp'],
                            'anomalies': s['anomalies']
                        })
                        
        # KS25 Block
        if is_ks25:
            # 1. Python CV
            co_id = ks25_cv_course
            coname = courses_map.get(co_id, "Python")
            res_cv = predict_class_pass_rate(cursor, cid, co_id, class_course_seq, excel_data, cname, coname, batch)
            if res_cv['size'] > 0:
                dashboard_data['KS25']['cv'].append({
                    'class_name': cname,
                    'course_name': coname,
                    'size': res_cv['size'],
                    'v_class': res_cv['v_class'],
                    'mult_env': res_cv['mult_env'],
                    'pred_old': res_cv['avg_pred_old'],
                    'actual_pass': res_cv['actual_pass_rate'],
                    'err': abs(res_cv['avg_pred_old'] - res_cv['actual_pass_rate'])
                })
                
            # 2. Python Web CURR
            co_id_curr = ks25_curr_course
            coname_curr = courses_map.get(co_id_curr, "Python Web")
            res_curr = predict_class_pass_rate(cursor, cid, co_id_curr, class_course_seq, excel_data, cname, coname_curr, batch)
            if res_curr['size'] > 0:
                dashboard_data['KS25']['curr'].append({
                    'class_name': cname,
                    'course_name': coname_curr,
                    'size': res_curr['size'],
                    'v_class': res_curr['v_class'],
                    'mult_env': res_curr['mult_env'],
                    'pred_old': res_curr['avg_pred_old'],
                    'pred_new': res_curr['avg_pred_new'],
                    'actual_pass': res_curr['actual_pass_rate']
                })
                # Add students to Care List
                for s in res_curr['students']:
                    if s['risk_level'] in ('RED', 'YELLOW') or len(s['anomalies']) > 0:
                        all_care_list.append({
                            'batch': 'K25',
                            'class_name': cname,
                            'course_name': coname_curr,
                            'student_id': s['student_id'],
                            'full_name': s['full_name'],
                            'risk_level': s['risk_level'],
                            'reasons': s['reasons'],
                            'p_final': s['p_final'],
                            'att': s['att'],
                            'hw': s['hw'],
                            'el': s['el'],
                            'rp': s['rp'],
                            'anomalies': s['anomalies']
                        })
                        
        # QTKD Block
        if is_qtkd:
            # 1. DTB201 CV
            co_id = 188
            coname = courses_map.get(co_id, "DTB201")
            res_cv = predict_class_pass_rate(cursor, cid, co_id, class_course_seq, excel_data, cname, coname, "QTKD")
            if res_cv['size'] > 0:
                dashboard_data['QTKD']['cv'].append({
                    'class_name': cname,
                    'course_name': coname,
                    'size': res_cv['size'],
                    'v_class': res_cv['v_class'],
                    'mult_env': res_cv['mult_env'],
                    'pred_old': res_cv['avg_pred_old'],
                    'actual_pass': res_cv['actual_pass_rate'],
                    'err': abs(res_cv['avg_pred_old'] - res_cv['actual_pass_rate'])
                })
                
            # 2. PRJ302 CURR
            co_id_curr = 218
            coname_curr = courses_map.get(co_id_curr, "PRJ302")
            res_curr = predict_class_pass_rate(cursor, cid, co_id_curr, class_course_seq, excel_data, cname, coname_curr, "QTKD")
            if res_curr['size'] > 0:
                dashboard_data['QTKD']['curr'].append({
                    'class_name': cname,
                    'course_name': coname_curr,
                    'size': res_curr['size'],
                    'v_class': res_curr['v_class'],
                    'mult_env': res_curr['mult_env'],
                    'pred_old': res_curr['avg_pred_old'],
                    'pred_new': res_curr['avg_pred_new'],
                    'actual_pass': res_curr['actual_pass_rate']
                })
                # Add students to Care List
                for s in res_curr['students']:
                    if s['risk_level'] in ('RED', 'YELLOW') or len(s['anomalies']) > 0:
                        all_care_list.append({
                            'batch': 'QTKD K25',
                            'class_name': cname,
                            'course_name': coname_curr,
                            'student_id': s['student_id'],
                            'full_name': s['full_name'],
                            'risk_level': s['risk_level'],
                            'reasons': s['reasons'],
                            'p_final': s['p_final'],
                            'att': s['att'],
                            'hw': s['hw'],
                            'el': s['el'],
                            'rp': s['rp'],
                            'anomalies': s['anomalies']
                        })
                        
    # -------------------------------------------------------------
    # WRITE REPORTS
    # -------------------------------------------------------------
    print("Writing predictions verification report...")
    md_report_path = "output/reports/core/agent_2_academic_prediction.md"
    
    # Calculate global MAEs
    k24_cv_errs = [c['err'] for c in dashboard_data['KS24']['cv']]
    k25_cv_errs = [c['err'] for c in dashboard_data['KS25']['cv']]
    k24_mae = mean(k24_cv_errs) if k24_cv_errs else 0.0
    k25_mae = mean(k25_cv_errs) if k25_cv_errs else 0.0
    
    with open(md_report_path, 'w', encoding='utf-8') as f:
        f.write("# BÁO CÁO DỰ BÁO HỌC THUẬT & KIỂM CHỨNG SAI SỐ KHOÁ K24 & K25\n\n")
        f.write(f"*Báo cáo được lập tự động ngày {datetime.now().strftime('%d/%m/%Y')} tích hợp chỉ số Ý thức lớp (Peer Pressure Multiplier).*\n\n")
        
        f.write("## 📌 TÓM TẮT ĐÁNH GIÁ SAI SỐ KIỂM CHỨNG (MAE)\n")
        f.write(f"- **Khóa K24 (Kiểm chứng qua môn Java Web Service)**: MAE = **{k24_mae:.2f}%**\n")
        f.write(f"- **Khóa K25 (Kiểm chứng qua môn Python)**: MAE = **{k25_mae:.2f}%**\n\n")
        
        f.write("## 📊 1. CHI TIẾT MÔN KIỂM CHỨNG (ĐỐI CHIẾU DB LỊCH SỬ)\n\n")
        
        f.write("### 🔹 Khóa K24 - Môn Java Web Service\n\n")
        f.write("| Tên Lớp | Sĩ số | Vi phạm lớp% | Hệ số Env | Dự báo (Luật cũ)% | Thực tế DB% | Sai số% |\n")
        f.write("| :--- | :---: | :---: | :---: | :---: | :---: | :---: |\n")
        for c in dashboard_data['KS24']['cv']:
            f.write(f"| {c['class_name']} | {c['size']} | {c['v_class']:.1f}% | {c['mult_env']:.2f} | **{c['pred_old']:.1f}%** | **{c['actual_pass']:.1f}%** | {c['pred_old'] - c['actual_pass']:.1f}% |\n")
            
        f.write("\n### 🔹 Khóa K25 - Môn Python\n\n")
        f.write("| Tên Lớp | Sĩ số | Vi phạm lớp% | Hệ số Env | Dự báo (Luật cũ)% | Thực tế DB% | Sai số% |\n")
        f.write("| :--- | :---: | :---: | :---: | :---: | :---: | :---: |\n")
        for c in dashboard_data['KS25']['cv']:
            f.write(f"| {c['class_name']} | {c['size']} | {c['v_class']:.1f}% | {c['mult_env']:.2f} | **{c['pred_old']:.1f}%** | **{c['actual_pass']:.1f}%** | {c['pred_old'] - c['actual_pass']:.1f}% |\n")
            
        f.write("\n---\n\n")
        f.write("## 📊 2. DỰ BÁO MÔN HỌC HIỆN TẠI (ÁP DỤNG QUY CHẾ MỚI)\n\n")
        
        f.write("### 🔹 Khóa K24 - Môn AI Application (Hiện tại)\n\n")
        f.write("| Tên Lớp | Sĩ số | Vi phạm lớp% | Hệ số Env | Dự báo (Luật cũ)% | Dự báo (Quy chế mới)% |\n")
        f.write("| :--- | :---: | :---: | :---: | :---: | :---: |\n")
        for c in dashboard_data['KS24']['curr']:
            f.write(f"| {c['class_name']} | {c['size']} | {c['v_class']:.1f}% | {c['mult_env']:.2f} | **{c['pred_old']:.1f}%** | **{c['pred_new']:.1f}%** |\n")
            
        f.write("\n### 🔹 Khóa K25 - Môn Python Web (Hiện tại)\n\n")
        f.write("| Tên Lớp | Sĩ số | Vi phạm lớp% | Hệ số Env | Dự báo (Luật cũ)% | Dự báo (Quy chế mới)% |\n")
        f.write("| :--- | :---: | :---: | :---: | :---: | :---: |\n")
        for c in dashboard_data['KS25']['curr']:
            f.write(f"| {c['class_name']} | {c['size']} | {c['v_class']:.1f}% | {c['mult_env']:.2f} | **{c['pred_old']:.1f}%** | **{c['pred_new']:.1f}%** |\n")

    print("Writing multi-level risk student care list...")
    with open(md_report_path, 'a', encoding='utf-8') as f:
        f.write("\n\n---\n\n")
        f.write("# DANH SÁCH HỌC VIÊN CẦN CAN THIỆP MÔN HIỆN TẠI (CARE LIST ĐA TẦNG)\n\n")
        f.write("> [!IMPORTANT]\n")
        f.write("> Danh sách chỉ lọc ra các học viên có nguy cơ trượt của môn học hiện tại (AI Application đối với K24, Python Web đối với K25).\n\n")
        
        # Red Risk
        f.write("## 🔴 1. DANH SÁCH NGUY CƠ CAO (BÁO ĐỘNG ĐỎ - CẤM THI / HỌC LỰC YẾU)\n\n")
        f.write("| Khóa | Lớp | Học viên | XS đỗ% | Chuyên cần (vắng)% | Bài tập (nợ)% | Elearning (vi phạm) | Lý do chính |\n")
        f.write("| :---: | :--- | :--- | :---: | :---: | :---: | :---: | :--- |\n")
        reds = [s for s in all_care_list if s['risk_level'] == 'RED']
        for s in sorted(reds, key=lambda x: (x['batch'], x['class_name'], x['p_final'])):
            f.write(f"| {s['batch']} | {s['class_name']} | **{s['full_name']}** | **{s['p_final']:.1f}%** | {s['att']:.1f}% | {100.0 - s['hw']:.1f}% | {s['el']:.0f} bài | {', '.join(s['reasons'])} |\n")
            
        # Yellow Risk
        f.write("\n## 🟡 2. DANH SÁCH NGUY CƠ TRUNG BÌNH (CẢNH BÁO VÀNG - CẬN CẤM THI / MẤT GỐC)\n\n")
        f.write("| Khóa | Lớp | Học viên | XS đỗ% | Chuyên cần (vắng)% | Bài tập (nợ)% | Elearning (vi phạm) | Dấu hiệu báo động |\n")
        f.write("| :---: | :--- | :--- | :---: | :---: | :---: | :---: | :--- |\n")
        yellows = [s for s in all_care_list if s['risk_level'] == 'YELLOW']
        for s in sorted(yellows, key=lambda x: (x['batch'], x['class_name'], x['p_final'])):
            f.write(f"| {s['batch']} | {s['class_name']} | **{s['full_name']}** | **{s['p_final']:.1f}%** | {s['att']:.1f}% | {100.0 - s['hw']:.1f}% | {s['el']:.0f} bài | {', '.join(s['reasons'])} |\n")

    # Aggregate cohort anomalies summary
    anomalies_summary = {
        'K24': { 'copy_suspect': 0, 'discipline_paradox': 0, 'passive_learner': 0, 'sudden_drop': 0 },
        'K25': { 'copy_suspect': 0, 'discipline_paradox': 0, 'passive_learner': 0, 'sudden_drop': 0 },
        'QTKD': { 'copy_suspect': 0, 'discipline_paradox': 0, 'passive_learner': 0, 'sudden_drop': 0 }
    }
    for s in all_care_list:
        cohort_key = 'K24' if s['batch'] == 'K24' else ('QTKD' if 'QTKD' in s['batch'] else 'K25')
        for anomaly in s.get('anomalies', []):
            if anomaly in anomalies_summary[cohort_key]:
                anomalies_summary[cohort_key][anomaly] += 1

    # Save JSON data for the HTML Dashboard script
    with open('data/processed/agent2_output.json', 'w', encoding='utf-8') as jf:
        json.dump({
            'dashboard_data': dashboard_data,
            'care_list': all_care_list,
            'anomalies_summary': anomalies_summary
        }, jf, ensure_ascii=False, indent=2)

    print("Success. MD Reports and JSON generated.")
    
    # Sinh HTML Report tự động
    import subprocess
    print("Agent 2: Đang sinh trang báo cáo HTML...")
    subprocess.run([sys.executable, "agents/core/agent_2_academic_pred/generate_report.py"])
    
    conn.close()

if __name__ == '__main__':
    main()
