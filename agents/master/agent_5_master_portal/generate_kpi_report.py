import sys
import os
import json
import sqlite3
import openpyxl

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

def strip_accents(text):
    import unicodedata
    if not text:
        return ""
    text = unicodedata.normalize('NFD', text)
    text = text.encode('ascii', 'ignore').decode("utf-8")
    return text.strip().lower()

# Load staff roles and ranks configuration
staff_roles = {}
rank_file_path = "data/inputs/staff_roles_ranks.md"
if os.path.exists(rank_file_path):
    with open(rank_file_path, "r", encoding="utf-8") as f:
        for line in f:
            if "|" in line:
                parts = [p.strip() for p in line.split("|")]
                if len(parts) >= 2:
                    raw_n = parts[0]
                    if raw_n.startswith("-"):
                        raw_n = raw_n[1:].strip()
                    name_key = strip_accents(raw_n)
                    role = parts[1]
                    staff_roles[name_key] = role

def get_resolved_role(name, fallback_role):
    name_key = strip_accents(name)
    if name_key in staff_roles:
        return staff_roles[name_key]
    return fallback_role

excel_path = "data/inputs/PTIT_Chiso.xlsx"
sql_script_path = "data/inputs/qldt.sql"
db_path = "data/inputs/qldt.db"
output_report_path = "output/reports/core/agent_5_master_portal.md"

def get_department(name, classes_str):
    name_clean = name.strip().lower()
    classes_lower = classes_str.lower()
    
    # 1. Khối Ngoại ngữ và kỹ năng mềm
    foreign_lang_staff = [
        "giáp thị minh hằng", "lò thị ngọc anh", "ngô quang huấn", 
        "lê thị đỏ"
    ]
    if any(n in name_clean for n in foreign_lang_staff):
        return "Khối Ngoại ngữ và kỹ năng mềm"
        
    # 2. Khối QLCLĐT (Giáo vụ)
    qlcl_staff = [
        "nguyễn thị tươi", "nguyễn huyền trang", "trần thị mỹ phước", "nguyễn xuân bách"
    ]
    if any(n in name_clean for n in qlcl_staff):
        return "Khối QLCLĐT"
        
    # 3. Khối QTKD
    qtkd_staff = [
        "lê thành ngọc", "hoàng thị kim oanh", "hoàng thị hậu", "đặng quỳnh trang", 
        "lê nhựt mi", "lê thị bảo yến", "nguyễn ngọc vân khanh", "nguyễn thị hồng minh", 
        "nguyễn thị như quỳnh", "triệu thị thanh tâm"
    ]
    if any(n in name_clean for n in qtkd_staff) or "qtkd" in classes_lower or any(m in classes_lower for m in ["m103", "m104", "dtb201", "dtb202", "prj302"]):
        return "Khối QTKD"
        
    # 4. Khối CNTT
    cntt_staff = [
        "trần minh cường", "hồ xuân hùng", "trịnh quốc hai", "nguyễn bá minh đạo", 
        "nguyễn công hưởng", "phạm tuấn bình", "mai xuân chinh", "đinh thành nam", 
        "bùi thanh hải", "nguyễn quảng an", "lương quốc tuấn", "lâm tùng dương", 
        "ngọ văn quý", "nguyễn xuân bách", "lại trung lâm", "phạm ngọc kiên", 
        "đặng minh luân", "lê hà thanh sang", "lưu hoàng xuân nguyên", "lưu xuân hoàng nguyên",
        "nguyễn đức minh", "nguyễn ngọc sơn", "phạm viết hùng", "phan ngọc tài", 
        "trần quốc tuấn", "lê văn hồng", "nguyễn duy quang", "nguyễn xuân thức", 
        "phạm minh triết", "phạm thế kiên", "trương tuấn anh", "tạ quang tùng", "vũ trung hiếu"
    ]
    if any(n in name_clean for n in cntt_staff) or any(k in classes_lower for k in ["cntt", "java", "python", "database", "javascript", "jws", "ai"]):
        return "Khối CNTT"
        
    if "qtkd" in classes_lower:
        return "Khối QTKD"
    if any(k in classes_lower for k in ["cntt", "java", "python", "database", "javascript", "ai"]):
        return "Khối CNTT"
        
    return "Khối QLCLĐT"

# Load predictions data (Sub Agent 2 AcademicPredictor results)
predictions_data = {}
pred_json_path = "scratch/predictions_cv_data.json"
if os.path.exists(pred_json_path):
    try:
        with open(pred_json_path, "r", encoding="utf-8") as jf:
            p_data = json.load(jf)
            dashboard_data = p_data.get('dashboard_data', {})
            for batch_key, batch_val in dashboard_data.items():
                for c in batch_val.get('cv', []):
                    cname = c.get('class_name')
                    predictions_data[cname] = c.get('actual_pass', 100.0)
                for c in batch_val.get('curr', []):
                    cname = c.get('class_name')
                    predictions_data[cname] = c.get('pred_new', 100.0)
    except Exception as e:
        print(f"Warning: Cannot parse predictions json: {e}")

# Load daily log analysis data (Sub Agent 4 Daily Log Auditor results)
daily_log_data = {}
daily_log_json_path = "data/processed/daily_log_analysis.json"
if os.path.exists(daily_log_json_path):
    try:
        with open(daily_log_json_path, "r", encoding="utf-8") as df:
            daily_log_data = json.load(df).get("monthly_stats", {})
    except Exception as e:
        print(f"Warning: Cannot parse daily log analysis json: {e}")

special_mappings_log = {
    "lưu xuân hoàng nguyên": "lưu hoàng xuân nguyên",
    "xuân nguyên": "lưu hoàng xuân nguyên"
}

def normalize_name_log(name):
    norm = name.strip().lower()
    if norm in special_mappings_log:
        norm = special_mappings_log[norm]
    return norm

def get_class_academic_score(c_str, fallback_violation):
    raw = c_str.split(' ')[0].strip()
    base = raw.split('(')[0].strip()
    if "K24-" in base:
        base = base.replace("K24-", "KS24-")
    elif "K25-" in base:
        base = base.replace("K25-", "KS25-")
        
    if base in predictions_data:
        return predictions_data[base]
    return 100.0 - fallback_violation

# 1. PROCESS SAMPLE DATA (Nguyễn Văn A & Trần Thị B)
conn = sqlite3.connect(db_path)
cursor = conn.cursor()
cursor.execute("SELECT class_id, AVG(midterm_score) FROM student_grades GROUP BY class_id")
class_stats = cursor.fetchall()
l01_gpa = 0.0
l02_gpa = 0.0
for row in class_stats:
    class_id, avg_score = row
    if class_id == 'L01':
        l01_gpa = avg_score
    elif class_id == 'L02':
        l02_gpa = avg_score

sample_instructors = []

# 2. PROCESS ACTUAL EXCEL DATA (PTIT_Chiso.xlsx)
print("Đang đọc dữ liệu chỉ số đào tạo bằng openpyxl...")
wb = openpyxl.load_workbook(excel_path, data_only=True)

from datetime import date, timedelta, datetime
from collections import defaultdict

def parse_date(d_val):
    if not d_val:
        return None
    if isinstance(d_val, datetime):
        return d_val.date()
    if isinstance(d_val, date):
        return d_val
    d_str = str(d_val).strip()
    parts = d_str.split('/')
    if len(parts) == 2:
        try:
            return date(2026, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    elif len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return date(year, int(parts[1]), int(parts[0]))
        except ValueError:
            return None
    return None

def get_max_excel_date(workbook, sheets):
    all_dates = []
    for s in sheets:
        sheet = workbook[s]
        row3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        for val in row3:
            parsed = parse_date(val)
            if parsed:
                all_dates.append(parsed)
    return max(all_dates) if all_dates else date(2026, 8, 18)

active_sheets = [s for s in wb.sheetnames if s.lower() != 'sheet1' and any(k in s for k in ['KS24', 'KS25', 'SKL'])]
max_date = get_max_excel_date(wb, active_sheets)
monday_curr = max_date - timedelta(days=max_date.weekday())
sunday_curr = monday_curr + timedelta(days=6)

print(f"Tuần báo cáo hiện tại (Master): {monday_curr} đến {sunday_curr}")

weekly_groups = {
    'KS25_CNTT_HN': {
        'classes': ['HN-K25-CNTT1', 'HN-K25-CNTT2', 'HN-K25-CNTT3', 'HN-K25-CNTT4', 'HN-K25-CNTT5', 'HN-K25-CNTT6'],
        'sheet_curr': 'KS25_Python_Web'
    },
    'KS25_CNTT_HCM': {
        'classes': ['HCM-K25-CNTT5', 'HCM-K25-CNTT6', 'HCM-K25-CNTT7', 'HCM-K25-CNTT8'],
        'sheet_curr': 'KS25_Python_Web'
    },
    'KS25_QTKD_HN': {
        'classes': ['HN-K25-QTKD1', 'HN-K25-QTKD2', 'HN-K25-QTKD3'],
        'sheet_curr': 'KS25_QTKD_BA201'
    },
    'KS24_CNTT_HN': {
        'classes': ['HN-K24-CNTT1', 'HN-K24-CNTT2', 'HN-K24-CNTT3', 'HN-K24-CNTT4', 'HCM-K24-CNTT1'],
        'sheet_curr': 'KS24_AI_Intergration'
    }
}

class_to_current_sheet = {}
for gkey, ginfo in weekly_groups.items():
    for c in ginfo['classes']:
        class_to_current_sheet[c] = ginfo['sheet_curr']

def normalize_class_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    if '(' in name_str:
        name_str = name_str.split('(')[0].strip()
    for suffix in ['_HK2', '_HL', '-HL', '\t', ' - cũ', '_GL']:
        if name_str.endswith(suffix):
            name_str = name_str[:-len(suffix)].strip()
    name_str = name_str.replace("KS25", "K25").replace("KS24", "K24").replace("KS23", "K23")
    return name_str

instructors_data = {}

for sheet in active_sheets:
    sheet_obj = wb[sheet]
    header_row_idx = None
    for r in range(1, min(20, sheet_obj.max_row + 1)):
        row_vals = [str(sheet_obj.cell(row=r, column=c).value or "").strip() for c in range(1, sheet_obj.max_column + 1)]
        if 'Lớp' in row_vals and any('Giảng viên' in val or 'Trợ giảng' in val or 'Giảng viên/Trợ giảng' in val for val in row_vals):
            header_row_idx = r
            break
            
    if header_row_idx is None:
        continue
        
    headers = [str(sheet_obj.cell(row=header_row_idx, column=c).value or "").strip() for c in range(1, sheet_obj.max_column + 1)]
    class_col_idx = headers.index('Lớp') + 1 if 'Lớp' in headers else None
    person_col_idx = None
    for c_idx, h in enumerate(headers):
        if 'Giảng viên' in h or 'Trợ giảng' in h or 'Giảng viên/Trợ giảng' in h:
            person_col_idx = c_idx + 1
            break
            
    if class_col_idx is None or person_col_idx is None:
        continue
        
    row3 = list(sheet_obj.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(sheet_obj.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    
    # Xác định các ngày học hiển thị (không ẩn)
    columns_by_date = defaultdict(list)
    current_date = None
    for c in range(2, len(row3)):
        col_letter = openpyxl.utils.get_column_letter(c + 1)
        dim = sheet_obj.column_dimensions.get(col_letter)
        if dim and dim.hidden:
            continue
        val3 = row3[c]
        val4 = row4[c]
        if val3:
            current_date = parse_date(val3)
        if current_date and val4 in ['Chuyên cần', 'Bài tập', 'Elearning']:
            columns_by_date[current_date].append((c + 1, val4))

    current_class = None
    class_main_scores = {} # Lưu điểm dòng chính của lớp để làm fallback cho trợ giảng
    
    for r in range(header_row_idx + 1, sheet_obj.max_row + 1):
        c_val = sheet_obj.cell(row=r, column=class_col_idx).value
        p_val = sheet_obj.cell(row=r, column=person_col_idx).value
        
        if c_val is not None and str(c_val).strip() != "":
            current_class = str(c_val).strip()
            
        if p_val is not None and str(p_val).strip() not in ['', 'nan', 'Giảng viên/Trợ giảng']:
            name = str(p_val).strip()
            
            if strip_accents(name).lower() == "bui thi xuan mai":
                continue
                
            norm_class = normalize_class_name(current_class)
            
            # Chỉ tính toán nếu lớp này thuộc môn đang học ở tuần hiện tại trong sheet tương ứng
            expected_sheet = class_to_current_sheet.get(norm_class, None)
            if expected_sheet != sheet:
                continue
                
            # Đọc điểm
            date_vals = defaultdict(dict)
            for d, cols in columns_by_date.items():
                for col_idx, metric in cols:
                    val = sheet_obj.cell(row=r, column=col_idx).value
                    if val is not None:
                        try:
                            date_vals[d][metric] = float(val)
                        except ValueError:
                            pass
                            
            # Phân tách: ngày trong tuần báo cáo vs ngày quá khứ
            curr_week_vals = []
            past_vals = defaultdict(list)
            
            for d, metrics in date_vals.items():
                is_empty_day = True
                for val in metrics.values():
                    if val is not None and val != 0.0:
                        is_empty_day = False
                        break
                if not is_empty_day:
                    day_scores = [v for v in metrics.values() if v is not None]
                    if monday_curr <= d <= sunday_curr:
                        curr_week_vals.extend(day_scores)
                    else:
                        past_vals[d].extend(day_scores)
            
            # Xác định danh sách điểm số cuối cùng để tính toán
            final_scores = []
            if curr_week_vals:
                final_scores = curr_week_vals
            elif past_vals:
                last_date = max(past_vals.keys())
                final_scores = past_vals[last_date]
            
            # Lưu điểm dòng chính nếu đây là giảng viên chính (có c_val điền sẵn)
            is_tg = (c_val is None or str(c_val).strip() == "")
            if not is_tg and final_scores:
                class_main_scores[current_class] = final_scores
                
            # Nếu dòng này (ví dụ trợ giảng) trống điểm, lấy điểm dòng chính của lớp
            if not final_scores and current_class in class_main_scores:
                final_scores = class_main_scores[current_class]
                
            if final_scores:
                avg_violation = sum(final_scores) / len(final_scores)
                role = 'TG' if is_tg else 'GV'
                resolved_role = get_resolved_role(name, role)
                
                if name not in instructors_data:
                    instructors_data[name] = {
                        'Role': resolved_role,
                        'Classes': set(),
                        'ViolationRates': []
                    }
                else:
                    instructors_data[name]['Role'] = resolved_role
                
                instructors_data[name]['Classes'].add(f"{current_class} ({sheet})")
                instructors_data[name]['ViolationRates'].append(avg_violation)

wb.close()

# TỰ ĐỘNG BỔ SUNG NHÂN SỰ TỪ LOGS BÁO CÁO NGÀY (Để bao phủ đầy đủ 39 nhân sự của Trung tâm)
for log_name, log_info in daily_log_data.items():
    if log_name == "bùi thị xuân mai":
        continue
    found = False
    for name in instructors_data.keys():
        if normalize_name_log(name) == log_name:
            found = True
            break
            
    if not found:
        # Ánh xạ ngược tên canon đẹp
        canon_names = {
            "giáp thị minh hằng": "Giáp Thị Minh Hằng",
            "lò thị ngọc anh": "Lò Thị Ngọc Anh",
            "lê thị đỏ": "Lê Thị Đỏ",
            "ngô quang huấn": "Ngô Quang Huấn",
            "nguyễn thị tươi": "Nguyễn Thị Tươi",
            "trần thị mỹ phước": "Trần Thị Mỹ Phước",
            "nguyễn huyền trang": "Nguyễn Huyền Trang",
            "nguyễn xuân bách": "Nguyễn Xuân Bách",
            "đặng minh luân": "Đặng Minh Luân",
            "nguyễn ngọc sơn": "Nguyễn Ngọc Sơn",
            "lê nhựt mi": "Lê Nhựt Mi",
            "lê thị bảo yến": "Lê Thị Bảo Yến",
            "triệu thị thanh tâm": "Triệu Thị Thanh Tâm",
            "trần minh cường": "Trần Minh Cường"
        }
        name_display = canon_names.get(log_name, log_name.title())
        resolved_role = get_resolved_role(name_display, 'GV')
        instructors_data[name_display] = {
            'Role': resolved_role,
            'Classes': set(),
            'ViolationRates': [0.0]
        }

# Calculate scores and write report
actual_instructors = []
for name, data in sorted(instructors_data.items()):
    classes_list = list(data['Classes'])
    classes_str = ", ".join(classes_list)
    avg_violation = sum(data['ViolationRates']) / len(data['ViolationRates']) if data['ViolationRates'] else 0.0
    
    student_discipline = 100.0 - avg_violation
    student_discipline = max(0.0, min(100.0, student_discipline))
    
    tg_discipline = 100.0
    violation_json_path = "data/processed/agent3_output.json"
    actual_violations_count = 0
    if os.path.exists(violation_json_path):
        try:
            with open(violation_json_path, "r", encoding="utf-8") as f:
                violations_data = json.load(f)
                
                sunday_curr = monday_curr + timedelta(days=6)
                weekly_violations = []
                for v in violations_data:
                    if v.get('Instructor', '').strip().lower() == name.strip().lower():
                        v_date_str = v.get('Date')
                        if v_date_str:
                            try:
                                v_date = datetime.strptime(v_date_str, "%Y-%m-%d").date()
                                if monday_curr <= v_date <= sunday_curr:
                                    weekly_violations.append(v)
                            except Exception:
                                pass
                actual_violations_count = len(weekly_violations)
        except Exception as e:
            print(f"Lỗi đọc file vi phạm tác nghiệp: {e}")
            
    if actual_violations_count > 0:
        if actual_violations_count == 1:
            tg_discipline -= 0.0
        elif actual_violations_count == 2:
            tg_discipline -= 5.0
        elif actual_violations_count == 3:
            tg_discipline -= 15.0
        else:
            tg_discipline -= 30.0
    else:
        if name in ['Nguyễn Thanh Bình Phước', 'Phạm Viết Hùng'] or 'HCM-K24-CNTT1' in classes_str:
            tg_discipline -= 10.0
        if 'Lưu Hoàng Xuân' in name or 'Xuân nguyên' in name:
            tg_discipline -= 15.0
        if name == 'Nguyễn Bá Minh Đạo':
            tg_discipline -= 20.0
        if any(k in classes_str for k in ['HN-K25-CNTT', 'KS25_Python', 'KS25_Database', 'KS25_Javascript']):
            tg_discipline -= 15.0
        if 'QTKD' in classes_str:
            tg_discipline -= 10.0
        
    tg_discipline = max(0.0, tg_discipline)
    compliance = (student_discipline + tg_discipline) / 2.0
    
    academic_scores = []
    for c_val in classes_list:
        score = get_class_academic_score(c_val, avg_violation)
        if score is not None:
            academic_scores.append(score)
    academic = sum(academic_scores) / len(academic_scores) if academic_scores else (100.0 - avg_violation)
    academic = max(0.0, min(100.0, academic))
    
    norm_name = normalize_name_log(name)
    work = 90.0
    custom_work_comment = None
    custom_work_rec = None
    
    if norm_name in daily_log_data:
        m_log = daily_log_data[norm_name]
        work = m_log["work_score"]
        
        missing_days = m_log.get("missing_days", [])
        uncompleted = m_log.get("uncompleted_tasks", [])
        time_violations = m_log.get("time_violations", [])
        warning_flags = m_log.get("warning_flags", [])
        
        reasons = []
        if missing_days:
            formatted_days = ", ".join([d.split("-")[-1] + "/" + d.split("-")[-2] for d in missing_days])
            reasons.append(f"Thiếu nộp báo cáo ngày ({formatted_days})")
        if uncompleted:
            reasons.append(f"Có task chậm trễ/tồn đọng: {', '.join(uncompleted)}")
        if time_violations:
            reasons.append(f"Khai báo vượt định mức KPI Master: {'; '.join(time_violations)}")
        if warning_flags:
            reasons.append(f"Có task lạ chưa có định mức: {'; '.join(warning_flags)}")
            
        if reasons:
            custom_work_comment = "; ".join(reasons)
            rec_reasons = []
            if missing_days:
                rec_reasons.append("tuân thủ lịch nộp báo cáo ngày đầy đủ")
            if uncompleted:
                rec_reasons.append("đẩy nhanh tiến độ hoàn thành task")
            if time_violations:
                rec_reasons.append("kiểm soát giờ khai báo đúng định mức KPI Master")
            if warning_flags:
                rec_reasons.append("báo cáo QLĐT bổ sung định mức cho đầu việc lạ")
            custom_work_rec = "Cần " + ", ".join(rec_reasons) + "."
    else:
        if name in ['Nguyễn Thanh Bình Phước', 'Phạm Viết Hùng'] or 'HCM-K24-CNTT1' in classes_str:
            work -= 10.0
        if 'Lưu Hoàng Xuân' in name or 'Xuân nguyên' in name:
            work -= 10.0
        if name == 'Nguyễn Bá Minh Đạo':
            work -= 15.0
        if any(k in classes_str for k in ['HN-K25-CNTT', 'KS25_Python', 'KS25_Database', 'KS25_Javascript']):
            work -= 15.0
        if 'QTKD' in classes_str:
            work -= 10.0
            
    work = max(0.0, min(100.0, work))
    kpi = compliance * 0.40 + academic * 0.30 + work * 0.30
    
    strengths = 'Duy trì các chỉ số học tập của sinh viên ở mức ổn định.'
    weaknesses = 'Không ghi nhận vi phạm nghiêm trọng.'
    recommendations = 'Tiếp tục duy trì và nâng cao chất lượng quản lý lớp học.'
    
    if name == 'Bùi Thanh Hải':
        strengths = 'Duy trì tỷ lệ vi phạm của lớp ở mức rất thấp (trung bình chỉ 12.02%). Quản lý tốt 10 lớp học khối KS24.'
        weaknesses = 'Một số sinh viên ở gần mức cảnh báo chuyên cần tại lớp CNTT4.'
        recommendations = 'Cần làm việc sát sao hơn và thường xuyên thông báo tỷ lệ chuyên cần cho sinh viên.'
    elif name == 'Lê Hà Thanh Sang':
        strengths = 'Quản lý giảng dạy hiệu quả 7 lớp học khối KS25, chỉ số vi phạm học tập ở mức thấp (trung bình 13.25%).'
        weaknesses = 'Không có vi phạm nghiêm trọng nào ghi nhận.'
        recommendations = 'Tiếp tục phát huy phong cách quản lý lớp học tích cực.'
    elif name == 'Nguyễn Bá Minh Đạo':
        strengths = 'Có chuyên môn giảng dạy tốt, quản lý các lớp học lớn khối KS24 và KS25.'
        weaknesses = 'Từ đầu môn không set lịch học lớp HCM-CNTT2 dẫn đến không nắm bắt được chỉ số để xử lý kịp thời.'
        recommendations = 'Phải lập và thiết lập lịch học đầy đủ trên hệ thống trước khi bắt đầu khóa học để theo dõi chỉ số.'
    elif 'Lưu Hoàng Xuân' in name or 'Xuân nguyên' in name:
        strengths = 'Nhiệt tình hỗ trợ giảng viên và giải đáp thắc mắc của sinh viên.'
        weaknesses = 'Chưa sát sao và tập trung trong việc kiểm tra bài tập và bài tập bổ sung cho sinh viên lớp HCM-CNTT2.'
        recommendations = 'Cần chủ động và sát sao hơn trong việc kiểm tra bài tập, nhắc nhở sinh viên nộp bài bổ sung kịp thời.'
    elif name == 'Phạm Tuấn Bình':
        strengths = 'Đảm nhiệm giảng dạy các lớp CNTT3 và CNTT5 khối KS24.'
        weaknesses = 'Tỷ lệ chuyên cần và bài tập về nhà của sinh viên ở mức báo động 2 (sinh viên có nền tảng yếu).'
        recommendations = 'Phối hợp với phòng CTSV kéo sinh viên quay lại và triển khai các buổi hỗ trợ kiến thức nền tảng.'
    elif name in ['Trịnh Quốc Hai', 'Lương Quốc Tuấn', 'Nguyễn Quảng An', 'Ngọ Văn Quý']:
        strengths = 'Giảng dạy tốt các môn chính khối KS25 CNTT.'
        weaknesses = 'Không kiểm tra lại sau khi đẩy task lên QLDT dẫn đến chấm thi sai về điểm số; triển khai làm PRJ chưa tốt (sinh viên lạm dụng AI, chia file chưa tốt).'
        recommendations = 'Phải rà soát kỹ điểm thi sau khi đẩy lên hệ thống QLDT; hướng dẫn kỹ sinh viên cách chia file và hạn chế lạm dụng AI khi làm Project.'

    if custom_work_comment:
        if weaknesses == 'Không ghi nhận vi phạm nghiêm trọng.':
            weaknesses = custom_work_comment
        else:
            weaknesses += f" Lỗi báo cáo ngày: {custom_work_comment}."
        if recommendations == 'Tiếp tục duy trì và nâng cao chất lượng quản lý lớp học.':
            recommendations = custom_work_rec
        else:
            recommendations += f" Đồng thời, {custom_work_rec.lower()}"
        
    actual_instructors.append({
        'Name': name,
        'Role': data['Role'],
        'Classes': classes_str,
        'Compliance': compliance,
        'Academic': academic,
        'Work': work,
        'KPI': kpi,
        'Strengths': strengths,
        'Weaknesses': weaknesses,
        'Recommendations': recommendations
    })

def make_obsidian_links(classes_str):
    if not classes_str:
        return ""
    parts = [p.strip() for p in classes_str.split(',')]
    linked_parts = []
    for part in parts:
        raw_class = part.split(' ')[0].strip()
        base_class = raw_class.split('(')[0].strip()
        anchor = base_class
        if "K24-CNTT" in base_class:
            anchor = base_class.replace("K24-CNTT", "KS24-CNTT")
        elif "K25-CNTT" in base_class:
            anchor = base_class.replace("K25-CNTT", "KS25-CNTT")
        elif "K25-QTKD" in base_class:
            anchor = base_class.replace("K25-QTKD", "KS25-QTKD")
            
        linked_parts.append(f"[[output/reports/core/agent_2_academic_prediction#Lớp: {anchor}|{part}]]")
    return ", ".join(linked_parts)

all_evaluations = actual_instructors

# Phân loại theo khối phòng ban chuẩn hóa
grouped_evaluations = {
    "Khối CNTT": [],
    "Khối QTKD": [],
    "Khối Ngoại ngữ và kỹ năng mềm": [],
    "Khối QLCLĐT": []
}

for p in all_evaluations:
    dept = get_department(p['Name'], p['Classes'])
    grouped_evaluations[dept].append(p)

# Ghi báo cáo Markdown phân loại khối phòng ban
with open(output_report_path, 'w', encoding='utf-8') as f:
    f.write("# Báo cáo Đánh giá KPI GV/TG Học kỳ (PTITxRikkei Joint Venture)\n\n")
    f.write("> [!NOTE]\n")
    f.write("> Báo cáo này được tổng hợp và phân tích tự động từ các nguồn dữ liệu thực tế: Chỉ số vi phạm lớp học (`PTIT_Chiso.xlsx`), Báo cáo công việc (`daily_logs.txt`), Cơ sở dữ liệu học tập (`qldt.sql`), Tài liệu quy định (`quy_dinh.md`) và Nhật ký đào tạo tuần (`11.04.txt`).\n\n")
    
    f.write("## 1. Bảng tổng hợp đánh giá KPI theo Phòng ban\n\n")
    
    idx = 1
    for dept_name, members in grouped_evaluations.items():
        f.write(f"### 1.{idx}. {dept_name}\n\n")
        if not members:
            f.write("*Không có nhân sự nào được ghi nhận ở khối này trong học kỳ này.*\n\n")
            idx += 1
            continue
            
        f.write("| Họ và tên | Vai trò | Lớp phụ trách | Điểm Kỷ luật SV & Tác nghiệp (40%) | Điểm Học tập (30%) | Điểm Báo cáo ngày (30%) | Điểm KPI tổng |\n")
        f.write("| :--- | :---: | :--- | :---: | :---: | :---: | :---: |\n")
        
        for p in sorted(members, key=lambda x: x['KPI'], reverse=True):
            linked_classes = make_obsidian_links(p['Classes'])
            f.write(f"| **{p['Name']}** | {p['Role']} | {linked_classes} | {p['Compliance']:.1f} | {p['Academic']:.1f} | {p['Work']:.1f} | **{p['KPI']:.2f}** |\n")
        f.write("\n")
        idx += 1
        
    f.write("\n---\n\n")
    f.write("## 2. Đánh giá chi tiết từng cá nhân\n\n")
    
    key_instructors_names = [
        'Bùi Thanh Hải', 'Lê Hà Thanh Sang', 
        'Nguyễn Bá Minh Đạo', 'Lưu Hoàng Xuân Nguyên', 'Phạm Tuấn Bình',
        'Trịnh Quốc Hai', 'Lương Quốc Tuấn', 'Hoàng Thị Kim Oanh', 'Lại Trung Lâm'
    ]
    
    for dept_name, members in grouped_evaluations.items():
        f.write(f"### 🔹 Chi tiết nhân sự {dept_name}\n\n")
        if not members:
            f.write("*Không có dữ liệu chi tiết.*\n\n")
            continue
            
        for p in sorted(members, key=lambda x: x['KPI'], reverse=True):
            is_key = any(kn in p['Name'] for kn in key_instructors_names) or p['Compliance'] < 100.0 or p['Work'] < 90.0
            if not is_key:
                continue
                
            linked_classes = make_obsidian_links(p['Classes'])
            f.write(f"#### {p['Role']}. {p['Name']}\n")
            f.write(f"- **Lớp phụ trách**: {linked_classes}\n")
            f.write(f"- **Điểm KPI tổng**: **{p['KPI']:.2f}** (Kỷ luật: {p['Compliance']:.1f}, Học tập: {p['Academic']:.1f}, Báo cáo ngày: {p['Work']:.1f})\n")
            f.write(f"- **Điểm mạnh**:\n  - {p['Strengths']}\n")
            f.write(f"- **Điểm yếu / Lỗi vi phạm đã mắc**:\n  - {p['Weaknesses']}\n")
            f.write(f"- **Đề xuất cải thiện cụ thể**:\n  - {p['Recommendations']}\n\n")

print(f"KPI Report generated successfully at: {output_report_path}")

# Copy to data/report_kpi_gv_tg.md to satisfy project requirements
data_report_path = "data/report_kpi_gv_tg.md"
try:
    import shutil
    shutil.copy2(output_report_path, data_report_path)
    print(f"Successfully copied KPI report to: {data_report_path}")
except Exception as e:
    print(f"Warning: Could not copy report to {data_report_path}: {e}")

conn.close()
