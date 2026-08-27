import os
import json
import re
import unicodedata
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

def normalize_name(name):
    if not name:
        return ""
    norm = name.strip().lower()
    return norm

def normalize_vietnamese_name(name):
    if not name:
        return ""
    name = " ".join(name.strip().split())
    name = name.lower()
    name = unicodedata.normalize('NFKD', name)
    name = "".join([c for c in name if not unicodedata.combining(c)])
    name = name.replace("đ", "d")
    return name

def load_staff_profiles(md_path):
    profiles = {}
    if not os.path.exists(md_path):
        print(f"Warning: File {md_path} not found.")
        return profiles
        
    current_group = "Khối CNTT"
    with open(md_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.startswith("## "):
                group_title = line[3:].strip()
                # Chuẩn hóa tên nhóm của Ngoại ngữ
                if "Ngoại ngữ" in group_title or "Ngoai ngu" in group_title:
                    current_group = "Khối Ngoại ngữ và Kỹ năng mềm"
                else:
                    current_group = group_title
            elif line.startswith("- "):
                parts = line[2:].split("|")
                if len(parts) >= 2:
                    name = parts[0].strip()
                    role = parts[1].strip()
                    rank_part = parts[2].strip() if len(parts) >= 3 else "Rank: N/A"
                    
                    rank_val = "3"  # Mặc định
                    rank_match = re.search(r'Rank:?\s*(\d+)', rank_part, re.IGNORECASE)
                    if rank_match:
                        rank_val = rank_match.group(1)
                    
                    norm = normalize_name(name)
                    profiles[norm] = {
                        "name": name,
                        "group": current_group,
                        "role": role,
                        "rank": rank_val
                    }
    return profiles

def load_kpi_masters():
    qtkd_master = {}
    cntt_master = {}
    
    # 1. QTKD KPI Master
    qtkd_path = r"C:\Users\DELL\Downloads\_Task Management_ QL Khối QTKD.xlsx"
    if os.path.exists(qtkd_path):
        try:
            wb = openpyxl.load_workbook(qtkd_path, data_only=True)
            if "KPI_MASTER" in wb.sheetnames:
                sheet = wb["KPI_MASTER"]
                for r in range(2, sheet.max_row + 1):
                    key = sheet.cell(row=r, column=6).value
                    std_time = sheet.cell(row=r, column=5).value
                    if key and std_time is not None:
                        qtkd_master[str(key).strip()] = float(std_time)
            wb.close()
        except Exception as e:
            print("Warning loading QTKD KPI Master:", e)
            
    # 2. CNTT KPI Master
    cntt_path = r"C:\Users\DELL\Downloads\Quản lý hiệu suất đào tạo.xlsx"
    if os.path.exists(cntt_path):
        try:
            wb = openpyxl.load_workbook(cntt_path, data_only=True)
            sheetname = "Cấu trúc KPI công việc GV. TG"
            if sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                for r in range(2, sheet.max_row + 1):
                    key = sheet.cell(row=r, column=7).value
                    std_time = sheet.cell(row=r, column=5).value
                    if key and std_time is not None:
                        cntt_master[str(key).strip()] = float(std_time)
            wb.close()
        except Exception as e:
            print("Warning loading CNTT KPI Master:", e)
            
    return qtkd_master, cntt_master

def get_task_unit(task_name):
    name_no_accents = normalize_vietnamese_name(task_name)
    
    if any(k in name_no_accents for k in ["giang day", "len lop", "day hoc", "day ly thuyet", "day slot", "day session", "day thuc hanh", "day cntt", "giang day session"]):
        return "Buổi học"
    if any(k in name_no_accents for k in ["soan slide", "soan bai", "soan giao an", "lam slide", "chuan bi slide", "chuan bi bai giang", "lam hoc lieu", "viet quiz", "soan de", "chuan bi giang day", "chuan bi bai", "chuan bi thuc hanh"]):
        return "Buổi học"
    if any(k in name_no_accents for k in ["cham bai", "cham thi", "cham btvn", "cham project", "cham prj", "cham van dap", "cham thuc hanh", "cham code", "cham bat tap", "cham bai tap"]):
        if "san pham" in name_no_accents or "project" in name_no_accents or "prj" in name_no_accents:
            return "Sản phẩm"
        if "van dap" in name_no_accents:
            return "Sinh viên"
        return "Bài"
    if "trong thi" in name_no_accents:
        return "Ca thi"
    if any(k in name_no_accents for k in ["video", "clip", "quay video", "lam video", "quay record", "record slot", "quay clip"]):
        return "Video"
    if any(k in name_no_accents for k in ["mindmap", "ban do tu duy"]):
        return "Mindmap"
    if any(k in name_no_accents for k in ["ho tro", "support", "fix bug", "huong dan sv", "huong dan hoc vien", "giai dap", "tu van", "mentor"]):
        return "Giờ"
    if any(k in name_no_accents for k in ["hop", "meeting", "giao ban", "sinh hoat", "pmo", "giao ban khoi"]):
        return "Giờ"
    if any(k in name_no_accents for k in ["nghi phep", "xin phep nghi", "phep nam", "off", "nghi le"]):
        return "Ngày"
    return "Lần"

def parse_kpi_master_keys(qtkd_master, cntt_master):
    old_kpi_list = []
    
    # 1. CNTT
    for key, std_time in cntt_master.items():
        parts = key.split('-')
        if len(parts) >= 3:
            rank = parts[-1].strip()
            role = parts[-2].strip()
            task_title = "-".join(parts[:-2]).strip()
            old_kpi_list.append({
                "group": "Khối CNTT",
                "task_title": task_title,
                "role": "Giảng viên" if "giảng" in role.lower() or "gv" in role.lower() or "giang" in role.lower() else "Trợ giảng",
                "rank": rank,
                "std_hours": std_time / 60.0
            })
            
    # 2. QTKD
    for key, std_time in qtkd_master.items():
        parts = key.split('-')
        if len(parts) >= 3:
            role = parts[0].strip()
            rank = parts[1].strip()
            task_title = parts[2].strip()
            old_kpi_list.append({
                "group": "Khối Quản trị Kinh doanh (QTKD)",
                "task_title": task_title,
                "role": "Giảng viên" if "giảng" in role.lower() or "gv" in role.lower() else "Trợ giảng",
                "rank": rank,
                "std_hours": std_time / 60.0
            })
            
    # Gom nhóm ma trận cũ theo (group, task_title, role)
    old_kpi_matrix = {}
    for old in old_kpi_list:
        m_key = (old["group"], old["task_title"], old["role"])
        if m_key not in old_kpi_matrix:
            old_kpi_matrix[m_key] = {str(i): 0.0 for i in range(1, 6)}
            
        r_str = old["rank"]
        if r_str in old_kpi_matrix[m_key]:
            old_kpi_matrix[m_key][r_str] = old["std_hours"]
            
    # Chuyển đổi sang phẳng
    old_kpi_matrix_flat = []
    for (g_group, g_title, g_role), ranks in old_kpi_matrix.items():
        valid_vals = [v for v in ranks.values() if v > 0]
        avg_hours = sum(valid_vals) / len(valid_vals) if valid_vals else 0.0
        
        old_kpi_matrix_flat.append({
            "group": g_group,
            "task_title": g_title,
            "role": g_role,
            "unit": get_task_unit(g_title),
            "r1": ranks["1"],
            "r2": ranks["2"],
            "r3": ranks["3"],
            "r4": ranks["4"],
            "r5": ranks["5"],
            "avg_hours": round(avg_hours, 2)
        })
        
    return old_kpi_matrix_flat

def match_kpi_standard_time(group, name, role, rank, task_title, kpi_master_qtkd, kpi_master_cntt):
    title_norm = task_title.strip().lower()
    is_qtkd = "QTKD" in group
    kpi_db = kpi_master_qtkd if is_qtkd else kpi_master_cntt
    
    role_norm = "giảng viên" if "giảng" in role.lower() or "gv" in role.lower() else "trợ giảng"
    try:
        rank_val = int(rank)
    except:
        rank_val = 3
        
    matched_task_type = None
    if any(k in title_norm for k in ["giảng dạy", "lên lớp", "day ly thuyet", "triển khai buổi học"]):
        matched_task_type = "Giảng dạy lý thuyết - Buổi học" if role_norm == "giảng viên" else "Triển khai buổi thực hành - Buổi"
    elif any(k in title_norm for k in ["chuẩn bị", "soạn slide", "soạn bài", "soạn giáo án"]):
        matched_task_type = "Chuẩn bị giảng dạy - Buổi học" if role_norm == "giảng viên" else "Chuẩn bị buổi thực hành - Buổi"
    elif "mindmap" in title_norm or "bản đồ tư duy" in title_norm:
        matched_task_type = "Làm mindmap bài học - Session"
    elif any(k in title_norm for k in ["support", "hỗ trợ", "fix bug", "sửa lỗi", "hướng dẫn"]):
        matched_task_type = "Báo cáo hỗ trợ SV.HV" if not is_qtkd else "Hỗ trợ học viên"
    elif any(k in title_norm for k in ["chấm bài", "chấm thi", "chấm thực hành"]):
        matched_task_type = "Chấm thi thực hành - Bài"
    elif "chấm sản phẩm" in title_norm or "chấm project" in title_norm or "chấm prj" in title_norm:
        matched_task_type = "Chấm thi sản phẩm - Sản phẩm"
    elif "vấn đáp" in title_norm or "chấm vấn đáp" in title_norm:
        matched_task_type = "Chấm thi vấn đáp - Sinh viên"
    elif "trông thi" in title_norm:
        matched_task_type = "Trông thi thực hành - Ca" if "thực hành" in title_norm else "Trông thi lý thuyết - Ca"
        
    if matched_task_type:
        for key, std_time in kpi_db.items():
            key_norm = key.lower()
            if is_qtkd:
                role_part = "giảng viên" if "giảng" in key_norm else "trợ giảng"
                rank_part = re.search(r'-(\d+)-', key_norm)
                rank_num = int(rank_part.group(1)) if rank_part else 3
                task_part = key_norm.split('-')[-1]
                if role_part == role_norm and rank_num == rank_val and (matched_task_type.lower() in key_norm or task_part in title_norm):
                    return std_time, matched_task_type, False
            else:
                role_part = "giảng viên" if "giảng" in key_norm else "trợ giảng"
                rank_part = re.search(r'-(\d+)$', key_norm)
                rank_num = int(rank_part.group(1)) if rank_part else 3
                task_part = key_norm.split('-')[0]
                if role_part == role_norm and rank_num == rank_val and (matched_task_type.lower() in key_norm or task_part in title_norm):
                    return std_time, matched_task_type, False

    for key, std_time in kpi_db.items():
        key_norm = key.lower()
        if role_norm in key_norm and f"-{rank_val}" in key_norm:
            task_name = key_norm.split('-')[0] if not is_qtkd else key_norm.split('-')[-1]
            if task_name in title_norm or title_norm in task_name:
                return std_time, key.split('-')[0] if not is_qtkd else key.split('-')[-1], False
                
    return None, "Đầu việc tự do/chưa định mức", True

def get_task_group(title):
    title_no_accents = normalize_vietnamese_name(title)
    
    if any(k in title_no_accents for k in ["nghi phep", "xin phep nghi", "phep nam", "off", "nghi le"]):
        return "Nghỉ phép"
    if any(k in title_no_accents for k in ["hop", "meeting", "giao ban", "sinh hoat", "pmo", "giao ban khoi"]):
        return "Họan giao ban & Sinh hoạt"
    if any(k in title_no_accents for k in ["giang day", "len lop", "day hoc", "day ly thuyet", "day slot", "day session", "day thuc hanh", "day cntt", "giang day session"]):
        return "Giảng dạy & Lên lớp"
    if any(k in title_no_accents for k in ["soan slide", "soan bai", "soan giao an", "lam slide", "chuan bi slide", "chuan bi bai giang", "lam hoc lieu", "viet quiz", "soan de", "chuan bi giang day", "chuan bi bai", "chuan bi thuc hanh"]):
        return "Soạn bài & Chuẩn bị giảng dạy"
    if any(k in title_no_accents for k in ["cham bai", "cham thi", "cham btvn", "cham project", "cham prj", "cham van dap", "cham thuc hanh", "cham code", "cham bat tap", "cham bai tap"]):
        return "Chấm bài & Chấm thi"
    if any(k in title_no_accents for k in ["video", "clip", "quay video", "lam video", "quay record", "record slot", "quay clip"]):
        return "Sản xuất Video học liệu"
    if any(k in title_no_accents for k in ["ho tro", "support", "fix bug", "huong dan sv", "huong dan hoc vien", "giai dap", "tu van", "mentor"]):
        return "Hỗ trợ học viên & Giải đáp"
    if any(k in title_no_accents for k in ["mindmap", "ban do tu duy"]):
        return "Làm mindmap bài học"
        
    return "Đầu việc tự do/Khác"

def clean_title(title):
    title = re.sub(r'\b\d{1,2}[/-]\d{1,2}\b', '', title)
    title = re.sub(r'\b[A-Za-z\d]+-[A-Za-z\d]+-[A-Za-z\d]+\b', '', title)
    title = re.sub(r'\b[A-Za-z\d]+-K\d{2}-\w+\b', '', title)
    title = re.sub(r'\bK\d{2}-\w+\b', '', title)
    title = re.sub(r'\bKS\d{2}-\w+\b', '', title)
    title = re.sub(r'\b(HN|HCM|K24|K25|KS24|KS25|CNTT|QTKD)\w*\b', '', title, flags=re.IGNORECASE)
    title = re.sub(r'[\-\[\]\(\)\:\+\,]', ' ', title)
    title = re.sub(r'\s+', ' ', title).strip()
    return title

def main():
    print("=" * 80)
    print("BẮT ĐẦU CHẠY PHÂN TÍCH KPI OPPORTUNITIES & ĐỀ XUẤT KPI MASTER")
    print("=" * 80)
    
    # 1. Load profiles & old KPI Masters
    staff_profiles = load_staff_profiles("data/inputs/staff_roles_ranks.md")
    kpi_master_qtkd, kpi_master_cntt = load_kpi_masters()
    
    # Parse KPI Master cũ thành danh sách phẳng ma trận
    old_kpi_matrix_flat = parse_kpi_master_keys(kpi_master_qtkd, kpi_master_cntt)
    
    # 2. Đọc raw logs từ daily_log_analysis.json
    daily_log_path = "data/processed/daily_log_analysis.json"
    if not os.path.exists(daily_log_path):
        print(f"Error: {daily_log_path} not found.")
        sys.exit(1)
        
    with open(daily_log_path, "r", encoding="utf-8") as f:
        log_data = json.load(f)
        
    raw_reports = log_data.get("raw_reports", {})
    
    task_records = []  # Thống kê Task thực tế
    new_kpi_matrix = {}  # Thống kê Đầu việc phát sinh mới dạng Ma trận
    detailed_free_tasks = {} # Chi tiết các đầu việc tự do cụ thể theo Rank
    
    for group, members in raw_reports.items():
        # Chuẩn hóa tên nhóm ngay từ đầu
        user_group = group
        if "Ngoại ngữ" in group or "Ngoai ngu" in group:
            user_group = "Khối Ngoại ngữ và Kỹ năng mềm"
        elif "QTKD" in group:
            user_group = "Khối Quản trị Kinh doanh (QTKD)"
        elif "CNTT" in group:
            user_group = "Khối CNTT"
        elif "QLCL" in group:
            user_group = "Khối QLCLĐT"
            
        for name, m_data in members.items():
            norm_name = normalize_name(name)
            profile = staff_profiles.get(norm_name, {
                "name": name,
                "group": user_group,
                "role": "Giảng viên",
                "rank": "3"
            })
            
            role = profile["role"]
            rank = profile["rank"]
            u_group = profile["group"]
            if "Ngoại ngữ" in u_group or "Ngoai ngu" in u_group:
                u_group = "Khối Ngoại ngữ và Kỹ năng mềm"
            
            reports = m_data.get("reports", {})
            for date, r in reports.items():
                if not r:
                    continue
                tasks = r.get("tasks", [])
                for t in tasks:
                    t_title = t.get("title", "")
                    if not t_title:
                        continue
                    t_hours = float(t.get("hours", 0.0))
                    
                    # Đối chiếu định mức cũ
                    std_time, matched_cat, is_wildcard = match_kpi_standard_time(
                        u_group, name, role, rank, t_title, kpi_master_qtkd, kpi_master_cntt
                    )
                    
                    status = "Đúng định mức"
                    std_hours = 0.0
                    if is_wildcard:
                        status = "Chưa có định mức"
                    else:
                        std_hours = std_time / 60.0
                        if t_hours > std_hours * 1.5:
                            status = "Vượt định mức"
                            
                    # Thêm vào danh sách Task thực tế
                    task_records.append({
                        "date": date,
                        "group": u_group,
                        "task_title": t_title,
                        "cleaned_title": clean_title(t_title),
                        "member": name,
                        "role": role,
                        "rank": rank,
                        "hours": t_hours,
                        "std_hours": std_hours,
                        "status": status
                    })
                    
                    # Gom nhóm ma trận cho việc phát sinh mới
                    if is_wildcard:
                        task_group_name = get_task_group(t_title)
                        m_key = (u_group, task_group_name)
                        if m_key not in new_kpi_matrix:
                            new_kpi_matrix[m_key] = {
                                "ranks": {str(i): [] for i in range(1, 6)},
                                "members": set(),
                                "total_count": 0
                            }
                        
                        r_str = str(rank)
                        if r_str in new_kpi_matrix[m_key]["ranks"]:
                            new_kpi_matrix[m_key]["ranks"][r_str].append(t_hours)
                        new_kpi_matrix[m_key]["members"].add(name)
                        new_kpi_matrix[m_key]["total_count"] += 1
                        
                        # Gom nhóm chi tiết việc tự do theo (group, cleaned_title, rank)
                        c_title = clean_title(t_title)
                        if not c_title:
                            c_title = t_title
                        det_key = (u_group, c_title, r_str)
                        if det_key not in detailed_free_tasks:
                            detailed_free_tasks[det_key] = {
                                "hours_list": [],
                                "members": set()
                             }
                        detailed_free_tasks[det_key]["hours_list"].append(t_hours)
                        detailed_free_tasks[det_key]["members"].add(name)
                        
    # Chuyển đổi new_kpi_matrix sang phẳng
    new_kpi_matrix_flat = []
    for (g_group, g_task), data in new_kpi_matrix.items():
        ranks_hours = {}
        valid_avgs = []
        for r in range(1, 6):
            r_str = str(r)
            hours_list = data["ranks"][r_str]
            if hours_list:
                avg = sum(hours_list) / len(hours_list)
                ranks_hours[f"r{r}"] = round(avg, 2)
                valid_avgs.append(avg)
            else:
                ranks_hours[f"r{r}"] = 0.0
                
        avg_hours = sum(valid_avgs) / len(valid_avgs) if valid_avgs else 0.0
        proposed_hours = round(avg_hours * 4) / 4
        proposed_hours = max(0.25, min(8.0, proposed_hours))
        
        new_kpi_matrix_flat.append({
            "group": g_group,
            "task_group": g_task,
            "unit": get_task_unit(g_task),
            "r1": ranks_hours["r1"] if ranks_hours["r1"] > 0 else proposed_hours,
            "r2": ranks_hours["r2"] if ranks_hours["r2"] > 0 else proposed_hours,
            "r3": ranks_hours["r3"] if ranks_hours["r3"] > 0 else proposed_hours,
            "r4": ranks_hours["r4"] if ranks_hours["r4"] > 0 else proposed_hours,
            "r5": ranks_hours["r5"] if ranks_hours["r5"] > 0 else proposed_hours,
            "avg_hours": round(avg_hours, 2),
            "count": data["total_count"],
            "members": ", ".join(list(data["members"])[:3])
        })
        
    # Phẳng hóa detailed_free_tasks
    detailed_free_tasks_flat = []
    for (d_group, d_title, d_rank), data in detailed_free_tasks.items():
        hours_list = data["hours_list"]
        avg = sum(hours_list) / len(hours_list)
        detailed_free_tasks_flat.append({
            "group": d_group,
            "task_title": d_title,
            "rank": d_rank,
            "unit": get_task_unit(d_title),
            "avg_hours": round(avg, 2),
            "count": len(hours_list),
            "members": ", ".join(list(data["members"])[:3])
        })
        
    # Sắp xếp kết quả
    task_records = sorted(task_records, key=lambda x: x["date"], reverse=True)
    new_kpi_matrix_flat = sorted(new_kpi_matrix_flat, key=lambda x: (x["group"], x["task_group"]))
    old_kpi_matrix_flat = sorted(old_kpi_matrix_flat, key=lambda x: (x["group"], x["task_title"]))
    detailed_free_tasks_flat = sorted(detailed_free_tasks_flat, key=lambda x: (x["group"], x["task_title"], x["rank"]))
    
    # 3. Xuất file JSON
    output_json = {
        "old_kpi_master": old_kpi_matrix_flat,
        "new_free_tasks": new_kpi_matrix_flat,
        "detailed_free_tasks": detailed_free_tasks_flat,
        "task_records": task_records
    }
    
    os.makedirs("data/processed", exist_ok=True)
    with open("data/processed/kpi_opportunities.json", "w", encoding="utf-8") as f:
        json.dump(output_json, f, indent=2, ensure_ascii=False)
    print("Đã tạo file data/processed/kpi_opportunities.json")
    
    # 4. Xuất file Excel đề xuất ma trận (proposed_kpi_master.xlsx)
    excel_path = "output/reports/proposed_kpi_master.xlsx"
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Xóa sheet mặc định
    
    groups_list = ["Khối CNTT", "Khối Quản trị Kinh doanh (QTKD)", "Khối Ngoại ngữ và Kỹ năng mềm"]
    
    for grp in groups_list:
        sheet_name = grp.replace("Khối ", "").replace(" và Kỹ năng mềm", "")
        ws = wb.create_sheet(title=sheet_name)
        
        # Phần A: Danh mục KPI Master cũ
        ws.append([f"PHẦN A: DANH MỤC KPI MASTER CŨ ({grp.upper()})"])
        ws.append(["Đầu công việc chuẩn", "Đơn vị tính", "Vai trò", "Định mức R1 (h)", "Định mức R2 (h)", "Định mức R3 (h)", "Định mức R4 (h)", "Định mức R5 (h)", "Trung bình (h)"])
        
        has_old = False
        for old in old_kpi_matrix_flat:
            if old["group"] == grp or (grp == "Khối Quản trị Kinh doanh (QTKD)" and old["group"].startswith("Khối QTKD")):
                ws.append([
                    old["task_title"],
                    old["unit"],
                    old["role"],
                    old["r1"],
                    old["r2"],
                    old["r3"],
                    old["r4"],
                    old["r5"],
                    old["avg_hours"]
                ])
                has_old = True
        
        if not has_old:
            ws.append(["Không có dữ liệu định mức cũ cho khối này"])
            
        # Thêm 2 dòng trống
        ws.append([])
        ws.append([])
        
        # Phần B: Đầu việc phát sinh thực tế đề xuất
        ws.append([f"PHẦN B: ĐẦU VIỆC PHÁT SINH MỚI ĐỀ XUẤT"])
        ws.append(["Đầu công việc phát sinh", "Đơn vị tính", "Đề xuất R1 (h)", "Đề xuất R2 (h)", "Đề xuất R3 (h)", "Đề xuất R4 (h)", "Đề xuất R5 (h)", "Đề xuất Trung bình (h)", "Tần suất", "Nhân sự khai báo đại diện"])
        
        has_new = False
        for new in new_kpi_matrix_flat:
            if new["group"] == grp or (grp == "Khối Quản trị Kinh doanh (QTKD)" and new["group"].startswith("Khối QTKD")):
                ws.append([
                    new["task_group"],
                    new["unit"],
                    new["r1"],
                    new["r2"],
                    new["r3"],
                    new["r4"],
                    new["r5"],
                    new["avg_hours"],
                    new["count"],
                    new["members"]
                ])
                has_new = True
                
        if not has_new:
            ws.append(["Không có công việc phát sinh mới"])
            
        # Thêm 2 dòng trống
        ws.append([])
        ws.append([])
        
        # Phần C: Chi tiết đầu việc tự do thực tế theo Rank
        ws.append([f"PHẦN C: CHI TIẾT ĐẦU VIỆC TỰ DO THỰC TẾ THEO RANK"])
        ws.append(["Đầu công việc tự do cụ thể", "Đơn vị tính", "Rank", "Giờ thực tế trung bình (h)", "Số phút trung bình", "Tần suất", "Nhân sự khai báo đại diện"])
        
        has_det = False
        for det in detailed_free_tasks_flat:
            if det["group"] == grp or (grp == "Khối Quản trị Kinh doanh (QTKD)" and det["group"].startswith("Khối QTKD")):
                ws.append([
                    det["task_title"],
                    det["unit"],
                    det["rank"],
                    det["avg_hours"],
                    int(det["avg_hours"] * 60),
                    det["count"],
                    det["members"]
                ])
                has_det = True
                
        if not has_det:
            ws.append(["Không có công việc tự do cụ thể"])
            
        # Tự động chỉnh độ rộng cột
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(excel_path)
    print(f"Đã tạo file Excel đề xuất cấu hình ma trận tại: {excel_path}")
    print("=" * 80)
    print("HOÀN THÀNH CHẠY SCRIPT ANALYZE_KPI_OPPORTUNITIES THÀNH CÔNG")
    print("=" * 80)

if __name__ == "__main__":
    main()
