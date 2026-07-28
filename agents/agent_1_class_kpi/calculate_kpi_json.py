import os
import sys
import json
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

def main():
    print("Agent 1: Bắt đầu phân tích kỷ luật sinh viên từ PTIT_Chiso.xlsx...")
    
    excel_path = "data/inputs/PTIT_Chiso.xlsx"
    output_json_path = "data/processed/agent1_output.json"
    
    if not os.path.exists(excel_path):
        print(f"Error: Không tìm thấy {excel_path}")
        sys.exit(1)
        
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    target_sheets = ['KS25_Python_Web', 'KS25_QTKD_PRJ302']
    
    instructors_data = {}
    
    for sheet in target_sheets:
        if sheet not in wb.sheetnames:
            continue
        sheet_obj = wb[sheet]
        
        header_row_idx = None
        for r in range(1, min(20, sheet_obj.max_row + 1)):
            row_vals = [str(sheet_obj.cell(row=r, column=c).value or "").strip() for c in range(1, sheet_obj.max_column + 1)]
            if 'Lớp' in row_vals and any('Giảng viên' in val or 'Trợ giảng' in val or 'Giảng viên/Trợ giảng' in val for val in row_vals):
                header_row_idx = r
                break
                
        if header_row_idx is None:
            continue
            
        headers = [str(sheet_obj.cell(row=header_row_idx, column=c).value or "").strip() for c in range(1, sheet_obj.max_column + 1)]
        class_col_idx = None
        person_col_idx = None
        for c_idx, h in enumerate(headers):
            if 'Lớp' in h:
                class_col_idx = c_idx + 1
            elif 'Giảng viên/Trợ giảng' in h or 'Giảng viên' in h or 'Trợ giảng' in h:
                person_col_idx = c_idx + 1
                
        if class_col_idx is None or person_col_idx is None:
            continue
            
        current_class = None
        for r in range(header_row_idx + 1, sheet_obj.max_row + 1):
            c_val = sheet_obj.cell(row=r, column=class_col_idx).value
            p_val = sheet_obj.cell(row=r, column=person_col_idx).value
            
            if c_val is not None and str(c_val).strip() != "":
                current_class = str(c_val).strip()
                
            if p_val is not None and str(p_val).strip() not in ['', 'nan', 'Giảng viên/Trợ giảng']:
                name = str(p_val).strip()
                
                numeric_vals = []
                for c in range(4, sheet_obj.max_column + 1):
                    val = sheet_obj.cell(row=r, column=c).value
                    if isinstance(val, (int, float)) and val is not None:
                        numeric_vals.append(float(val))
                        
                if not numeric_vals:
                    continue
                    
                avg_violation = sum(numeric_vals) / len(numeric_vals)
                is_tg = (c_val is None or str(c_val).strip() == "")
                role = 'TG' if is_tg else 'GV'
                
                if name not in instructors_data:
                    instructors_data[name] = {
                        'Role': role,
                        'Classes': set(),
                        'ViolationRates': []
                    }
                
                instructors_data[name]['Classes'].add(f"{current_class} ({sheet})")
                instructors_data[name]['ViolationRates'].append(avg_violation)

    wb.close()
    
    results = {}
    for name, data in sorted(instructors_data.items()):
        avg_violation = sum(data['ViolationRates']) / len(data['ViolationRates']) if data['ViolationRates'] else 0.0
        student_discipline = 100.0 - avg_violation
        student_discipline = max(0.0, min(100.0, student_discipline))
        
        results[name] = {
            "name": name,
            "role": data['Role'],
            "classes": list(data['Classes']),
            "avg_violation_rate": round(avg_violation, 2),
            "student_discipline_score": round(student_discipline, 1)
        }
        
    os.makedirs("data", exist_ok=True)
    with open(output_json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=4)
    print(f"Agent 1: Đã lưu dữ liệu JSON tại {output_json_path}")

if __name__ == "__main__":
    main()
